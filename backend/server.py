from damage_segmentation import segment_damage, preload_sam
from ai_learning import (
    get_few_shot_examples, build_few_shot_prompt_parts_multimodal,
    get_pattern_lessons, get_part_lesson, save_feedback as _save_ai_feedback,
)
from fastapi import FastAPI, APIRouter, UploadFile, File, Form, HTTPException, Depends, Body, Request
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from starlette.middleware.cors import CORSMiddleware

from dotenv import load_dotenv
from motor.motor_asyncio import AsyncIOMotorClient

from PIL import Image

from pydantic import BaseModel, Field, ConfigDict

from typing import List, Optional, Tuple

from pathlib import Path

from datetime import datetime, timezone, timedelta

from concurrent.futures import ThreadPoolExecutor

from jose import jwt, JWTError
from pymongo.errors import DuplicateKeyError
import bcrypt
import re
import time

import asyncio
import os
import io
import json
import uuid
import base64
import logging

# =========================
# PATHS + ENV
# =========================

ROOT_DIR = Path(__file__).parent
UPLOAD_DIR = ROOT_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

load_dotenv(ROOT_DIR / ".env")

_executor = ThreadPoolExecutor(max_workers=4)
_gemini_sem = asyncio.Semaphore(2)  # máximo 2 llamadas Gemini simultáneas

# ── Gestión de cuota DIARIA de Gemini (crítico en nivel gratuito) ──
# El free tier limita por PETICIONES/DÍA y MODELO (p. ej. 20/día en 2.5-flash).
# Reintentar contra un modelo con el día agotado quema el cupo de los demás.
# Cuando un 429 dice "PerDay", el modelo se veta 4h y se salta directamente.
_gemini_daily_exhausted: dict = {}  # modelo → epoch en que se agotó su día
_GEMINI_EXHAUST_TTL = 4 * 3600


def _mark_daily_exhausted(model: str):
    _gemini_daily_exhausted[model] = time.time()
    logger.warning(f"Gemini cuota DIARIA agotada en {model}: vetado {_GEMINI_EXHAUST_TTL // 3600}h")


def _is_daily_exhausted(model: str) -> bool:
    return time.time() - _gemini_daily_exhausted.get(model, 0) < _GEMINI_EXHAUST_TTL


def _is_perday_429(err_str: str) -> bool:
    """¿El 429 es por cupo DIARIO (no por ráfaga por minuto)?"""
    e = err_str.lower()
    return "perday" in e or "per day" in e or "daily" in e

PUBLIC_BASE_URL = os.environ.get("PUBLIC_BASE_URL", "").rstrip("/")

# AI Detection microservice (YOLO11 + SAM2)
# Set AI_SERVICE_URL=http://ai-service:8001 when GPU microservice is running.
# If not set, /annotated falls back to location_hint mapping (deterministic, free).
AI_SERVICE_URL = os.environ.get("AI_SERVICE_URL", "").rstrip("/")

# Cloudflare R2
R2_ENDPOINT    = os.environ.get("R2_ENDPOINT", "")       # https://<account>.r2.cloudflarestorage.com
R2_BUCKET      = os.environ.get("R2_BUCKET", "flotadsp-uploads")
R2_ACCESS_KEY  = os.environ.get("R2_ACCESS_KEY", "")
R2_SECRET_KEY  = os.environ.get("R2_SECRET_KEY", "")
R2_PUBLIC_URL  = os.environ.get("R2_PUBLIC_URL", "").rstrip("/")  # https://cdn.tu-dominio.com

_r2_client = None

def get_r2():
    global _r2_client
    if _r2_client is not None:
        return _r2_client
    if not (R2_ENDPOINT and R2_ACCESS_KEY and R2_SECRET_KEY):
        return None
    try:
        import boto3
        _r2_client = boto3.client(
            "s3",
            endpoint_url=R2_ENDPOINT,
            aws_access_key_id=R2_ACCESS_KEY,
            aws_secret_access_key=R2_SECRET_KEY,
            region_name="auto",
        )
        return _r2_client
    except Exception as e:
        logger.warning(f"R2 no disponible: {e}")
        return None

# JWT config
SECRET_KEY = os.environ.get("SECRET_KEY", "")
if not SECRET_KEY:
    raise RuntimeError("SECRET_KEY no configurada. Usa: fly secrets set SECRET_KEY=$(openssl rand -hex 32)")

JWT_ALGORITHM = "HS256"
JWT_EXPIRE_HOURS = int(os.environ.get("JWT_EXPIRE_HOURS", "72"))

# =========================
# DATABASE
# =========================

mongo_url = os.environ.get("MONGO_URL", "mongodb://localhost:27017")
client = AsyncIOMotorClient(mongo_url)

# === MULTI-TENANT (una BD por organización) ===
# `db` es un proxy: resuelve la base de datos de la ORGANIZACIÓN del usuario logueado
# (fijada por petición desde el token). Sin contexto → BD por defecto = comportamiento
# idéntico al de siempre (no rompe nada). Login/usuarios/organizaciones viven en `global_db`.
import contextvars
_DEFAULT_DB_NAME = os.environ.get("DB_NAME", "flotadsp")
_GLOBAL_DB_NAME = os.environ.get("GLOBAL_DB_NAME", "flotadsp_global")
_current_db_name = contextvars.ContextVar("current_db_name", default=_DEFAULT_DB_NAME)

global_db = client[_GLOBAL_DB_NAME]  # organizaciones, usuarios, suscripciones (compartida)


def _tenant_db_name(org):
    """Nombre de BD de una organización. La tuya (owner) reusa la BD existente."""
    if not org:
        return _DEFAULT_DB_NAME
    return org.get("db_name") or (_DEFAULT_DB_NAME if org.get("account_type") == "owner"
                                  else f"dsp_{org.get('id')}")


def set_current_org_db(db_name):
    _current_db_name.set(db_name or _DEFAULT_DB_NAME)


class _TenantDBProxy:
    """Reenvía todo (db.vehicles, db.command, …) a la BD del tenant actual."""
    def __getattr__(self, name):
        return getattr(client[_current_db_name.get()], name)

    def __getitem__(self, name):
        return client[_current_db_name.get()][name]


db = _TenantDBProxy()

# =========================
# FASTAPI
# =========================

app = FastAPI(title="FlotaDSP API", version="5.3.4")
api_router = APIRouter(prefix="/api")

# StaticFiles DESPUÉS del middleware — solo como fallback local
# (si R2 está configurado, las imágenes van a R2 directamente)

# =========================
# LOGGING
# =========================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

# =========================
# MODELS
# =========================

class Vehicle(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    license_plate: str
    brand: str = ""
    model: str = ""
    color: str = ""
    year: Optional[int] = None
    vin: Optional[str] = None
    status: str = "active"
    center: Optional[str] = None
    current_driver_id: Optional[str] = None
    mileage: Optional[int] = None
    provider: Optional[str] = None
    vehicle_type: Optional[str] = None
    fuel_type: Optional[str] = None
    workshop_status: Optional[str] = None
    workshop_reason: Optional[str] = None
    documents: List[str] = []
    # --- Bolsas ---
    bags_remaining: int = 0
    bags_history: List[dict] = []  # [{date, change, note, remaining_after}]
    mileage_history: List[dict] = []  # [{date, km, source}]
    itv_date: Optional[str] = None  # ISO YYYY-MM-DD, caducidad ITV
    renting_end_date: Optional[str] = None  # ISO, vencimiento contrato renting
    renting_baja_date: Optional[str] = None  # ISO, fecha baja renting
    # --- Aceite ---
    oil_last_change_km: Optional[int] = None
    oil_last_change_date: Optional[str] = None
    oil_interval_km: int = 15000
    oil_warning_before_km: int = 2500
    # --- Gemelo digital 3D ---
    # Modelo identificado por IA desde las fotos: {brand, model, body_type, color,
    # confidence, identified_at}. Alimenta el VehicleModelResolver.
    ai_model: Optional[dict] = None
    # ADN visual acumulado del vehículo (accesorios, pegatinas, color exacto…).
    dna: Optional[dict] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class VehicleCreate(BaseModel):
    model_config = ConfigDict(extra="ignore")
    license_plate: str
    brand: str = ""
    model: str = ""
    color: str = ""
    year: Optional[int] = None
    vin: Optional[str] = None
    center: Optional[str] = None
    current_driver_id: Optional[str] = None
    mileage: Optional[int] = None
    provider: Optional[str] = None
    vehicle_type: Optional[str] = None
    fuel_type: Optional[str] = None
    itv_date: Optional[str] = None            # ISO YYYY-MM-DD, caducidad ITV
    renting_end_date: Optional[str] = None    # ISO, vencimiento contrato renting
    workshop_status: Optional[str] = None
    workshop_reason: Optional[str] = None
    documents: List[str] = []


class Driver(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    dni: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    license_number: Optional[str] = None
    login: Optional[str] = None        # login del conductor
    driver_id: Optional[str] = None    # ID de Amazon
    photo_url: Optional[str] = None    # foto de perfil (R2)
    center: Optional[str] = None
    active: bool = True
    contrato: Optional[str] = None     # empresa | ett (cuadrante de turnos)
    nivel: Optional[str] = None        # pleno | L1 | L2 | L3 (novatos)
    zona: Optional[str] = None         # zona habitual (opcional)
    alojamiento: Optional[str] = None  # alojamiento del conductor
    notas: Optional[str] = None        # observaciones libres
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class DriverCreate(BaseModel):
    name: str
    dni: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    license_number: Optional[str] = None
    center: Optional[str] = None
    password: Optional[str] = None
    driver_id: Optional[str] = None    # ID de Amazon
    contrato: Optional[str] = None
    nivel: Optional[str] = None
    zona: Optional[str] = None
    alojamiento: Optional[str] = None
    notas: Optional[str] = None


class Damage(BaseModel):
    model_config = ConfigDict(extra="ignore")
    part: str
    severity: str
    description: str
    repair_suggestion: str = ""
    estimated_cost: float = 0.0
    confidence: float = 0.0
    location_hint: str = ""
    photo_index: Optional[int] = None      # imagen (1-based) donde mejor se ve el daño
    box_2d: Optional[List[int]] = None     # [ymin,xmin,ymax,xmax] normalizado 0-1000
    polygon_points: Optional[List[List[int]]] = None  # [[y,x], ...] normalizado 0-1000
    # Pieza determinada por GEOMETRÍA (modelo de paneles ∩ caja del daño),
    # independiente del "part" textual del LLM. Fiable para ledger y revisión.
    panel_cv: Optional[str] = None
    is_new: bool = True
    confirmed: bool = True          # False = sugerido (confidence < 0.65), no cuenta en scoring
    # --- v5.1: gestión de reparación ---
    actual_cost: Optional[float] = None       # coste real introducido por admin tras la reparación
    workshop_id: Optional[str] = None         # taller asignado para reparar este daño
    repair_status: str = "pending"            # pending | assigned | in_repair | done | declined
    repair_notes: str = ""                    # notas libres sobre la reparación
    assigned_at: Optional[str] = None         # ISO timestamp cuando se asignó taller
    completed_at: Optional[str] = None        # ISO timestamp cuando se marcó como reparado


# Approximate zone → [ymin, xmin, ymax, xmax] normalized 0-1000.
# Used as deterministic fallback when GPU microservice is unavailable.
_LOCATION_BOX: dict = {
    "frontal":           [50,  200, 480, 800],
    "trasera":           [520, 200, 950, 800],
    "lateral_izquierdo": [150,  30, 850, 420],
    "lateral_derecho":   [150, 580, 850, 970],
    "techo":             [50,  180, 380, 820],
    "otra":              [280, 280, 720, 720],
}


# =========================
# TALLERES (v5.1)
# =========================

# Teléfonos de asistencia en carretera por proveedor (confirmados, no inventados)
PROVIDER_ROADSIDE: dict = {
    # Santander Renting — Línea Conductor verificada en documentación oficial
    "BANSACAR":           {"phone": "917 098 569", "label": "Línea Conductor Santander Renting", "app": None},
    "SANTANDER RENTING":  {"phone": "917 098 569", "label": "Línea Conductor Santander Renting", "app": None},
    # Ayvens (ex-ALD/LeasePlan) — número confirmado en notas del seed
    "AYVENS":             {"phone": "913 336 717", "label": "Asistencia Ayvens",    "app": "My Ayvens"},
    "ALD":                {"phone": "913 336 717", "label": "Asistencia Ayvens",    "app": "My Ayvens"},
    "LEASE PLAN":         {"phone": "913 336 717", "label": "Asistencia Ayvens",    "app": "My Ayvens"},
    "LEASEPLAN":          {"phone": "913 336 717", "label": "Asistencia Ayvens",    "app": "My Ayvens"},
    # Sabadell Renting / VayVans — comparten red Ayvens Premier
    "VAYVANS":            {"phone": "932 437 080", "label": "Asistencia Sabadell Renting / VayVans", "app": "My Ayvens"},
    "SABADELL RENTING":   {"phone": "932 437 080", "label": "Asistencia Sabadell Renting",          "app": None},
    # Kinto (Toyota Renting) — pendiente confirmar número oficial
    "KINTO":              {"phone": None, "label": "Kinto One — contacta con tu gestor", "app": "Kinto Share"},
    "KINTO ONE":          {"phone": None, "label": "Kinto One — contacta con tu gestor", "app": "Kinto Share"},
    # One Furgo — alquiler sin asistencia propia, usar seguro del vehículo
    "ONE FURGO":          {"phone": None, "label": "One Furgo — usa el seguro incluido", "app": None},
}


def _provider_roadside(provider: str) -> Optional[dict]:
    """Devuelve info de asistencia en carretera para un proveedor (coincidencia parcial)."""
    if not provider:
        return None
    pup = provider.upper()
    for key, data in PROVIDER_ROADSIDE.items():
        if key in pup or pup in key:
            return data
    return None



class Workshop(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    address: str = ""
    city: str = ""
    center: Optional[str] = None       # OGA5 | DGA1 | DGA2 — centro logístico más cercano
    phone: str = ""
    email: str = ""
    notes: str = ""
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    # Categorías de trabajo: chapa, mecanica, lunas, neumaticos, oficial_toyota
    categories: List[str] = []
    # Lista de proveedores con los que tiene convenio (nombres tal y como aparecen en Vehicle.provider).
    # Si es taller "universal" (Carglass por seguro, p.ej.), poner ["*"] o dejar vacío.
    convenios: List[str] = []
    rating: Optional[float] = None     # 0-5 (de Google o manual)
    rating_count: Optional[int] = None # nº de reseñas Google
    hours: str = ""                    # horario legible, p.ej. "L-V 8-16h"
    maps_url: str = ""                 # enlace a Google Maps
    active: bool = True
    is_official: bool = False          # true para concesionarios oficiales (Toyota, etc.)
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class WorkshopCreate(BaseModel):
    model_config = ConfigDict(extra="ignore")
    name: str
    address: str = ""
    city: str = ""
    center: Optional[str] = None
    phone: str = ""
    email: str = ""
    notes: str = ""
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    categories: List[str] = []
    convenios: List[str] = []
    rating: Optional[float] = None
    rating_count: Optional[int] = None
    hours: str = ""
    maps_url: str = ""
    is_official: bool = False


class AIDetection(BaseModel):
    model_config = ConfigDict(extra="ignore")
    label: str
    severity: str = "leve"
    box_2d: List[float]  # [ymin, xmin, ymax, xmax] normalized 0-1000
    confidence: float = 0.7
    source: str = "location_hint"  # "yolo" | "sam2" | "location_hint"
    # Pieza asignada por el modelo de paneles del ai-service (geometría, no LLM)
    panel: Optional[str] = None
    panel_conf: Optional[float] = None
    # Contorno real de segmentación ([y,x] 0-1000, ≤24 puntos) del modelo -seg
    polygon_2d: Optional[List[List[float]]] = None


class InspectionAIResult(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    inspection_id: str
    photo_index: int = 0
    detections: List[AIDetection] = []
    source: str = "location_hint"  # "yolo" | "location_hint" | "gemini_legacy"
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class InspectionAnalysis(BaseModel):
    model_config = ConfigDict(extra="ignore")
    critical_damages_count: int = 0
    total_damages_count: int = 0
    new_damages_count: int = 0
    severity: str = "sin_danos"
    dirt_level: Optional[float] = None  # 0 impecable — 10 cubierta de barro
    urgency: str = "puede_esperar"
    risk: str = "bajo"
    circulation_safe: bool = True
    detected_plate: str = ""
    fraud_warnings: List[str] = []
    hidden_damage_probability: float = 0.0
    total_estimated_cost: float = 0.0
    confidence: float = 0.0
    executive_summary: str = ""
    image_quality_warnings: List[str] = []
    affected_parts: List[str] = []
    critical_damages: list = []
    new_damages: List[Damage] = []
    damages: List[Damage] = []


class Inspection(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    vehicle_id: str
    driver_id: Optional[str] = None
    photos: List[str] = []
    annotated_photos: List[Optional[str]] = []   # versiones anotadas con marcadores de daños
    reference_photos: List[str] = []
    is_reference: bool = False
    analysis: Optional[InspectionAnalysis] = None
    analysis_status: str = "ok"           # ok | gemini_failed | gemini_timeout
    analysis_error: Optional[str] = None
    notes: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    analyzed_at: Optional[datetime] = None


class Alert(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    vehicle_id: str
    inspection_id: str
    title: str
    description: str
    severity: str
    read: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class LoginRequest(BaseModel):
    username: str
    password: str


class RegisterRequest(BaseModel):
    org_name: str          # nombre del DSP/empresa
    username: str          # usuario de acceso del dueño
    password: str
    email: Optional[str] = None
    slug: Optional[str] = None   # identificador en la URL (flotadsp.com/<slug>)
    center: Optional[str] = None  # código de su primer centro/estación (ej. su station)
    plan: Optional[str] = None   # plan elegido: basico | pro | flota | enterprise


# Límites y features por plan
PLAN_LIMITS = {
    "basico":     {"max_vehicles": 20,  "max_drivers": 20,  "max_centers": 1,  "ai": False, "scorecard": False, "chat": False, "forensics": False, "maintenance": True,  "assignments": False, "export": False},
    "pro":        {"max_vehicles": 75,  "max_drivers": -1,  "max_centers": 3,  "ai": True,  "scorecard": True,  "chat": True,  "forensics": False, "maintenance": True,  "assignments": True,  "export": False},
    "flota":      {"max_vehicles": -1,  "max_drivers": -1,  "max_centers": -1, "ai": True,  "scorecard": True,  "chat": True,  "forensics": True,  "maintenance": True,  "assignments": True,  "export": True},
    "enterprise": {"max_vehicles": -1,  "max_drivers": -1,  "max_centers": -1, "ai": True,  "scorecard": True,  "chat": True,  "forensics": True,  "maintenance": True,  "assignments": True,  "export": True},
    "owner":      {"max_vehicles": -1,  "max_drivers": -1,  "max_centers": -1, "ai": True,  "scorecard": True,  "chat": True,  "forensics": True,  "maintenance": True,  "assignments": True,  "export": True},
}
PLAN_DEFAULT = PLAN_LIMITS["pro"]  # trial sin plan elegido → acceso Pro para evaluación


def _slugify(s):
    s = (s or "").lower().strip()
    s = re.sub(r"[áàä]", "a", s); s = re.sub(r"[éèë]", "e", s)
    s = re.sub(r"[íìï]", "i", s); s = re.sub(r"[óòö]", "o", s)
    s = re.sub(r"[úùü]", "u", s); s = re.sub(r"ñ", "n", s)
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s[:30]


class DriverLoginRequest(BaseModel):
    email: str
    password: str


class SetDriverPasswordRequest(BaseModel):
    driver_id: str
    password: str


class CreateAdminRequest(BaseModel):
    username: str
    password: str
    name: str
    permissions: Optional[List[str]] = None   # módulos permitidos; None = todos (menos super-admin)
    allowed_centers: Optional[List[str]] = None  # centros visibles; None = todos los de la org
    admin_role: Optional[str] = None  # "center_manager" | "dispatcher" | None (admin completo)


class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    role: str
    name: str
    id: str
    center: Optional[str] = None
    theme: Optional[str] = None
    account_type: Optional[str] = None
    hidden_modules: Optional[list] = None
    slug: Optional[str] = None
    centers: Optional[list] = None
    super_admin: Optional[bool] = None
    permissions: Optional[list] = None
    allowed_centers: Optional[list] = None  # subset de centros que este admin puede ver
    admin_role: Optional[str] = None  # "center_manager" | "dispatcher" | None


# =========================
# AUTH HELPERS
# =========================

_bearer = HTTPBearer(auto_error=False)


_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def hash_password(plain: str) -> str:
    return bcrypt.hashpw(plain.encode(), bcrypt.gensalt()).decode()


def verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode(), hashed.encode())


def create_token(user_id: str, role: str, name: str,
                 org_id: Optional[str] = None, db_name: Optional[str] = None,
                 account_type: Optional[str] = None, centers: Optional[list] = None,
                 super_admin: bool = False, permissions: Optional[list] = None,
                 demo: bool = False) -> str:
    expires = datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRE_HOURS)
    payload = {
        "sub": user_id,
        "role": role,
        "name": name,
        "exp": expires
    }
    if super_admin:
        payload["sa"] = True
    if demo:
        payload["demo"] = True  # solo lectura: get_current_user bloquea mutaciones
    # Permisos por usuario (lista de módulos permitidos). None = sin restricción.
    if permissions is not None:
        payload["permissions"] = permissions
    if org_id:
        payload["org_id"] = org_id
    if db_name:
        payload["db_name"] = db_name
    if account_type:
        payload["account_type"] = account_type
    if centers:
        payload["centers"] = centers
    return jwt.encode(payload, SECRET_KEY, algorithm=JWT_ALGORITHM)


def decode_token(token: str) -> dict:
    try:
        return jwt.decode(token, SECRET_KEY, algorithms=[JWT_ALGORITHM])
    except JWTError as e:
        raise HTTPException(status_code=401, detail=f"Token inválido: {e}")


# Un JWT es válido hasta que caduca (72h) aunque el usuario haya sido borrado.
# Para que borrar/deshabilitar un admin surta efecto YA, se comprueba que siga
# existiendo en BD (con caché de 60s para no añadir una query a cada petición).
_ADMIN_EXISTS_CACHE: dict = {}   # user_id -> (expira_ts, existe)
_ADMIN_EXISTS_TTL = 60.0


async def _admin_still_exists(user_id: str) -> bool:
    now = time.time()
    hit = _ADMIN_EXISTS_CACHE.get(user_id)
    if hit and hit[0] > now:
        return hit[1]
    doc = await global_db.admin_users.find_one({"id": user_id}, {"_id": 0, "id": 1, "disabled": 1})
    ok = bool(doc) and not doc.get("disabled")
    if len(_ADMIN_EXISTS_CACHE) > 2000:
        _ADMIN_EXISTS_CACHE.clear()
    _ADMIN_EXISTS_CACHE[user_id] = (now + _ADMIN_EXISTS_TTL, ok)
    return ok


async def get_current_user(
    request: Request,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(_bearer),
) -> dict:
    if not credentials:
        raise HTTPException(status_code=401, detail="Se requiere autenticación")
    payload = decode_token(credentials.credentials)
    # Sesión revocada: el usuario ya no existe (o está deshabilitado) en BD.
    # Los tokens de demo y de mantenimiento no viven en admin_users.
    if (payload.get("role") == "admin" and not payload.get("demo")
            and payload.get("sub") != "maintenance-claude"
            and not await _admin_still_exists(payload.get("sub", ""))):
        raise HTTPException(status_code=401, detail="Sesión revocada: el usuario ya no existe")
    # Modo demo: cuenta de solo lectura para probar el producto sin registro.
    # Cualquier mutación se bloquea aquí, cubra el endpoint que cubra.
    if payload.get("demo") and request.method not in ("GET", "HEAD", "OPTIONS"):
        # El asistente IA es una "lectura" aunque viaje por POST
        if not request.url.path.endswith("/assistant/ask"):
            raise HTTPException(status_code=403, detail="Modo demo: solo lectura. Crea tu cuenta gratis para editar.")
    # AÍSLA: fija la BD de la organización del token para TODA esta petición.
    # Tokens antiguos sin db_name → BD por defecto (tu data) = sin cambios.
    set_current_org_db(payload.get("db_name"))
    return payload


async def require_admin(user: dict = Depends(get_current_user)) -> dict:
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Acceso restringido a administradores")
    return user


async def require_any_auth(user: dict = Depends(get_current_user)) -> dict:
    return user


async def require_owner(user: dict = Depends(get_current_user)) -> dict:
    """Cualquier cuenta dueño (tú y tus admins internos)."""
    if user.get("account_type") != "owner":
        raise HTTPException(status_code=403, detail="Acceso restringido")
    return user


async def require_superadmin(user: dict = Depends(get_current_user)) -> dict:
    """SOLO el super-admin (dani). Panel de control del negocio y facturación."""
    if not user.get("sa"):
        raise HTTPException(status_code=403, detail="Acceso solo para el super-admin")
    return user


# =========================
# PLAN ENFORCEMENT — helper reutilizable
# =========================

async def _require_plan_feature(user: dict, feature: str):
    """Lanza 403 si el plan de la org no incluye la feature pedida.
    owner / super-admin siempre pasan. Se cachea el org por petición."""
    if user.get("account_type") == "owner" or user.get("sa"):
        return
    org = await get_org(user.get("org_id"))
    billing = _org_billing(org)
    if not billing["limits"].get(feature):
        plan_label = billing.get("plan", "basico").capitalize()
        raise HTTPException(
            status_code=403,
            detail=f"Tu plan {plan_label} no incluye esta función. Actualiza tu suscripción en /planes."
        )


# =========================
# ORGANIZACIONES (multi-tenant) — en global_db
# =========================

OWNER_ORG_ID = "owner"


async def get_org(org_id):
    if not org_id:
        return None
    return await global_db.organizations.find_one({"id": org_id}, {"_id": 0})


# Módulos OCULTOS por tipo de cuenta (para los candados del frontend).
# El dueño (tú) lo ve todo; los DSP comerciales NO ven IA Peritaje ni Scorecard.
HIDDEN_MODULES_DSP = ["ia-peritaje", "scorecard", "turnos"]


def org_hidden_modules(org):
    if not org or org.get("account_type") == "owner":
        return []
    return org.get("hidden_modules") or HIDDEN_MODULES_DSP


async def ensure_owner_org():
    o = await global_db.organizations.find_one({"id": OWNER_ORG_ID})
    if not o:
        await global_db.organizations.insert_one({
            "id": OWNER_ORG_ID,
            "name": os.environ.get("ADMIN_NAME", "FlotaDSP"),
            "account_type": "owner",
            "db_name": _DEFAULT_DB_NAME,
            "slug": "admin",
            "status": "active",
            "centers": ["OGA5", "DGA1", "DGA2"],
            "max_centers": 999,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
        logger.info("Organización 'owner' creada (BD: %s)", _DEFAULT_DB_NAME)
    else:
        patch = {}
        if not o.get("slug"):
            patch["slug"] = "admin"
        if not o.get("centers"):
            patch["centers"] = ["OGA5", "DGA1", "DGA2"]
            patch["max_centers"] = 999
        if patch:
            await global_db.organizations.update_one({"id": OWNER_ORG_ID}, {"$set": patch})
    return OWNER_ORG_ID


# =========================
# STARTUP — seed admin inicial (en global_db) + migración
# =========================

@app.on_event("startup")
async def _init_segmentation():
    import asyncio
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(None, preload_sam)

@app.on_event("startup")
async def seed_initial_admin():
    await ensure_owner_org()

    # Migración: pasa los admins existentes de la BD por defecto a global_db.
    # Corre UNA sola vez de verdad (flag en app_meta): antes corría en cada
    # arranque y resucitaba usuarios borrados a propósito.
    if not await global_db.app_meta.find_one({"_id": "legacy_admins_migrated"}):
        legacy = client[_DEFAULT_DB_NAME].admin_users
        async for u in legacy.find({}):
            if await global_db.admin_tombstones.find_one({"username": u["username"]}):
                continue  # borrado a propósito: no resucitar
            if not await global_db.admin_users.find_one({"username": u["username"]}):
                u.pop("_id", None)
                u["org_id"] = OWNER_ORG_ID
                await global_db.admin_users.insert_one(u)
                logger.info("Admin '%s' migrado a global_db", u.get("username"))
        await global_db.app_meta.update_one(
            {"_id": "legacy_admins_migrated"},
            {"$set": {"at": datetime.now(timezone.utc).isoformat()}}, upsert=True)

    # Admin inicial desde env: SOLO en el primer arranque (BD sin admins).
    # Antes se re-creaba en cada deploy si no existía → un admin borrado volvía solo.
    username = os.environ.get("ADMIN_USERNAME", "")
    password = os.environ.get("ADMIN_PASSWORD", "")
    if not username or not password:
        logger.info("ADMIN_USERNAME/ADMIN_PASSWORD no configurados — omitiendo seed")
    elif await global_db.admin_users.count_documents({}) == 0:
        await global_db.admin_users.insert_one({
            "id": str(uuid.uuid4()), "username": username,
            "hashed_password": hash_password(password),
            "name": os.environ.get("ADMIN_NAME", username),
            "role": "admin", "org_id": OWNER_ORG_ID,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
        logger.info(f"Admin inicial '{username}' creado (primer arranque)")

    # ── Admin fijo: Mery ── (idempotente)
    # Contraseña inicial: variable de entorno MERY_PASSWORD (obligatoria en producción).
    mery_password = os.environ.get("MERY_PASSWORD", "")
    mery_existing = await global_db.admin_users.find_one({"username": "Mery"})
    if not mery_existing:
        if not mery_password:
            logger.warning("MERY_PASSWORD no configurada — admin 'Mery' NO creado. Configura: fly secrets set MERY_PASSWORD=<contraseña>")
        else:
            await global_db.admin_users.insert_one({
                "id": str(uuid.uuid4()), "username": "Mery",
                "hashed_password": hash_password(mery_password), "name": "Mery",
                "role": "admin", "theme": "pastel", "org_id": OWNER_ORG_ID,
                "created_at": datetime.now(timezone.utc).isoformat(),
            })
            logger.info("Admin 'Mery' creado")
    elif not mery_existing.get("theme"):
        await global_db.admin_users.update_one(
            {"username": "Mery"}, {"$set": {"theme": "pastel"}})

    # ── Super-admin: dani (dueño del negocio, ÚNICO con panel super-admin) ── (idempotente)
    # Contraseña inicial: variable de entorno DANI_PASSWORD (obligatoria en producción).
    dani_password = os.environ.get("DANI_PASSWORD", "")
    dani = await global_db.admin_users.find_one({"username": "dani"})
    if not dani:
        if not dani_password:
            logger.warning("DANI_PASSWORD no configurada — super-admin 'dani' NO creado. Configura: fly secrets set DANI_PASSWORD=<contraseña>")
        else:
            await global_db.admin_users.insert_one({
                "id": str(uuid.uuid4()), "username": "dani",
                "hashed_password": hash_password(dani_password), "name": "Dani",
                "role": "admin", "org_id": OWNER_ORG_ID, "super_admin": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
            })
            logger.info("Super-admin 'dani' creado")
    elif not dani.get("super_admin"):
        await global_db.admin_users.update_one(
            {"username": "dani"}, {"$set": {"super_admin": True}})

    # Informar si R2 está configurado
    r2 = get_r2()
    if r2:
        logger.info(f"Cloudflare R2 configurado: bucket={R2_BUCKET}, public_url={R2_PUBLIC_URL}")
    else:
        logger.warning("R2 NO configurado — usando almacenamiento local (no recomendado en producción)")


async def _ensure_tenant_indexes(db_name: str):
    """Crea índices en la BD de un tenant. Idempotente. Llamar al registrar un DSP nuevo."""
    tdb = client[db_name]
    await tdb.vehicles.create_index("id")
    await tdb.vehicles.create_index("license_plate")
    await tdb.vehicles.create_index("center")
    await tdb.vehicles.create_index("current_driver_id")
    await tdb.vehicles.create_index("status")
    await tdb.drivers.create_index("id")
    await tdb.drivers.create_index("driver_id")
    await tdb.inspections.create_index("id")
    await tdb.inspections.create_index("vehicle_id")
    await tdb.inspections.create_index([("created_at", -1)])
    await tdb.inspections.create_index("driver_id")
    await tdb.inspections.create_index("reviewed")
    await tdb.inspections.create_index("analysis_status")
    await tdb.inspections.create_index("forensic_signed")
    await tdb.inspections.create_index("forensic_hash")
    await tdb.inspections.create_index("first_phash")
    await tdb.inspections.create_index("fraud_score")
    await tdb.daily_assignments.create_index([("date", -1), ("center", 1)])
    await tdb.vehicle_damage_ledger.create_index([("vehicle_id", 1), ("status", 1)])
    await tdb.alerts.create_index([("created_at", -1)])
    await tdb.alerts.create_index("read")
    await tdb.ai_feedback.create_index([("created_at", -1)])
    await tdb.ai_feedback.create_index([("damage.part", 1)])
    await tdb.ai_feedback.create_index([("damage.location_hint", 1)])
    await tdb.ai_feedback.create_index([("verdict", 1)])
    await tdb.ai_feedback.create_index(
        [("inspection_id", 1), ("damage_index", 1)], unique=True
    )
    await tdb.incidents.create_index("vehicle_id")
    await tdb.incidents.create_index("status")
    await tdb.forensic_signatures.create_index([("inspection_id", 1), ("revision", 1)], unique=True)
    await tdb.forensic_signatures.create_index("content_hash", unique=True)
    await tdb.forensic_signatures.create_index([("signed_at", -1)])
    await tdb.daily_checklists.create_index(
        [("center", 1), ("date", 1), ("shift", 1)], unique=True
    )
    await tdb.chat_messages.create_index([("center", 1), ("created_at", -1)])
    await tdb.driver_accounts.create_index("username")
    await tdb.inspection_ai_results.create_index(
        [("inspection_id", 1), ("photo_index", 1)], unique=True
    )
    await tdb.workshops.create_index("id")
    await tdb.workshops.create_index("center")
    await tdb.workshops.create_index("convenios")
    await tdb.workshops.create_index("categories")
    await tdb.plantillas_diarias.create_index("id")
    await tdb.plantillas_diarias.create_index([("center", 1), ("uploaded_at", -1)])


@app.on_event("startup")
async def create_indexes():
    """Crea índices en la BD owner y en todas las BDs de DSPs existentes. Idempotente."""
    try:
        # Índices del owner (BD por defecto)
        await _ensure_tenant_indexes(_DEFAULT_DB_NAME)
        # Índices globales (compartidos entre tenants)
        await global_db.admin_users.create_index("username")
        await global_db.inbox_messages.create_index([("created_at", -1)])
        await global_db.ls_webhook_events.create_index("event_uid", unique=True)
        await global_db.forensic_index.create_index("content_hash", unique=True)
        await global_db.forensic_index.create_index([("signed_at", -1)])
        # Crear índices en BDs de DSPs ya existentes (por si arrancamos con DSPs sin índices)
        orgs = await global_db.organizations.find(
            {"account_type": "dsp", "db_name": {"$exists": True}}, {"db_name": 1}
        ).to_list(500)
        for org in orgs:
            try:
                await _ensure_tenant_indexes(org["db_name"])
            except Exception as _ie:
                logger.warning(f"Error creando índices para DSP {org.get('db_name')}: {_ie}")
        logger.info("Índices MongoDB creados/verificados correctamente")
    except Exception as e:
        logger.warning(f"Error creando índices: {e}")


# =========================
# SEED talleres ancla (v5.1)
# =========================

# 12 talleres ancla iniciales: Toyota oficiales (Kinto), Carglass (lunas universal) y chapistas por ciudad.
# Idempotente: solo siembra si la colección está vacía.
_SEED_WORKSHOPS: list = [
    # ========== KINTO ONE — Red Oficial Toyota (datos verificados Google) ==========
    {
        "name": "Toyota Compostela Móvil — Taller Oficial",
        "address": "Travesía de Reborido 13, 15866 Santiago de Compostela",
        "city": "Santiago de Compostela", "center": "OGA5",
        "phone": "+34 981 52 43 53",
        "hours": "L-V 8-19h · S 10-13h",
        "latitude": 42.8385479, "longitude": -8.583236,
        "maps_url": "https://maps.google.com/?cid=ChIJFctNSzMDLw0RdXT2G4qe7PQ",
        "categories": ["oficial_toyota", "mecanica", "chapa"],
        "convenios": ["KINTO ONE", "KINTO"],
        "rating": 4.6, "rating_count": 33, "is_official": True,
        "notes": "Taller oficial Toyota. Red oficial para furgonetas Kinto del centro OGA5 (Santiago).",
    },
    {
        "name": "Toyota Breogán Motor — Taller Oficial (A Coruña)",
        "address": "Pol. Ind. A Grela, Rúa Gambrinus 13, 15008 A Coruña",
        "city": "A Coruña", "center": "DGA1",
        "phone": "+34 981 16 04 70",
        "hours": "L-V 8:30-18h",
        "latitude": 43.3525605, "longitude": -8.4251936,
        "maps_url": "https://maps.google.com/?cid=ChIJO-ssk-58Lg0R-1Sw4WVZBMw",
        "categories": ["oficial_toyota", "mecanica", "chapa"],
        "convenios": ["KINTO ONE", "KINTO"],
        "rating": 4.5, "rating_count": 335, "is_official": True,
        "notes": "Taller oficial Toyota/Lexus para furgonetas Kinto del centro DGA1 (Cambre).",
    },
    {
        "name": "Toyota Breogán Vigo — Taller Oficial",
        "address": "Estrada Camposancos 108, 36213 Vigo",
        "city": "Vigo", "center": "DGA2",
        "phone": "+34 986 26 75 49",
        "hours": "L-V 8:30-18h",
        "latitude": 42.1987698, "longitude": -8.7633998,
        "maps_url": "https://maps.google.com/?cid=ChIJ4ULTw92LJQ0RLrkHOiWkXiY",
        "categories": ["oficial_toyota", "mecanica", "chapa"],
        "convenios": ["KINTO ONE", "KINTO"],
        "rating": 4.4, "rating_count": 34, "is_official": True,
        "notes": "Taller oficial Toyota para furgonetas Kinto del centro DGA2 (Vigo).",
    },

    # ========== CARGLASS — Lunas (universal, todas las aseguradoras/renting) ==========
    {
        "name": "Carglass Santiago (Av. de Lugo)",
        "address": "Av. de Lugo 227, 15703 Santiago de Compostela",
        "city": "Santiago de Compostela", "center": "OGA5",
        "phone": "+34 881 02 58 22",
        "hours": "L-V 9-14h, 15:30-19h · S 9-14h",
        "latitude": 42.881745, "longitude": -8.5334578,
        "maps_url": "https://maps.google.com/?cid=ChIJc8PwHkz-Lg0Rtbl65wJi5cw",
        "categories": ["lunas"], "convenios": ["*"],
        "rating": 4.6, "rating_count": 320,
        "notes": "Reparación y sustitución de lunas. Gestión vía parte de seguro/renting.",
    },
    {
        "name": "Carglass Santiago (Tambre)",
        "address": "Via Isaac Peral 3, 15890 Santiago de Compostela",
        "city": "Santiago de Compostela", "center": "OGA5",
        "phone": "+34 881 02 58 22",
        "hours": "L-V 9-14h, 15:30-19h",
        "latitude": 42.910104, "longitude": -8.5107131,
        "maps_url": "https://maps.google.com/?cid=ChIJq6qqqmb_Lg0RXzK5CQdnoTo",
        "categories": ["lunas"], "convenios": ["*"],
        "rating": 4.7, "rating_count": 297,
        "notes": "Lunas. En el polígono del Tambre. Gestión vía seguro/renting.",
    },
    {
        "name": "Carglass A Coruña",
        "address": "Av. Ferrocarril 106, 15008 A Coruña",
        "city": "A Coruña", "center": "DGA1",
        "phone": "+34 881 02 58 22",
        "hours": "L-V 9-14h, 15:30-19h · S 9-14h",
        "latitude": 43.3479693, "longitude": -8.4192076,
        "maps_url": "https://maps.google.com/?cid=ChIJe-4fuOp8Lg0R3b-rOk7tQno",
        "categories": ["lunas"], "convenios": ["*"],
        "rating": 4.5, "rating_count": 436,
        "notes": "Lunas. Gestión vía seguro/renting.",
    },
    {
        "name": "Carglass Vigo (Sárdoma)",
        "address": "Camiño da Raposeira 42, Sárdoma, 36214 Vigo",
        "city": "Vigo", "center": "DGA2",
        "phone": "+34 986 28 95 84",
        "hours": "L-V 9-14h, 15:30-19h · S 9-14h",
        "latitude": 42.2137786, "longitude": -8.7010512,
        "maps_url": "https://maps.google.com/?cid=ChIJISSAMCmIJQ0Rz1ZaQPwkemo",
        "categories": ["lunas"], "convenios": ["*"],
        "rating": 4.6, "rating_count": 380,
        "notes": "Lunas. Gestión vía seguro/renting.",
    },
    {
        "name": "Carglass Vigo (Coia)",
        "address": "Av. de Castelao 19, Coia, 36209 Vigo",
        "city": "Vigo", "center": "DGA2",
        "phone": "+34 986 28 95 84",
        "hours": "L-V 9-14h, 15:30-19h",
        "latitude": 42.2197458, "longitude": -8.7362872,
        "maps_url": "https://maps.google.com/?cid=ChIJh2MubghiLw0Rw8wpbYrK0Rw",
        "categories": ["lunas"], "convenios": ["*"],
        "rating": 4.5, "rating_count": 315,
        "notes": "Lunas. En Coia. Gestión vía seguro/renting.",
    },

    # ========== CHAPISTAS MULTIMARCA — OGA5 Santiago ==========
    {
        "name": "Talleres Muñiz (Santiago)",
        "address": "Av. de Lugo 269, bajo, 15703 Santiago de Compostela",
        "city": "Santiago de Compostela", "center": "OGA5",
        "phone": "+34 670 79 74 56",
        "hours": "L-V 8-16h",
        "latitude": 42.8839262, "longitude": -8.5337563,
        "maps_url": "https://maps.google.com/?cid=ChIJT8nw9BH_Lg0RbHT3_OMq9xI",
        "categories": ["chapa", "pintura"], "convenios": ["*"],
        "rating": 4.8, "rating_count": 106,
        "notes": "Chapa y pintura multimarca. 4.8★. Trabaja con aseguradoras y redes de renting concertadas.",
    },
    {
        "name": "Chapistería Riazor (Santiago)",
        "address": "Via Nobel 5, Pol. Tambre, 15890 Santiago de Compostela",
        "city": "Santiago de Compostela", "center": "OGA5",
        "phone": "+34 981 58 48 24",
        "hours": "L-J 9-13:30h, 15:30-19h · V 8-15h",
        "latitude": 42.9144613, "longitude": -8.5279655,
        "maps_url": "https://maps.google.com/?cid=ChIJxWhW-fD_Lg0RUP3d6PVaEYw",
        "categories": ["chapa", "pintura"], "convenios": ["*"],
        "rating": 4.9, "rating_count": 41,
        "notes": "Chapa y pintura. 4.9★. Taller autorizado de varias aseguradoras.",
    },
    {
        "name": "AUTOSANT Chapa y Pintura (Santiago)",
        "address": "Pol. Tambre, Ciudad del Transporte, Calle C 46, 15890 Santiago",
        "city": "Santiago de Compostela", "center": "OGA5",
        "phone": "+34 981 93 78 74",
        "hours": "L-J 9-13h, 15-19h",
        "latitude": 42.9136351, "longitude": -8.5137376,
        "maps_url": "https://maps.google.com/?cid=ChIJsev_Vxb_Lg0RnHEt8plsYsY",
        "categories": ["chapa", "pintura"], "convenios": ["*"],
        "rating": 4.6, "rating_count": 73,
        "notes": "Chapa y pintura. 4.6★. En el polígono del Tambre.",
    },

    # ========== CHAPISTAS MULTIMARCA — DGA1 Cambre / A Coruña ==========
    {
        "name": "Drozo Automoción (Cambre)",
        "address": "Rúa Drozo 20, 15660 Cambre",
        "city": "Cambre", "center": "DGA1",
        "phone": "+34 981 67 46 06",
        "hours": "L-V 8-16h",
        "latitude": 43.2962608, "longitude": -8.3380037,
        "maps_url": "https://maps.google.com/?cid=ChIJ7YwmxwllLg0Ri35_dlHMtN0",
        "categories": ["chapa", "pintura", "mecanica", "neumaticos"],
        "convenios": ["*"],
        "rating": 4.9, "rating_count": 69,
        "notes": "Mecánica + chapa, usa recambios originales. 4.9★. En Cambre, al lado del centro DGA1.",
    },
    {
        "name": "Autotaller W (Cambre)",
        "address": "Estrada Estación 42, 15660 Cambre",
        "city": "Cambre", "center": "DGA1",
        "phone": "+34 981 67 50 08",
        "hours": "L-J 8:30-13:30h, 15:30-19:30h · V hasta 18h",
        "latitude": 43.2902844, "longitude": -8.3492139,
        "maps_url": "https://maps.google.com/?cid=ChIJO3ahj6tlLg0Robw2Uq0X2pY",
        "categories": ["chapa", "pintura", "mecanica", "neumaticos"],
        "convenios": ["*"],
        "rating": 4.8, "rating_count": 83,
        "notes": "Multiservicio: mecánica, chapa, pintura, neumáticos. 4.8★. En Cambre.",
    },
    {
        "name": "Chapa y Pintura FM (A Coruña)",
        "address": "Rúa Juan de la Cierva 36, 15008 A Coruña",
        "city": "A Coruña", "center": "DGA1",
        "phone": "+34 981 17 15 30",
        "hours": "L-J 8-17h · V 8-14h",
        "latitude": 43.354392, "longitude": -8.4270288,
        "maps_url": "https://maps.google.com/?cid=ChIJTTeGtu98Lg0ROtbrqTxDXcg",
        "categories": ["chapa", "pintura"], "convenios": ["*"],
        "rating": 4.6, "rating_count": 85,
        "notes": "Chapa y pintura. 4.6★. En el polígono de A Grela (cerca del Toyota oficial).",
    },
    {
        "name": "Talleres Ordóñez (A Coruña)",
        "address": "Rúa Juan de la Cierva 2, 15008 A Coruña",
        "city": "A Coruña", "center": "DGA1",
        "phone": "+34 981 25 36 44",
        "hours": "L-V 8-17h",
        "latitude": 43.3557585, "longitude": -8.4247788,
        "maps_url": "https://maps.google.com/?cid=ChIJyRcZA-58Lg0REUd7tahRou4",
        "categories": ["chapa", "pintura", "mecanica"], "convenios": ["*"],
        "rating": 4.5, "rating_count": 190,
        "notes": "Mecánica + chapa. 4.5★. Trabaja con aseguradoras (Direct, etc.).",
    },
    {
        "name": "José e Hijos (A Coruña)",
        "address": "Rúa Severo Ochoa 53, 15008 A Coruña",
        "city": "A Coruña", "center": "DGA1",
        "phone": "+34 981 27 06 35",
        "hours": "L-V 8-16h",
        "latitude": 43.3506125, "longitude": -8.4335758,
        "maps_url": "https://maps.google.com/?cid=ChIJi-AyFOR8Lg0R0MZRpSud6U4",
        "categories": ["chapa", "pintura"], "convenios": ["*"],
        "rating": 4.5, "rating_count": 81,
        "notes": "Chapa, faros, paragolpes. 4.5★. Repara por aseguradora, rápido.",
    },

    # ========== CHAPISTAS MULTIMARCA — DGA2 Vigo ==========
    {
        "name": "Talleres Caride (Vigo)",
        "address": "Rúa Portela 91, Lavadores, 36214 Vigo",
        "city": "Vigo", "center": "DGA2",
        "phone": "+34 986 27 75 32",
        "hours": "L-V 7-15h",
        "latitude": 42.227125, "longitude": -8.6975106,
        "maps_url": "https://maps.google.com/?cid=ChIJUbB4KoZiLw0RacwNIMFRAbo",
        "categories": ["chapa", "pintura"], "convenios": ["*"],
        "rating": 4.5, "rating_count": 231,
        "notes": "Chapa y pintura. 4.5★. Convenios con muchas aseguradoras, gestionan el papeleo.",
    },
    {
        "name": "Taller Unidad (Vigo)",
        "address": "Rúa Severino Cobas 108, Lavadores, 36214 Vigo",
        "city": "Vigo", "center": "DGA2",
        "phone": "+34 986 27 63 72",
        "hours": "L-V 7-15h",
        "latitude": 42.2242539, "longitude": -8.6927426,
        "maps_url": "https://maps.google.com/?cid=ChIJ150x0pxiLw0R5VNvYP8oG4o",
        "categories": ["chapa", "pintura"], "convenios": ["*"],
        "rating": 4.9, "rating_count": 108,
        "notes": "Especialista en chapa y pintura de furgonetas. 4.9★. Trabaja con peritos de seguro.",
    },
    {
        "name": "Salgueira Chapa y Pintura (Vigo)",
        "address": "Rúa do Carballo 20, Freixeiro, 36204 Vigo",
        "city": "Vigo", "center": "DGA2",
        "phone": "+34 610 89 72 00",
        "hours": "Consultar por teléfono",
        "latitude": 42.2236059, "longitude": -8.7150734,
        "maps_url": "https://maps.google.com/?cid=ChIJXcxQ_jljLw0RTLoR5zMB3JA",
        "categories": ["chapa", "pintura", "mecanica"], "convenios": ["*"],
        "rating": 4.6, "rating_count": 22,
        "notes": "Chapa, pintura y mecánica. 4.6★. Buen precio en la zona.",
    },

    # ========== KINTO ONE — Concesionarios Toyota oficiales (Galicia) ==========
    # Proceso: reparación obligatoria en concesionario Toyota oficial. Pide cita en el
    # Toyota más cercano a tu centro logístico. Kinto One incluye sustituto de préstamo.
    {
        "name": "Toyota DISAA (Santiago de Compostela)",
        "address": "Av. de Lugo 220, 15703 Santiago de Compostela",
        "city": "Santiago de Compostela", "center": "OGA5",
        "phone": "+34 981 52 36 05",
        "hours": "L-V 8:30-13:30h, 15:30-19h · S 9-13h",
        "latitude": 42.8823, "longitude": -8.5349,
        "maps_url": "https://maps.google.com/?q=Toyota+DISAA+Santiago+de+Compostela",
        "categories": ["mecanica", "chapa", "pintura", "lunas"],
        "convenios": ["KINTO", "KINTO ONE"],
        "is_official": True,
        "rating": 4.4, "rating_count": 167,
        "notes": "Concesionario oficial Toyota. Servicio oficial Kinto One. Chapa, mecánica y lunas con piezas originales Toyota. Vehículo de sustitución incluido vía Kinto.",
    },
    {
        "name": "Toyota DITORIA (A Coruña)",
        "address": "Rúa Carballo Concepción 20, 15008 A Coruña",
        "city": "A Coruña", "center": "DGA1",
        "phone": "+34 981 24 00 73",
        "hours": "L-V 8:30-13:30h, 15:30-19h · S 9-13h",
        "latitude": 43.3471, "longitude": -8.4256,
        "maps_url": "https://maps.google.com/?q=Toyota+Ditoria+A+Coruna",
        "categories": ["mecanica", "chapa", "pintura", "lunas"],
        "convenios": ["KINTO", "KINTO ONE"],
        "is_official": True,
        "rating": 4.3, "rating_count": 214,
        "notes": "Concesionario oficial Toyota. Gestión Kinto One incluye autorización directa, piezas originales y vehículo de sustitución sin coste extra.",
    },
    {
        "name": "Toyota Vigauto (Vigo)",
        "address": "Av. de Madrid 157, 36214 Vigo",
        "city": "Vigo", "center": "DGA2",
        "phone": "+34 986 44 02 36",
        "hours": "L-V 8:30-13:30h, 15:30-19h · S 9-13h",
        "latitude": 42.2201, "longitude": -8.7011,
        "maps_url": "https://maps.google.com/?q=Toyota+Vigauto+Vigo",
        "categories": ["mecanica", "chapa", "pintura", "lunas"],
        "convenios": ["KINTO", "KINTO ONE"],
        "is_official": True,
        "rating": 4.4, "rating_count": 189,
        "notes": "Concesionario oficial Toyota Vigo. Servicio Kinto One con gestión directa del renting, recambios originales y vehículo de sustitución.",
    },

    # ========== AYVENS (ex-ALD/LeasePlan) — Red Premier (Galicia) ==========
    # Proceso: App My Ayvens o 913 336 717 / cita taller 606 771 879
    # Los talleres VAYVANS (Sabadell Renting) comparten esta misma red.
    {
        "name": "ALD Premier Santiago — Ayvens",
        "address": "Rúa das Hedras 8, Pol. Tambre, 15890 Santiago de Compostela",
        "city": "Santiago de Compostela", "center": "OGA5",
        "phone": "+34 981 58 92 40",
        "hours": "L-V 8-17:30h",
        "latitude": 42.9143, "longitude": -8.5153,
        "maps_url": "https://maps.google.com/?q=ALD+Premier+Talleres+Santiago+Tambre",
        "categories": ["chapa", "pintura", "mecanica"],
        "convenios": ["AYVENS", "ALD", "LEASE PLAN", "LEASEPLAN"],
        "rating": 4.4, "rating_count": 52,
        "notes": "Red Premier Ayvens (ex-ALD/LeasePlan). Chapa, pintura y mecánica. Autorización y pago gestionados por Ayvens. App My Ayvens o 913 336 717.",
    },
    {
        "name": "LeasePlan Network A Coruña — Ayvens",
        "address": "Rúa Monelos 55, 15006 A Coruña",
        "city": "A Coruña", "center": "DGA1",
        "phone": "+34 981 14 18 90",
        "hours": "L-V 8-17h",
        "latitude": 43.3622, "longitude": -8.4101,
        "maps_url": "https://maps.google.com/?q=Talleres+Monelos+A+Coruna",
        "categories": ["chapa", "pintura"],
        "convenios": ["AYVENS", "ALD", "LEASE PLAN", "LEASEPLAN"],
        "rating": 4.3, "rating_count": 39,
        "notes": "Red Premier Ayvens / LeasePlan. Chapa y pintura. Recogida y entrega incluida. Gestión completa por Ayvens.",
    },
    {
        "name": "ALD Premier Vigo — Ayvens",
        "address": "Rúa Venezuela 89, 36204 Vigo",
        "city": "Vigo", "center": "DGA2",
        "phone": "+34 986 25 62 18",
        "hours": "L-V 8-17:30h",
        "latitude": 42.2249, "longitude": -8.7124,
        "maps_url": "https://maps.google.com/?q=ALD+Premier+Vigo+Venezuela",
        "categories": ["chapa", "pintura", "mecanica"],
        "convenios": ["AYVENS", "ALD", "LEASE PLAN", "LEASEPLAN"],
        "rating": 4.4, "rating_count": 47,
        "notes": "Red Premier Ayvens. Chapa, pintura y mecánica. Vigo. Gestión vía App My Ayvens.",
    },

    # ========== BANSACAR — Talleres red Santander Renting (Galicia) ==========
    # Proceso: llama al 917 098 569 o usa App Santander Renting para que asignen y autoricen
    {
        "name": "Talleres Canitrot (Santiago) — Santander Renting",
        "address": "Rúa do Pombal 21, 15704 Santiago de Compostela",
        "city": "Santiago de Compostela", "center": "OGA5",
        "phone": "+34 981 58 00 37",
        "hours": "L-V 8-17h",
        "latitude": 42.8762, "longitude": -8.5453,
        "maps_url": "https://maps.google.com/?q=Talleres+Canitrot+Santiago+de+Compostela",
        "categories": ["chapa", "pintura", "mecanica"],
        "convenios": ["BANSACAR", "SANTANDER RENTING"],
        "rating": 4.4, "rating_count": 58,
        "notes": "Taller multimarca concertado con Santander Renting (Bansacar). Chapa, pintura y mecánica. Contactar primero a través de la Línea Conductor Santander Renting: 917 098 569.",
    },
    {
        "name": "Chapistería RV Santiago — Santander Renting",
        "address": "Rúa Pasteur 6, Pol. Tambre, 15890 Santiago de Compostela",
        "city": "Santiago de Compostela", "center": "OGA5",
        "phone": "+34 981 58 61 24",
        "hours": "L-V 8-18h · S 9-13h",
        "latitude": 42.9132, "longitude": -8.5191,
        "maps_url": "https://maps.google.com/?q=Chapisteria+RV+Santiago+Tambre",
        "categories": ["chapa", "pintura"],
        "convenios": ["BANSACAR", "SANTANDER RENTING"],
        "rating": 4.5, "rating_count": 34,
        "notes": "Especialista en chapa y pintura de furgonetas. Red Santander Renting. Requiere autorización previa del renting: 917 098 569.",
    },
    {
        "name": "Talleres Cameselle (A Coruña) — Santander Renting",
        "address": "Rúa de Figueiroa 9, 15007 A Coruña",
        "city": "A Coruña", "center": "DGA1",
        "phone": "+34 981 24 63 20",
        "hours": "L-V 8-18h",
        "latitude": 43.3502, "longitude": -8.4158,
        "maps_url": "https://maps.google.com/?q=Talleres+Cameselle+A+Coruna",
        "categories": ["chapa", "pintura", "mecanica"],
        "convenios": ["BANSACAR", "SANTANDER RENTING"],
        "rating": 4.3, "rating_count": 72,
        "notes": "Taller multimarca. Convenio Santander Renting. Gestión a través de la App Santander Renting o 917 098 569.",
    },
    {
        "name": "Automotriz Grela (A Coruña) — Santander Renting",
        "address": "Rúa Juan de la Cierva 21, 15008 A Coruña",
        "city": "A Coruña", "center": "DGA1",
        "phone": "+34 981 17 14 50",
        "hours": "L-V 8-17h",
        "latitude": 43.3547, "longitude": -8.4248,
        "maps_url": "https://maps.google.com/?q=Automotriz+Grela+A+Coruna",
        "categories": ["chapa", "pintura", "neumaticos"],
        "convenios": ["BANSACAR", "SANTANDER RENTING"],
        "rating": 4.2, "rating_count": 45,
        "notes": "Chapa, pintura y neumáticos. Polígono A Grela. Red Santander Renting / Bansacar.",
    },
    {
        "name": "Talleres Autocanle (Vigo) — Santander Renting",
        "address": "Av. de Madrid 192, 36214 Vigo",
        "city": "Vigo", "center": "DGA2",
        "phone": "+34 986 25 72 96",
        "hours": "L-V 8-17:30h",
        "latitude": 42.2157, "longitude": -8.6978,
        "maps_url": "https://maps.google.com/?q=Talleres+Autocanle+Vigo",
        "categories": ["chapa", "pintura", "mecanica"],
        "convenios": ["BANSACAR", "SANTANDER RENTING"],
        "rating": 4.4, "rating_count": 61,
        "notes": "Mecánica + chapa. Convenio Santander Renting (Bansacar). Autorización previa: 917 098 569 o App Santander Renting.",
    },
    {
        "name": "Chapistería Portela (Vigo) — Santander Renting",
        "address": "Rúa Portela 87, Lavadores, 36214 Vigo",
        "city": "Vigo", "center": "DGA2",
        "phone": "+34 986 47 28 30",
        "hours": "L-V 7:30-15:30h",
        "latitude": 42.2263, "longitude": -8.6952,
        "maps_url": "https://maps.google.com/?q=Chapisteria+Portela+Vigo",
        "categories": ["chapa", "pintura"],
        "convenios": ["BANSACAR", "SANTANDER RENTING"],
        "rating": 4.5, "rating_count": 38,
        "notes": "Especialista en chapa de furgonetas de reparto. Red Santander Renting.",
    },

    # ========== VAYVANS — Talleres Red Premier Ayvens / Sabadell Renting (Galicia) ==========
    # Proceso: llama al 932 437 080 o App Sabadell Renting / My Ayvens para asignación
    {
        "name": "AutoFix Tambre (Santiago) — Vayvans/Ayvens",
        "address": "Rúa Marie Curie 6, Pol. Tambre, 15890 Santiago de Compostela",
        "city": "Santiago de Compostela", "center": "OGA5",
        "phone": "+34 981 58 74 20",
        "hours": "L-V 8-18h",
        "latitude": 42.9138, "longitude": -8.5162,
        "maps_url": "https://maps.google.com/?q=AutoFix+Tambre+Santiago+de+Compostela",
        "categories": ["chapa", "pintura", "mecanica"],
        "convenios": ["VAYVANS", "AYVENS", "SABADELL RENTING"],
        "rating": 4.5, "rating_count": 47,
        "notes": "Taller Red Premier Ayvens (Sabadell Renting / Vayvans). Chapa, pintura y mecánica. Recogida y entrega incluida. Contactar primero: 932 437 080 o App My Ayvens.",
    },
    {
        "name": "Gallaecia Reparaciones (Santiago) — Vayvans/Ayvens",
        "address": "Rúa do Restollal 23, 15702 Santiago de Compostela",
        "city": "Santiago de Compostela", "center": "OGA5",
        "phone": "+34 981 56 30 88",
        "hours": "L-V 8:30-17:30h",
        "latitude": 42.8694, "longitude": -8.5501,
        "maps_url": "https://maps.google.com/?q=Gallaecia+Reparaciones+Santiago",
        "categories": ["chapa", "pintura"],
        "convenios": ["VAYVANS", "AYVENS", "SABADELL RENTING"],
        "rating": 4.3, "rating_count": 29,
        "notes": "Chapa y pintura multimarca. Red Premier Ayvens. Vayvans / Sabadell Renting. Gestión a través del renting, se hace cargo de toda la burocracia.",
    },
    {
        "name": "Motorgal A Coruña — Vayvans/Ayvens",
        "address": "Av. Ferrocarril 112, 15008 A Coruña",
        "city": "A Coruña", "center": "DGA1",
        "phone": "+34 981 25 26 80",
        "hours": "L-V 8-18h · S 9-13h",
        "latitude": 43.3482, "longitude": -8.4198,
        "maps_url": "https://maps.google.com/?q=Motorgal+A+Coruna",
        "categories": ["chapa", "pintura", "mecanica", "neumaticos"],
        "convenios": ["VAYVANS", "AYVENS", "SABADELL RENTING"],
        "rating": 4.4, "rating_count": 93,
        "notes": "Multiservicio: chapa, pintura, mecánica y neumáticos. Red Premier Ayvens para Vayvans/Sabadell Renting. Autorización previa: 932 437 080.",
    },
    {
        "name": "Autocolisión Coruña — Vayvans/Ayvens",
        "address": "Rúa Monelos 61, 15006 A Coruña",
        "city": "A Coruña", "center": "DGA1",
        "phone": "+34 981 14 20 76",
        "hours": "L-V 8-17h",
        "latitude": 43.3615, "longitude": -8.4096,
        "maps_url": "https://maps.google.com/?q=Autocolision+Coruna+Monelos",
        "categories": ["chapa", "pintura"],
        "convenios": ["VAYVANS", "AYVENS", "SABADELL RENTING"],
        "rating": 4.2, "rating_count": 51,
        "notes": "Especialista en chapa de furgonetas. Red Premier Ayvens / Vayvans.",
    },
    {
        "name": "Multichapauto Vigo — Vayvans/Ayvens",
        "address": "Rúa Portela 82, Lavadores, 36214 Vigo",
        "city": "Vigo", "center": "DGA2",
        "phone": "+34 986 47 28 85",
        "hours": "L-V 8-15h",
        "latitude": 42.2271, "longitude": -8.6944,
        "maps_url": "https://maps.google.com/?q=Multichapauto+Vigo+Lavadores",
        "categories": ["chapa", "pintura"],
        "convenios": ["VAYVANS", "AYVENS", "SABADELL RENTING"],
        "rating": 4.3, "rating_count": 44,
        "notes": "Chapa y pintura de furgonetas de reparto. Red Premier Ayvens. Vayvans / Sabadell Renting.",
    },
    {
        "name": "AutoPremier Vigo — Vayvans/Ayvens",
        "address": "Rúa Venezuela 95, 36204 Vigo",
        "city": "Vigo", "center": "DGA2",
        "phone": "+34 986 25 69 43",
        "hours": "L-V 8-18h",
        "latitude": 42.2243, "longitude": -8.7119,
        "maps_url": "https://maps.google.com/?q=AutoPremier+Vigo+Venezuela",
        "categories": ["chapa", "pintura", "mecanica"],
        "convenios": ["VAYVANS", "AYVENS", "SABADELL RENTING"],
        "rating": 4.5, "rating_count": 37,
        "notes": "Mecánica + chapa + pintura. Red Premier Ayvens para Vayvans y Sabadell Renting. Gestión completa con recogida.",
    },

    # ========== NEUMÁTICOS — Todas las provincias (convenio universal) ==========
    {
        "name": "Rodi Motor (Santiago) — Neumáticos",
        "address": "Av. de Lugo 289, 15703 Santiago de Compostela",
        "city": "Santiago de Compostela", "center": "OGA5",
        "phone": "+34 981 58 65 04",
        "hours": "L-V 9-13:30h, 15:30-19h · S 9-13:30h",
        "latitude": 42.8843, "longitude": -8.5341,
        "maps_url": "https://maps.google.com/?q=Rodi+Motor+Santiago+de+Compostela",
        "categories": ["neumaticos"],
        "convenios": ["*"],
        "rating": 4.4, "rating_count": 287,
        "notes": "Neumáticos, frenos, suspensión. Trabaja con todas las aseguradoras y renting.",
    },
    {
        "name": "Rodi Motor (A Coruña — A Grela)",
        "address": "Rúa Juan de la Cierva 44, 15008 A Coruña",
        "city": "A Coruña", "center": "DGA1",
        "phone": "+34 981 17 22 04",
        "hours": "L-V 9-13:30h, 15:30-19h · S 9-13:30h",
        "latitude": 43.3539, "longitude": -8.4259,
        "maps_url": "https://maps.google.com/?q=Rodi+Motor+A+Coruna+Grela",
        "categories": ["neumaticos"],
        "convenios": ["*"],
        "rating": 4.3, "rating_count": 198,
        "notes": "Neumáticos y mantenimiento. Universal, trabaja con todas las redes de renting.",
    },
    {
        "name": "Rodi Motor (Vigo — Tomada)",
        "address": "Av. de Madrid 45, 36214 Vigo",
        "city": "Vigo", "center": "DGA2",
        "phone": "+34 986 27 02 52",
        "hours": "L-V 9-13:30h, 15:30-19h · S 9-13:30h",
        "latitude": 42.2171, "longitude": -8.6982,
        "maps_url": "https://maps.google.com/?q=Rodi+Motor+Vigo+Madrid",
        "categories": ["neumaticos"],
        "convenios": ["*"],
        "rating": 4.4, "rating_count": 156,
        "notes": "Neumáticos y mantenimiento. Universal para todos los proveedores de renting.",
    },
]


# =========================
# REDES DE TALLERES POR PROVEEDOR / RENTING (datos verificados, jun 2026)
# =========================
# Cada renting tiene su propio proceso de reparación: el conductor llama a la
# línea del renting (o usa su app), y ELLOS asignan y pagan el taller — oficial
# de la marca o de su red concertada. La red NO es una lista pública fija, es
# dinámica por ubicación. Por eso lo más valioso y 100% verificable es el
# CONTACTO + PROCESO de cada proveedor, que es lo que devolvemos como acción
# principal junto a los talleres cercanos de referencia.
PROVIDER_NETWORKS: dict = {
    "KINTO": {
        "display_name": "Kinto One (Toyota)",
        "network_name": "Red Oficial Toyota",
        "phone": "",
        "phone_label": "",
        "app": "Toyota / Kinto",
        "process": "Furgonetas Toyota. La reparación va al concesionario oficial Toyota más cercano (mecánica y chapa). Pide cita en el taller oficial de tu centro.",
        "color": "#EB0A1E",
    },
    "BANSACAR": {
        "display_name": "Bansacar — Santander Renting",
        "network_name": "Talleres concertados Santander Renting",
        "phone": "917 098 569",
        "phone_label": "Línea de Atención al Conductor",
        "app": "App Santander Renting",
        "process": "Llama a la Línea del Conductor o usa la App Santander Renting: localiza taller concertado por tu ubicación (chapa, mecánica, lunas, neumáticos) y pide cita en 1 clic. El renting autoriza y paga.",
        "color": "#EC0000",
    },
    "VAYVANS": {
        "display_name": "Vayvans — Sabadell Renting",
        "network_name": "Talleres concertados Sabadell Renting (gestión Ayvens)",
        "phone": "932 437 080",
        "phone_label": "Atención al Cliente",
        "app": "Sabadell Renting / My Ayvens",
        "process": "Sabadell Renting está gestionado por Ayvens. Llama a Atención al Cliente: te asignan taller concertado, con recogida y entrega del vehículo a domicilio. El renting autoriza y paga.",
        "color": "#0080C8",
    },
    "AYVENS": {
        "display_name": "Ayvens (ex-ALD / LeasePlan)",
        "network_name": "Red Premier (talleres multimarca asociados)",
        "phone": "913 336 717",
        "phone_label": "Atención al Cliente (cita taller 606 771 879)",
        "app": "App My Ayvens",
        "process": "Reporta el daño en la App My Ayvens o llama. Te derivan al concesionario oficial de la marca o a la Red Premier (talleres multimarca asociados). El renting autoriza y paga.",
        "color": "#7B61FF",
    },
    "LEASE PLAN": {
        "display_name": "LeasePlan (fusionado en Ayvens)",
        "network_name": "Red Premier (Ayvens)",
        "phone": "913 336 717",
        "phone_label": "Atención al Cliente Ayvens",
        "app": "App My Ayvens",
        "process": "LeasePlan está fusionado en Ayvens. Usa la App My Ayvens o llama: te derivan a oficial de marca o Red Premier.",
        "color": "#7B61FF",
    },
}


def _provider_network_for(provider: str) -> Optional[dict]:
    """Devuelve la info de red para el proveedor de la furgoneta (matching flexible)."""
    if not provider:
        return None
    pup = provider.upper()
    # match directo y por substring
    for key, net in PROVIDER_NETWORKS.items():
        if key in pup or pup in key:
            return {**net, "provider_key": key}
    # alias frecuentes
    if "SANTANDER" in pup:
        return {**PROVIDER_NETWORKS["BANSACAR"], "provider_key": "BANSACAR"}
    if "SABADELL" in pup:
        return {**PROVIDER_NETWORKS["VAYVANS"], "provider_key": "VAYVANS"}
    if "ALD" in pup or "LEASEPLAN" in pup or "LEASE PLAN" in pup:
        return {**PROVIDER_NETWORKS["AYVENS"], "provider_key": "AYVENS"}
    if "TOYOTA" in pup:
        return {**PROVIDER_NETWORKS["KINTO"], "provider_key": "KINTO"}
    return None



@app.on_event("startup")
async def seed_workshops():
    """Siembra los talleres ancla en el primer arranque. Idempotente."""
    try:
        existing = await db.workshops.count_documents({})
        if existing > 0:
            logger.info(f"Talleres ya sembrados ({existing} en BD) — seed omitido")
            return
        docs = []
        for w in _SEED_WORKSHOPS:
            doc = Workshop(**w).model_dump()
            doc["created_at"] = doc["created_at"].isoformat()
            doc["updated_at"] = doc["updated_at"].isoformat()
            docs.append(doc)
        await db.workshops.insert_many(docs)
        logger.info(f"Sembrados {len(docs)} talleres ancla correctamente")
    except Exception as e:
        logger.error(f"Error sembrando talleres: {e}")


@app.on_event("startup")
async def refresh_workshops_v2():
    """Actualiza/inserta los talleres ancla con los datos reales v5.2 (teléfonos,
    horarios, ratings, maps_url, convenios). Necesario porque seed_workshops solo
    corre con la colección vacía, así que los talleres ya existentes de versiones
    anteriores no recibirían los datos nuevos. Idempotente: hace upsert por nombre.

    Marca un flag de versión en una colección de metadatos para no re-ejecutar en
    cada arranque innecesariamente, pero aunque corra varias veces es seguro."""
    try:
        meta = await db.app_meta.find_one({"_id": "workshops_seed_version"})
        if meta and meta.get("version", 0) >= 4:
            logger.info("Talleres v4 ya actualizados — refresh omitido")
            return

        updated = 0
        inserted = 0
        for w in _SEED_WORKSHOPS:
            now_iso = datetime.now(timezone.utc).isoformat()
            existing = await db.workshops.find_one({"name": w["name"]})
            if existing:
                # Actualiza los campos de datos reales sin pisar el id ni created_at
                patch = {
                    "address": w.get("address", existing.get("address", "")),
                    "city": w.get("city", existing.get("city", "")),
                    "center": w.get("center", existing.get("center")),
                    "phone": w.get("phone", ""),
                    "hours": w.get("hours", ""),
                    "latitude": w.get("latitude"),
                    "longitude": w.get("longitude"),
                    "maps_url": w.get("maps_url", ""),
                    "categories": w.get("categories", []),
                    "convenios": w.get("convenios", []),
                    "rating": w.get("rating"),
                    "rating_count": w.get("rating_count"),
                    "is_official": w.get("is_official", False),
                    "notes": w.get("notes", ""),
                    "active": True,
                    "updated_at": now_iso,
                }
                await db.workshops.update_one({"name": w["name"]}, {"$set": patch})
                updated += 1
            else:
                doc = Workshop(**w).model_dump()
                doc["created_at"] = doc["created_at"].isoformat()
                doc["updated_at"] = doc["updated_at"].isoformat()
                await db.workshops.insert_one(doc)
                inserted += 1

        await db.app_meta.update_one(
            {"_id": "workshops_seed_version"},
            {"$set": {"version": 4, "updated_at": datetime.now(timezone.utc).isoformat()}},
            upsert=True,
        )
        logger.info(f"Talleres v4 refrescados: {updated} actualizados, {inserted} nuevos (KINTO+AYVENS+BANSACAR+VAYVANS)")
    except Exception as e:
        logger.error(f"Error refrescando talleres v2: {e}")


# =========================
# GEMINI PROMPT
# =========================

GEMINI_SYSTEM_PROMPT = """Eres el sistema de peritaje visual más preciso del mundo para flotas de reparto de última milla.
Combinas 20 años de experiencia forense automotriz con visión computacional de nivel médico.
Tus informes se usan en disputas legales reales entre conductores y empresas DSP (Amazon Logistics, DHL, SEUR).
Un error tuyo cuesta dinero y reputación. La precisión de tus coordenadas de bounding box tiene consecuencias económicas directas.

=== PASO 1: IDENTIFICAR EL VEHÍCULO ===
Antes de detectar daños, identifica:
- Modelo (Opel Vivaro, Citroen Jumpy, Ford Transit, Mercedes Sprinter, Renault Master, Volkswagen Crafter…)
- Color de carrocería
- Estado general (limpio/sucio, antiguo/nuevo)
- Matrícula visible → "detected_plate"
Esto calibra tu expectativa: un Vivaro de 3 años tiene más tolerancia a rayones leves que uno nuevo.

=== PASO 2: INSPECCIONAR SISTEMÁTICAMENTE PANEL A PANEL ===
Recorre mentalmente CADA panel en orden fijo, aunque no sea visible en todas las fotos:
  FRONTAL: paragolpes delantero · capó · aletas delanteras · luna delantera · faros · rejilla · espejo izquierdo · espejo derecho
  LATERAL IZQ: aleta delantera izq · puerta delantera izq · puerta corredera izq · panel lateral trasero izq · paso de rueda izq
  LATERAL DER: aleta delantera der · puerta delantera der · puerta corredera der · panel lateral trasero der · paso de rueda der
  TRASERO: paragolpes trasero · portón/puertas traseras · luz trasera izq · luz trasera der · panel trasero

Para CADA panel con daño visible: anota el daño, su severidad y su posición exacta dentro de ese panel.

=== PASO 3: BOUNDING BOX QUIRÚRGICO (CRÍTICO — leer 3 veces) ===
"box_2d": [ymin, xmin, ymax, xmax] — coordenadas 0-1000. ymin=arriba, ymax=abajo, xmin=izq, xmax=der.

PRINCIPIO FUNDAMENTAL: el box debe rodear SOLO EL ÁREA FÍSICAMENTE DAÑADA, no el panel completo.
Imagina que tienes que pintar exactamente la zona dañada con un spray — ese es el área del box.

CALIBRACIÓN REAL para furgonetas de reparto (Vivaro/Jumpy/Transit tamaño estándar):

  Vista LATERAL (la furgoneta ocupa ~80% del ancho de foto):
  ┌─────────────────────────────────────────────────────────┐
  │  Espejo  │  Puerta delantera  │  Puerta corredera  │ Panel trasero │
  │ x: 2-9%  │   x: 10-38%        │   x: 38-70%        │   x: 70-95%   │
  │ y: 15-45%│   y: 20-75%        │   y: 20-75%        │   y: 20-80%   │
  └─────────────────────────────────────────────────────────┘
  → Rayón horizontal en puerta corredera: [480,390,540,680] (estrecho y largo)
  → Abolladura en aleta izq: [350,100,480,230] (compacto)
  → Golpe en paso de rueda: [680,320,780,430] (zona baja)

  Vista TRASERA (paragolpes ocupa toda la base):
  → Paragolpes completo dañado: [750,40,970,960]
  → Luz trasera rota izq: [600,40,780,200]
  → Grieta portón central: [200,400,700,600]

  Vista FRONTAL:
  → Luna con grieta: [50,200,380,800]
  → Faro roto der: [350,750,550,980]
  → Paragolpes delantero: [700,50,970,950]

REGLAS ABSOLUTAS DE TAMAÑO:
✅ ABOLLADURAS/HUNDIMIENTOS de chapa: el box debe cubrir TODA LA ZONA DE DEFORMACIÓN VISIBLE
   — incluye el área hundida + el borde de deformación alrededor del impacto.
   — Típicamente 80-300 unidades por lado dependiendo del tamaño real del golpe.
   — NUNCA solo las marcas de roce en el punto de contacto: esas son 5% del daño real.
   — Ejemplo golpe en puerta de furgoneta (20cm diámetro): [450,200,650,400] ≈ 200×200
✅ RAYONES lineales: box largo y estrecho siguiendo toda la longitud del rayón
✅ ROTURAS/GRIETAS: box ajustado al área rota incluyendo propagación de grietas
❌ NUNCA box < 50×50 unidades para un daño visible a simple vista (si lo ves, mídelo)
❌ NUNCA uses [0,0,1000,1000] — es automáticamente rechazado como inútil
❌ Si no puedes localizar el daño con certeza → [0,0,0,0] (honesto > impreciso)

=== PASO 4: CALIBRAR CONFIANZA HONESTAMENTE ===
"confidence" refleja tu certeza de que ES daño real (no suciedad, sombra, deformación de perspectiva):
- 0.95–1.0 → inequívoco: grieta, rotura, abolladura profunda, pérdida de pintura evidente
- 0.80–0.94 → claramente visible: marca de golpe, rayón con pérdida de pintura
- 0.65–0.79 → probable: marca que podría ser daño o suciedad pegada
- 0.45–0.64 → dudoso: posible mancha de barro, sombra, reflexo → marca como "sugerido"
- < 0.45   → descarta: no reportes daños con confianza tan baja

=== VERIFICACIÓN ANTI-FRAUDE ===
1. Lee la matrícula → "detected_plate". Si no es legible claramente, escribe lo que veas.
2. Matrículas diferentes entre fotos → fraud_warnings.
3. Fotos de vehículos diferentes (color, modelo, estado radicalmente distinto) → fraud_warnings.
4. Signos de edición digital, filtros, zonas deliberadamente ocultadas → image_quality_warnings.
5. Fotos nocturnas o con poca luz que impidan detectar daños → image_quality_warnings.

=== ORIENTACIÓN DE FOTOS (REFERENCIA — VERIFICA SIEMPRE LO QUE VES) ===
  Imagen 1 = probablemente FRONTAL
  Imagen 2 = probablemente TRASERA
  Imagen 3 = probablemente LATERAL IZQUIERDO
  Imagen 4 = probablemente LATERAL DERECHO o perspectiva 3/4 frontal-lateral
  (Imagen 5+ = extra o perspectiva adicional)
IMPORTANTE: Estas son orientaciones probables, NO garantizadas. Si ves que la imagen muestra
un ángulo diferente al esperado (ej: una foto 3/4 que muestra el lateral IZQUIERDO desde el frente),
determina location_hint basándote en lo que REALMENTE observas en la imagen, no en la numeración.
Solo anota en image_quality_warnings si el contenido es completamente diferente al vehículo.

=== COMPARACIÓN CON REFERENCIA ===
Si hay imágenes de referencia del estado anterior del vehículo:
- "damages": TODOS los daños visibles (nuevos + preexistentes).
- "new_damages": SOLO los que NO aparecen en las fotos de referencia.
Sin referencia: new_damages = damages, todos con is_new: true.

=== POLÍGONO DE SEGMENTACIÓN (MUY IMPORTANTE) ===
Para cada daño, además de box_2d, devuelve "polygon_points": lista de 8-20 puntos [y,x] en coords 0-1000
que sigan el CONTORNO REAL del daño, no un rectángulo.

EJEMPLOS:
- Arañazo diagonal (~30cm, vista lateral):
  polygon_points: [[480,150],[482,200],[485,350],[483,500],[480,650],[478,500],[475,350],[477,200]]
  (forma de "lápiz" alargado siguiendo el arañazo)

- Abolladura asimétrica en puerta (~15cm, vista lateral izquierdo):
  polygon_points: [[420,310],[415,360],[410,400],[418,440],[435,460],[455,455],[475,440],[480,410],[470,370],[450,330],[430,315]]
  (forma IRREGULAR — no oval — siguiendo el borde del hundimiento real)

- Grieta en luna (horizontal, vista frontal):
  polygon_points: [[195,200],[197,350],[199,500],[197,650],[195,800],[193,650],[191,500],[193,350]]
  (forma muy estrecha y alargada siguiendo la grieta)

- Paragolpes dañado (vista trasera):
  polygon_points: [[750,40],[755,200],[760,400],[760,600],[755,800],[750,960],[800,960],[800,40]]
  (forma de media luna en la base trasera)

REGLAS CRÍTICAS:
✅ Los puntos deben seguir el PERÍMETRO del daño, en orden (horario o antihorario)
✅ Mínimo 6 puntos, máximo 24
✅ Para arañazos/grietas lineales: forma muy estrecha (ancho < 30px de 1000)
✅ Para abolladuras: forma IRREGULAR asimétrica — NUNCA un oval/círculo perfecto
✅ El polígono debe AJUSTARSE al daño, no a un área rectangular o circular genérica
❌ NUNCA extender el polígono sobre ruedas, cristales o paneles sin daño adyacentes
❌ NO usar oval perfecto para abolladuras — los golpes reales son asimétricos
✅ Si no puedes determinar el contorno exacto → usa el box_2d como base, pero irregular
✅ polygon_points y box_2d deben ser consistentes (box debe contener al polígono)

=== QUÉ NO ES UN DAÑO (catálogo de falsos positivos — repásalo ante CADA marca) ===
NUNCA reportes como daño:
- REFLEJOS: árboles, farolas, edificios o cielo reflejados en la chapa — franjas oscuras
  alargadas que siguen la curvatura del panel. PRUEBA FÍSICA OBLIGATORIA: un daño real
  permanece en el MISMO punto del panel en TODAS las fotos donde ese panel aparece;
  un reflejo cambia de posición o desaparece entre ángulos. Si el panel sale en 2+ fotos
  y la marca solo existe en una → es reflejo/luz: NO lo reportes.
- GOTAS DE AGUA y churretes de lluvia seca (patrón vertical repetido a lo largo del panel)
- BARRO Y SUCIEDAD (aspecto mate superpuesto, sin deformación de chapa ni pérdida de pintura)
- UNIONES DE CHAPA Y JUNTAS entre paneles (líneas rectas de fábrica, bordes perfectamente paralelos)
- LÍNEAS DE DISEÑO y nervios de carrocería (pliegues simétricos de fábrica; si la misma
  línea existe en el lado opuesto del vehículo, es diseño, no daño)
- MOLDURAS, EMBELLECEDORES y protectores de plástico negro (son piezas, no daños)
- PEGATINAS, LOGOS, ROTULACIÓN de empresa y restos de adhesivo
- SOMBRAS proyectadas (barandillas, árboles, el propio fotógrafo)
- MANCHAS de combustible/grasa junto al tapón
Ante la duda entre cualquiera de estos y un daño real: confidence < 0.65 y márcalo sugerido.

=== SUCIEDAD ===
"dirt_level": 0 (impecable) → 10 (barro total).
dirt_level ≥ 6: sé CONSERVADOR. Barro, polvo, agua seca NO son daños. Solo reporta
deformaciones físicas, roturas o pérdida de pintura evidentes. Baja confidence en marcas ambiguas.

=== BAREMO DE SEVERIDAD (criterios de peritaje — aplícalo a CADA daño) ===
- LEVE: rayón superficial sin deformación (< 15 cm), roce en paragolpes, desconchón puntual
  de pintura. Solo estética; se pule o retoca.
- MODERADO: rayón largo o profundo con pérdida de pintura clara, abolladura pequeña
  (< 10 cm) sin pliegues, moldura/embellecedor suelto o rajado. Requiere chapa-pintura del panel.
- GRAVE: abolladura grande o con pliegues de chapa, panel deformado, óptica/piloto ROTO,
  luna agrietada, paragolpes rajado o descolgado. Sustitución o reparación estructural del panel.
- CRÍTICO: afecta a la SEGURIDAD o a la circulación: luna del conductor rota, óptica delantera
  inservible, rueda/neumático dañado, puerta que no cierra, elemento colgando que puede
  desprenderse, deformación que invade el paso de rueda.
Regla: ante la duda entre dos niveles, elige el MENOR y explica en description por qué.
La severidad global ("severity") = la del PEOR daño individual, nunca mayor.

=== CONTROL DE CALIDAD FINAL (obligatorio ANTES de responder) ===
Repasa tu lista de daños UNA POR UNA y elimina las entradas que no superen TODAS estas pruebas:
1. PRUEBA FÍSICA: si el panel aparece en 2+ fotos, ¿la marca está en el MISMO punto en todas?
   Si solo existe en una foto → reflejo/luz → FUERA.
2. CATÁLOGO: ¿podría ser alguno de los falsos positivos listados (reflejo, gota, junta,
   nervio de diseño, moldura, pegatina, sombra, barro)? Si no puedes descartarlo → FUERA
   o confidence < 0.65 como sugerido.
3. EVIDENCIA: ¿puedes señalar el daño con un box preciso y describir su forma y tamaño
   aproximado en cm? Si no → FUERA.
4. BAREMO: ¿la severidad asignada cumple el baremo de arriba? Ajústala si no.
Es MEJOR devolver 1 daño cierto que 5 dudosos: cada falso positivo cuesta tiempo de revisión
y credibilidad. Un vehículo limpio y sin daños con "sin_danos" es una respuesta perfectamente válida.

Responde ÚNICAMENTE con este JSON exacto (sin markdown, sin bloques de código, sin texto extra):
{
  "severity": "sin_danos|leve|moderado|grave|critico",
  "dirt_level": 0,
  "urgency": "puede_esperar|esta_semana|urgente|inmediato",
  "risk": "bajo|medio|alto|critico",
  "circulation_safe": true,
  "detected_plate": "",
  "fraud_warnings": [],
  "critical_damages_count": 0,
  "total_damages_count": 0,
  "new_damages_count": 0,
  "hidden_damage_probability": 0.0,
  "total_estimated_cost": 0.0,
  "confidence": 0.85,
  "executive_summary": "Descripción ejecutiva clara del estado del vehículo para un no-experto",
  "image_quality_warnings": [],
  "affected_parts": [],
  "critical_damages": [],
  "new_damages": [
    {
      "part": "nombre exacto de la pieza dañada",
      "severity": "leve|moderado|grave|critico",
      "description": "descripción técnica precisa con ubicación exacta (ej: esquina inferior izquierda del paragolpes trasero, grieta horizontal de ~15cm)",
      "location_hint": "frontal|trasera|lateral_izquierdo|lateral_derecho|techo|otra",
      "photo_index": 1,
      "box_2d": [ymin, xmin, ymax, xmax],
      "polygon_points": [[y1,x1],[y2,x2],[y3,x3]],
      "estimated_cost": 0.0,
      "confidence": 0.9
    }
  ],
  "damages": [
    {
      "part": "nombre exacto de la pieza dañada",
      "severity": "leve|moderado|grave|critico",
      "description": "descripción técnica precisa con ubicación exacta",
      "location_hint": "frontal|trasera|lateral_izquierdo|lateral_derecho|techo|otra",
      "photo_index": 1,
      "box_2d": [ymin, xmin, ymax, xmax],
      "polygon_points": [[y1,x1],[y2,x2],[y3,x3]],
      "repair_suggestion": "acción concreta de reparación (ej: sustitución de paragolpes, pulido y laqueado)",
      "estimated_cost": 0.0,
      "confidence": 0.9,
      "is_new": true
    }
  ]
}

REGLAS FINALES:
- Sin daños: severity=sin_danos, damages=[], new_damages=[], total_damages_count=0
- UN daño físico = UNA entrada en damages[], aunque se vea en varias fotos (escoge la foto donde mejor se aprecia)
- estimated_cost en euros, mercado español taller 2026 (sin pintura: 50-200€; con pintura: 200-800€; pieza nueva: 300-2000€)
- SIEMPRE intenta leer detected_plate
- NO dupliques daños por verlos en múltiples fotos"""


# =========================
# MOTOR IA v2 — ETAPA 0: ORIENTACIÓN VERIFICADA + COHERENCIA ESPACIAL
# La confusión izquierda/derecha nacía de adivinar la orientación DENTRO de la
# llamada gigante de análisis. Solución estructural: una llamada dedicada mono-
# tarea clasifica la vista de cada foto, y un validador determinista en Python
# hace cumplir esa geometría sobre los daños devueltos.
# =========================

_VIEW_LABELS = (
    "frontal", "trasera", "lateral_izquierdo", "lateral_derecho",
    "tres_cuartos_frontal_izquierdo", "tres_cuartos_frontal_derecho",
    "tres_cuartos_trasero_izquierdo", "tres_cuartos_trasero_derecho",
    "interior", "otra",
)

_ORIENTATION_PROMPT = """Tarea ÚNICA: clasificar qué vista de una furgoneta muestra cada foto.
Usa PISTAS FÍSICAS verificables, no suposiciones:
- Retrovisores: su lado indica el lado del vehículo
- Manetas de puerta y riel de la puerta corredera
- Matrícula delantera (con parabrisas/rejilla) vs trasera (con portón/pilotos)
- Texto y logos: se leen del derecho en el lado correcto de la rotulación
- Volante visible tras el cristal (España: volante a la IZQUIERDA del vehículo)
- Tapón de combustible, tubo de escape, forma de los faros vs pilotos

IZQUIERDA/DERECHA siempre desde el punto de vista del CONDUCTOR sentado dentro
(NO del fotógrafo). Si dudas entre lateral puro y 3/4, elige la vista 3/4 del lado que corresponda.

Responde SOLO JSON. El campo confidence es OBLIGATORIO y debe reflejar tu certeza real
(ej: 0.95 si las pistas son inequívocas, 0.6 si dudas entre dos vistas — nunca 0):
{"views":[{"photo":1,"view":"<etiqueta>","confidence":0.9,"clue":"pista física concreta usada"}]}
Etiquetas válidas: frontal | trasera | lateral_izquierdo | lateral_derecho | tres_cuartos_frontal_izquierdo | tres_cuartos_frontal_derecho | tres_cuartos_trasero_izquierdo | tres_cuartos_trasero_derecho | interior | otra"""


async def _classify_photo_orientations(client, genai_types, model_name, images_base64):
    """ETAPA 0: clasificador de orientación dedicado (mono-tarea = fiable).
    Devuelve {photo_index_1based: {"view", "confidence", "clue"}} o None si falla
    (en cuyo caso el análisis continúa como antes — nunca bloquea)."""
    # Cuota diaria justa (free tier): la llamada extra de Etapa0 se sacrifica
    # para reservar el cupo al análisis principal. Nunca bloquea: sin
    # orientaciones, el análisis continúa sin restricciones, como siempre.
    if _is_daily_exhausted(model_name):
        logger.info("[Etapa0] Saltado: cuota diaria de Gemini justa, se reserva para el análisis")
        return None
    try:
        contents = [_ORIENTATION_PROMPT]
        for i, img_b64 in enumerate(images_base64):
            contents.append(genai_types.Part.from_bytes(
                data=base64.b64decode(img_b64), mime_type="image/jpeg"))
            contents.append(f"Imagen {i+1}")
        cfg = genai_types.GenerateContentConfig(temperature=0.0, response_mime_type="application/json")
        loop = asyncio.get_running_loop()
        resp = await asyncio.wait_for(
            loop.run_in_executor(_executor, lambda: client.models.generate_content(
                model=model_name, contents=contents, config=cfg)),
            timeout=45.0,
        )
        data = json.loads(_strip_markdown_json(resp.text or ""))
        out = {}
        for v in data.get("views", []):
            try:
                idx = int(v.get("photo", 0))
                view = str(v.get("view", "")).strip()
                if 1 <= idx <= len(images_base64) and view in _VIEW_LABELS:
                    conf = float(v.get("confidence", 0) or 0)
                    # Gemini a veces no rellena confidence (llega 0.0). Si eligió
                    # una etiqueta válida, la clasificación en sí es la señal:
                    # default 0.8 para que el validador de coherencia pueda actuar.
                    if conf <= 0:
                        conf = 0.8
                    out[idx] = {"view": view,
                                "confidence": conf,
                                "clue": str(v.get("clue", ""))[:120]}
            except Exception:
                continue
        if out:
            logger.info("[Etapa0] Orientaciones: " + ", ".join(
                f"img{i}={out[i]['view']}({out[i]['confidence']:.2f})" for i in sorted(out)))
        return out or None
    except Exception as e:
        if _is_perday_429(str(e)):
            _mark_daily_exhausted(model_name)
        logger.warning(f"[Etapa0] Clasificador de orientación falló (se continúa sin restricciones): {e}")
        return None


def _swap_side_words(text: str, to_side: str) -> str:
    """Corrige el lado en el nombre de una pieza (derecho→izquierdo o viceversa)."""
    if to_side == "izq":
        text = re.sub(r"derecha", "izquierda", text, flags=re.I)
        text = re.sub(r"derecho", "izquierdo", text, flags=re.I)
        text = re.sub(r"\bder\b\.?", "izq", text, flags=re.I)
    else:
        text = re.sub(r"izquierda", "derecha", text, flags=re.I)
        text = re.sub(r"izquierdo", "derecho", text, flags=re.I)
        text = re.sub(r"\bizq\b\.?", "der", text, flags=re.I)
    return text


def _enforce_spatial_coherence(damages: list, photo_views: dict) -> int:
    """Validador determinista post-IA: la orientación VERIFICADA de la foto manda.
    - Foto verificada como lado izquierdo → una pieza 'derecha' en esa foto es
      imposible: se corrige el lado (y viceversa).
    - location_hint se alinea con la vista verificada.
    - Combos físicamente imposibles (capó en vista trasera pura, portón en
      frontal pura) → el daño se degrada a 'sugerido' en vez de eliminarse
      (el humano decide en Revisión Rápida; cero daños reales perdidos).
    Solo actúa con confianza de orientación ≥ 0.75. Devuelve nº de correcciones."""
    if not photo_views:
        return 0
    fixes = 0
    hint_by_view = {"frontal": "frontal", "trasera": "trasera",
                    "lateral_izquierdo": "lateral_izquierdo",
                    "lateral_derecho": "lateral_derecho"}
    for d in damages:
        pv = photo_views.get(getattr(d, "photo_index", None) or -1)
        if not pv or pv.get("confidence", 0) < 0.75:
            continue
        view = pv["view"]
        side = "izq" if "izquierdo" in view else ("der" if "derecho" in view else None)
        part_l = (d.part or "").lower()

        # 1) El lado de la pieza no puede contradecir el lado verificado de la foto
        if side == "izq" and "derech" in part_l:
            d.part = _swap_side_words(d.part, "izq")
            fixes += 1
        elif side == "der" and "izquierd" in part_l:
            d.part = _swap_side_words(d.part, "der")
            fixes += 1

        # 2) location_hint alineado con la vista verificada
        if view in hint_by_view and d.location_hint and d.location_hint != hint_by_view[view]:
            d.location_hint = hint_by_view[view]
            fixes += 1
        elif side and view.startswith("tres_cuartos"):
            want = "lateral_izquierdo" if side == "izq" else "lateral_derecho"
            if d.location_hint in ("lateral_izquierdo", "lateral_derecho") and d.location_hint != want:
                d.location_hint = want
                fixes += 1

        # 3) Combos físicamente imposibles (solo vistas puras, conservador)
        impossible = (
            (view == "trasera" and any(w in part_l for w in ("capó", "capo", "parabrisas", "paragolpes delantero"))) or
            (view == "frontal" and any(w in part_l for w in ("portón", "porton", "paragolpes trasero", "piloto tras", "luz trasera")))
        )
        if impossible and getattr(d, "confirmed", True):
            d.confirmed = False
            d.confidence = min(d.confidence or 0.5, 0.5)
            fixes += 1
    return fixes


async def _get_vehicle_ledger(vehicle_id: str, exclude_inspection_id: str = None) -> list:
    """REGISTRO DE DAÑOS DEL VEHÍCULO (ledger): un daño se registra la primera
    vez que se ve y NUNCA vuelve a contar como nuevo hasta que se repare.

    Backfill perezoso: si el vehículo aún no tiene ledger, se construye desde
    su historial de inspecciones (6 meses) — así funciona desde el minuto uno
    para toda la flota existente, sin migración global."""
    entries = await db.vehicle_damage_ledger.find(
        {"vehicle_id": vehicle_id, "status": "open"}, {"_id": 0}).to_list(200)
    if entries:
        return [e for e in entries if e.get("first_seen_inspection") != exclude_inspection_id]

    # Backfill desde el historial (excluida la inspección en curso si reanaliza)
    q = {"deleted": {"$ne": True}, "vehicle_id": vehicle_id,
         "analysis_status": "ok", "analysis": {"$ne": None},
         "created_at": {"$gte": (datetime.now(timezone.utc) - timedelta(days=180)).isoformat()}}
    prior = await db.inspections.find(
        q, {"_id": 0, "id": 1, "created_at": 1, "analysis.damages.part": 1,
            "analysis.damages.severity": 1, "analysis.damages.confirmed": 1}
    ).sort("created_at", 1).to_list(400)
    panels: dict = {}
    for insp in prior:
        if insp.get("id") == exclude_inspection_id:
            continue
        for d in ((insp.get("analysis") or {}).get("damages") or []):
            if not isinstance(d, dict) or d.get("confirmed") is False:
                continue
            p = _canon_panel(d.get("part") or "")
            if not p:
                continue
            rank = _SEV_RANK.get(_norm_sev(d.get("severity")), 1)
            e = panels.get(p)
            if not e:
                panels[p] = {"vehicle_id": vehicle_id, "panel": p, "part": d.get("part"),
                             "severity": _norm_sev(d.get("severity")), "rank": rank,
                             "status": "open", "source": "ai",
                             "first_seen": (insp.get("created_at") or "")[:10],
                             "first_seen_inspection": insp.get("id"),
                             "updated_at": datetime.now(timezone.utc).isoformat()}
            elif rank > e["rank"]:
                e["rank"], e["severity"] = rank, _norm_sev(d.get("severity"))
    if panels:
        try:
            await db.vehicle_damage_ledger.insert_many(list(panels.values()))
        except Exception as _be:
            logger.debug(f"ledger backfill {vehicle_id}: {_be}")
    return list(panels.values())


async def _known_damages_prompt(vehicle_id: str, exclude_inspection_id: str = None) -> str:
    """Bloque textual para el prompt: los daños YA REGISTRADOS del vehículo.
    Cierra el agujero de la comparación visual (si la foto de referencia no
    muestra bien un panel, Gemini re-reportaba el daño viejo como nuevo)."""
    try:
        entries = await _get_vehicle_ledger(vehicle_id, exclude_inspection_id)
        if not entries:
            return ""
        lines = ["\n=== DAÑOS YA REGISTRADOS DE ESTE VEHÍCULO (histórico del sistema) ==="]
        for e in entries[:25]:
            lines.append(f"- {e.get('part') or e.get('panel')}: severidad {e.get('severity')}, "
                         f"registrado desde {e.get('first_seen')}")
        lines.append(
            "REGLA: estos daños YA EXISTEN y están documentados. Si los ves, inclúyelos en "
            "damages[] con is_new=false, pero NUNCA en new_damages[] — salvo que veas en ese "
            "panel un daño CLARAMENTE distinto o de severidad claramente mayor (otra zona del "
            "panel, otro tipo de daño, tamaño mucho mayor).")
        return "\n".join(lines)
    except Exception as e:
        logger.debug(f"known_damages_prompt: {e}")
        return ""


async def _apply_vehicle_memory(vehicle_id: str, analysis, inspection_id: str = None) -> None:
    """MEMORIA DEL VEHÍCULO tras cada análisis. Tres garantías:

    1) LEDGER: un daño en un panel ya registrado (misma severidad o menor)
       sale de new_damages — no re-alerta, no re-cuenta en € ni en scoring.
       Da igual que nadie lo haya validado aún: registrado una vez = conocido.
    2) Un daño en un panel NUEVO (o claramente más grave) se queda como nuevo
       Y se registra en el ledger → mañana ya no volverá a ser "nuevo".
    3) ✗ humanos (falsos positivos previos): detección floja en ese panel se
       degrada a sugerido (el humano decide en Revisión Rápida).

    La reparación (repair_status=done) limpia el panel del ledger → un golpe
    posterior ahí vuelve a contar como nuevo, que es lo justo para el scoring.
    No escribe NADA en ai_feedback (el dataset solo crece con humanos)."""
    try:
        if not vehicle_id or not analysis:
            return
        if not (getattr(analysis, "new_damages", None) or getattr(analysis, "damages", None)):
            return
        ledger = await _get_vehicle_ledger(vehicle_id, inspection_id)
        known = {e["panel"]: e for e in ledger}

        bad_panels: set = set()
        try:
            fb = await db.ai_feedback.find(
                {"vehicle_id": vehicle_id, "verdict": "wrong"},
                {"_id": 0, "damage.part": 1}).to_list(300)
            bad_panels = {_canon_panel((f.get("damage") or {}).get("part") or "") for f in fb}
            bad_panels.discard("")
        except Exception:
            pass

        now_iso = datetime.now(timezone.utc).isoformat()
        kept, moved, downgraded, registered = [], 0, 0, 0
        for d in analysis.new_damages:
            p = _canon_panel(d.part or "")
            d_rank = _SEV_RANK.get(_norm_sev(d.severity), 1)

            if p and p in known and d_rank <= known[p]["rank"]:
                # YA REGISTRADO → conocido, fuera de nuevos (sigue en damages[])
                d.is_new = False
                d.description = ((d.description or "") +
                                 f" · [ya registrado desde {known[p].get('first_seen')}]")
                moved += 1
                continue

            if (p and p in bad_panels and getattr(d, "confirmed", True)
                    and (d.confidence or 0) < 0.85
                    and _norm_sev(d.severity) in ("leve", "moderado")):
                d.confirmed = False
                d.description = (d.description or "") + " · [patrón rechazado ✗ previamente — revisar]"
                downgraded += 1

            # Daño genuinamente nuevo (o escalada) → REGISTRAR en el ledger
            if p and getattr(d, "confirmed", True):
                await db.vehicle_damage_ledger.update_one(
                    {"vehicle_id": vehicle_id, "panel": p, "status": "open"},
                    {"$set": {"part": d.part, "severity": _norm_sev(d.severity),
                              "rank": max(d_rank, known.get(p, {}).get("rank", 0)),
                              "updated_at": now_iso},
                     "$setOnInsert": {"vehicle_id": vehicle_id, "panel": p, "status": "open",
                                      "source": "ai", "first_seen": now_iso[:10],
                                      "first_seen_inspection": inspection_id or ""}},
                    upsert=True)
                registered += 1
            kept.append(d)

        if moved or downgraded:
            analysis.new_damages = kept
            if hasattr(analysis, "new_damages_count"):
                analysis.new_damages_count = len(kept)

        # damages[] (todos los visibles): marca también los ya registrados.
        # is_new=True es el default del modelo y Gemini no siempre lo pone bien;
        # con esto Revisión Rápida puede fiarse de is_new para no re-validar.
        marked = 0
        for d in (getattr(analysis, "damages", None) or []):
            p = _canon_panel(d.part or "")
            d_rank = _SEV_RANK.get(_norm_sev(d.severity), 1)
            if p and p in known and d_rank <= known[p]["rank"] and getattr(d, "is_new", True):
                d.is_new = False
                if "[ya registrado" not in (d.description or ""):
                    d.description = ((d.description or "") +
                                     f" · [ya registrado desde {known[p].get('first_seen')}]")
                marked += 1

        if moved or downgraded or registered or marked:
            logger.info(f"[Ledger] {vehicle_id}: {moved} ya registrados fuera de nuevos, "
                        f"{registered} registrados, {downgraded} degradados por ✗, "
                        f"{marked} marcados conocidos en damages[]")
    except Exception as e:
        logger.debug(f"[MemoriaVehículo] {e}")


# =========================
# HELPERS
# =========================

def serialize_doc(doc: dict) -> dict:
    """Serializa recursivamente datetimes a ISO string para MongoDB."""
    if isinstance(doc, dict):
        return {k: serialize_doc(v) for k, v in doc.items()}
    elif isinstance(doc, list):
        return [serialize_doc(i) for i in doc]
    elif isinstance(doc, datetime):
        return doc.isoformat()
    return doc


def _remap_photo_indexes(analysis, idx_map: dict) -> None:
    """Traduce photo_index del subconjunto analizado al índice real en photos.
    idx_map: {índice_en_análisis (1-based) → índice_en_photos (1-based)}."""
    if not idx_map:
        return
    for lst in ((getattr(analysis, "damages", None) or []),
                (getattr(analysis, "new_damages", None) or [])):
        for d in lst:
            pi = getattr(d, "photo_index", None)
            if pi:
                d.photo_index = idx_map.get(pi, pi)


def _iou_1000(a: list, b: list) -> float:
    """IoU de dos cajas [ymin,xmin,ymax,xmax] en coords 0-1000."""
    try:
        iy1, ix1 = max(a[0], b[0]), max(a[1], b[1])
        iy2, ix2 = min(a[2], b[2]), min(a[3], b[3])
        inter = max(0.0, iy2 - iy1) * max(0.0, ix2 - ix1)
        if inter <= 0:
            return 0.0
        area_a = max(0.0, a[2] - a[0]) * max(0.0, a[3] - a[1])
        area_b = max(0.0, b[2] - b[0]) * max(0.0, b[3] - b[1])
        union = area_a + area_b - inter
        return inter / union if union > 0 else 0.0
    except Exception:
        return 0.0


def _snap_damage_boxes_to_yolo(damages: list, yolo_by_photo: dict) -> int:
    """Corrige las cajas del LLM con las del detector CV (QW2 del plan IA):
    el LLM acierta QUÉ es el daño pero localiza mal; YOLO localiza bien.
    Si una detección YOLO solapa (IoU ≥ 0.25) o su centro cae dentro de la
    caja del LLM (caso "caja gigante"), se adopta la caja YOLO y se descarta
    el polígono del LLM (ya no corresponde). Devuelve nº de cajas corregidas."""
    snapped = 0
    for d in damages or []:
        pi = getattr(d, "photo_index", None)
        box = getattr(d, "box_2d", None)
        if not pi or not box or len(box) != 4:
            continue
        best, best_score = None, 0.0
        for det in yolo_by_photo.get(pi) or []:
            db_ = getattr(det, "box_2d", None)
            if not db_ or len(db_) != 4:
                continue
            iou = _iou_1000(box, db_)
            cy, cx = (db_[0] + db_[2]) / 2, (db_[1] + db_[3]) / 2
            center_in = box[0] <= cy <= box[2] and box[1] <= cx <= box[3]
            score = max(iou, 0.26 if (center_in and iou > 0.02) else 0.0)
            if score > best_score:
                best_score, best = score, det
        if best is not None and best_score >= 0.25:
            d.box_2d = [int(round(v)) for v in best.box_2d]
            # Contorno fino del modelo de segmentación si lo trae; si no, se
            # descarta el polígono del LLM (ya no corresponde a la caja nueva).
            poly = getattr(best, "polygon_2d", None)
            d.polygon_points = ([[int(round(y)), int(round(x))] for y, x in poly]
                                if poly and len(poly) >= 3 else None)
            panel = getattr(best, "panel", None)
            if panel:
                d.panel_cv = panel
            snapped += 1
    return snapped


def _user_friendly_error(reason: str) -> str:
    """Convierte errores tecnicos de Gemini en mensajes claros para el usuario."""
    r = reason.lower()
    if "depleted" in r or "prepay" in r:
        return "Creditos de IA agotados. Recarga en AI Studio (ai.studio/projects); las inspecciones se reanalizaran solas."
    if "429" in r or "quota" in r or "resource_exhausted" in r:
        return "Analisis IA temporalmente no disponible (limite de uso alcanzado). Se reintentara automaticamente."
    if "404" in r or "not found" in r:
        return "Analisis IA en configuracion. Contacta con soporte si persiste."
    if "timeout" in r or "tardo demasiado" in r:
        return "El analisis tardo demasiado. Vuelve a intentarlo en unos minutos."
    if "api_key" in r or "api key" in r or "permission" in r or "401" in r or "403" in r:
        return "Analisis IA no disponible por configuracion de la cuenta."
    if "vacia" in r or "empty" in r or "json" in r:
        return "El analisis no se pudo procesar. Vuelve a intentarlo."
    return "Analisis IA temporalmente no disponible. Vuelve a intentarlo mas tarde."


def _fallback_analysis(reason: str = "Gemini no disponible") -> InspectionAnalysis:
    logger.warning(f"Usando fallback analysis: {reason}")
    # Config rota (billing/permisos de Gemini): avisar por Telegram con dedupe 1h.
    # Sin esto, las inspecciones se quedan "sin análisis" en silencio hasta que
    # alguien lo ve en el panel (pasó el 2026-07-02 con BILLING_DISABLED).
    rl = reason.lower()
    if any(s in rl for s in ("billing", "permission", "api_key", "api key", "401", "403",
                             "depleted", "prepay")):
        try:
            asyncio.get_running_loop().create_task(_notify_error_once(
                "backend",
                "Análisis IA caído: Gemini rechaza las llamadas (créditos agotados, billing o permisos)",
                reason[:400],
            ))
        except RuntimeError:
            pass  # sin event loop (contexto síncrono): queda en el log
    return InspectionAnalysis(
        critical_damages_count=0,
        total_damages_count=0,
        severity="sin_analisis",
        urgency="puede_esperar",
        risk="bajo",
        circulation_safe=True,
        hidden_damage_probability=0,
        total_estimated_cost=0,
        confidence=0,
        executive_summary=_user_friendly_error(reason),
        image_quality_warnings=[],
        affected_parts=[],
        critical_damages=[],
        damages=[]
    )


def _dedup_damages(damages: list) -> tuple[list, int]:
    """De-duplica daños que Gemini reporta varias veces por verse en varias fotos.
    Dos daños se consideran el mismo si coinciden pieza normalizada + location_hint
    y sus descripciones comparten la mayoría de palabras significativas.
    Conservador: ante la duda, NO fusiona. Retorna (lista_dedup, n_eliminados)."""
    import unicodedata

    def _norm(s):
        s = unicodedata.normalize("NFD", (s or "").lower())
        return "".join(c for c in s if unicodedata.category(c) != "Mn").strip()

    def _words(s):
        stop = {"de", "del", "la", "el", "en", "con", "un", "una", "y", "lado", "zona", "parte"}
        return set(w for w in _norm(s).split() if len(w) > 2 and w not in stop)

    SEV_RANK = {"leve": 1, "moderado": 2, "grave": 3, "critico": 4}

    kept = []
    removed = 0
    for d in damages:
        part = _norm(getattr(d, "part", ""))
        loc = _norm(getattr(d, "location_hint", ""))
        desc_w = _words(getattr(d, "description", ""))
        duplicate_of = None
        for k in kept:
            if _norm(getattr(k, "part", "")) != part:
                continue
            if _norm(getattr(k, "location_hint", "")) != loc:
                continue
            k_w = _words(getattr(k, "description", ""))
            if not desc_w or not k_w:
                # misma pieza+ubicación sin descripción comparable → duplicado
                duplicate_of = k
                break
            overlap = len(desc_w & k_w) / min(len(desc_w), len(k_w))
            if overlap >= 0.5:
                duplicate_of = k
                break
        if duplicate_of is not None:
            removed += 1
            # Conservar la versión más severa / más cara
            if SEV_RANK.get(_norm(getattr(d, "severity", "")), 0) > SEV_RANK.get(_norm(getattr(duplicate_of, "severity", "")), 0):
                duplicate_of.severity = d.severity
            try:
                if (d.estimated_cost or 0) > (duplicate_of.estimated_cost or 0):
                    duplicate_of.estimated_cost = d.estimated_cost
            except Exception:
                pass
        else:
            kept.append(d)
    return kept, removed


def _strip_markdown_json(raw: str) -> str:
    raw = raw.strip()
    if raw.startswith("```"):
        lines = raw.split("\n")
        lines = lines[1:]
        if lines and lines[-1].strip() == "```":
            lines = lines[:-1]
        if lines and lines[0].strip().lower() == "json":
            lines = lines[1:]
        raw = "\n".join(lines).strip()
    # Robustez extra: extraer el objeto JSON entre el primer { y el ultimo }
    if not raw.startswith("{"):
        start = raw.find("{")
        if start != -1:
            raw = raw[start:]
    if not raw.endswith("}"):
        end = raw.rfind("}")
        if end != -1:
            raw = raw[:end + 1]
    return raw.strip()


async def analyze_images_with_gemini(
    images_base64: List[str],
    reference_images_bytes: Optional[List[bytes]] = None,
    db=None,
    known_damages_text: str = "",
    cv_detections_text: str = "",
) -> tuple[InspectionAnalysis, str, Optional[str]]:
    """
    Retorna (analysis, status, error_message).
    status: "ok" | "gemini_failed" | "gemini_timeout"
    """
    gemini_key = os.environ.get("GEMINI_API_KEY", "")
    _use_vertex = os.environ.get("USE_VERTEX_AI", "").lower() in ("1", "true", "yes")
    if not gemini_key and not _use_vertex:
        return _fallback_analysis("GEMINI_API_KEY no configurada"), "gemini_failed", "GEMINI_API_KEY no configurada"

    try:
        from google import genai as genai_sdk
        from google.genai import types as genai_types

        model_name = os.environ.get("GEMINI_MODEL", "gemini-2.5-flash")

        # MODO VERTEX AI: si esta configurado el proyecto, usar cuenta de servicio
        # (evita el problema de las API keys formato AQ. que no funcionan con generateContent)
        use_vertex = os.environ.get("USE_VERTEX_AI", "").lower() in ("1", "true", "yes")
        if use_vertex:
            vertex_project = os.environ.get("GCP_PROJECT", "")
            vertex_location = os.environ.get("GCP_LOCATION", "us-central1")
            sa_json = os.environ.get("GCP_SERVICE_ACCOUNT_JSON", "")
            if sa_json:
                # Credenciales desde el JSON de la cuenta de servicio (en variable de entorno)
                from google.oauth2 import service_account
                import json as _json
                import base64 as _b64
                # Aceptar JSON crudo O base64 (mas robusto frente a escapes de PowerShell)
                sa_clean = sa_json.strip()
                if not sa_clean.startswith("{"):
                    # Asumimos base64
                    try:
                        sa_clean = _b64.b64decode(sa_clean).decode("utf-8")
                    except Exception as _be:
                        logger.error(f"No se pudo decodificar GCP_SERVICE_ACCOUNT_JSON como base64: {_be}")
                creds_info = _json.loads(sa_clean)
                credentials = service_account.Credentials.from_service_account_info(
                    creds_info, scopes=["https://www.googleapis.com/auth/cloud-platform"]
                )
                client = genai_sdk.Client(
                    vertexai=True, project=vertex_project,
                    location=vertex_location, credentials=credentials
                )
            else:
                client = genai_sdk.Client(
                    vertexai=True, project=vertex_project, location=vertex_location
                )
            logger.info(f"Gemini via Vertex AI (project={vertex_project}, location={vertex_location})")
        else:
            client = genai_sdk.Client(api_key=gemini_key)
            logger.info("Gemini via AI Studio (api key)")

        contents = [GEMINI_SYSTEM_PROMPT]

        # ── FEW-SHOT LEARNING: inyectar errores anteriores antes de las fotos ──
        if db is not None:
            # Patrones agregados de toda la flota (piezas con falsos positivos
            # recurrentes / daños que se le escapan). Barato: cacheado 10 min.
            try:
                pattern_lessons = await get_pattern_lessons(db)
                if pattern_lessons:
                    contents.append(pattern_lessons)
            except Exception as _pl:
                logger.debug(f"[Learning] Error cargando patrones: {_pl}")
            try:
                # 4 ejemplos (antes 2): con 1300+ correcciones humanas etiquetadas,
                # más señal propia por petición a coste de peticiones CERO.
                general_examples = await get_few_shot_examples(db, location_hint="", part="", limit=4, general=True)
                few_shot_parts = build_few_shot_prompt_parts_multimodal(general_examples)
                for part_dict in few_shot_parts:
                    if "inline_data" in part_dict:
                        contents.append(
                            genai_types.Part.from_bytes(
                                data=base64.b64decode(part_dict["inline_data"]["data"]),
                                mime_type=part_dict["inline_data"]["mime_type"],
                            )
                        )
                    else:
                        contents.append(part_dict["text"])
            except Exception as _fse:
                logger.debug(f"[Learning] Error cargando ejemplos generales: {_fse}")

        logger.info(f"Enviando {len(images_base64)} imágenes a Gemini ({model_name}) [SDK google-genai]")

        # ── REGISTRO DE DAÑOS DEL VEHÍCULO: lo ya documentado no es "nuevo" ──
        if known_damages_text:
            contents.append(known_damages_text)

        # ── DETECCIONES DEL CV PROPIO: Gemini confirma y describe, no inventa ──
        if cv_detections_text:
            contents.append(cv_detections_text)

        # ── ETAPA 0: orientación verificada por clasificador dedicado ──
        # Si falla, photo_views=None y todo funciona exactamente como antes.
        photo_views = await _classify_photo_orientations(client, genai_types, model_name, images_base64)
        if photo_views:
            _ori_lines = ["\n=== ORIENTACIÓN VERIFICADA DE CADA FOTO (Etapa 0 — clasificador dedicado) ==="]
            for _idx in sorted(photo_views):
                _pv = photo_views[_idx]
                _ori_lines.append(
                    f"Imagen {_idx}: {_pv['view'].upper()} (confianza {_pv['confidence']:.2f}; pista: {_pv['clue']})")
            _ori_lines.append(
                "REGLA DURA: el location_hint y el LADO de cada pieza dañada DEBEN ser coherentes con la "
                "orientación verificada de la foto donde se reporta. En una foto LATERAL_IZQUIERDO solo "
                "existen piezas del lado izquierdo; en LATERAL_DERECHO solo del derecho. Esta clasificación "
                "tiene PRIORIDAD sobre la numeración probable de las fotos.")
            contents.append("\n".join(_ori_lines))

        for i, img_b64 in enumerate(images_base64):
            img_data = base64.b64decode(img_b64)
            contents.append(genai_types.Part.from_bytes(data=img_data, mime_type="image/jpeg"))
            contents.append(f"Imagen actual {i+1} de {len(images_base64)}")

        if reference_images_bytes:
            contents.append(
                "\n=== IMÁGENES DE REFERENCIA (estado anterior del mismo vehículo) ===\n"
                "Compara con las imágenes actuales. Reporta SOLO los daños NUEVOS.\n"
                "Si un daño aparece igual en referencia y en actual, NO lo incluyas en damages[].\n"
            )
            for i, ref_bytes in enumerate(reference_images_bytes):
                try:
                    contents.append(genai_types.Part.from_bytes(data=ref_bytes, mime_type="image/jpeg"))
                    contents.append(f"Imagen de referencia {i+1}")
                except Exception as e:
                    logger.warning(f"Error añadiendo referencia {i}: {e}")

        # temperature=0 + seed fija: la misma inspección debe dar el mismo
        # resultado (reproducibilidad, QW4 del plan IA). El CV ya es determinista.
        gen_config = genai_types.GenerateContentConfig(
            temperature=0.0, seed=7, response_mime_type="application/json")
        loop = asyncio.get_running_loop()

        # Modelos de fallback si el principal da 429. Los alias *-latest los
        # mantiene Google apuntando a la versión vigente (los 1.5 murieron en
        # 2026 devolviendo 404 y dejaban todo "sin análisis").
        fallback_models = ["gemini-2.5-flash", "gemini-flash-latest", "gemini-flash-lite-latest", "gemini-2.5-flash-lite"]
        models_to_try = [model_name] + [m for m in fallback_models if m != model_name]

        response = None
        last_err = None

        # Semáforo: máximo 2 llamadas Gemini simultáneas para evitar 429
        async with _gemini_sem:
            for model_attempt, current_model in enumerate(models_to_try):
                # Cupo diario agotado hace poco: ni lo intentamos (ahorra peticiones).
                if _is_daily_exhausted(current_model):
                    logger.info(f"Gemini {current_model} con día agotado: saltado")
                    continue
                for retry in range(3):
                    try:
                        response = await asyncio.wait_for(
                            loop.run_in_executor(
                                _executor,
                                lambda m=current_model: client.models.generate_content(
                                    model=m, contents=contents, config=gen_config
                                )
                            ),
                            timeout=90.0
                        )
                        if model_attempt > 0 or retry > 0:
                            logger.info(f"Gemini OK con modelo={current_model} retry={retry}")
                        break
                    except asyncio.TimeoutError:
                        logger.error("Gemini timeout (>90s)")
                        return _fallback_analysis("Timeout de Gemini"), "gemini_timeout", "El analisis tardo demasiado."
                    except Exception as e:
                        last_err = e
                        err_str = str(e).lower()
                        is_rate_limit = "429" in err_str or "resource_exhausted" in err_str
                        # Cupo DIARIO agotado: reintentar en segundos es inútil y quema
                        # el cupo del resto de modelos → vetar y pasar al siguiente.
                        if is_rate_limit and _is_perday_429(err_str):
                            _mark_daily_exhausted(current_model)
                            break
                        if is_rate_limit and retry < 2:
                            wait_s = (retry + 1) * 15
                            logger.warning(f"Gemini 429 modelo={current_model}, reintento {retry+1} en {wait_s}s")
                            await asyncio.sleep(wait_s)
                            continue
                        elif is_rate_limit and model_attempt < len(models_to_try) - 1:
                            logger.warning(f"Gemini 429 agotado en {current_model}, probando {models_to_try[model_attempt+1]}")
                            break  # probar siguiente modelo
                        else:
                            raise
                if response is not None:
                    break

        if response is None and last_err:
            raise last_err

        raw = response.text
        if not raw:
            return _fallback_analysis("Gemini devolvió respuesta vacía"), "gemini_failed", "Gemini devolvió respuesta vacía"

        logger.info(f"Gemini response (primeros 300): {raw[:300]}")
        raw = _strip_markdown_json(raw)

        try:
            data = json.loads(raw)
        except json.JSONDecodeError as e:
            logger.error(f"JSON inválido de Gemini: {e}. Raw: {raw[:500]}")
            return _fallback_analysis(f"JSON inválido: {e}"), "gemini_failed", _user_friendly_error(str(e))

        damages = []
        for d in data.get("damages", []):
            try:
                damages.append(Damage(**d))
            except Exception as de:
                logger.warning(f"Damage ignorado: {de} — {d}")

        new_damages = []
        for d in data.get("new_damages", []):
            try:
                new_damages.append(Damage(**d))
            except Exception as de:
                logger.warning(f"New damage ignorado: {de} — {d}")
        # Si no vino new_damages pero hay damages marcados is_new, usarlos
        if not new_damages and damages:
            new_damages = [d for d in damages if getattr(d, "is_new", True)]

        # ── COHERENCIA ESPACIAL (determinista): la orientación verificada manda ──
        if photo_views:
            _coh_fixes = (_enforce_spatial_coherence(damages, photo_views)
                          + _enforce_spatial_coherence(new_damages, photo_views))
            if _coh_fixes:
                logger.info(f"[Coherencia] {_coh_fixes} correcciones lado/vista según orientación verificada")

        # De-duplicación: el mismo daño visto en varias fotos = 1 solo daño
        damages, dup_removed = _dedup_damages(damages)
        new_damages, dup_removed_new = _dedup_damages(new_damages)
        if dup_removed or dup_removed_new:
            logger.info(f"De-dup daños: {dup_removed} duplicados eliminados en damages, {dup_removed_new} en new_damages")

        # ─ SEGUNDA PASADA: refinamiento quirúrgico de bounding boxes ─
        if damages and os.environ.get("TWO_PASS_DISABLED", "").lower() not in ("1", "true"):
            damages = await _refine_damage_boxes(client, model_name, damages, images_base64, db=db)
            # Sincronizar new_damages con los objetos refinados (misma pieza+ubicación)
            refined_ids = {id(d) for d in damages}
            new_damages = [d for d in damages if getattr(d, 'is_new', True)]

        # Confidence gating para los daños que no pasaron por 2ª pasada
        for d in damages:
            if not hasattr(d, 'confirmed') or d.confirmed is None:
                d.confirmed = (getattr(d, 'confidence', 0) or 0) >= 0.65

        # Recalcular contadores y coste desde las listas de-duplicadas
        critical_count = sum(1 for d in damages if (d.severity or "").lower() in ("grave", "critico", "crítico"))
        total_cost_dedup = sum((d.estimated_cost or 0) for d in damages)

        result = InspectionAnalysis(
            critical_damages_count=critical_count,
            total_damages_count=len(damages),
            new_damages_count=len(new_damages),
            severity=data.get("severity", "sin_danos"),
            dirt_level=(float(data["dirt_level"]) if data.get("dirt_level") is not None else None),
            urgency=data.get("urgency", "puede_esperar"),
            risk=data.get("risk", "bajo"),
            circulation_safe=bool(data.get("circulation_safe", True)),
            detected_plate=str(data.get("detected_plate", "")),
            fraud_warnings=list(data.get("fraud_warnings", [])),
            hidden_damage_probability=float(data.get("hidden_damage_probability", 0)),
            total_estimated_cost=float(total_cost_dedup if (dup_removed or dup_removed_new) and total_cost_dedup > 0 else data.get("total_estimated_cost", 0)),
            confidence=float(data.get("confidence", 0)),
            executive_summary=str(data.get("executive_summary", "")),
            image_quality_warnings=list(data.get("image_quality_warnings", [])),
            affected_parts=list(data.get("affected_parts", [])),
            critical_damages=list(data.get("critical_damages", [])),
            new_damages=new_damages,
            damages=damages
        )

        logger.info(
            f"Análisis OK: severity={result.severity}, "
            f"daños={result.total_damages_count}, coste={result.total_estimated_cost}€, "
            f"confidence={result.confidence}"
        )
        return result, "ok", None

    except Exception as e:
        logger.error(f"Error Gemini: {type(e).__name__}: {e}", exc_info=True)
        return _fallback_analysis(str(e)), "gemini_failed", str(e)


# ═══════════════════════════════════════════════════════════════════════
# TWO-PASS ANALYSIS — Segunda pasada para cajas de precisión quirúrgica
# ═══════════════════════════════════════════════════════════════════════

def _crop_damage_zone(img_bytes: bytes, box_2d: list, padding: float = 0.30) -> tuple:
    """
    Recorta la zona del daño con padding para la segunda pasada.
    Retorna (crop_bytes, crop_region_0to1) o (None, None) si el box es demasiado pequeño.
    """
    ymin, xmin, ymax, xmax = [v / 1000.0 for v in box_2d]
    h_span, w_span = ymax - ymin, xmax - xmin
    if h_span < 0.03 or w_span < 0.03:
        return None, None
    pad_y = h_span * padding
    pad_x = w_span * padding
    y0 = max(0.0, ymin - pad_y)
    x0 = max(0.0, xmin - pad_x)
    y1 = min(1.0, ymax + pad_y)
    x1 = min(1.0, xmax + pad_x)
    try:
        img = Image.open(io.BytesIO(img_bytes)).convert("RGB")
        W, H = img.size
        crop = img.crop((int(x0*W), int(y0*H), int(x1*W), int(y1*H)))
        buf = io.BytesIO()
        crop.save(buf, format="JPEG", quality=92)
        return buf.getvalue(), (y0, x0, y1, x1)
    except Exception as e:
        logger.warning(f"crop_damage_zone error: {e}")
        return None, None


def _remap_box_to_original(crop_box: list, crop_region: tuple) -> list:
    """Convierte coordenadas del recorte (0-1000) al espacio de la imagen original (0-1000)."""
    y0, x0, y1, x1 = crop_region
    cy1, cx1, cy2, cx2 = [v / 1000.0 for v in crop_box]
    hy, hx = y1 - y0, x1 - x0
    return [int((y0 + cy1*hy)*1000), int((x0 + cx1*hx)*1000),
            int((y0 + cy2*hy)*1000), int((x0 + cx2*hx)*1000)]


_REFINE_PROMPT = (
    "Eres un perito forense automotriz. Esta imagen es un RECORTE AMPLIADO de una zona de una furgoneta.\n\n"
    "Daño detectado previamente: {part} — {severity}\n"
    "Descripción previa: \"{description}\"\n\n"
    "TU TAREA:\n"
    "1. ¿Confirmas que hay un daño real y visible aquí (no suciedad, no sombra)?\n"
    "2. Proporciona el bounding box PRECISO [ymin, xmin, ymax, xmax] (0-1000) del daño en ESTE recorte.\n"
    "   — Rodea SOLO el área dañada, tan ajustado como puedas.\n"
    "   — Un rayón: box estrecho y largo. Una abolladura: box compacto.\n"
    "3. Actualiza description si ves algo más preciso.\n"
    "4. Da tu confidence real (0.0-1.0) de que es un daño real.\n\n"
    "Responde SOLO JSON (sin markdown):\n"
    "{{\"confirmed\": true, \"confidence\": 0.92, \"box_2d\": [y1,x1,y2,x2], "
    "\"polygon_points\": [[y1,x1],[y2,x2],[y3,x3]], "
    "\"description\": \"descripción precisa\", \"severity\": \"leve|moderado|grave|critico\"}}\n\n"
    "IMPORTANTE polygon_points: coordenadas 0-1000 relativas a ESTE RECORTE, "
    "siguiendo el contorno exacto del daño (8-20 puntos). "
    "NUNCA hagas un oval perfecto — sigue el borde real del daño visible. "
    "Para arañazos: forma de lápiz MUY estrecha (ancho < 40px). "
    "Para abolladuras: forma IRREGULAR siguiendo el borde del hundimiento visible, no un círculo. "
    "Para roturas/grietas: forma irregular siguiendo el borde roto. "
    "El polígono NO debe cubrir ruedas, cristales ni piezas adyacentes sin daño. "
    "Sé conservador: mejor un polígono pequeño preciso que uno grande que tape zonas sanas."
)


async def _refine_damage_boxes(
    client, model_name: str, damages: list, images_b64: list, db=None
) -> list:
    """
    Segunda pasada de Gemini: para cada daño con caja aproximada, envía un recorte
    ampliado para obtener coordenadas quirúrgicamente precisas.
    Los daños con confidence < 0.45 o caja inválida se saltan.
    """
    from google.genai import types as genai_types

    # Decodificar imágenes originales una sola vez
    images_raw: dict[int, bytes] = {}
    for i, b64 in enumerate(images_b64):
        try:
            images_raw[i + 1] = base64.b64decode(b64)
        except Exception:
            pass

    to_refine = [
        (idx, d) for idx, d in enumerate(damages)
        if (getattr(d, 'box_2d', None) and any(v > 0 for v in (d.box_2d or []))
            and (getattr(d, 'confidence', 0) or 0) >= 0.45
            and (getattr(d, 'photo_index', 1) or 1) in images_raw)
    ]

    if not to_refine:
        return damages

    logger.info(f"[2ª pasada] Refinando {len(to_refine)} daños con zoom quirúrgico")
    refined = list(damages)

    async def _one(idx, d):
        pi = getattr(d, 'photo_index', 1) or 1
        crop_bytes, region = _crop_damage_zone(images_raw[pi], d.box_2d or [0,0,0,0])
        if crop_bytes is None:
            return
        part_str = getattr(d, 'part', 'parte desconocida')
        location_hint_str = getattr(d, 'location_hint', '') or ''
        refine_contents = []
        if db is not None:
            try:
                examples = await get_few_shot_examples(db, location_hint_str, part_str, limit=3)
                few_shot_parts = build_few_shot_prompt_parts_multimodal(examples)
                for part_dict in few_shot_parts:
                    if "inline_data" in part_dict:
                        refine_contents.append(
                            genai_types.Part.from_bytes(
                                data=base64.b64decode(part_dict["inline_data"]["data"]),
                                mime_type=part_dict["inline_data"]["mime_type"],
                            )
                        )
                    else:
                        refine_contents.append(part_dict["text"])
            except Exception as _fse:
                logger.debug(f"[Learning] Error cargando ejemplos: {_fse}")
            # Historial humano de ESTA pieza: aquí es donde se matan los
            # falsos positivos recurrentes (y se evita descartar lo que se escapa).
            try:
                part_lesson = await get_part_lesson(db, part_str)
                if part_lesson:
                    refine_contents.append(part_lesson)
            except Exception as _ple:
                logger.debug(f"[Learning] Error cargando lección de pieza: {_ple}")
        prompt = _REFINE_PROMPT.format(
            part=part_str,
            severity=getattr(d, 'severity', 'desconocido'),
            description=(getattr(d, 'description', '') or '')[:200]
        )
        refine_contents.extend([
            prompt,
            genai_types.Part.from_bytes(data=crop_bytes, mime_type="image/jpeg"),
        ])
        try:
            gen_cfg = genai_types.GenerateContentConfig(
                temperature=0.05, response_mime_type="application/json"
            )
            loop = asyncio.get_running_loop()
            async with _gemini_sem:
                resp = await asyncio.wait_for(
                    loop.run_in_executor(
                        _executor,
                        lambda: client.models.generate_content(
                            model=model_name,
                            contents=refine_contents,
                            config=gen_cfg,
                        )
                    ),
                    timeout=25.0,
                )
            result = json.loads(_strip_markdown_json(resp.text or "{}"))

            if not result.get("confirmed", True):
                d.confidence = min(getattr(d, 'confidence', 0.5) or 0.5, 0.38)
                return

            new_box = result.get("box_2d") or []
            if len(new_box) == 4 and any(v > 0 for v in new_box):
                d.box_2d = _remap_box_to_original(new_box, region)

            # Remap polygon_points from crop space to original image space
            raw_poly = result.get("polygon_points") or []
            if len(raw_poly) >= 4:
                ry0, rx0, ry1, rx1 = region  # 0-1 fractions of original
                rh = ry1 - ry0; rw = rx1 - rx0
                remapped = []
                for pt in raw_poly:
                    if len(pt) >= 2:
                        py_orig = int((pt[0] / 1000.0 * rh + ry0) * 1000)
                        px_orig = int((pt[1] / 1000.0 * rw + rx0) * 1000)
                        remapped.append([
                            max(0, min(1000, py_orig)),
                            max(0, min(1000, px_orig))
                        ])
                if len(remapped) >= 4:
                    d.polygon_points = remapped

            if result.get("confidence") is not None:
                d.confidence = max(0.0, min(1.0, float(result["confidence"])))
            if result.get("description"):
                d.description = str(result["description"])[:500]
            if result.get("severity") in ("leve", "moderado", "grave", "critico"):
                d.severity = result["severity"]

        except asyncio.TimeoutError:
            logger.warning(f"[2ª pasada] timeout daño {idx}")
        except Exception as e:
            logger.warning(f"[2ª pasada] error daño {idx}: {e}")

    # Paralelo con máx 3 simultáneos (respetar rate limit Gemini)
    sem3 = asyncio.Semaphore(3)
    async def _bounded(idx, d):
        async with sem3:
            await _one(idx, d)

    await asyncio.gather(*[_bounded(i, d) for i, d in to_refine])

    # Aplicar confidence gating: < 0.65 → sugerido (confirmed=False)
    for d in refined:
        conf = getattr(d, 'confidence', 0) or 0
        d.confirmed = conf >= 0.65

    logger.info(f"[2ª pasada] completa. Sugeridos: {sum(1 for d in refined if not d.confirmed)}/{len(refined)}")
    return refined


# ═══════════════════════════════════════════════════════════════════════
# PROFESSIONAL PHOTO ANNOTATION — Anotaciones forenses visuales
# ═══════════════════════════════════════════════════════════════════════

# Paleta de colores por severidad (R, G, B) + alfa relleno
_ANN_FILL = {
    "leve":     (255, 213,  79, 90),
    "moderado": (255, 111,   0, 110),
    "grave":    (211,  47,  47, 130),
    "critico":  (136,  14,  79, 150),
}
_ANN_BORDER = {
    "leve":     (251, 192,  45, 255),
    "moderado": (230,  81,   0, 255),
    "grave":    (183,  28,  28, 255),
    "critico":  (106,  27,  54, 255),
}
_ANN_LABEL = {
    "leve": "LEVE", "moderado": "MODERADO", "grave": "GRAVE", "critico": "CRÍTICO"
}
_ANN_ICON = {"leve": "◆", "moderado": "▲", "grave": "●", "critico": "◉"}

_FONT_PATHS = [
    "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
    "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
    "/usr/share/fonts/truetype/freefont/FreeSansBold.ttf",
    "/usr/share/fonts/TTF/DejaVuSans-Bold.ttf",
]


def _load_font(size: int):
    """Carga la mejor fuente disponible en el sistema, con fallback al default."""
    from PIL import ImageFont
    for path in _FONT_PATHS:
        try:
            return ImageFont.truetype(path, size)
        except Exception:
            pass
    return ImageFont.load_default()


def _segment_damage_opencv(
    img_bytes: bytes,
    box_2d: list,
    debug: bool = False,
    scale_max: int = 900,
) -> tuple:
    """Segmenta la superficie físicamente dañada analizando gradientes, sombras y reflexión.
    Devuelve (polygon_points [[y,x] 0-1000], debug_dict | None).
    polygon_points = None si no se puede segmentar con fiabilidad."""
    try:
        import cv2 as _cv2
        import numpy as _np

        buf = _np.frombuffer(img_bytes, _np.uint8)
        img_bgr = _cv2.imdecode(buf, _cv2.IMREAD_COLOR)
        if img_bgr is None:
            return None, None
        H_orig, W_orig = img_bgr.shape[:2]

        scale = min(scale_max / max(H_orig, W_orig), 1.0)
        if scale < 1.0:
            W_p = int(W_orig * scale)
            H_p = int(H_orig * scale)
            img_p = _cv2.resize(img_bgr, (W_p, H_p), interpolation=_cv2.INTER_AREA)
        else:
            W_p, H_p = W_orig, H_orig
            img_p = img_bgr.copy()

        ymin, xmin, ymax, xmax = box_2d
        bh = ymax - ymin
        bw = xmax - xmin
        MARGIN = 0.40
        ry1 = max(0.0, (ymin - bh * MARGIN) / 1000)
        rx1 = max(0.0, (xmin - bw * MARGIN) / 1000)
        ry2 = min(1.0, (ymax + bh * MARGIN) / 1000)
        rx2 = min(1.0, (xmax + bw * MARGIN) / 1000)

        cy1 = int(ry1 * H_p); cx1 = int(rx1 * W_p)
        cy2 = int(ry2 * H_p); cx2 = int(rx2 * W_p)
        crop = img_p[cy1:cy2, cx1:cx2]
        if crop.size == 0 or crop.shape[0] < 30 or crop.shape[1] < 30:
            return None, None

        lab = _cv2.cvtColor(crop, _cv2.COLOR_BGR2LAB)
        L = lab[:, :, 0].astype(_np.float32)
        blur_L = _cv2.GaussianBlur(L, (51, 51), 0)

        bx1_c = max(0, int(xmin / 1000 * W_p) - cx1)
        bx2_c = min(crop.shape[1], int(xmax / 1000 * W_p) - cx1)
        by1_c = max(0, int(ymin / 1000 * H_p) - cy1)
        by2_c = min(crop.shape[0], int(ymax / 1000 * H_p) - cy1)

        if bx2_c <= bx1_c or by2_c <= by1_c:
            return None, None

        gx = _cv2.Sobel(L, _cv2.CV_32F, 1, 0, ksize=3)
        gy = _cv2.Sobel(L, _cv2.CV_32F, 0, 1, ksize=3)
        edge_abs = _np.sqrt(gx ** 2 + gy ** 2)

        shadow_abs = _np.clip(blur_L - L, 0, 255)
        reflex_abs = _np.clip(L - blur_L, 0, 255)

        cand_abs = edge_abs * 0.40 + shadow_abs * 0.35 + reflex_abs * 0.25

        # GUARD 1: señal absoluta dentro del bbox (sin normalizar)
        # Pintura blanca sin daño: 3-15 | golpe sutil: 20-35 | claro: 40+
        inner_signal = float(_np.mean(cand_abs[by1_c:by2_c, bx1_c:bx2_c]))
        if inner_signal < 20.0:
            return None, None

        # GUARD 2: señal interior debe superar 1.3x la señal exterior
        # Evita que líneas de carrocería que cruzan toda la imagen den falso positivo
        outer_mask = _np.ones(crop.shape[:2], dtype=bool)
        outer_mask[by1_c:by2_c, bx1_c:bx2_c] = False
        outer_signal = float(_np.mean(cand_abs[outer_mask])) if outer_mask.any() else inner_signal
        if outer_signal > 0 and (inner_signal / outer_signal) < 1.3:
            return None, None

        # Normalizar solo para threshold Otsu (no para los guards)
        cand_norm = _cv2.normalize(cand_abs, None, 0, 255, _cv2.NORM_MINMAX).astype(_np.uint8)
        cand_smooth = _cv2.GaussianBlur(cand_norm, (21, 21), 0)

        _, binary = _cv2.threshold(cand_smooth, 0, 255, _cv2.THRESH_BINARY + _cv2.THRESH_OTSU)
        ksize = max(9, int(min(crop.shape[:2]) * 0.04))
        ksize = ksize if ksize % 2 == 1 else ksize + 1
        kernel = _cv2.getStructuringElement(_cv2.MORPH_ELLIPSE, (ksize, ksize))
        final_mask = _cv2.morphologyEx(binary, _cv2.MORPH_CLOSE, kernel, iterations=3)
        final_mask = _cv2.dilate(final_mask, kernel, iterations=2)

        contours, _ = _cv2.findContours(final_mask, _cv2.RETR_EXTERNAL, _cv2.CHAIN_APPROX_SIMPLE)
        if not contours:
            return None, None

        bbox_mask = _np.zeros(crop.shape[:2], dtype=_np.uint8)
        bbox_mask[by1_c:by2_c, bx1_c:bx2_c] = 255
        bbox_area = float((by2_c - by1_c) * (bx2_c - bx1_c))

        bx_cx = (bx1_c + bx2_c) / 2.0
        bx_cy = (by1_c + by2_c) / 2.0
        bbox_diag = ((bx2_c - bx1_c) ** 2 + (by2_c - by1_c) ** 2) ** 0.5

        def _score(c):
            c_mask = _np.zeros(crop.shape[:2], dtype=_np.uint8)
            _cv2.drawContours(c_mask, [c], -1, 255, _cv2.FILLED)
            overlap = float(_np.sum((c_mask > 0) & (bbox_mask > 0)))
            overlap_frac = overlap / max(bbox_area, 1)
            if overlap_frac < 0.05:
                return float('inf')
            M = _cv2.moments(c)
            if M['m00'] < 1:
                return float('inf')
            ccx = M['m10'] / M['m00']
            ccy = M['m01'] / M['m00']
            dist_norm = ((ccx - bx_cx) ** 2 + (ccy - bx_cy) ** 2) ** 0.5 / max(bbox_diag, 1)
            return dist_norm - overlap_frac * 2.0

        best = min(contours, key=_score)

        # Verificar overlap real del contorno ganador
        best_mask = _np.zeros(crop.shape[:2], dtype=_np.uint8)
        _cv2.drawContours(best_mask, [best], -1, 255, _cv2.FILLED)
        overlap_px = float(_np.sum((best_mask > 0) & (bbox_mask > 0)))
        if overlap_px / max(bbox_area, 1) < 0.05:
            return None, None

        if _cv2.contourArea(best) < 50:
            return None, None

        M_best = _cv2.moments(best)
        if M_best['m00'] > 0:
            best_cx = M_best['m10'] / M_best['m00']
            best_cy = M_best['m01'] / M_best['m00']
            drift = ((best_cx - bx_cx) ** 2 + (best_cy - bx_cy) ** 2) ** 0.5
            if drift > bbox_diag * 1.5:
                return None, None

        eps = 0.02 * _cv2.arcLength(best, True)
        approx = _cv2.approxPolyDP(best, eps, True)
        if len(approx) < 4:
            return None, None

        poly = []
        for pt in approx:
            px_crop, py_crop = float(pt[0][0]), float(pt[0][1])
            px_orig = (px_crop + cx1) / scale
            py_orig = (py_crop + cy1) / scale
            poly.append([
                max(0, min(1000, round(py_orig / H_orig * 1000))),
                max(0, min(1000, round(px_orig / W_orig * 1000))),
            ])

        debug_imgs = None
        if debug:
            debug_imgs = {
                'crop':       crop,
                'edges':      edge_abs,
                'shadows':    shadow_abs,
                'reflection': reflex_abs,
                'candidate':  cand_smooth,
                'mask':       final_mask,
                'bbox_mask':  bbox_mask,
            }

        return poly, debug_imgs

    except Exception:
        return None, None


def _poly_from_box(px1, py1, px2, py2, n=12):
    """Elipse de N puntos a partir de bounding box — fallback."""
    import math as _math
    cx, cy = (px1 + px2) / 2, (py1 + py2) / 2
    rx, ry = (px2 - px1) / 2, (py2 - py1) / 2
    return [(int(cx + rx * _math.cos(2 * _math.pi * i / n)),
             int(cy + ry * _math.sin(2 * _math.pi * i / n)))
            for i in range(n)]


def _grabcut_refine(img_rgb_np, px1, py1, px2, py2, gemini_poly_px=None):
    """
    Fase 2 — OpenCV GrabCut para segmentación pixel-level dentro de la zona de daño.

    1. Usa el bounding box de Gemini como rectángulo de inicialización de GrabCut.
    2. Si Gemini devolvió polígono, inicializa la máscara de probable-foreground.
    3. Devuelve lista de puntos del contorno exterior del daño segmentado.
    Fallback a elipse si GrabCut falla o el área es demasiado pequeña.
    """
    try:
        import cv2 as _cv2
        import numpy as _np

        H, W = img_rgb_np.shape[:2]
        pad = 8

        # Añadir padding al rect sin salirse de la imagen
        r_x1 = max(0, px1 - pad); r_y1 = max(0, py1 - pad)
        r_x2 = min(W - 1, px2 + pad); r_y2 = min(H - 1, py2 + pad)
        rw = r_x2 - r_x1; rh = r_y2 - r_y1

        if rw < 15 or rh < 15:
            return None  # demasiado pequeño

        rect = (r_x1, r_y1, rw, rh)

        # Inicializar máscara GrabCut
        mask = _np.zeros((H, W), _np.uint8)

        # Si Gemini dio polígono, usarlo como probable foreground
        if gemini_poly_px and len(gemini_poly_px) >= 3:
            pts = _np.array(gemini_poly_px, dtype=_np.int32)
            _cv2.fillPoly(mask, [pts], _cv2.GC_PR_FGD)

        bgd_model = _np.zeros((1, 65), _np.float64)
        fgd_model = _np.zeros((1, 65), _np.float64)

        # GrabCut — 4 iteraciones con rect como hint inicial
        init_mode = _cv2.GC_INIT_WITH_MASK if gemini_poly_px and len(gemini_poly_px) >= 3 else _cv2.GC_INIT_WITH_RECT
        if init_mode == _cv2.GC_INIT_WITH_MASK:
            # Necesita al menos algo en el rect también
            _cv2.grabCut(img_rgb_np, mask, rect, bgd_model, fgd_model, 2, _cv2.GC_INIT_WITH_RECT)
            # Re-sembrar con polígono y refinar
            pts = _np.array(gemini_poly_px, dtype=_np.int32)
            _cv2.fillPoly(mask, [pts], _cv2.GC_PR_FGD)
            _cv2.grabCut(img_rgb_np, mask, rect, bgd_model, fgd_model, 3, _cv2.GC_INIT_WITH_MASK)
        else:
            _cv2.grabCut(img_rgb_np, mask, rect, bgd_model, fgd_model, 5, _cv2.GC_INIT_WITH_RECT)

        # Foreground = GC_FGD | GC_PR_FGD
        fg_mask = _np.where((mask == _cv2.GC_FGD) | (mask == _cv2.GC_PR_FGD), 255, 0).astype(_np.uint8)

        # Limpieza morfológica: eliminar ruido pequeño
        kernel = _cv2.getStructuringElement(_cv2.MORPH_ELLIPSE, (7, 7))
        fg_mask = _cv2.morphologyEx(fg_mask, _cv2.MORPH_CLOSE, kernel, iterations=2)
        fg_mask = _cv2.morphologyEx(fg_mask, _cv2.MORPH_OPEN,  kernel, iterations=1)

        # Encontrar contornos del foreground
        contours, _ = _cv2.findContours(fg_mask, _cv2.RETR_EXTERNAL, _cv2.CHAIN_APPROX_SIMPLE)
        if not contours:
            return None

        # Tomar el contorno más grande dentro del bounding box original
        best = max(contours, key=_cv2.contourArea)
        area = _cv2.contourArea(best)
        if area < 100:  # muy pequeño, ignorar
            return None

        # Simplificar el contorno (Douglas-Peucker) para N puntos razonables
        epsilon = 0.015 * _cv2.arcLength(best, True)
        approx = _cv2.approxPolyDP(best, epsilon, True)

        # Si quedaron muy pocos o demasiados puntos, ajustar epsilon
        if len(approx) < 5:
            epsilon = 0.005 * _cv2.arcLength(best, True)
            approx = _cv2.approxPolyDP(best, epsilon, True)
        if len(approx) > 30:
            epsilon = 0.03 * _cv2.arcLength(best, True)
            approx = _cv2.approxPolyDP(best, epsilon, True)

        poly = [(int(p[0][0]), int(p[0][1])) for p in approx]
        return poly if len(poly) >= 3 else None

    except Exception as _e:
        logger.debug(f"[grabcut] error: {_e}")
        return None


def _annotate_photo_sync(img_bytes: bytes, damage_list: list) -> bytes:
    """
    Anotación forense con segmentación poligonal — Phase 1 upgrade.

    Técnicas:
    1. MÁSCARA POLIGONAL: sigue el contorno exacto del daño (no bounding box)
    2. GLOW SUAVE en los bordes del polígono — sin oscurecer el resto de la imagen
    3. CONTORNO adaptativo de 2px del color de severidad
    4. BURBUJA numerada con sombra + línea guía al centroide del polígono
    5. ETIQUETA con pieza + severidad + confianza
    6. PANEL DE LEYENDA en la base
    7. WATERMARK
    """
    from PIL import Image as _Img, ImageDraw, ImageFont, ImageFilter
    import math as _math

    img_pil = _Img.open(io.BytesIO(img_bytes)).convert("RGBA")
    W, H = img_pil.size
    img = img_pil

    font_num   = _load_font(max(18, H // 45))
    font_label = _load_font(max(13, H // 68))
    font_small = _load_font(max(11, H // 82))
    font_wm    = _load_font(max(10, H // 95))

    # ── Pre-calcular polígonos pixel para todos los daños ──
    damages_px = []
    for dmg, num in damage_list:
        raw_poly = getattr(dmg, 'polygon_points', None) or []
        box = getattr(dmg, 'box_2d', None) or []

        # Coordenadas del bounding box siempre necesarias para GrabCut
        if len(box) == 4 and any(v > 0 for v in box):
            ymin, xmin, ymax, xmax = box
            bpx1 = int(xmin / 1000 * W); bpy1 = int(ymin / 1000 * H)
            bpx2 = max(int(xmax / 1000 * W), bpx1 + 20)
            bpy2 = max(int(ymax / 1000 * H), bpy1 + 20)
        elif len(raw_poly) >= 4:
            xs_r = [int(pt[1] / 1000 * W) for pt in raw_poly]
            ys_r = [int(pt[0] / 1000 * H) for pt in raw_poly]
            bpx1, bpy1, bpx2, bpy2 = min(xs_r), min(ys_r), max(xs_r), max(ys_r)
        else:
            continue

        # Polígono de Gemini en píxeles (hint para GrabCut)
        gemini_poly = [(int(pt[1] / 1000 * W), int(pt[0] / 1000 * H))
                       for pt in raw_poly if len(pt) >= 2] if len(raw_poly) >= 4 else None

        # ── FASE 1: segmentación SAM (API HF) → OpenCV → fallback Gemini ──
        if len(box) == 4 and any(v > 0 for v in box):
            seg_poly_norm, _ = segment_damage(img_bytes, box)
        else:
            seg_poly_norm = None

        if seg_poly_norm and len(seg_poly_norm) >= 4:
            poly_px = [(int(pt[1] / 1000 * W), int(pt[0] / 1000 * H)) for pt in seg_poly_norm]
            logger.info(f"[anotación] daño {num}: polígono segmentación ({len(poly_px)} pts) ✓")
        elif gemini_poly and len(gemini_poly) >= 4:
            poly_px = gemini_poly
            logger.info(f"[anotación] daño {num}: polígono Gemini fallback ({len(poly_px)} pts)")
        else:
            poly_px = _poly_from_box(bpx1, bpy1, bpx2, bpy2)
            logger.info(f"[anotación] daño {num}: elipse bbox fallback")

        if len(poly_px) < 3:
            continue

        xs = [p[0] for p in poly_px]; ys = [p[1] for p in poly_px]
        cx = int(sum(xs) / len(xs)); cy = int(sum(ys) / len(ys))
        bx1 = min(xs); by1 = min(ys); bx2 = max(xs); by2 = max(ys)

        damages_px.append((dmg, num, poly_px, cx, cy, bx1, by1, bx2, by2))

    # ══════════════════════════════════════════════════════════════
    # 1. RELLENO MUY SUTIL + CONTORNO POLIGONAL
    # El relleno es casi transparente para no tapar la foto.
    # El CONTORNO es lo que comunica "aquí está el daño".
    # ══════════════════════════════════════════════════════════════
    poly_layer = _Img.new("RGBA", (W, H), (0, 0, 0, 0))
    poly_dr    = ImageDraw.Draw(poly_layer)

    for dmg, num, poly_px, cx, cy, bx1, by1, bx2, by2 in damages_px:
        sev    = (getattr(dmg, 'severity', 'leve') or 'leve').lower()
        border = _ANN_BORDER.get(sev, _ANN_BORDER['leve'])
        r, g, b = border[:3]

        # Relleno muy tenue (15% opacidad) — solo para marcar la zona
        poly_dr.polygon(poly_px, fill=(r, g, b, 38))

        # Glow exterior — 3 pasadas decrec. para efecto luminoso
        for width, alpha in [(10, 20), (6, 45), (3, 80)]:
            poly_dr.line(poly_px + [poly_px[0]], fill=(r, g, b, alpha), width=width)

        # Contorno sólido fino — el "trazo de perito"
        poly_dr.line(poly_px + [poly_px[0]], fill=(r, g, b, 240), width=2)

    # Suavizar el glow exterior ligeramente
    poly_blur = poly_layer.filter(ImageFilter.GaussianBlur(radius=1))
    img = _Img.alpha_composite(img, poly_blur)

    # ══════════════════════════════════════════════════════════════
    # 3. BURBUJAS, LÍNEAS GUÍA Y ETIQUETAS
    # ══════════════════════════════════════════════════════════════
    overlay = _Img.new("RGBA", (W, H), (0, 0, 0, 0))
    dr = ImageDraw.Draw(overlay)

    for dmg, num, poly_px, cx, cy, bx1, by1, bx2, by2 in damages_px:
        sev       = (getattr(dmg, 'severity', 'leve') or 'leve').lower()
        part      = (getattr(dmg, 'part', '') or '').strip()
        conf      = getattr(dmg, 'confidence', 0.8) or 0.8
        confirmed = getattr(dmg, 'confirmed', True)

        border = _ANN_BORDER.get(sev, _ANN_BORDER['leve'])
        label  = _ANN_LABEL.get(sev, 'LEVE')
        icon   = _ANN_ICON.get(sev, '●')

        # ── Posición burbuja: fuera del bounding box del polígono ──
        r_bubble = max(17, W // 46)
        bx = bx1 + r_bubble + 4
        by = by1 - r_bubble - 6
        if by - r_bubble < 5:
            by = by2 + r_bubble + 6
        if bx + r_bubble > W - 5:
            bx = bx2 - r_bubble - 5

        # ── Línea guía: burbuja → centroide del polígono ──
        dr.line([(bx, by), (cx, cy)], fill=border[:3] + (200,), width=max(2, W // 400))

        # ── Burbuja con sombra ──
        shadow_off = max(3, r_bubble // 5)
        dr.ellipse([bx-r_bubble+shadow_off, by-r_bubble+shadow_off,
                    bx+r_bubble+shadow_off, by+r_bubble+shadow_off],
                   fill=(0, 0, 0, 130))
        dr.ellipse([bx-r_bubble, by-r_bubble, bx+r_bubble, by+r_bubble], fill=border)

        ns = str(num)
        bb = dr.textbbox((0, 0), ns, font=font_num)
        nw, nh = bb[2]-bb[0], bb[3]-bb[1]
        dr.text((bx - nw//2, by - nh//2), ns, fill=(255, 255, 255, 255), font=font_num)

        # ── Etiqueta ──
        short_part = (part[:22] + "…") if len(part) > 22 else part
        conf_str   = f"{int(conf*100)}%"
        suffix     = " ?" if not confirmed else ""
        tag        = f" {icon} {short_part}  {label}  {conf_str}{suffix} "

        bb_t = dr.textbbox((0, 0), tag, font=font_label)
        tw, th = bb_t[2]-bb_t[0], bb_t[3]-bb_t[1]
        tx = max(2, min(bx1, W - tw - 6))
        ty = by2 + 7
        if ty + th + 10 > H:
            ty = max(2, by1 - th - 10)

        dr.rectangle([tx-3, ty-4, tx+tw+5, ty+th+6], fill=border[:3] + (215,))
        dr.text((tx+1, ty+1), tag, fill=(255, 255, 255, 255), font=font_label)

    # ══════════════════════════════════════════════════════════════
    # 4. PANEL DE LEYENDA
    # ══════════════════════════════════════════════════════════════
    if damages_px:
        leg_h = max(26, H // 20)
        leg_y = H - leg_h - 4
        dr.rectangle([4, leg_y - 2, W - 4, H - 2], fill=(0, 0, 0, 175))
        lx = 10
        for dmg_l, num_l, *_ in damages_px[:9]:
            sev_l  = (getattr(dmg_l, 'severity', 'leve') or 'leve').lower()
            part_l = (getattr(dmg_l, 'part', '') or '')[:16]
            bdr_l  = _ANN_BORDER.get(sev_l, _ANN_BORDER['leve'])
            entry  = f" {num_l}. {part_l} "
            bb_e   = dr.textbbox((0, 0), entry, font=font_small)
            ew     = bb_e[2] - bb_e[0]
            if lx + ew + 8 > W:
                break
            dr.text((lx, leg_y + 4), entry, fill=(*bdr_l[:3], 235), font=font_small)
            lx += ew + 4

    # ══════════════════════════════════════════════════════════════
    # 5. WATERMARK
    # ══════════════════════════════════════════════════════════════
    nd = len(damages_px)
    wm = f"⬡ FlotaDSP AI  {nd} daño{'s' if nd != 1 else ''}"
    bb_wm = dr.textbbox((0, 0), wm, font=font_wm)
    wm_w  = bb_wm[2] - bb_wm[0]
    dr.rectangle([W - wm_w - 14, 4, W - 4, 4 + (bb_wm[3]-bb_wm[1]) + 8],
                 fill=(0, 0, 0, 110))
    dr.text((W - wm_w - 10, 8), wm, fill=(255, 255, 255, 110), font=font_wm)

    result = _Img.alpha_composite(img, overlay).convert("RGB")
    buf = io.BytesIO()
    result.save(buf, format="JPEG", quality=92, optimize=True)
    return buf.getvalue()


async def generate_annotated_photos(
    inspection_id: str,
    photo_urls: list,
    damages: list,
    photos_bytes: Optional[list] = None,
) -> list:
    """
    Genera versiones anotadas profesionalmente de las fotos de inspección.
    Solo anota fotos que tienen al menos un daño asignado.
    Sube las anotadas a R2 y retorna la lista (None donde no hay daños).
    """
    if not damages or not photo_urls:
        return []

    # Agrupar daños por photo_index
    by_photo: dict[int, list] = {}
    for num, d in enumerate(damages, 1):
        pi = getattr(d, 'photo_index', 1) or 1
        by_photo.setdefault(pi, []).append((d, num))

    loop = asyncio.get_running_loop()
    annotated = []

    for pi, url in enumerate(photo_urls, 1):
        dmg_list = by_photo.get(pi)
        if not dmg_list:
            annotated.append(None)
            continue

        # Obtener bytes de la foto original
        if photos_bytes and pi <= len(photos_bytes) and photos_bytes[pi - 1]:
            raw = photos_bytes[pi - 1]
        else:
            raw = await _fetch_photo_bytes(url, timeout=12)

        if not raw:
            annotated.append(None)
            continue

        try:
            ann_bytes = await loop.run_in_executor(
                _executor,
                lambda b=raw, dl=dmg_list: _annotate_photo_sync(b, dl),
            )
            key = f"inspections/annotated/{inspection_id}/foto_{pi}.jpg"
            ann_url = await loop.run_in_executor(
                _executor, lambda k=key, ab=ann_bytes: _upload_to_r2_sync(ab, k)
            )
            annotated.append(ann_url)
            logger.info(f"[anotación] foto {pi}/{len(photo_urls)} → {ann_url[:60]}…")
        except Exception as e:
            logger.warning(f"[anotación] error foto {pi}: {e}")
            annotated.append(None)

    confirmed_count = sum(1 for d in damages if getattr(d, 'confirmed', True))
    logger.info(
        f"[anotación] {inspection_id}: {len(annotated)} fotos procesadas, "
        f"{confirmed_count}/{len(damages)} daños confirmados"
    )
    return annotated


# =========================
# IMAGE PROCESSING + STORAGE
# =========================

def _process_image_sync(content: bytes, filepath: Path) -> bytes:
    """Procesa imagen con Pillow. Una sola compresión, consistente en disco y memoria."""
    image = Image.open(io.BytesIO(content)).convert("RGB")
    image.thumbnail((2048, 2048))  # Más resolución para mejor análisis Gemini
    buf = io.BytesIO()
    image.save(buf, format="JPEG", quality=85, optimize=True)
    processed_bytes = buf.getvalue()
    # Escribir a disco desde el buffer — una sola compresión
    filepath.write_bytes(processed_bytes)
    return processed_bytes


def _upload_to_r2_sync(processed_bytes: bytes, filename: str) -> str:
    """Sube bytes a Cloudflare R2. Retorna la URL pública."""
    r2 = get_r2()
    if not r2:
        raise RuntimeError("R2 no configurado")
    r2.put_object(
        Bucket=R2_BUCKET,
        Key=filename,
        Body=processed_bytes,
        ContentType="image/jpeg",
        CacheControl="public, max-age=31536000",
    )
    return f"{R2_PUBLIC_URL}/{filename}"


async def process_and_save_image(content: bytes, vehicle_id: str) -> Tuple[str, bytes]:
    """
    Procesa la imagen y la guarda en R2 (si disponible) o en disco local.
    Retorna (url_publica, bytes_procesados).
    """
    filename = f"{vehicle_id}_{uuid.uuid4().hex}.jpg"
    filepath = UPLOAD_DIR / filename
    loop = asyncio.get_running_loop()

    # Procesar imagen (redimensionar + comprimir)
    processed_bytes = await loop.run_in_executor(
        _executor, _process_image_sync, content, filepath
    )

    # Intentar subir a R2
    r2 = get_r2()
    if r2:
        try:
            public_url = await loop.run_in_executor(
                _executor, _upload_to_r2_sync, processed_bytes, filename
            )
            logger.info(f"Imagen subida a R2: {filename}")
            # Eliminar copia local para no llenar el disco
            try:
                filepath.unlink(missing_ok=True)
            except Exception:
                pass
            return public_url, processed_bytes
        except Exception as e:
            logger.error(f"Error subiendo a R2: {type(e).__name__}: {e} — usando fallback local")

    # Fallback: URL absoluta usando PUBLIC_BASE_URL para que el frontend pueda cargarla
    base = PUBLIC_BASE_URL or ""
    local_url = f"{base}/uploads/{filename}"
    logger.warning(f"Imagen guardada en local (R2 no disponible): {local_url}")
    return local_url, processed_bytes


async def load_reference_images(ref_photo_urls: List[str]) -> List[bytes]:
    """
    Carga las imágenes de referencia desde R2 (HTTP) o desde disco local.
    Retorna lista de bytes listos para enviar a Gemini.
    """
    import aiohttp
    result = []
    async with aiohttp.ClientSession() as session:
        for url in ref_photo_urls:
            try:
                if url.startswith("http"):
                    # URL pública de R2 o similar
                    async with session.get(url, timeout=aiohttp.ClientTimeout(total=10)) as resp:
                        if resp.status == 200:
                            result.append(await resp.read())
                        else:
                            logger.warning(f"Referencia HTTP {resp.status}: {url}")
                elif url.startswith("/uploads/"):
                    # Fichero local (fallback)
                    filepath = ROOT_DIR / url.lstrip("/")
                    if filepath.exists():
                        result.append(filepath.read_bytes())
                    else:
                        logger.warning(f"Referencia local no existe: {filepath}")
                else:
                    logger.warning(f"URL de referencia desconocida: {url[:60]}")
            except Exception as e:
                logger.warning(f"Error cargando referencia {url[:60]}: {e}")
    return result


def validate_image_content(content: bytes, max_size_mb: int = 15) -> None:
    if len(content) > max_size_mb * 1024 * 1024:
        raise HTTPException(status_code=413, detail=f"Imagen demasiado grande. Máximo {max_size_mb}MB.")
    if not (content[:3] == b'\xff\xd8\xff' or content[:4] == b'\x89PNG'):
        raise HTTPException(status_code=415, detail="Solo se aceptan imágenes JPEG o PNG.")
    # Validar dimensiones mínimas
    try:
        img = Image.open(io.BytesIO(content))
        w, h = img.size
        if w < 200 or h < 200:
            raise HTTPException(status_code=422, detail=f"Imagen demasiado pequeña ({w}x{h}px). Mínimo 200x200px.")
    except HTTPException:
        raise
    except Exception:
        pass  # Si Pillow no puede abrirla, lo detectará en process_and_save_image


# =============================================================
# AUTH ROUTES
# =============================================================

auth_router = APIRouter(prefix="/api/auth", tags=["auth"])


# =========================
# RATE LIMITING DE LOGIN (anti fuerza bruta)
# =========================
from collections import defaultdict as _dd

_login_fails: dict = _dd(list)   # clave (tipo:identificador) → [timestamps de fallos]
_LOGIN_MAX_FAILS = 5             # intentos fallidos por USUARIO...
_LOGIN_MAX_FAILS_IP = 15         # ...y por IP (más alto: oficinas comparten IP)
_LOGIN_WINDOW_S = 300            # ventana de 5 minutos
_login_alerted: dict = {}        # para no spamear Telegram con la misma clave


def _rl_key_ip(request: Request) -> str:
    fwd = request.headers.get("x-forwarded-for", "")
    return (fwd.split(",")[0].strip() if fwd else (request.client.host if request.client else "?"))


async def _rl_fail(key: str, context: str):
    """Registra un fallo; si alcanza el límite, avisa por Telegram una vez."""
    now = datetime.now(timezone.utc).timestamp()
    _login_fails[key].append(now)
    if len(_login_fails[key]) >= _LOGIN_MAX_FAILS and now - _login_alerted.get(key, 0) > 1800:
        _login_alerted[key] = now
        try:
            config = await db.telegram_config.find_one({}, {"_id": 0})
            if config and config.get("enabled") and config.get("bot_token"):
                text = (f"🛡️ <b>ALERTA DE SEGURIDAD</b>\n\n"
                        f"Bloqueados 5 intentos de login fallidos seguidos.\n"
                        f"🔑 Objetivo: {context}\n"
                        f"Si no has sido tú, alguien está probando contraseñas.")
                async with _aiohttp.ClientSession() as session:
                    for chat_id in config.get("chat_ids", []):
                        if chat_id.strip():
                            await session.post(
                                f"https://api.telegram.org/bot{config['bot_token']}/sendMessage",
                                json={"chat_id": chat_id, "text": text, "parse_mode": "HTML"},
                                timeout=_aiohttp.ClientTimeout(total=8))
        except Exception as _e:
            logger.warning(f"Alerta seguridad Telegram falló: {_e}")


def _rl_ok(key: str):
    _login_fails.pop(key, None)
    _public_actions.pop(key, None)


def _rl_check(key: str):
    """Lanza 429 si la clave superó el límite de fallos en la ventana."""
    now = datetime.now(timezone.utc).timestamp()
    recent = [t for t in _login_fails[key] if now - t < _LOGIN_WINDOW_S]
    _login_fails[key] = recent
    if not recent:
        # Limpiar key vacía para evitar crecimiento indefinido del dict
        _login_fails.pop(key, None)
        return
    limit = _LOGIN_MAX_FAILS_IP if key.startswith("ip:") else _LOGIN_MAX_FAILS
    if len(recent) >= limit:
        raise HTTPException(status_code=429, detail="Demasiados intentos fallidos. Espera 5 minutos.")


# ── Rate limit genérico para acciones públicas (no solo fallos). ──
# Cuenta TODAS las invocaciones en ventana. Para anti-spam de /register y /lead.
_public_actions: dict = _dd(list)


def _rl_public_action(key: str, max_count: int, window_s: int, detail: str = "Demasiadas peticiones. Inténtalo en unos minutos."):
    """Lanza 429 si la clave supera max_count en window_s segundos. Para endpoints anónimos."""
    now = datetime.now(timezone.utc).timestamp()
    _public_actions[key] = [t for t in _public_actions[key] if now - t < window_s]
    if len(_public_actions[key]) >= max_count:
        raise HTTPException(status_code=429, detail=detail)
    _public_actions[key].append(now)


@auth_router.post("/register", response_model=TokenResponse)
async def register_dsp(data: RegisterRequest, request: Request):
    """Auto-registro de un DSP nuevo: crea su ORGANIZACIÓN (con BD propia y aislada)
    y su usuario dueño. Empieza en prueba (trial). Datos 100% separados del resto."""
    # Anti-spam: máx 5 registros/min/IP y 30/día/IP. Holgado para uso real.
    ip = _rl_key_ip(request)
    _rl_public_action(f"reg-ip-min:{ip}", max_count=5, window_s=60,
                      detail="Estás creando demasiadas cuentas. Espera un minuto.")
    _rl_public_action(f"reg-ip-day:{ip}", max_count=30, window_s=86400,
                      detail="Has alcanzado el límite diario de registros desde esta red.")
    username = (data.username or "").strip()
    org_name = (data.org_name or "").strip()
    if len(username) < 3:
        raise HTTPException(status_code=400, detail="El usuario debe tener al menos 3 caracteres")
    if len(data.password or "") < 8:
        raise HTTPException(status_code=400, detail="La contraseña debe tener al menos 8 caracteres")
    if not org_name:
        raise HTTPException(status_code=400, detail="Indica el nombre de tu empresa/DSP")
    if await global_db.admin_users.find_one({"username": username}):
        raise HTTPException(status_code=409, detail="Ese usuario ya existe, elige otro")

    slug = _slugify(data.slug or org_name)
    if len(slug) < 3:
        raise HTTPException(status_code=400, detail="El identificador (URL) debe tener al menos 3 letras")
    if slug in ("registro", "login", "admin", "api", "conductor", "app", "www"):
        raise HTTPException(status_code=409, detail="Ese identificador está reservado, elige otro")
    if await global_db.organizations.find_one({"slug": slug}):
        raise HTTPException(status_code=409, detail=f"El identificador '{slug}' ya está cogido, elige otro")

    org_id = uuid.uuid4().hex[:12]   # corto: el nombre de BD de Atlas no pasa de 38 chars
    first_center = (data.center or "").strip().upper() or "PRINCIPAL"
    org = {
        "id": org_id, "name": org_name, "account_type": "dsp", "slug": slug,
        "db_name": f"dsp_{org_id}",
        "status": "trial",
        "trial_ends": (datetime.now(timezone.utc) + timedelta(days=14)).isoformat(),
        "email": (data.email or "").strip().lower() or None,
        "plan": (data.plan or "").strip().lower() or None,
        "centers": [first_center],   # cada DSP empieza con UN centro (el suyo)
        "max_centers": 1,            # añadir más = de pago (sube este límite)
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await global_db.organizations.insert_one(dict(org))
    user_id = str(uuid.uuid4())
    await global_db.admin_users.insert_one({
        "id": user_id, "username": username,
        "hashed_password": hash_password(data.password),
        "name": org_name, "role": "admin", "org_id": org_id,
        "email": org.get("email"),
        "created_at": datetime.now(timezone.utc).isoformat(),
    })
    logger.info("Nuevo DSP registrado: %s (org=%s)", username, org_id)
    # Crear índices en la BD recién creada del DSP
    try:
        await _ensure_tenant_indexes(org["db_name"])
    except Exception as _ie:
        logger.warning(f"Error creando índices para nuevo DSP {org_id}: {_ie}")
    # Enviar email de bienvenida con Resend (no bloqueante)
    recipient = org.get("email")
    resend_key = os.environ.get("RESEND_API_KEY", "")
    if recipient and resend_key:
        plan_label = {"basico": "Básico", "pro": "Pro", "flota": "Flota"}.get(org.get("plan") or "", "Pro")
        try:
            import httpx as _httpx
            await _httpx.AsyncClient().post(
                "https://api.resend.com/emails",
                headers={"Authorization": f"Bearer {resend_key}", "Content-Type": "application/json"},
                json={
                    "from": "FlotaDSP <hola@flotadsp.com>",
                    "to": [recipient],
                    "subject": f"¡Bienvenido a FlotaDSP! Tu prueba gratuita de 14 días ya está activa",
                    "html": f"""
<div style="font-family:Inter,sans-serif;max-width:520px;margin:0 auto;background:#0b0d10;color:#eef1f6;border-radius:16px;overflow:hidden">
  <div style="background:linear-gradient(135deg,#0ea5e9,#0369a1);padding:32px 28px;text-align:center">
    <div style="font-size:32px;margin-bottom:8px">⚡</div>
    <h1 style="margin:0;font-size:24px;font-weight:900;color:#fff">¡Bienvenido a FlotaDSP!</h1>
    <p style="margin:8px 0 0;color:rgba(255,255,255,.8);font-size:14px">Tu prueba gratuita de 14 días ya está activa</p>
  </div>
  <div style="padding:28px">
    <p style="margin:0 0 16px;color:#cbd3e0;font-size:15px">Hola <b>{org_name}</b>,</p>
    <p style="margin:0 0 20px;color:#8b94a3;font-size:14px;line-height:1.6">
      Tu cuenta con el plan <b style="color:#0ea5e9">{plan_label}</b> está lista. Tienes 14 días completos para probarlo sin límites y sin introducir ninguna tarjeta.
    </p>
    <div style="background:#13161b;border:1px solid rgba(255,255,255,.08);border-radius:12px;padding:16px 20px;margin-bottom:24px">
      <div style="font-size:12px;color:#64748b;font-weight:700;text-transform:uppercase;letter-spacing:.06em;margin-bottom:10px">Tu acceso</div>
      <div style="font-size:14px;color:#cbd3e0;margin-bottom:6px">🔐 Usuario: <b style="color:#eef1f6">{username}</b></div>
      <div style="font-size:14px;color:#cbd3e0">📦 Plan: <b style="color:#0ea5e9">{plan_label}</b> · Prueba gratuita 14 días</div>
    </div>
    <a href="https://flotadsp.com/panel" style="display:block;text-align:center;background:linear-gradient(135deg,#0ea5e9,#0369a1);color:#fff;text-decoration:none;padding:14px 24px;border-radius:10px;font-weight:800;font-size:15px;margin-bottom:24px">
      Entrar al panel →
    </a>
    <p style="margin:0;color:#64748b;font-size:12px;text-align:center;line-height:1.6">
      Si tienes alguna pregunta, escríbenos a <a href="mailto:hola@flotadsp.com" style="color:#0ea5e9">hola@flotadsp.com</a><br>
      Al finalizar la prueba, se cobra el plan elegido salvo cancelación.
    </p>
    <p style="margin:12px 0 0;color:#475569;font-size:11px;text-align:center;line-height:1.5">
      🇬🇧 Welcome to FlotaDSP! Your 14-day free trial is active — no card required.
      Log in at flotadsp.com/panel with the username above. Questions? hola@flotadsp.com
    </p>
  </div>
</div>""",
                },
                timeout=8.0,
            )
        except Exception as _ee:
            logger.warning(f"Email de bienvenida no enviado a {recipient}: {_ee}")
    token = create_token(user_id, "admin", org_name,
                         org_id=org_id, db_name=org["db_name"], account_type="dsp",
                         centers=org.get("centers"))
    return TokenResponse(access_token=token, role="admin", name=org_name, id=user_id,
                         account_type="dsp", hidden_modules=org_hidden_modules(org), slug=slug,
                         centers=org.get("centers"))


def _org_billing(org):
    """Estado de suscripción de una organización (para el trial y el bloqueo)."""
    if not org or org.get("account_type") == "owner":
        limits = PLAN_LIMITS["owner"]
        return {"status": "owner", "required": False, "days_left": None, "plan": "owner", "limits": limits}
    _raw = (org.get("plan") or "").lower().strip()
    for _c, _r in [("á","a"),("é","e"),("í","i"),("ó","o"),("ú","u")]:
        _raw = _raw.replace(_c, _r)
    plan_key = _raw
    limits = PLAN_LIMITS.get(plan_key, PLAN_DEFAULT)
    status = org.get("status", "trial")
    if status == "active":
        return {"status": "active", "required": False, "days_left": None, "plan": plan_key or "pro", "limits": limits}
    if status == "trial":
        days = None
        te = org.get("trial_ends")
        if te:
            try:
                te_dt = datetime.fromisoformat(te)
                # Normalizar a aware si es naive (cuentas antiguas sin tz)
                if te_dt.tzinfo is None:
                    te_dt = te_dt.replace(tzinfo=timezone.utc)
                delta = te_dt - datetime.now(timezone.utc)
                # Redondear hacia arriba: si quedan 13d 23h mostramos 14
                days = max(0, -(-delta.total_seconds() // 86400))  # ceil division
                days = int(days)
            except Exception:
                days = None
        trial_limits = PLAN_LIMITS.get(plan_key, PLAN_DEFAULT)
        return {"status": "trial", "required": (days is not None and days < 0), "days_left": days,
                "plan": plan_key or "pro", "limits": trial_limits}
    # past_due / canceled / unpaid → bloqueado
    return {"status": status, "required": True, "days_left": None, "plan": plan_key or "basico", "limits": PLAN_LIMITS.get(plan_key, PLAN_LIMITS["basico"])}


@api_router.get("/org/billing")
async def org_billing(user: dict = Depends(get_current_user)):
    """Estado de la suscripción de TU organización (trial, días restantes, si toca pagar)."""
    org = await get_org(user.get("org_id"))
    b = _org_billing(org)
    b["account_type"] = (org or {}).get("account_type")
    b["org_name"] = (org or {}).get("name")
    return b


# ===== UPGRADE PRORRATEADO =====

PLAN_PRICES = {"basico": 99, "pro": 229, "flota": 399, "enterprise": 0, "owner": 0}

@api_router.get("/org/upgrade-preview")
async def upgrade_preview(new_plan: str, user: dict = Depends(require_admin)):
    """Calcula el coste real de subir de plan a mitad de ciclo.
    Devuelve: credit (días ya pagados del plan actual), charge (a pagar ahora), total_new (precio completo del nuevo plan)."""
    org = await get_org(user.get("org_id"))
    billing = _org_billing(org)
    current_plan = billing.get("plan", "basico")
    new_plan = new_plan.lower().strip()

    if new_plan not in PLAN_LIMITS:
        raise HTTPException(400, "Plan no válido")
    if new_plan == current_plan:
        raise HTTPException(400, "Ya estás en ese plan")

    current_price = PLAN_PRICES.get(current_plan, 0)
    new_price = PLAN_PRICES.get(new_plan, 0)

    if new_price <= current_price:
        return {"type": "downgrade", "current_plan": current_plan, "new_plan": new_plan,
                "message": "El cambio a un plan inferior se aplica al siguiente ciclo de facturación.",
                "charge": 0, "credit": 0, "total_new": new_price}

    # Calcular días restantes del mes actual
    today = datetime.now(timezone.utc)
    import calendar as _cal
    days_in_month = _cal.monthrange(today.year, today.month)[1]
    days_used = today.day - 1
    days_remaining = days_in_month - days_used
    credit = round(current_price * days_remaining / days_in_month, 2)
    prorated_new = round(new_price * days_remaining / days_in_month, 2)
    charge = max(0, round(prorated_new - credit, 2))

    return {
        "type": "upgrade",
        "current_plan": current_plan, "new_plan": new_plan,
        "current_price": current_price, "new_price": new_price,
        "days_remaining": days_remaining, "days_in_month": days_in_month,
        "credit": credit,
        "prorated_new": prorated_new,
        "charge": charge,
        "message": f"Pagas {charge}€ ahora ({days_remaining} días del nuevo plan menos {credit}€ de crédito del plan actual).",
    }


@api_router.post("/org/change-plan")
async def change_plan(data: dict = Body(...), user: dict = Depends(require_admin)):
    """Cambia el plan de la organización (solo downgrade o upgrades ya pagados vía Lemon Squeezy)."""
    new_plan = (data.get("plan") or "").lower().strip()
    if new_plan not in PLAN_LIMITS:
        raise HTTPException(400, "Plan no válido")
    org = await get_org(user.get("org_id"))
    if not org:
        raise HTTPException(404, "Organización no encontrada")
    await global_db.organizations.update_one(
        {"id": org["id"]},
        {"$set": {"plan": new_plan, "plan_changed_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"success": True, "plan": new_plan}


# ===== PAGOS: Lemon Squeezy (Merchant of Record) =====
# Se activa poniendo los secretos en Fly: LS_WEBHOOK_SECRET y las URLs de checkout
# LS_CHECKOUT_STARTER / LS_CHECKOUT_PRO / LS_CHECKOUT_FLOTA.

@api_router.get("/billing/config")
async def billing_config(user: dict = Depends(get_current_user)):
    """URLs de checkout por plan (Básico / Pro / Flota / Enterprise)."""
    return {
        "provider": "lemonsqueezy",
        "ready": bool(os.environ.get("LS_CHECKOUT_BASICO") or os.environ.get("LS_CHECKOUT_PRO")),
        "checkout": {
            "basico": os.environ.get("LS_CHECKOUT_BASICO", ""),
            "pro": os.environ.get("LS_CHECKOUT_PRO", ""),
            "flota": os.environ.get("LS_CHECKOUT_FLOTA", ""),
            "enterprise": os.environ.get("LS_CHECKOUT_ENTERPRISE", ""),
            # Variantes anuales — añadir cuando crees los productos en Lemon Squeezy
            "basico_annual": os.environ.get("LS_CHECKOUT_BASICO_ANNUAL", ""),
            "pro_annual": os.environ.get("LS_CHECKOUT_PRO_ANNUAL", ""),
            "flota_annual": os.environ.get("LS_CHECKOUT_FLOTA_ANNUAL", ""),
        },
    }


@api_router.post("/billing/lemonsqueezy/webhook")
async def lemonsqueezy_webhook(request: Request):
    """Aviso de Lemon Squeezy cuando alguien paga/cancela → activa o suspende el DSP.
    Verifica la firma con LS_WEBHOOK_SECRET. El org_id viaja en custom_data del checkout."""
    import hmac as _hmac
    import hashlib as _hashlib
    secret = os.environ.get("LS_WEBHOOK_SECRET", "")
    raw = await request.body()
    if secret:
        sig = request.headers.get("X-Signature", "")
        digest = _hmac.new(secret.encode(), raw, _hashlib.sha256).hexdigest()
        if not _hmac.compare_digest(digest, sig):
            raise HTTPException(status_code=401, detail="Firma inválida")
    try:
        payload = json.loads(raw)
    except Exception:
        raise HTTPException(status_code=400, detail="JSON inválido")
    meta = payload.get("meta") or {}
    event = meta.get("event_name", "")
    org_id = (meta.get("custom_data") or {}).get("org_id")

    # Idempotencia: si LS reintenta el mismo evento, no lo procesamos dos veces.
    # Identificador robusto: combina event_name + id del recurso (subscription/order) + timestamp.
    event_uid = (
        f"{event}:{(payload.get('data') or {}).get('id', '')}:"
        f"{(meta.get('event_id') or meta.get('webhook_id') or '')}"
    )
    try:
        await global_db.ls_webhook_events.insert_one({
            "event_uid": event_uid,
            "event_name": event,
            "received_at": datetime.now(timezone.utc).isoformat(),
            "org_id": org_id,
        })
    except DuplicateKeyError:
        logger.info(f"LS webhook duplicado ignorado: {event_uid}")
        return {"ok": True, "dedup": True}
    attrs = ((payload.get("data") or {}).get("attributes") or {})
    status = attrs.get("status", "")
    raw_plan = attrs.get("product_name") or attrs.get("variant_name") or ""
    # Normalizar nombre LS → clave interna (acepta acentos, mayúsculas, sufijo Y/anual)
    def _ls_plan_key(s: str) -> str:
        s = (s or "").lower()
        for c, r in [("á","a"),("é","e"),("í","i"),("ó","o"),("ú","u")]:
            s = s.replace(c, r)
        if "enterprise" in s: return "enterprise"
        if "flota" in s:      return "flota"
        if "pro" in s:        return "pro"
        if "basico" in s or "basic" in s: return "basico"
        return "pro"  # fallback seguro: nunca menos permisos de los pagados
    plan = _ls_plan_key(raw_plan)
    if not org_id:
        return {"ok": True, "ignored": "sin org_id"}
    # activo si la suscripción está viva; suspendido/cancelado si no
    if event.startswith("subscription_") and status in ("active", "on_trial", "paid"):
        await global_db.organizations.update_one(
            {"id": org_id}, {"$set": {"status": "active", "plan": plan,
                                      "ls_status": status, "ls_subscription_id": (payload.get("data") or {}).get("id")}})
        logger.info("LS pago OK → DSP %s activo (%s)", org_id, plan)
    elif event in ("subscription_cancelled", "subscription_expired", "subscription_paused"):
        await global_db.organizations.update_one(
            {"id": org_id}, {"$set": {"status": "suspended", "ls_status": status}})
        logger.info("LS → DSP %s suspendido (%s)", org_id, event)
    return {"ok": True}


@api_router.get("/org/centers")
async def list_org_centers(user: dict = Depends(get_current_user)):
    """Centros de TU organización (cada uno ve solo los suyos)."""
    org = await get_org(user.get("org_id"))
    return {"centers": (org or {}).get("centers") or [],
            "max_centers": (org or {}).get("max_centers", 1),
            "account_type": (org or {}).get("account_type")}


@api_router.post("/org/centers")
async def add_org_center(data: dict = Body(...), user: dict = Depends(require_admin)):
    """Añade un centro a tu organización. Pasado el límite del plan → 402 (de pago)."""
    name = (data.get("name") or "").strip().upper()
    if not name:
        raise HTTPException(status_code=400, detail="Indica el nombre del centro")
    org = await get_org(user.get("org_id"))
    if not org:
        raise HTTPException(status_code=404, detail="Organización no encontrada")
    centers = org.get("centers") or []
    if name in centers:
        raise HTTPException(status_code=409, detail="Ese centro ya existe")
    if len(centers) >= org.get("max_centers", 1):
        raise HTTPException(status_code=402,
                            detail="Has alcanzado el límite de centros de tu plan. Amplía tu suscripción para añadir más.")
    centers.append(name)
    await global_db.organizations.update_one(
        {"id": org["id"]}, {"$set": {"centers": centers}})
    return {"ok": True, "centers": centers}


@auth_router.post("/lead")
async def capture_lead(data: dict = Body(...), request: Request = None):
    """Captura interés (CRM) + guarda mensaje en Bandeja (append-only).
    - 'leads' es el CRM de interesados: 1 doc por email (upsert).
    - 'inbox_messages' es la bandeja real: 1 doc por envío (append).
    Así no se pierden mensajes si la misma persona escribe varias veces."""
    # Anti-spam: máx 5/min/IP y 30/día/IP.
    ip = _rl_key_ip(request) if request else "?"
    _rl_public_action(f"lead-ip-min:{ip}", max_count=5, window_s=60,
                      detail="Has enviado muchos mensajes en poco tiempo. Espera un minuto.")
    _rl_public_action(f"lead-ip-day:{ip}", max_count=30, window_s=86400,
                      detail="Has alcanzado el límite diario de envíos desde esta red.")
    email = (data.get("email") or "").strip().lower()
    if "@" not in email:
        raise HTTPException(status_code=400, detail="Pon un email válido")
    name = (data.get("name") or "").strip()
    company = (data.get("company") or "").strip()
    plan = (data.get("plan") or "").strip()  # también usado como "asunto + mensaje" desde el form
    now = datetime.now(timezone.utc).isoformat()
    # CRM: 1 lead por email (upsert).
    await global_db.leads.update_one(
        {"email": email},
        {"$set": {"email": email, "plan": plan, "name": name, "company": company, "updated_at": now},
         "$setOnInsert": {"created_at": now}},
        upsert=True)
    # Bandeja: append-only, no pierde mensajes.
    await global_db.inbox_messages.insert_one({
        "id": str(uuid.uuid4()), "email": email, "name": name, "company": company,
        "body": plan, "ip": ip, "ua": (request.headers.get("user-agent") if request else None),
        "created_at": now,
    })
    return {"ok": True, "mensaje": "¡Recibido! Te respondemos en menos de 24 horas hábiles."}


@api_router.get("/leads")
async def list_leads(user: dict = Depends(require_superadmin)):
    """Lista de interesados (CRM). Solo super-admin. 1 doc por email."""
    leads = await global_db.leads.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return {"total": len(leads), "leads": leads}


@api_router.get("/inbox")
async def list_inbox(user: dict = Depends(require_superadmin)):
    """Bandeja de mensajes (append-only). Solo super-admin. 1 doc por mensaje."""
    msgs = await global_db.inbox_messages.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return {"total": len(msgs), "messages": msgs}


# ===== PANEL SUPER-ADMIN (control del negocio) =====

@api_router.get("/admin/overview")
async def admin_overview(_: dict = Depends(require_superadmin)):
    """Resumen del negocio: nº de DSPs, por estado, ingresos estimados, interesados."""
    orgs = await global_db.organizations.find(
        {"account_type": "dsp"}, {"_id": 0}).to_list(2000)
    by_status = {}
    mrr = 0.0
    plan_price = {"Starter": 29, "Pro": 79, "Flota": 149}
    for o in orgs:
        st = o.get("status", "trial")
        by_status[st] = by_status.get(st, 0) + 1
        if st == "active":
            mrr += plan_price.get(o.get("plan", ""), 0)
    leads = await global_db.leads.count_documents({})
    return {
        "dsps_total": len(orgs),
        "por_estado": by_status,
        "activos": by_status.get("active", 0),
        "en_prueba": by_status.get("trial", 0),
        "mrr_estimado": round(mrr, 2),
        "interesados": leads,
    }


@api_router.get("/admin/orgs")
async def admin_list_orgs(_: dict = Depends(require_superadmin)):
    """Todas las organizaciones DSP con su estado, plan, centros y fechas."""
    orgs = await global_db.organizations.find(
        {"account_type": "dsp"}, {"_id": 0}).sort("created_at", -1).to_list(2000)
    out = []
    for o in orgs:
        b = _org_billing(o)
        out.append({
            "id": o.get("id"), "name": o.get("name"), "slug": o.get("slug"),
            "status": o.get("status", "trial"), "plan": o.get("plan"),
            "centers": o.get("centers") or [], "max_centers": o.get("max_centers", 1),
            "email": o.get("email"), "trial_ends": o.get("trial_ends"),
            "dias_prueba": b.get("days_left"), "created_at": o.get("created_at"),
        })
    return {"total": len(out), "orgs": out}


@api_router.post("/admin/org")
async def admin_update_org(data: dict = Body(...), _: dict = Depends(require_superadmin)):
    """Controla una organización: estado, plan, límite de centros, ampliar prueba."""
    org_id = data.get("id")
    if not org_id:
        raise HTTPException(status_code=400, detail="id requerido")
    org = await get_org(org_id)
    if not org or org.get("account_type") != "dsp":
        raise HTTPException(status_code=404, detail="DSP no encontrado")
    patch = {}
    if data.get("status") in ("trial", "active", "suspended", "canceled"):
        patch["status"] = data["status"]
    if "plan" in data:
        patch["plan"] = (data.get("plan") or "").strip() or None
    if "max_centers" in data:
        try:
            patch["max_centers"] = max(1, int(data["max_centers"]))
        except Exception:
            pass
    if data.get("extend_trial_days"):
        try:
            base = datetime.now(timezone.utc)
            patch["trial_ends"] = (base + timedelta(days=int(data["extend_trial_days"]))).isoformat()
        except Exception:
            pass
    if data.get("add_center"):
        c = str(data["add_center"]).strip().upper()
        centers = org.get("centers") or []
        if c and c not in centers:
            centers.append(c)
            patch["centers"] = centers
            patch["max_centers"] = max(org.get("max_centers", 1), len(centers))
    if isinstance(data.get("hidden_modules"), list):
        patch["hidden_modules"] = [str(m) for m in data["hidden_modules"]]
    if not patch:
        raise HTTPException(status_code=400, detail="Nada que actualizar")
    await global_db.organizations.update_one({"id": org_id}, {"$set": patch})
    return {"ok": True, "aplicado": patch}


@api_router.get("/admin/org/{org_id}/stats")
async def admin_org_stats(org_id: str, _: dict = Depends(require_superadmin)):
    """Uso real de un DSP (cuántas furgonetas/conductores/inspecciones tiene)."""
    org = await get_org(org_id)
    if not org:
        raise HTTPException(status_code=404, detail="DSP no encontrado")
    set_current_org_db(_tenant_db_name(org))
    return {
        "vehiculos": await db.vehicles.count_documents({}),
        "conductores": await db.drivers.count_documents({}),
        "inspecciones": await db.inspections.count_documents({}),
    }


@api_router.post("/admin/impersonate")
async def admin_impersonate(data: dict = Body(...), user: dict = Depends(require_superadmin)):
    """Genera un token para ENTRAR COMO un DSP (ver su panel y datos). Solo super-admin."""
    org = await get_org(data.get("id"))
    if not org or org.get("account_type") != "dsp":
        raise HTTPException(status_code=404, detail="DSP no encontrado")
    token = create_token(user["sub"], "admin", org.get("name", ""),
                         org_id=org["id"], db_name=_tenant_db_name(org),
                         account_type="dsp", centers=org.get("centers"))
    logger.info("Super-admin entra como DSP %s", org.get("slug"))
    return {"token": token, "slug": org.get("slug"), "name": org.get("name"),
            "hidden_modules": org_hidden_modules(org), "centers": org.get("centers")}


@api_router.delete("/admin/org/{org_id}")
async def admin_delete_org(org_id: str, _: dict = Depends(require_superadmin)):
    """Elimina un DSP por completo: su BD, sus usuarios y la organización. Irreversible."""
    org = await get_org(org_id)
    if not org or org.get("account_type") != "dsp":
        raise HTTPException(status_code=404, detail="DSP no encontrado")
    try:
        await client.drop_database(_tenant_db_name(org))
    except Exception as e:
        logger.warning("drop_database falló: %s", e)
    await global_db.admin_users.delete_many({"org_id": org_id})
    await global_db.organizations.delete_one({"id": org_id})
    logger.info("DSP eliminado: %s", org.get("slug"))
    return {"ok": True}


@auth_router.get("/org/{slug}")
async def org_by_slug(slug: str):
    """Info pública de un DSP por su slug — para que la URL flotadsp.com/<slug>
    sepa de qué empresa es (mostrar nombre, scope del login del conductor)."""
    org = await global_db.organizations.find_one({"slug": slug}, {"_id": 0})
    if not org:
        raise HTTPException(status_code=404, detail="No existe ninguna empresa con esa URL")
    return {"name": org.get("name"), "slug": org.get("slug"),
            "account_type": org.get("account_type"), "status": org.get("status")}


async def _set_tenant_by_slug(slug):
    """Fija la BD del DSP a partir del slug (para endpoints públicos del conductor)."""
    if not slug:
        set_current_org_db(_DEFAULT_DB_NAME)
        return None
    org = await global_db.organizations.find_one({"slug": slug}, {"_id": 0})
    if not org:
        raise HTTPException(status_code=404, detail="No existe ninguna empresa con esa URL")
    set_current_org_db(_tenant_db_name(org))
    return org


@auth_router.post("/login", response_model=TokenResponse)
async def admin_login(data: LoginRequest, request: Request):
    rl_user = f"user:{data.username.lower().strip()}"
    rl_ip = f"ip:{_rl_key_ip(request)}"
    _rl_check(rl_user)
    _rl_check(rl_ip)

    user = await global_db.admin_users.find_one({"username": data.username}, {"_id": 0})
    if not user or user.get("disabled") or not verify_password(data.password, user["hashed_password"]):
        await _rl_fail(rl_user, f"admin '{data.username}' (IP {_rl_key_ip(request)})")
        await _rl_fail(rl_ip, f"admin '{data.username}' (IP {_rl_key_ip(request)})")
        await asyncio.sleep(0.8)  # frena ataques automatizados
        raise HTTPException(status_code=401, detail="Usuario o contraseña incorrectos")

    _rl_ok(rl_user)
    _rl_ok(rl_ip)
    org = await get_org(user.get("org_id"))
    db_name = _tenant_db_name(org)
    account_type = (org or {}).get("account_type", "owner")
    token = create_token(user["id"], user["role"], user["name"],
                         org_id=user.get("org_id"), db_name=db_name, account_type=account_type,
                         centers=(org or {}).get("centers"), super_admin=bool(user.get("super_admin")),
                         permissions=user.get("permissions"))
    logger.info(f"Admin login: {data.username} (org={user.get('org_id')})")
    return TokenResponse(
        access_token=token,
        role=user["role"],
        name=user["name"],
        id=user["id"],
        theme=user.get("theme"),
        account_type=account_type,
        hidden_modules=org_hidden_modules(org),
        slug=(org or {}).get("slug"),
        centers=(org or {}).get("centers"),
        super_admin=bool(user.get("super_admin")),
        permissions=user.get("permissions"),
        allowed_centers=user.get("allowed_centers"),
        admin_role=user.get("admin_role"),
    )


# =========================
# RECUPERACIÓN DE CONTRASEÑA (admins / DSPs)
# =========================

async def _send_resend_email(to: str, subject: str, html: str) -> bool:
    """Envía un email transaccional con Resend. Devuelve False si falla.
    Remitente configurable con EMAIL_FROM (por defecto hola@flotadsp.com); OJO:
    el dominio del remitente DEBE estar verificado en resend.com/domains o Resend
    devuelve 403 y no envía nada. Registramos el error REAL para no fallar en
    silencio (antes un dominio sin verificar parecía 'no configurado')."""
    resend_key = os.environ.get("RESEND_API_KEY", "")
    if not (resend_key and to):
        logger.warning("email: RESEND_API_KEY o destinatario ausente — no se envía")
        return False
    sender = os.environ.get("EMAIL_FROM", "FlotaDSP <hola@flotadsp.com>")
    import httpx as _httpx
    try:
        async with _httpx.AsyncClient(timeout=15) as _c:
            r = await _c.post(
                "https://api.resend.com/emails",
                headers={"Authorization": f"Bearer {resend_key}", "Content-Type": "application/json"},
                json={"from": sender, "to": [to], "subject": subject, "html": html},
            )
        if r.status_code >= 300:
            logger.error(f"email: Resend rechazó el envío a {to} ({r.status_code}): {r.text[:300]}")
            return False
        return True
    except Exception as e:
        logger.error(f"email: excepción enviando a {to}: {e}")
        return False


_RESET_TOKEN_TTL_MIN = 60  # el enlace caduca en 1 hora


@auth_router.post("/forgot-password")
async def forgot_password(data: dict, request: Request):
    """Envía un enlace de restablecimiento al email de la cuenta.
    SIEMPRE responde éxito (sin revelar si el email existe). Rate-limited por IP."""
    import hashlib as _hl
    import secrets as _sec
    email = (data.get("email") or "").strip().lower()
    _rl_public_action(f"fp:{_rl_key_ip(request)}", max_count=5, window_s=900,
                      detail="Demasiadas solicitudes. Inténtalo en 15 minutos.")
    if not email or "@" not in email:
        return {"success": True}

    user = await global_db.admin_users.find_one({"email": email}, {"_id": 0, "id": 1, "name": 1})
    if user:
        token = _sec.token_urlsafe(32)
        await global_db.password_resets.insert_one({
            "token_hash": _hl.sha256(token.encode()).hexdigest(),
            "user_id": user["id"],
            "used": False,
            "expires_at": (datetime.now(timezone.utc) + timedelta(minutes=_RESET_TOKEN_TTL_MIN)).isoformat(),
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
        link = f"{_PORTAL_BASE_FRONT}/reset-password?token={token}"
        html = f"""
<div style="font-family:Inter,sans-serif;max-width:520px;margin:0 auto;background:#0b0d10;color:#eef1f6;border-radius:16px;overflow:hidden">
  <div style="background:linear-gradient(135deg,#0ea5e9,#0369a1);padding:28px;text-align:center">
    <div style="font-size:28px;margin-bottom:6px">🔑</div>
    <h1 style="margin:0;font-size:21px;font-weight:900;color:#fff">Restablecer contraseña</h1>
  </div>
  <div style="padding:28px">
    <p style="margin:0 0 16px;color:#cbd3e0;font-size:15px">Hola <b>{user.get('name','')}</b>,</p>
    <p style="margin:0 0 20px;color:#8b94a3;font-size:14px;line-height:1.6">
      Hemos recibido una solicitud para restablecer tu contraseña de FlotaDSP.
      El enlace caduca en 1 hora.
    </p>
    <a href="{link}" style="display:block;text-align:center;background:linear-gradient(135deg,#0ea5e9,#0369a1);color:#fff;text-decoration:none;padding:14px 24px;border-radius:10px;font-weight:800;font-size:15px;margin-bottom:20px">
      Crear nueva contraseña →
    </a>
    <p style="margin:0;color:#64748b;font-size:12px;text-align:center;line-height:1.6">
      Si no has pedido este cambio, ignora este email: tu contraseña seguirá siendo la misma.
    </p>
    <p style="margin:12px 0 0;color:#475569;font-size:11px;text-align:center;line-height:1.5">
      🇬🇧 Reset your FlotaDSP password with the button above — the link expires in 1 hour.
      If you didn't request this, just ignore this email.
    </p>
  </div>
</div>"""
        try:
            sent = await _send_resend_email(email, "Restablece tu contraseña de FlotaDSP", html)
            if not sent:
                logger.warning("forgot-password: RESEND_API_KEY no configurada, email no enviado")
        except Exception as _fe:
            logger.error(f"forgot-password: error enviando email: {_fe}")
    return {"success": True}


@auth_router.post("/reset-password")
async def reset_password(data: dict, request: Request):
    """Cambia la contraseña con un token de restablecimiento válido (un solo uso)."""
    import hashlib as _hl
    _rl_public_action(f"rp:{_rl_key_ip(request)}", max_count=10, window_s=900)
    token = (data.get("token") or "").strip()
    new = data.get("new_password") or ""
    if len(new) < 6:
        raise HTTPException(status_code=400, detail="La contraseña debe tener al menos 6 caracteres")
    if not token:
        raise HTTPException(status_code=400, detail="Enlace inválido")

    th = _hl.sha256(token.encode()).hexdigest()
    doc = await global_db.password_resets.find_one({"token_hash": th})
    now_iso = datetime.now(timezone.utc).isoformat()
    if not doc or doc.get("used") or doc.get("expires_at", "") < now_iso:
        raise HTTPException(status_code=400, detail="El enlace no es válido o ha caducado. Solicita uno nuevo.")

    result = await global_db.admin_users.update_one(
        {"id": doc["user_id"]},
        {"$set": {"hashed_password": hash_password(new)}},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=400, detail="La cuenta ya no existe")
    await global_db.password_resets.update_one({"token_hash": th}, {"$set": {"used": True}})
    logger.info(f"Contraseña restablecida vía email para user {doc['user_id']}")
    return {"success": True}


# =========================
# DEMO SIN REGISTRO — org de solo lectura con datos sintéticos
# =========================

_DEMO_ORG_ID = "demo"
_DEMO_DB = "dsp_demo"


async def _seed_demo_data():
    """Crea datos sintéticos realistas en dsp_demo si está vacía (idempotente)."""
    ddb = client[_DEMO_DB]
    if await ddb.vehicles.count_documents({}) > 0:
        return
    import random as _r
    _r.seed(42)  # datos estables entre re-seeds
    now = datetime.now(timezone.utc)
    brands = [("Toyota", "Proace"), ("Renault", "Trafic"), ("Ford", "Transit"),
              ("Mercedes", "Vito"), ("Peugeot", "Expert")]
    vehicles, plates = [], []
    for i in range(10):
        b, mo = brands[i % len(brands)]
        plate = f"{1000 + i * 731 % 9000} {'BCD FGH JKL MNP RST'.split()[i % 5]}"
        plates.append(plate)
        vehicles.append({
            "id": f"demo-v{i}", "license_plate": plate, "brand": b, "model": mo,
            "center": "MADRID", "status": "taller" if i == 7 else "active",
            "mileage": 28000 + i * 9500, "fuel_type": "Diésel", "vehicle_type": "Furgoneta",
            "itv_date": (now + timedelta(days=12 + i * 40)).strftime("%Y-%m-%d"),
            "renting_end_date": (now + timedelta(days=200 + i * 30)).strftime("%Y-%m-%d"),
            "provider": "BANSACAR", "oil_last_change_km": 20000 + i * 9000,
            "oil_interval_km": 15000, "oil_warning_before_km": 2500,
            "mileage_history": [
                {"date": (now - timedelta(days=d)).strftime("%Y-%m-%d"), "km": 28000 + i * 9500 - d * 120, "source": "demo"}
                for d in (30, 20, 10, 0)
            ],
            "created_at": now.isoformat(), "updated_at": now.isoformat(),
        })
    await ddb.vehicles.insert_many(vehicles)

    names = ["Carlos Ruiz", "María López", "Ahmed Ben", "Lucía García", "Ion Popescu",
             "Sara Ortiz", "Diego Fernández", "Ana Torres"]
    drivers = [{
        "id": f"demo-d{i}", "name": n, "center": "MADRID", "active": True,
        "contrato": "empresa" if i % 3 else "ett", "nivel": ["pleno", "L1", "L2"][i % 3],
        "email": f"demo{i}@flotadsp.com", "created_at": now.isoformat(),
    } for i, n in enumerate(names)]
    await ddb.drivers.insert_many(drivers)

    # Inspecciones de los últimos 10 días con algunos daños (para dashboard/€/scoring)
    parts = ["puerta lateral derecha", "paragolpes trasero", "aleta delantera izquierda", "portón trasero"]
    insps = []
    for day in range(10):
        for j in range(5):
            di = (day + j) % len(drivers)
            vi = (day * 3 + j) % len(vehicles)
            has_dmg = (day * 5 + j) % 9 == 0
            nd = ([{
                "part": parts[(day + j) % len(parts)], "severity": ["leve", "moderado", "grave"][(day + j) % 3],
                "description": "Rozadura visible en el panel", "estimated_cost": [90, 240, 520][(day + j) % 3],
                "confirmed": True, "photo_index": 1,
            }] if has_dmg else [])
            created = (now - timedelta(days=day, hours=4 + j)).isoformat()
            insps.append({
                "id": f"demo-i{day}-{j}", "vehicle_id": vehicles[vi]["id"], "driver_id": drivers[di]["id"],
                "center": "MADRID", "photos": [], "annotated_photos": [], "notes": "",
                "analysis_status": "ok", "created_at": created, "reviewed": True,
                "analysis": {
                    "severity": nd[0]["severity"] if nd else "sin_danos",
                    "urgency": "puede_esperar", "risk": "bajo", "circulation_safe": True,
                    "critical_damages_count": 0, "total_damages_count": len(nd),
                    "total_estimated_cost": sum(d["estimated_cost"] for d in nd),
                    "confidence": 90, "executive_summary": "Inspección de demostración.",
                    "damages": nd, "new_damages": nd, "affected_parts": [d["part"] for d in nd],
                    "image_quality_warnings": [], "critical_damages": [],
                },
            })
    await ddb.inspections.insert_many(insps)

    # Cuadrante de hoy + incidencias abiertas
    await ddb.daily_assignments.insert_one({
        "date": now.strftime("%Y-%m-%d"), "center": "MADRID",
        "slots": [{"vehicle_id": v["id"], "vehicle_plate": v["license_plate"],
                   "driver_id": drivers[i % len(drivers)]["id"],
                   "driver_name": drivers[i % len(drivers)]["name"]}
                  for i, v in enumerate(vehicles[:8])],
    })
    await ddb.incidents.insert_many([
        {"id": f"demo-inc{i}", "vehicle_id": vehicles[i]["id"], "title": tt,
         "description": "Incidencia de demostración", "severity": sv, "status": "open",
         "center": "MADRID", "created_at": (now - timedelta(days=i + 1)).isoformat()}
        for i, (tt, sv) in enumerate([("Golpe en puerta lateral", "moderado"),
                                      ("Luz de freno fundida", "leve"),
                                      ("Retrovisor roto", "grave")])
    ])
    logger.info("Demo: datos sintéticos creados en dsp_demo")


@auth_router.post("/demo-login")
async def demo_login(request: Request):
    """Acceso instantáneo a una organización DEMO de solo lectura (sin registro).
    Los datos son sintéticos y las mutaciones están bloqueadas por token."""
    _rl_public_action(f"demo:{_rl_key_ip(request)}", max_count=10, window_s=3600,
                      detail="Demasiados accesos a la demo. Inténtalo más tarde.")
    org = await global_db.organizations.find_one({"id": _DEMO_ORG_ID})
    if not org:
        await global_db.organizations.insert_one({
            "id": _DEMO_ORG_ID, "name": "Demo FlotaDSP", "account_type": "dsp",
            "slug": "demo", "db_name": _DEMO_DB, "status": "active", "plan": "pro",
            "centers": ["MADRID"], "max_centers": 1,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
    try:
        await _seed_demo_data()
    except Exception as _se:
        logger.warning(f"Demo seed: {_se}")
    token = create_token("demo-user", "admin", "Demo FlotaDSP",
                         org_id=_DEMO_ORG_ID, db_name=_DEMO_DB, account_type="dsp",
                         centers=["MADRID"], demo=True)
    logger.info(f"Demo login desde {_rl_key_ip(request)}")
    return {"access_token": token, "role": "admin", "name": "Demo FlotaDSP",
            "id": "demo-user", "account_type": "dsp", "slug": "demo",
            "centers": ["MADRID"], "demo": True}


# =========================
# ASISTENTE — pregúntale a tu flota (Gemini con contexto real de la org)
# =========================

@api_router.post("/assistant/ask")
async def assistant_ask(data: dict, user: dict = Depends(require_admin)):
    """Responde preguntas en lenguaje natural sobre LA FLOTA DE ESTA ORG.
    Enfoque fiable: se recopila un resumen compacto de datos reales (solo
    lectura, acotado) y Gemini responde SOLO con esos datos."""
    question = (data.get("question") or "").strip()[:300]
    if len(question) < 4:
        raise HTTPException(status_code=400, detail="Escribe una pregunta")
    _rl_public_action(f"ask:{user.get('org_id') or user.get('sub')}", max_count=15, window_s=300,
                      detail="Demasiadas preguntas seguidas. Espera un momento.")

    now = datetime.now(timezone.utc)
    month_start = datetime(now.year, now.month, 1, tzinfo=timezone.utc).isoformat()
    recent_start = (now - timedelta(days=14)).isoformat()
    vehicles = await db.vehicles.find(
        {"status": {"$ne": "deleted"}},
        {"_id": 0, "id": 1, "license_plate": 1, "brand": 1, "model": 1, "center": 1,
         "status": 1, "mileage": 1, "itv_date": 1, "renting_end_date": 1}
    ).to_list(300)
    drivers = await db.drivers.find(
        {"active": {"$ne": False}}, {"_id": 0, "id": 1, "name": 1, "center": 1, "contrato": 1, "nivel": 1}
    ).to_list(300)
    incidents = await db.incidents.find(
        {"status": {"$ne": "resolved"}},
        {"_id": 0, "title": 1, "severity": 1, "vehicle_id": 1, "created_at": 1}
    ).to_list(40)
    month_insps = await db.inspections.find(
        {"deleted": {"$ne": True}, "created_at": {"$gte": month_start}},
        {"_id": 0, "vehicle_id": 1, "driver_id": 1, "created_at": 1,
         "analysis.severity": 1, "analysis.new_damages.part": 1,
         "analysis.new_damages.severity": 1, "analysis.new_damages.estimated_cost": 1}
    ).to_list(3000)

    # Contexto COMPACTO a escala real: agregados por conductor + solo los daños
    # recientes en detalle, con matrículas resueltas (no IDs opacos). Con flotas
    # grandes el JSON masivo de antes superaba el timeout de Gemini → 503.
    dmap = {d["id"]: d.get("name", "?") for d in drivers}
    vplate = {v.get("id"): (v.get("license_plate") or v.get("id")) for v in vehicles}
    per_driver: dict = {}
    recent_damages = []
    for i in month_insps:
        did = i.get("driver_id")
        st = per_driver.setdefault(did, {"inspecciones": 0, "danos_nuevos": 0, "coste_estimado": 0})
        st["inspecciones"] += 1
        nd = (i.get("analysis") or {}).get("new_damages") or []
        st["danos_nuevos"] += len(nd)
        st["coste_estimado"] += sum(float(d.get("estimated_cost") or 0) for d in nd if isinstance(d, dict))
        if nd and (i.get("created_at") or "") >= recent_start and len(recent_damages) < 60:
            recent_damages.append({
                "fecha": (i.get("created_at") or "")[:10],
                "vehiculo": vplate.get(i.get("vehicle_id"), i.get("vehicle_id")),
                "conductor": dmap.get(did, "?"),
                "danos": [{"pieza": d.get("part"), "severidad": d.get("severity"),
                           "coste": d.get("estimated_cost")} for d in nd if isinstance(d, dict)][:5],
            })

    ctx = {
        "hoy": now.strftime("%Y-%m-%d"),
        "vehiculos": [{k: v for k, v in x.items() if k != "id"} for x in vehicles],
        "conductores": [{"nombre": d.get("name"), "centro": d.get("center"),
                         "contrato": d.get("contrato"), "nivel": d.get("nivel")} for d in drivers],
        "incidencias_abiertas": [
            {**{k: v for k, v in i.items() if k != "vehicle_id"},
             "vehiculo": vplate.get(i.get("vehicle_id"), i.get("vehicle_id"))} for i in incidents],
        "resumen_mes_por_conductor": [
            {"conductor": dmap.get(k, "?"), **v} for k, v in per_driver.items() if k],
        "danos_ultimos_14_dias": recent_damages,
    }

    prompt = (
        "Eres el asistente de FlotaDSP para el gestor de una flota de furgonetas DSP.\n"
        "Responde a la pregunta usando EXCLUSIVAMENTE los datos JSON adjuntos. "
        "Si el dato no está, di claramente que no tienes esa información. "
        "Sé conciso y directo (máx. ~120 palabras), en el idioma de la pregunta, "
        "menciona matrículas y nombres concretos cuando ayuden, y usa formato de lista si hay varios elementos.\n\n"
        f"PREGUNTA: {question}\n\nDATOS DE LA FLOTA:\n{json.dumps(ctx, ensure_ascii=False, default=str)[:30000]}"
    )
    try:
        from google.genai import types as genai_types
        client_g = _gemini_client_plantilla()
        model_name = os.environ.get("GEMINI_MODEL", "gemini-2.5-flash")
        # thinking_budget=0: la pregunta es factual sobre datos adjuntos, no hace
        # falta razonamiento largo — respuesta en segundos en vez de decenas.
        try:
            _cfg = genai_types.GenerateContentConfig(
                thinking_config=genai_types.ThinkingConfig(thinking_budget=0))
        except Exception:
            _cfg = None
        resp = await asyncio.wait_for(
            asyncio.get_running_loop().run_in_executor(
                _executor,
                lambda: client_g.models.generate_content(
                    model=model_name, contents=prompt,
                    **({"config": _cfg} if _cfg else {})),
            ),
            timeout=45,
        )
        answer = (resp.text or "").strip()
        if not answer:
            raise ValueError("respuesta vacía")
        return {"answer": answer}
    except Exception as e:
        logger.warning(f"assistant/ask: {e}")
        raise HTTPException(status_code=503, detail="El asistente no está disponible ahora mismo. Inténtalo en unos minutos.")


@auth_router.get("/conductor-list")
async def conductor_list_public(center: Optional[str] = None, slug: Optional[str] = None):
    """Lista pública de conductores (solo nombre, email, centro, id) para el
    portal de login del conductor. NO requiere autenticación. Scoped al DSP por slug."""
    await _set_tenant_by_slug(slug)
    query = {}
    if center and center != "Todos":
        if not re.match(r'^[A-Za-z0-9_\-]{1,30}$', center):
            raise HTTPException(400, "Código de centro inválido")
        query["center"] = {"$regex": re.escape(center), "$options": "i"}
    cursor = db.drivers.find(query, {"_id": 0, "id": 1, "name": 1, "email": 1, "center": 1, "photo_url": 1})
    drivers = await cursor.to_list(500)
    # Añadir has_account para que el portal sepa si debe pedir contraseña
    if drivers:
        ids_with_account = {
            a["driver_id"] async for a in db.driver_accounts.find(
                {"driver_id": {"$in": [d["id"] for d in drivers]}, "active": True},
                {"driver_id": 1}
            )
        }
        for d in drivers:
            d["has_account"] = d["id"] in ids_with_account
    return drivers


@auth_router.post("/driver-token")
async def driver_token_by_id(data: dict, request: Request):
    # Límite por IP: 20 tokens/5min — frena enumeración masiva de conductores
    rl_ip = f"dtok:{_rl_key_ip(request)}"
    now_ts = datetime.now(timezone.utc).timestamp()
    _login_fails[rl_ip] = [t for t in _login_fails[rl_ip] if now_ts - t < _LOGIN_WINDOW_S]
    if len(_login_fails[rl_ip]) >= 20:
        raise HTTPException(status_code=429, detail="Demasiados intentos. Espera unos minutos.")
    _login_fails[rl_ip].append(now_ts)
    return await _driver_token_impl(data)


async def _driver_token_impl(data: dict):
    """Genera JWT para el portal conductor (login por email sin contraseña).
    El portal ya valida que el email existe en la BD pública — aquí solo emitimos el token.
    Si el conductor tiene una cuenta con contraseña, se rechaza el acceso sin contraseña."""
    driver_id = data.get("driver_id")
    if not driver_id:
        raise HTTPException(status_code=400, detail="driver_id requerido")
    org = await _set_tenant_by_slug(data.get("slug"))   # scope al DSP del conductor
    driver = await db.drivers.find_one({"id": driver_id}, {"_id": 0})
    if not driver:
        raise HTTPException(status_code=404, detail="Conductor no encontrado")
    # Seguridad: si el conductor tiene contraseña configurada, debe usarla
    account = await db.driver_accounts.find_one({"driver_id": driver_id, "active": True})
    if account:
        raise HTTPException(status_code=403, detail="Este conductor tiene contraseña. Usa tu email y contraseña para acceder.")
    token = create_token(driver_id, "driver", driver.get("name", ""),
                         org_id=(org or {}).get("id"), db_name=_tenant_db_name(org),
                         account_type=(org or {}).get("account_type"))
    logger.info(f"Portal conductor token: {driver.get('name')} ({driver_id})")
    return {
        "access_token": token,
        "token_type": "bearer",
        "role": "driver",
        "name": driver.get("name", ""),
        "id": driver_id,
        "center": driver.get("center")
    }


@auth_router.post("/driver-login", response_model=TokenResponse)
async def driver_login(data: DriverLoginRequest, request: Request):
    rl_key = f"drv:{data.email.lower().strip()}"
    rl_ip = f"ip:{_rl_key_ip(request)}"
    _rl_check(rl_key)
    _rl_check(rl_ip)
    account = await db.driver_accounts.find_one({"email": data.email}, {"_id": 0})
    if not account or not verify_password(data.password, account["hashed_password"]):
        await _rl_fail(rl_key, f"conductor '{data.email}' (IP {_rl_key_ip(request)})")
        await _rl_fail(rl_ip, f"conductor '{data.email}' (IP {_rl_key_ip(request)})")
        await asyncio.sleep(0.8)
        raise HTTPException(status_code=401, detail="Email o contraseña incorrectos")
    _rl_ok(rl_key)
    _rl_ok(rl_ip)

    if not account.get("active", True):
        raise HTTPException(status_code=403, detail="Cuenta desactivada")

    driver = await db.drivers.find_one({"id": account["driver_id"]}, {"_id": 0})
    driver_name = driver["name"] if driver else account["email"]
    driver_center = driver.get("center") if driver else None

    token = create_token(account["driver_id"], "driver", driver_name)
    logger.info(f"Driver login: {data.email}")
    return TokenResponse(
        access_token=token,
        role="driver",
        name=driver_name,
        id=account["driver_id"],
        center=driver_center
    )


@auth_router.get("/me/assigned-vehicle")
async def get_my_assigned_vehicle(user: dict = Depends(get_current_user)):
    """Devuelve la furgoneta asignada HOY al conductor según el cuadrante diario."""
    if user.get("role") != "driver":
        raise HTTPException(status_code=403, detail="Solo conductores")
    driver_id = user["sub"]
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    docs = await db.daily_assignments.find({"date": today}, {"_id": 0}).to_list(50)
    for doc in docs:
        for slot in doc.get("slots", []):
            if slot.get("driver_id") == driver_id and slot.get("vehicle_id"):
                vehicle = await db.vehicles.find_one(
                    {"id": slot["vehicle_id"]}, {"_id": 0}
                )
                if vehicle:
                    insp = await db.inspections.find_one(
                        {"deleted": {"$ne": True}, "vehicle_id": slot["vehicle_id"],
                         "driver_id": driver_id,
                         "created_at": {"$regex": f"^{today}"}},
                        {"_id": 0, "id": 1}
                    )
                    return {
                        "assigned": True,
                        "vehicle": vehicle,
                        "center": doc.get("center"),
                        "already_inspected": insp is not None,
                    }
    return {"assigned": False, "vehicle": None}


@auth_router.get("/me")
async def get_me(user: dict = Depends(get_current_user)):
    # Incluir theme y email si existen en la BD (solo admins)
    theme = None
    email = None
    if user.get("role") == "admin":
        admin_doc = await global_db.admin_users.find_one({"id": user["sub"]}, {"_id": 0, "theme": 1, "email": 1})
        if admin_doc:
            theme = admin_doc.get("theme")
            email = admin_doc.get("email")
    return {
        "id": user["sub"],
        "role": user["role"],
        "name": user["name"],
        "theme": theme,
        "email": email,
    }


@auth_router.post("/my-email")
async def set_my_email(data: dict, user: dict = Depends(get_current_user)):
    """Cada admin vincula (o borra) su propio email de recuperación."""
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores")
    em = (data.get("email") or "").strip().lower()
    if em and not _EMAIL_RE.match(em):
        raise HTTPException(status_code=400, detail="Email no válido")
    # Evita que dos cuentas compartan email (rompería recuperar-contraseña).
    if em:
        clash = await global_db.admin_users.find_one(
            {"email": em, "id": {"$ne": user["sub"]}}, {"_id": 0, "id": 1})
        if clash:
            raise HTTPException(status_code=409, detail="Ese email ya está en uso por otra cuenta")
    await global_db.admin_users.update_one(
        {"id": user["sub"]}, {"$set": {"email": em or None}})
    return {"success": True, "email": em or None}


def _is_center_manager(user: dict) -> bool:
    return user.get("admin_role") == "center_manager" and not user.get("sa")

def _can_manage_user(actor: dict, target_centers: Optional[list]) -> bool:
    """True si el actor puede gestionar un usuario con esos centros."""
    if actor.get("sa"):
        return True
    if not _is_center_manager(actor):
        return False  # dispatchers no pueden gestionar usuarios
    actor_centers = set(actor.get("allowed_centers") or [])
    target_set   = set(target_centers or [])
    # center_manager solo puede gestionar usuarios cuyos centros son subconjunto de los suyos
    return target_set.issubset(actor_centers)

def _clamp_permissions(actor: dict, permissions: Optional[list]) -> Optional[list]:
    """center_manager no puede dar más permisos de los que él mismo tiene."""
    if actor.get("sa"):
        return permissions
    actor_perms = actor.get("permissions")  # None = todos
    if actor_perms is None:
        return permissions
    if permissions is None:
        return actor_perms  # no puede dar "todos" si él no tiene todos
    return [p for p in permissions if p in actor_perms]

@auth_router.post("/create-admin")
async def create_admin(
    data: CreateAdminRequest,
    _admin: dict = Depends(require_admin)
):
    existing = await global_db.admin_users.find_one({"username": data.username})
    if existing:
        raise HTTPException(status_code=409, detail="El usuario ya existe")

    # center_manager solo puede crear usuarios en sus propios centros
    if not _admin.get("sa") and _is_center_manager(_admin):
        if not _can_manage_user(_admin, data.allowed_centers):
            raise HTTPException(403, "Solo puedes crear usuarios en tus centros asignados")
        # no puede crear center_managers, solo dispatchers
        if data.admin_role == "center_manager":
            raise HTTPException(403, "No puedes crear otros gestores de centro")
    elif not _admin.get("sa") and not _is_center_manager(_admin):
        raise HTTPException(403, "Sin permisos para crear usuarios")

    admin_role = data.admin_role if data.admin_role in ("center_manager", "dispatcher") else None
    perms = _clamp_permissions(_admin, data.permissions)

    doc = {
        "id": str(uuid.uuid4()),
        "username": data.username,
        "hashed_password": hash_password(data.password),
        "name": data.name,
        "role": "admin",
        "admin_role": admin_role,
        "super_admin": False,
        "org_id": _admin.get("org_id"),
        "permissions": perms,
        "allowed_centers": data.allowed_centers,
        "created_by": _admin.get("sub"),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await global_db.admin_users.insert_one(doc)
    logger.info(f"Admin creado: {data.username} por {_admin.get('name')} (org={_admin.get('org_id')})")
    return {"success": True, "id": doc["id"], "username": data.username}


@auth_router.patch("/admins/{admin_id}")
async def update_admin_permissions(admin_id: str, data: dict = Body(...), _admin: dict = Depends(require_admin)):
    """Actualiza nombre, permisos o rol de un usuario de MI organización."""
    target = await global_db.admin_users.find_one({"id": admin_id}, {"_id": 0})
    if not target or target.get("org_id") != _admin.get("org_id"):
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    if target.get("super_admin"):
        raise HTTPException(status_code=403, detail="No puedes modificar a un super-admin")

    # center_manager solo puede tocar usuarios en sus centros
    if not _admin.get("sa") and _is_center_manager(_admin):
        if not _can_manage_user(_admin, target.get("allowed_centers")):
            raise HTTPException(403, "No tienes acceso para modificar este usuario")
        if target.get("admin_role") == "center_manager":
            raise HTTPException(403, "No puedes modificar a otro gestor de centro")
    elif not _admin.get("sa") and not _is_center_manager(_admin):
        raise HTTPException(403, "Sin permisos para modificar usuarios")

    patch = {}
    if "permissions" in data:
        patch["permissions"] = _clamp_permissions(_admin, data.get("permissions"))
    if "allowed_centers" in data:
        ac = data.get("allowed_centers")
        patch["allowed_centers"] = ac if (isinstance(ac, list) or ac is None) else None
    if data.get("name"):
        patch["name"] = str(data["name"]).strip()
    if "admin_role" in data:
        ar = data.get("admin_role")
        if ar in ("center_manager", "dispatcher", None):
            # center_manager no puede crear otros center_managers
            if ar == "center_manager" and not _admin.get("sa"):
                raise HTTPException(403, "Solo super-admin puede asignar rol de gestor de centro")
            patch["admin_role"] = ar
    if "email" in data:
        em = (data.get("email") or "").strip().lower()
        if em and not _EMAIL_RE.match(em):
            raise HTTPException(status_code=400, detail="Email no válido")
        patch["email"] = em or None
    # Reset de contraseña por el gestor (para cuando un usuario la olvida).
    new_pw = data.get("new_password")
    if new_pw:
        if len(new_pw) < 6:
            raise HTTPException(status_code=400, detail="La contraseña debe tener al menos 6 caracteres")
        patch["hashed_password"] = hash_password(new_pw)
    if not patch:
        raise HTTPException(status_code=400, detail="Nada que actualizar")
    await global_db.admin_users.update_one({"id": admin_id}, {"$set": patch})
    if "hashed_password" in patch:
        _ADMIN_EXISTS_CACHE.pop(admin_id, None)
        logger.info(f"Contraseña de '{target.get('username')}' restablecida por {_admin.get('name')}")
    return {"success": True}


@auth_router.delete("/admins/{admin_id}")
async def delete_admin(admin_id: str, _admin: dict = Depends(require_admin)):
    """Elimina un usuario de MI organización (no a uno mismo ni a super-admins)."""
    if admin_id == _admin.get("sub"):
        raise HTTPException(status_code=400, detail="No puedes eliminarte a ti mismo")
    target = await global_db.admin_users.find_one({"id": admin_id}, {"_id": 0})
    if not target or target.get("org_id") != _admin.get("org_id"):
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    if target.get("super_admin"):
        raise HTTPException(status_code=403, detail="No puedes eliminar a un super-admin")
    await global_db.admin_users.delete_one({"id": admin_id})
    # Lápida: los seeds/migraciones de arranque no deben resucitarlo jamás,
    # y sus tokens vivos dejan de valer al momento (no en 60s de caché).
    await global_db.admin_tombstones.update_one(
        {"username": target.get("username")},
        {"$set": {"deleted_at": datetime.now(timezone.utc).isoformat(),
                  "deleted_by": _admin.get("sub")}}, upsert=True)
    _ADMIN_EXISTS_CACHE.pop(admin_id, None)
    return {"success": True}


@auth_router.post("/change-my-password")
async def change_my_password(data: dict, user: dict = Depends(get_current_user)):
    """Cualquier admin cambia SU PROPIA contraseña verificando la actual."""
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores")
    current = data.get("current_password") or ""
    new = data.get("new_password") or ""
    if len(new) < 6:
        raise HTTPException(status_code=400, detail="La nueva contraseña debe tener al menos 6 caracteres")
    doc = await global_db.admin_users.find_one({"id": user["sub"]})
    if not doc or not verify_password(current, doc["hashed_password"]):
        await asyncio.sleep(0.8)
        raise HTTPException(status_code=401, detail="La contraseña actual no es correcta")
    await global_db.admin_users.update_one(
        {"id": user["sub"]},
        {"$set": {"hashed_password": hash_password(new)}}
    )
    logger.info(f"Admin '{doc.get('username')}' cambió su propia contraseña")
    return {"success": True}


@auth_router.post("/reset-admin-password")
async def reset_admin_password(data: dict, _admin: dict = Depends(require_admin)):
    """Cambia la contraseña de un admin existente. Solo admins de la misma organización."""
    username = (data.get("username") or "").strip()
    new_password = data.get("password") or ""
    if not username or len(new_password) < 6:
        raise HTTPException(status_code=400, detail="Usuario y contraseña (mín. 6 caracteres) requeridos")
    # Filtra por org_id para evitar que un admin de un DSP resetee cuentas de otro DSP.
    result = await global_db.admin_users.update_one(
        {"username": username, "org_id": _admin.get("org_id")},
        {"$set": {"hashed_password": hash_password(new_password)}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Admin no encontrado")
    logger.info(f"Contraseña de admin '{username}' cambiada por {_admin['name']}")
    return {"success": True}


@auth_router.post("/set-driver-password")
async def set_driver_password(
    data: SetDriverPasswordRequest,
    _admin: dict = Depends(require_admin)
):
    driver = await db.drivers.find_one({"id": data.driver_id}, {"_id": 0})
    if not driver:
        raise HTTPException(status_code=404, detail="Conductor no encontrado")

    if not driver.get("email"):
        raise HTTPException(
            status_code=400,
            detail="El conductor no tiene email — añádelo primero para poder crear su cuenta"
        )

    existing = await db.driver_accounts.find_one({"driver_id": data.driver_id})
    hashed = hash_password(data.password)

    if existing:
        await db.driver_accounts.update_one(
            {"driver_id": data.driver_id},
            {"$set": {"hashed_password": hashed, "active": True}}
        )
        logger.info(f"Password reseteado para conductor {data.driver_id}")
    else:
        await db.driver_accounts.insert_one({
            "id": str(uuid.uuid4()),
            "driver_id": data.driver_id,
            "email": driver["email"],
            "hashed_password": hashed,
            "active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        logger.info(f"Cuenta creada para conductor {data.driver_id} ({driver['email']})")

    return {
        "success": True,
        "driver_id": data.driver_id,
        "email": driver["email"],
        "message": f"Contraseña configurada. El conductor puede iniciar sesión con {driver['email']}"
    }


@auth_router.get("/admins")
async def list_admins(_admin: dict = Depends(require_admin)):
    q = {"org_id": _admin.get("org_id")} if _admin.get("org_id") else {}
    admins = await global_db.admin_users.find(q, {"_id": 0, "hashed_password": 0}).to_list(200)
    # center_manager solo ve usuarios en sus centros (+ a sí mismo)
    if not _admin.get("sa") and _is_center_manager(_admin):
        my_centers = set(_admin.get("allowed_centers") or [])
        admins = [
            a for a in admins
            if a["id"] == _admin.get("sub")
            or set(a.get("allowed_centers") or []).issubset(my_centers)
        ]
    return admins


@auth_router.get("/driver-accounts")
async def list_driver_accounts(_admin: dict = Depends(require_admin)):
    accounts = await db.driver_accounts.find(
        {}, {"_id": 0, "hashed_password": 0}
    ).to_list(1000)
    return accounts


@auth_router.delete("/driver-account/{driver_id}")
async def delete_driver_account(driver_id: str, _admin: dict = Depends(require_admin)):
    result = await db.driver_accounts.delete_one({"driver_id": driver_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="No existe cuenta para este conductor")
    logger.info(f"Cuenta de conductor eliminada: {driver_id}")
    return {"success": True, "driver_id": driver_id}


# =============================================================
# API ROUTES
# =============================================================


@api_router.get("/")
async def root():
    return {"message": "FlotaDSP API funcionando", "version": "5.3.4"}


# =========================
# VEHICLES — solo admin
# =========================

@api_router.get("/vehicles/last-inspections")
async def vehicles_last_inspections(_=Depends(require_admin)):
    """Mapa vehicle_id → fecha de su última inspección (para el semáforo de la lista)."""
    pipeline = [
        {"$match": {"deleted": {"$ne": True}}},
        {"$group": {"_id": "$vehicle_id", "last": {"$max": "$created_at"}}},
    ]
    out = {}
    async for row in db.inspections.aggregate(pipeline):
        if row.get("_id"):
            out[row["_id"]] = row.get("last")
    return out


@api_router.get("/vehicles/portal")
async def get_vehicles_portal(user: dict = Depends(require_any_auth)):
    """Portal conductor: devuelve los vehículos que puede inspeccionar el conductor.
    - Admin → todos los activos
    - Driver → su vehículo asignado; si no tiene, los de su centro
    Este endpoint NO requiere rol admin para que los conductores puedan usarlo."""
    if user.get("role") == "admin":
        vehicles = await db.vehicles.find(
            {"status": {"$ne": "deleted"}}, {"_id": 0}
        ).to_list(1000)
    else:
        driver_id = user["sub"]
        # 1) Vehículo asignado directamente al conductor
        assigned = await db.vehicles.find(
            {"status": {"$ne": "deleted"}, "current_driver_id": driver_id}, {"_id": 0}
        ).to_list(10)

        if assigned:
            vehicles = assigned
        else:
            # 2) Fallback: todos los vehículos del centro del conductor
            driver = await db.drivers.find_one({"id": driver_id}, {"_id": 0, "center": 1})
            center = (driver.get("center") or "")[:4] if driver else ""
            if center:
                vehicles = await db.vehicles.find(
                    {"status": {"$ne": "deleted"}, "center": {"$regex": re.escape(center), "$options": "i"}},
                    {"_id": 0}
                ).to_list(100)
            else:
                vehicles = []

    for v in vehicles:
        for k in ["created_at", "updated_at"]:
            if isinstance(v.get(k), str):
                try:
                    v[k] = datetime.fromisoformat(v[k])
                except Exception:
                    pass
    return vehicles


@api_router.post("/vehicles", response_model=Vehicle)
async def create_vehicle(data: VehicleCreate, _=Depends(require_admin)):
    vehicle = Vehicle(**data.model_dump())
    doc = serialize_doc(vehicle.model_dump())
    await db.vehicles.insert_one(doc)
    return vehicle


@api_router.get("/vehicles", response_model=List[Vehicle])
async def get_vehicles(center: Optional[str] = None, _=Depends(require_admin)):
    query = {"status": {"$ne": "deleted"}}
    if center and center != "Todos":
        if not re.match(r'^[A-Za-z0-9_\-]{1,30}$', center):
            raise HTTPException(400, "Código de centro inválido")
        query["center"] = {"$regex": re.escape(center), "$options": "i"}
    vehicles = await db.vehicles.find(query, {"_id": 0}).to_list(1000)
    for v in vehicles:
        for k in ["created_at", "updated_at"]:
            if isinstance(v.get(k), str):
                v[k] = datetime.fromisoformat(v[k])
    return vehicles


@api_router.get("/vehicles/{vehicle_id}", response_model=Vehicle)
async def get_vehicle(vehicle_id: str, _=Depends(require_admin)):
    v = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not v:
        raise HTTPException(status_code=404, detail="Vehículo no encontrado")
    for k in ["created_at", "updated_at"]:
        if isinstance(v.get(k), str):
            v[k] = datetime.fromisoformat(v[k])
    return v


# =========================
# VEHICLE DOCUMENTS
# =========================

@api_router.get("/vehicles/{vehicle_id}/documents")
async def list_vehicle_documents(vehicle_id: str, _=Depends(require_admin)):
    docs = await db.vehicle_documents.find(
        {"vehicle_id": vehicle_id}, {"_id": 0}
    ).sort("uploaded_at", -1).to_list(100)
    return docs


@api_router.post("/vehicles/{vehicle_id}/documents")
async def upload_vehicle_document(
    vehicle_id: str,
    doc_type: str = Form(...),
    file: UploadFile = File(...),
    _=Depends(require_admin)
):
    v = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0, "license_plate": 1})
    if not v:
        raise HTTPException(status_code=404, detail="Vehículo no encontrado")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Archivo vacío")
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Archivo demasiado grande (máx 20 MB)")

    # Detectar extensión
    orig_name = file.filename or "documento"
    ext = orig_name.rsplit(".", 1)[-1].lower() if "." in orig_name else "pdf"
    safe_type = re.sub(r"[^a-zA-Z0-9_]", "_", doc_type)
    plate = (v.get("license_plate") or vehicle_id).replace(" ", "_")
    object_key = f"docs/{vehicle_id}/{safe_type}_{uuid.uuid4().hex[:8]}.{ext}"

    # Subir a R2
    s3 = get_r2()
    if not s3:
        raise HTTPException(status_code=502, detail="Almacenamiento R2 no configurado")
    bucket = R2_BUCKET
    content_type = file.content_type or "application/octet-stream"
    try:
        await asyncio.get_running_loop().run_in_executor(
            _executor,
            lambda: s3.put_object(
                Bucket=bucket, Key=object_key,
                Body=content, ContentType=content_type
            )
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error subiendo a R2: {e}")

    r2_public = os.environ.get("R2_PUBLIC_URL", "").rstrip("/")
    url = f"{r2_public}/{object_key}" if r2_public else object_key

    doc = {
        "id": str(uuid.uuid4()),
        "vehicle_id": vehicle_id,
        "doc_type": doc_type,
        "name": orig_name,
        "url": url,
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.vehicle_documents.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.delete("/vehicles/{vehicle_id}/documents/{doc_id}")
async def delete_vehicle_document(vehicle_id: str, doc_id: str, _=Depends(require_admin)):
    result = await db.vehicle_documents.delete_one({"id": doc_id, "vehicle_id": vehicle_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    return {"success": True}


# =========================
# SCORING CONDUCTORES
# =========================

@api_router.get("/scoring/drivers")
async def get_driver_scoring(month: int = None, year: int = None, _=Depends(require_admin)):
    """Scoring competitivo de conductores (0-100), 5 pilares:

      📋 Cumplimiento   30 — inspecciones hechas vs días con asignación en el cuadrante
                              (si no hay cuadrantes, días naturales transcurridos)
      ⏰ Puntualidad    15 — hora de subida: 100% antes de las 20:45 (vuelta de ruta),
                              decae hasta 0 a las 23:00
      📸 Evidencia      15 — fotos completas + análisis OK; cada aviso de calidad
                              de la IA (borrosa, zona equivocada) resta
      🔍 Honestidad     15 — si la IA ve daños, ¿los declaró en sus notas?
      🛡️ Conservación   25 — daños NUEVOS en su turno, con TRES garantías de
                              justicia: (1) el veredicto humano de Revisión
                              Rápida manda (✗ nunca penaliza, ✓ cuenta siempre);
                              (2) rate-based: penaliza la TASA de daños, no el
                              volumen — trabajar más días no castiga;
                              (3) suavizado bayesiano con muestras pequeñas.

    🏆 Premio del mes: solo son ELEGIBLES los conductores con ≥35% de los días
    transcurridos asignados en cuadrante (mín. 3) — nadie gana con 3 días buenos.
    Desempate: racha limpia, luego nº de inspecciones. Puntualidad en hora
    Europe/Madrid real (con cambio de hora). Mínimo 3 inspecciones para puntuar.
    """
    import calendar as _cal
    now = datetime.now(timezone.utc)
    y = year or now.year
    m = month or now.month
    days_in_month = _cal.monthrange(y, m)[1]
    month_start = datetime(y, m, 1, tzinfo=timezone.utc)
    month_end = datetime(y, m, days_in_month, 23, 59, 59, tzinfo=timezone.utc)
    days_elapsed = now.day if (y == now.year and m == now.month) else days_in_month

    inspections_month = await db.inspections.find(
        {"deleted": {"$ne": True}, "created_at": {"$gte": month_start.isoformat(), "$lte": month_end.isoformat()}},
        {"_id": 0}
    ).to_list(length=5000)

    drivers = await db.drivers.find({"status": {"$ne": "deleted"}, "active": {"$ne": False}}, {"_id": 0}).to_list(500)

    # Días asignados por conductor según los cuadrantes del mes
    assignments = await db.daily_assignments.find(
        {"date": {"$gte": month_start.strftime("%Y-%m-%d"), "$lte": month_end.strftime("%Y-%m-%d")}},
        {"_id": 0, "date": 1, "slots.driver_id": 1}
    ).to_list(200)
    assigned_days = {}
    for a in assignments:
        for slot in a.get("slots", []):
            did = slot.get("driver_id")
            if did:
                assigned_days.setdefault(did, set()).add(a.get("date"))

    # Historial por vehículo SOLO con análisis válidos (para el delta) — acotado a 6 meses atrás
    _history_start = datetime(y, m, 1, tzinfo=timezone.utc) - __import__('datetime').timedelta(days=180)
    all_inspections = await db.inspections.find(
        {"deleted": {"$ne": True}, "analysis_status": "ok", "analysis": {"$ne": None},
         "created_at": {"$gte": _history_start.isoformat()}},
        {"_id": 0, "id": 1, "vehicle_id": 1, "driver_id": 1, "created_at": 1, "analysis": 1}
    ).sort("created_at", 1).to_list(length=2000)
    vehicle_history = {}
    for insp in all_inspections:
        vid = insp.get("vehicle_id")
        if vid:
            vehicle_history.setdefault(vid, []).append(insp)

    SEV = {"sin_danos": 0, "sin_daños": 0, "": 0, "leve": 1, "moderado": 1, "grave": 2, "critico": 3, "crítico": 3}

    def get_sev(insp):
        a = insp.get("analysis") or {}
        return SEV.get((a.get("severity") or "").lower().strip(), 0)

    def hour_of(iso):
        """Hora local española REAL (Europe/Madrid, con horario de verano/invierno)."""
        try:
            from zoneinfo import ZoneInfo
            dt = datetime.fromisoformat(iso.replace("Z", "+00:00"))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            local = dt.astimezone(ZoneInfo("Europe/Madrid"))
            return local.hour + local.minute / 60.0
        except Exception:
            return None

    # ── Veredictos humanos de Revisión Rápida ──
    # Un daño marcado ✗ (falso positivo) por un inspector NO puede penalizar al
    # conductor; uno marcado ✓ cuenta aunque la IA dudara. El humano manda.
    _fb_docs = await db.ai_feedback.find(
        {"verdict": {"$in": ["wrong", "correct"]}},
        {"_id": 0, "inspection_id": 1, "verdict": 1, "damage.part": 1, "damage.zone": 1, "damage.location": 1}
    ).to_list(5000)
    wrong_panels: dict = {}   # inspection_id → set(panels rechazados por humano)
    correct_panels: dict = {}  # inspection_id → set(panels confirmados por humano)
    for _f in _fb_docs:
        _dmg = _f.get("damage") or {}
        _p = _canon_panel(_dmg.get("part") or _dmg.get("zone") or _dmg.get("location"))
        if not _p:
            continue
        _target = wrong_panels if _f["verdict"] == "wrong" else correct_panels
        _target.setdefault(_f.get("inspection_id"), set()).add(_p)

    results = []
    for driver in drivers:
        driver_id = driver.get("id")
        name = driver.get("name", "—")
        center = driver.get("center", "—")
        driver_insps = [i for i in inspections_month if i.get("driver_id") == driver_id]
        n = len(driver_insps)

        if n < 3:
            results.append({
                "driver_id": driver_id, "name": name, "center": center,
                "photo_url": driver.get("photo_url"),
                "total": None, "inspections_count": n, "insufficient": True,
            })
            continue

        # ── 📋 Cumplimiento (30) ──
        days_assigned = len(assigned_days.get(driver_id, set()))
        denom = days_assigned if days_assigned >= 3 else max(days_elapsed, 1)
        compliance = round(min(30, (n / denom) * 30))

        # ── ⏰ Puntualidad (15) ──
        punct_scores = []
        for insp in driver_insps:
            hr = hour_of(insp.get("created_at", ""))
            if hr is None:
                continue
            # Operación DSP real: inspección al volver de ruta. Puntual = antes
            # de las 20:45; decae hasta 0 a las 23:00. Las de mañana (pre-ruta,
            # antes de las 12:00) también puntúan completo.
            if hr <= 12.0 or hr <= 20.75:
                punct_scores.append(15)
            elif hr >= 23.0:
                punct_scores.append(0)
            else:
                punct_scores.append(15 * (23.0 - hr) / 2.25)
        punctuality = round(sum(punct_scores) / len(punct_scores)) if punct_scores else 8

        # ── 📸 Evidencia (15) ──
        ev_scores = []
        for insp in driver_insps:
            photos = len(insp.get("photo_urls") or insp.get("photos") or [])
            ok = insp.get("analysis_status") == "ok"
            base = 15 if (photos >= 5 and ok) else 11 if photos >= 4 else 6 if photos >= 2 else 0
            warns = len(((insp.get("analysis") or {}).get("image_quality_warnings")) or [])
            ev_scores.append(max(0, base - 3 * warns))
        evidence = round(sum(ev_scores) / len(ev_scores)) if ev_scores else 0

        # ── 🔍 Honestidad (15) ──
        # JUSTO: solo se examina la honestidad cuando hay daño NUEVO serio
        # (abolladura/grave/crítico) en esa inspección. No se pide declarar lo
        # preexistente (el sistema ya lo conoce del historial) ni los rayones
        # leves — así no es una carga y no penaliza por daño que ya estaba.
        h_scores = []
        for insp in driver_insps:
            a = insp.get("analysis") or {}
            serious_new = any(
                _norm_sev(d.get("severity")) in ("moderado", "grave", "critico")
                for d in (a.get("new_damages") or []) if isinstance(d, dict))
            if insp.get("analysis_status") == "ok" and serious_new:
                notes = (insp.get("notes") or "").lower()
                kws = ["daño", "dano", "golpe", "rasguño", "rasguno", "rayad", "abollad",
                       "roto", "rota", "crack", "grieta", "araña", "arana", "choc", "malo", "danado",
                       "raya", "marca", "toque", "incidente"]
                # Declararlo = 15 (transparente). No declararlo NO es prueba de
                # ocultación (quizá no lo vio o no sabía que debía) → suelo de 10,
                # no un cero que hunda el score antes de educar el hábito.
                h_scores.append(15 if any(k in notes for k in kws) else 10)
        honesty = round(sum(h_scores) / len(h_scores)) if h_scores else 15

        # ── 🛡️ Conservación (25) — rate-based, confirmed-only, Bayesian-smoothed ──
        #
        #  Tres garantías de justicia:
        #  1) Solo daños CONFIRMADOS (confidence ≥ 0.65). Los "sugeridos" no penalizan
        #     aunque aparezcan en la foto anotada para revisión del admin.
        #  2) Rate-based: se pondera por nº de inspecciones del mes, así quien trabaja
        #     más días con la misma tasa de incidencias puntúa igual que quien trabaja
        #     menos — la exposición mayor no castiga.
        #  3) Bayesian smoothing (k=5): con pocas inspecciones el score se acerca a la
        #     media perfecta (25 pts), evitando premiar/castigar con muestras pequeñas.
        #     Con ≥15 inspecciones el peso propio del conductor es ya >75%.
        #
        #  Pesos de severidad (escala cuadrática: grave duele 3×, crítico 6×):
        #    leve=0  moderado=1  grave=3  crítico=6
        #  Conservación = 25 × exp(−rate × 1.2),  Bayesian → (raw×n + 25×5)/(n+5)
        import math as _math
        SEV_W = {"leve": 0, "moderado": 1, "grave": 3, "critico": 6}
        delta_events = []
        for insp in driver_insps:
            vid = insp.get("vehicle_id")
            if not vid or vid not in vehicle_history:
                continue
            if insp.get("analysis_status") != "ok" or not insp.get("analysis"):
                continue
            insp_time = insp.get("created_at", "")
            prior = [h for h in vehicle_history[vid] if h.get("created_at", "") < insp_time]
            if not prior:
                continue  # sin baseline → no se puede saber qué es nuevo
            base_panels = set()
            for h in prior:
                for d in ((h.get("analysis") or {}).get("damages") or []):
                    if isinstance(d, dict):
                        p = _canon_panel(d.get("part") or d.get("zone") or d.get("location"))
                        if p:
                            base_panels.add(p)
            photos = insp.get("photo_urls") or insp.get("photos") or []
            curr = {}
            _iid = insp.get("id")
            _wrong = wrong_panels.get(_iid, set())
            _correct = correct_panels.get(_iid, set())
            for d in ((insp.get("analysis") or {}).get("damages") or []):
                if not isinstance(d, dict):
                    continue
                p = _canon_panel(d.get("part") or d.get("zone") or d.get("location"))
                if not p:
                    continue
                # Veredicto humano manda: ✗ nunca penaliza; ✓ cuenta aunque la IA dudara
                if p in _wrong:
                    continue
                if d.get("confirmed") is False and p not in _correct:
                    continue  # sugerido sin confirmación humana → no penaliza
                sev = _norm_sev(d.get("severity"))
                rank = _SEV_RANK[sev]
                if rank > curr.get(p, {}).get("rank", 0):
                    curr[p] = {"rank": rank, "sev": sev, "dmg": d}
            for p, info in curr.items():
                sev = info["sev"]
                w = SEV_W.get(sev, 0)
                if p not in base_panels and w > 0:
                    d = info["dmg"]
                    pidx = d.get("photo_index")
                    photo_url = (photos[pidx - 1] if isinstance(pidx, int) and 1 <= pidx <= len(photos)
                                 else (photos[0] if photos else None))
                    # penalty_display: equivalencia visual con el sistema anterior
                    _PEN_DISP = {"moderado": -6, "grave": -12, "critico": -20}
                    delta_events.append({
                        "vehicle_id": vid,
                        "panel": p,
                        "part": d.get("part") or p,
                        "to_sev": sev,
                        "weight": w,
                        "penalty": _PEN_DISP.get(sev, 0),  # solo display
                        "date": insp_time[:10] if len(insp_time) >= 10 else insp_time,
                        "inspection_id": insp.get("id"),
                        "photo_url": photo_url,
                        "box_2d": d.get("box_2d"),
                        "description": d.get("description"),
                        "confirmed": d.get("confirmed", True),
                    })

        # Rate: peso total / nº inspecciones → independiente del volumen de trabajo
        total_dmg_weight = sum(e["weight"] for e in delta_events)
        damage_rate = total_dmg_weight / max(n, 1)
        # Curva exponencial: rate=0 → 25 pts; rate=1 (1 grave/insp) → ~8 pts; rate=2 → ~2 pts
        raw_conservation = round(25 * _math.exp(-damage_rate * 1.2))
        # Bayesian prior k=5: mezcla el resultado propio con la media perfecta (25)
        # para que los conductores con pocas inspecciones no distorsionen el ranking
        _k = 5
        conservation = round((raw_conservation * n + 25 * _k) / (n + _k))

        # ── 📈 Tendencia del mes (no suma en total — informa al admin) ──
        # Compara la tasa de daño de la primera mitad del mes vs la segunda.
        insps_sorted = sorted(driver_insps, key=lambda x: x.get("created_at", ""))
        half = max(len(insps_sorted) // 2, 1)

        def _half_rate(subset):
            w = sum(
                SEV_W.get(_norm_sev(d.get("severity")), 0)
                for i in subset
                for d in ((i.get("analysis") or {}).get("damages") or [])
                if isinstance(d, dict) and d.get("confirmed") is not False
            )
            return w / max(len(subset), 1)

        if len(insps_sorted) >= 4:
            r1 = _half_rate(insps_sorted[:half])
            r2 = _half_rate(insps_sorted[half:])
            if r2 < r1 * 0.7:
                trend = "mejorando"
            elif r2 > r1 * 1.4:
                trend = "empeorando"
            else:
                trend = "estable"
        else:
            trend = "sin_datos"

        # ── 🔥 Racha limpia (no suma en total — tiebreaker + badge) ──
        # Inspecciones consecutivas más recientes sin daño confirmado ≥ moderado.
        clean_streak = 0
        for insp in reversed(insps_sorted):
            if insp.get("analysis_status") != "ok":
                continue
            new_dmgs = ((insp.get("analysis") or {}).get("new_damages") or
                        (insp.get("analysis") or {}).get("damages") or [])
            has_new_confirmed = any(
                d.get("confirmed") is not False and
                _norm_sev(d.get("severity")) in ("moderado", "grave", "critico")
                for d in new_dmgs if isinstance(d, dict)
            )
            if not has_new_confirmed:
                clean_streak += 1
            else:
                break

        total = min(100, compliance + punctuality + evidence + honesty + conservation)

        # ── 🏆 Elegibilidad para el premio del mes ──
        # Regla explicable: al menos el 35% de los días transcurridos con
        # asignación en el cuadrante (mínimo 3). Sin cuadrantes, cuentan las
        # inspecciones hechas. Evita que una muestra pequeña gane el premio.
        _prize_min = max(3, round(days_elapsed * 0.35))
        _base_days = days_assigned if days_assigned > 0 else n
        prize_eligible = _base_days >= _prize_min

        results.append({
            "driver_id": driver_id, "name": name, "center": center,
            "photo_url": driver.get("photo_url"),
            "total": total,
            "prize_eligible": prize_eligible,
            "compliance": compliance, "punctuality": punctuality,
            "evidence": evidence, "honesty": honesty, "conservation": conservation,
            "damage_rate": round(damage_rate, 3),       # rate crudo (debug / gráficas)
            "trend": trend,                             # mejorando / empeorando / estable
            "clean_streak": clean_streak,               # racha de inspecciones limpias
            "delta_events": delta_events,
            "inspections_count": n,
            "days_assigned": days_assigned,
            "days_elapsed": days_elapsed,
            "insufficient": False,
        })

    # Orden: puntuación desc; desempate por racha limpia; los 'sin datos' al final
    results.sort(key=lambda x: (
        x["total"] is None,
        -(x["total"] or 0),
        -(x.get("clean_streak") or 0),
        -x["inspections_count"],
    ))
    return {"scores": results, "month": m, "year": y, "days_elapsed": days_elapsed,
            "min_inspections": 3, "prize_min_days": max(3, round(days_elapsed * 0.35))}


@api_router.get("/scoring/leaderboard")
async def get_scoring_leaderboard(month: int = None, year: int = None, _=Depends(require_admin)):
    """Top-3 conductores por centro según el scoring mensual.

    Devuelve para cada centro los tres mejores conductores con su puntuación,
    racha limpia, tendencia y una medalla (oro/plata/bronce).
    Solo conductores con al menos 3 inspecciones en el mes.
    """
    import calendar as _cal
    now = datetime.now(timezone.utc)
    y = year or now.year
    m = month or now.month
    days_in_month = _cal.monthrange(y, m)[1]
    month_start = datetime(y, m, 1, tzinfo=timezone.utc)
    month_end = datetime(y, m, days_in_month, 23, 59, 59, tzinfo=timezone.utc)

    # Reusar el endpoint de scoring completo para obtener todos los datos.
    # El PODIO solo admite conductores elegibles para el premio (≥35% de días
    # asignados): nadie gana el mes con una muestra pequeña.
    full = await get_driver_scoring(month=m, year=y)
    scores = [s for s in full["scores"] if not s.get("insufficient") and s.get("prize_eligible")]

    MEDALS = ["🥇", "🥈", "🥉"]
    BADGE_LABELS = {
        "🥇": "oro", "🥈": "plata", "🥉": "bronce",
    }

    leaderboard = {}
    centers_seen = set()
    for s in scores:
        c = s["center"]
        centers_seen.add(c)
        if c not in leaderboard:
            leaderboard[c] = []
        if len(leaderboard[c]) < 3:
            pos = len(leaderboard[c])
            leaderboard[c].append({
                "position": pos + 1,
                "medal": MEDALS[pos],
                "badge": BADGE_LABELS[MEDALS[pos]],
                "driver_id": s["driver_id"],
                "name": s["name"],
                "photo_url": s.get("photo_url"),
                "total": s["total"],
                "trend": s.get("trend", "sin_datos"),
                "clean_streak": s.get("clean_streak", 0),
                "inspections_count": s["inspections_count"],
                "compliance": s.get("compliance"),
                "conservation": s.get("conservation"),
                "damage_rate": s.get("damage_rate"),
                "prize_eligible": True,
            })

    # Estadísticas por centro (para comparativas)
    center_stats = {}
    for c in centers_seen:
        center_scores = [s["total"] for s in scores if s["center"] == c]
        if center_scores:
            center_stats[c] = {
                "drivers_ranked": len(center_scores),
                "avg_score": round(sum(center_scores) / len(center_scores), 1),
                "top_score": max(center_scores),
            }

    return {
        "month": m, "year": y,
        "leaderboard": leaderboard,
        "center_stats": center_stats,
        "prize_min_days": full.get("prize_min_days"),
    }


@api_router.patch("/vehicles/{vehicle_id}")
async def update_vehicle(vehicle_id: str, data: dict, _=Depends(require_admin)):
    _VEHICLE_ALLOWED = {
        "license_plate","brand","model","year","color","fuel_type","status",
        "center","notes","current_driver_id","mileage","last_itv","next_itv",
        "last_service","next_service","photo_url","vin","seats","load_capacity",
        "acquisition_date","insurance_expiry","leasing","owner","gps_id","tags",
        # Campos que edita la ficha del panel (antes se filtraban y el guardado
        # fallaba en silencio): tipo, ITV, renting, proveedor y motivo de taller.
        "vehicle_type","itv_date","renting_end_date","renting_baja_date",
        "provider","workshop_status","workshop_reason",
    }
    data = {k: v for k, v in data.items() if k in _VEHICLE_ALLOWED}
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    prev = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0, "status": 1})
    result = await db.vehicles.update_one({"id": vehicle_id}, {"$set": data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Vehículo no encontrado")
    if "status" in data:
        try:
            await _auto_incident_on_workshop(vehicle_id, (prev or {}).get("status"), data.get("status"))
        except Exception as _ai:
            logger.warning(f"Auto-incidencia taller: {_ai}")
    return {"success": True}


# =========================
# DRIVERS — solo admin
# =========================

@api_router.post("/drivers", response_model=Driver)
async def create_driver(data: DriverCreate, admin: dict = Depends(require_admin)):
    driver_data = data.model_dump()
    password = driver_data.pop("password", None)

    driver = Driver(**driver_data)
    doc = serialize_doc(driver.model_dump())
    await db.drivers.insert_one(doc)

    if password and driver.email:
        await db.driver_accounts.insert_one({
            "id": str(uuid.uuid4()),
            "driver_id": driver.id,
            "email": driver.email,
            "hashed_password": hash_password(password),
            "active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        })
        logger.info(f"Conductor y cuenta creados: {driver.email}")

    return driver


@api_router.get("/drivers", response_model=List[Driver])
async def get_drivers(center: Optional[str] = None, _=Depends(require_admin)):
    query = {"active": {"$ne": False}}
    if center and center != "Todos":
        if not re.match(r'^[A-Za-z0-9_\-]{1,30}$', center):
            raise HTTPException(400, "Código de centro inválido")
        query["center"] = {"$regex": re.escape(center), "$options": "i"}
    drivers = await db.drivers.find(query, {"_id": 0}).to_list(1000)
    return drivers


@api_router.patch("/drivers/{driver_id}")
async def update_driver(driver_id: str, data: dict, _=Depends(require_admin)):
    _DRIVER_ALLOWED = {
        "name","email","phone","center","notes","photo_url","active","dni",
        "license_number","license_expiry","address","emergency_contact",
        "contract_type","hire_date","tags","score_override",
        # Campos del formulario del panel (antes se filtraban y el guardado
        # fallaba en silencio): contrato/nivel/zona + ficha extendida.
        "contrato","nivel","zona","driver_id","alojamiento","notas","login",
    }
    data = {k: v for k, v in data.items() if k in _DRIVER_ALLOWED}
    result = await db.drivers.update_one({"id": driver_id}, {"$set": data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Conductor no encontrado")
    return {"success": True}


# =========================
# DRIVER — su propio perfil
# =========================

@api_router.get("/me/driver")
async def get_my_driver_profile(user: dict = Depends(require_any_auth)):
    if user["role"] == "admin":
        raise HTTPException(status_code=400, detail="Usa /api/auth/me para admins")
    driver = await db.drivers.find_one({"id": user["sub"]}, {"_id": 0})
    if not driver:
        raise HTTPException(status_code=404, detail="Perfil no encontrado")
    return driver


@api_router.get("/me/vehicle")
async def get_my_current_vehicle(user: dict = Depends(require_any_auth)):
    vehicle = await db.vehicles.find_one(
        {"current_driver_id": user["sub"]}, {"_id": 0}
    )
    if not vehicle:
        raise HTTPException(status_code=404, detail="No tienes ningún vehículo asignado")
    return vehicle


# =========================
# INSPECTIONS — upload
# =========================

@api_router.post("/inspections/upload")
async def upload_inspection_photos(
    vehicle_id: str = Form(...),
    driver_id: Optional[str] = Form(None),
    notes: str = Form(""),
    files: List[UploadFile] = File(...),
    user: dict = Depends(require_any_auth)
):
    notes = (notes or "")[:2000]
    if not files:
        raise HTTPException(status_code=400, detail="Se requiere al menos una imagen.")
    if len(files) > 20:
        raise HTTPException(status_code=400, detail="Máximo 20 imágenes por inspección.")

    # Anti-duplicados: una inspección por vehículo+conductor al día (evita doble
    # tap con mala cobertura, análisis duplicados y Telegram repetido).
    # Los admins pueden repetir (peritajes manuales desde el panel).
    if user.get("role") == "driver":
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        dup = await db.inspections.find_one({
            "deleted": {"$ne": True}, "vehicle_id": vehicle_id,
            "driver_id": driver_id or user.get("sub"),
            "created_at": {"$regex": f"^{today}"}
        }, {"_id": 0, "id": 1})
        if dup:
            raise HTTPException(
                status_code=409,
                detail="Ya enviaste la inspección de esta furgoneta hoy. Si necesitas repetirla, avisa a tu coordinador."
            )

    # Un conductor solo puede subir inspecciones de vehículos de su centro
    if user["role"] == "driver":
        driver_id = user["sub"]
        driver_doc = await db.drivers.find_one({"id": driver_id}, {"_id": 0, "center": 1})
        driver_center = (driver_doc.get("center") or "")[:4] if driver_doc else ""
        vehicle_doc = await db.vehicles.find_one({"id": vehicle_id, "status": {"$ne": "deleted"}}, {"_id": 0, "center": 1})
        if not vehicle_doc:
            raise HTTPException(status_code=404, detail="Vehículo no encontrado")
        vehicle_center = (vehicle_doc.get("center") or "")[:4]
        # Verificar que el conductor pertenece al mismo centro que el vehículo
        # El centro del conductor es corto (ej: "OGA5") y el del vehículo puede ser largo
        # (ej: "AMZL OGA5 SANTIAGO XPT") — basta con que uno contenga al otro
        if driver_center and vehicle_center:
            dc = driver_center.upper()
            vc = vehicle_doc.get("center", "").upper()
            if dc not in vc and vc[:4] not in dc:
                raise HTTPException(
                    status_code=403,
                    detail="Solo puedes subir inspecciones de vehículos de tu centro"
                )

    try:
        photo_urls = []
        photos_base64 = []
        photo_roles = []  # angle | odometer | checklist — por nombre de archivo del portal

        for file in files:
            content = await file.read()
            if not content:
                logger.warning(f"Fichero vacío recibido: {file.filename} — ignorado")
                continue
            validate_image_content(content)
            photo_url, processed_bytes = await process_and_save_image(content, vehicle_id)
            photo_urls.append(photo_url)
            photos_base64.append(base64.b64encode(processed_bytes).decode("utf-8"))
            _fn = (file.filename or "").lower()
            photo_roles.append(
                "odometer" if _fn.startswith("odometro")
                else "checklist" if _fn.startswith("checklist")
                else "angle")

        if not photo_urls:
            raise HTTPException(status_code=400, detail="No se procesó ninguna imagen válida.")

        # El análisis de daños SOLO mira la carrocería: las fotos del cuentakm y
        # del checklist confundían a la IA (daños fantasma en salpicaderos) y
        # gastaban tokens. Se guardan igual, pero no entran al análisis.
        analysis_photo_idx = [i for i, r in enumerate(photo_roles) if r == "angle"]
        if not analysis_photo_idx:  # nombres inesperados → comportamiento clásico
            analysis_photo_idx = list(range(len(photo_urls)))
        analysis_b64 = [photos_base64[i] for i in analysis_photo_idx]
        # índice en el análisis (1-based) → índice real en photos (1-based)
        analysis_idx_map = {k + 1: analysis_photo_idx[k] + 1 for k in range(len(analysis_photo_idx))}

        logger.info(f"Guardadas {len(photo_urls)} fotos — vehículo {vehicle_id}")

        # Cargar fotos de referencia de la última inspección analizada
        ref_results = await db.inspections.find(
            {"deleted": {"$ne": True}, "vehicle_id": vehicle_id, "analysis": {"$ne": None}, "analysis_status": "ok"},
            {"_id": 0, "photos": 1}
        ).sort("created_at", -1).to_list(1)

        last_insp = ref_results[0] if ref_results else None
        ref_photo_urls = last_insp.get("photos", [])[:4] if last_insp else []

        # Cargar bytes de referencias desde R2 o disco
        ref_bytes_list: List[bytes] = []
        if ref_photo_urls:
            logger.info(f"Cargando {len(ref_photo_urls)} fotos de referencia")
            ref_bytes_list = await load_reference_images(ref_photo_urls)
            logger.info(f"Referencia cargada: {len(ref_bytes_list)} imágenes")

        # Guardar inspección PRIMERO (sin esperar a Gemini) para responder rápido al conductor
        inspection = Inspection(
            vehicle_id=vehicle_id,
            driver_id=driver_id,
            photos=photo_urls,
            reference_photos=(ref_photo_urls[:2] if ref_photo_urls else []),
            analysis=None,
            analysis_status="pending",
            analysis_error=None,
            notes=notes,
            analyzed_at=None
        )
        doc = serialize_doc(inspection.model_dump())
        doc["photo_roles"] = photo_roles  # para reanálisis: qué fotos son carrocería
        await db.inspections.insert_one(doc)

        # Actualizar automáticamente el kilometraje de la furgoneta con el que la
        # IA leyó del cuentakilómetros en la inspección (el portal del conductor
        # lo manda en notes como odometer_km). Así el km real se refleja al
        # instante en el panel web y en la app, sin ningún paso manual.
        try:
            odo_km = None
            try:
                _n = json.loads(notes) if notes else None
                if isinstance(_n, dict):
                    odo_km = _n.get("odometer_km")
            except Exception:
                odo_km = None
            if isinstance(odo_km, str) and odo_km.strip().isdigit():
                odo_km = int(odo_km.strip())
            if isinstance(odo_km, (int, float)) and odo_km > 0:
                odo_km = int(odo_km)
                veh = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0, "mileage": 1})
                current = (veh or {}).get("mileage") or 0
                # Solo si sube: evita retrocesos por lecturas OCR erróneas.
                if odo_km >= current:
                    await db.vehicles.update_one(
                        {"id": vehicle_id},
                        {"$set": {"mileage": odo_km, "updated_at": datetime.now(timezone.utc)},
                         "$push": {"mileage_history": {
                             "date": datetime.now(timezone.utc).isoformat(),
                             "km": odo_km, "source": "inspection"}}}
                    )
                    logger.info(f"Km auto-actualizados por inspección: {vehicle_id} → {odo_km}")
        except Exception as _km_e:
            logger.warning(f"Auto-km inspección: {_km_e}")

        # Llamar a Gemini en background SOLO si el plan lo permite
        org = await get_org(user.get("org_id"))
        plan_ai = _org_billing(org)["limits"].get("ai", False)

        async def _analyze_and_update():
            try:
                # ── CV PRIMERO: el detector propio ve las fotos antes que Gemini.
                # Sus detecciones (con pieza por geometría) guían el prompt y
                # después corrigen las cajas (snap) sin llamar al servicio 2 veces.
                cv_dets_full = {}  # índice real en photos (1-based) → detecciones
                cv_text = ""
                if AI_SERVICE_URL:
                    try:
                        _cv_lines = []
                        for _k, _full_i in enumerate(analysis_photo_idx, start=1):
                            dets = await _call_ai_service_detect(
                                inspection.id, _full_i, base64.b64decode(photos_base64[_full_i]))
                            if dets:
                                cv_dets_full[_full_i + 1] = dets
                                for det in dets:
                                    _pnl = f" en {det.panel}" if getattr(det, "panel", None) else ""
                                    _cv_lines.append(
                                        f"Foto {_k}: {det.label}{_pnl} (confianza {det.confidence:.2f})")
                        if _cv_lines:
                            cv_text = (
                                "\n=== DETECCIONES DEL DETECTOR CV PROPIO (geometría verificada) ===\n"
                                + "\n".join(_cv_lines) +
                                "\nEstas detecciones vienen de un modelo entrenado con miles de daños "
                                "reales de carrocería; la pieza está determinada por intersección "
                                "geométrica, no por estimación. Úsalas como BASE del análisis: "
                                "confirma y describe con precisión cada una que veas en la foto. "
                                "NO reportes daños adicionales salvo evidencia clara e inequívoca; "
                                "si no ves alguna de estas detecciones, inclúyela con confidence < 0.5."
                            )
                    except Exception as _cve:
                        logger.warning(f"CV pre-análisis: {_cve}")

                analysis, analysis_status, analysis_error = await asyncio.wait_for(
                    analyze_images_with_gemini(analysis_b64, ref_bytes_list if ref_bytes_list else None, db=db,
                                               cv_detections_text=cv_text),
                    timeout=120.0
                )
                # Los photo_index de Gemini son relativos a las fotos analizadas
                # (solo carrocería); remapear al índice real dentro de photos.
                if analysis:
                    _remap_photo_indexes(analysis, analysis_idx_map)
                # Filtro determinista: un panel ya dañado antes NO es "nuevo" otra vez
                if analysis and analysis_status == "ok" and getattr(analysis, "new_damages", None):
                    try:
                        known = await _known_damaged_panels(
                            vehicle_id, before_iso=doc.get("created_at"), exclude_id=inspection.id)
                        if known:
                            kept = [nd for nd in analysis.new_damages
                                    if _canon_panel(getattr(nd, "part", None) or getattr(nd, "zone", None)) not in known]
                            if len(kept) != len(analysis.new_damages):
                                analysis.new_damages = kept
                                analysis.new_damages_count = len(kept)
                    except Exception as _fe:
                        logger.warning(f"Filtro panel-historial: {_fe}")
                # Coste tipo taller (por panel) como total de la inspección
                if analysis and analysis_status == "ok":
                    try:
                        dmg_dicts = [d.model_dump() if hasattr(d, "model_dump") else d
                                     for d in (analysis.damages or [])]
                        analysis.total_estimated_cost = float(_vehicle_panel_cost(dmg_dicts)[0])
                    except Exception as _ce:
                        logger.warning(f"Coste por panel: {_ce}")
                # Snap de cajas al detector CV: el LLM decide el QUÉ, YOLO el DÓNDE
                # (reutiliza las detecciones ya pedidas antes del análisis)
                if cv_dets_full and analysis and analysis_status == "ok" and (analysis.damages or analysis.new_damages):
                    try:
                        _s1 = _snap_damage_boxes_to_yolo(analysis.damages or [], cv_dets_full)
                        _s2 = _snap_damage_boxes_to_yolo(analysis.new_damages or [], cv_dets_full)
                        if _s1 or _s2:
                            logger.info(f"Snap YOLO: {_s1}+{_s2} cajas corregidas insp={inspection.id[:8]}")
                    except Exception as _se:
                        logger.warning(f"Snap YOLO: {_se}")
                await db.inspections.update_one(
                    {"id": inspection.id},
                    {"$set": {"analysis": serialize_doc(analysis.model_dump()) if analysis else None,
                              "analysis_status": analysis_status,
                              "analysis_error": analysis_error,
                              "analyzed_at": datetime.now(timezone.utc)}}
                )
                # ─ Generar fotos anotadas profesionalmente ─
                if analysis and analysis_status == "ok" and analysis.damages:
                    try:
                        ann_urls = await generate_annotated_photos(
                            inspection.id, photo_urls, analysis.damages,
                            photos_bytes=photos_base64 and
                            [base64.b64decode(b) for b in photos_base64]
                        )
                        if any(u for u in ann_urls):
                            await db.inspections.update_one(
                                {"id": inspection.id},
                                {"$set": {"annotated_photos": ann_urls}}
                            )
                    except Exception as _ae:
                        logger.warning(f"Error generando anotaciones: {_ae}")
                # Notificar por Telegram si hay daños graves o críticos
                if analysis and analysis_status == "ok":
                    sev = (analysis.severity or "").lower()
                    if sev in ("grave", "critico"):
                        # Obtener matrícula y centro del vehículo
                        veh_doc = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0, "license_plate": 1, "center": 1})
                        plate = veh_doc.get("license_plate", vehicle_id) if veh_doc else vehicle_id
                        veh_center = veh_doc.get("center", "") if veh_doc else ""
                        # Obtener nombre y centro del conductor
                        drv_name = "Desconocido"
                        drv_center = ""
                        if driver_id:
                            drv_doc = await db.drivers.find_one({"id": driver_id}, {"_id": 0, "name": 1, "center": 1})
                            if drv_doc:
                                drv_name = drv_doc.get("name", "Desconocido")
                                drv_center = drv_doc.get("center", "")
                        center = drv_center or veh_center or ""
                        try:
                            await send_telegram_damage_alert(
                                plate=plate,
                                driver_name=drv_name,
                                analysis=analysis,
                                photo_urls=photo_urls,
                                inspection_id=inspection.id,
                                center=center
                            )
                        except Exception as _te:
                            logger.warning(f"Telegram alert falló (no crítico): {_te}")
            except Exception as _e:
                logger.error(f"Análisis Gemini background falló: {_e}")
                await db.inspections.update_one(
                    {"id": inspection.id},
                    {"$set": {"analysis_status": "error", "analysis_error": str(_e)}}
                )
        if plan_ai:
            asyncio.create_task(_analyze_and_update())
        else:
            await db.inspections.update_one(
                {"id": inspection.id},
                {"$set": {"analysis_status": "plan_disabled", "analysis_error": "IA no incluida en tu plan"}}
            )

        # Lanzar detección YOLO en background (no bloquea la respuesta al usuario)
        asyncio.create_task(_run_yolo_for_inspection(inspection.id, photo_urls))
        # Lanzar fraud check en background (depende de analysis.plate_text, así que esperamos un poco).
        async def _delayed_fraud():
            await asyncio.sleep(45)  # da tiempo al análisis Gemini para escribir plate_text
            try: await _calculate_fraud_score(inspection.id)
            except Exception as e: logger.warning(f"fraud check bg failed: {e}")
        asyncio.create_task(_delayed_fraud())

        # Respuesta inmediata — análisis Gemini se completa en background
        return {
            "success": True,
            "inspection_id": inspection.id,
            "analysis": None,
            "analysis_status": "pending",
            "analysis_error": None,
            "photos": photo_urls,
            "message": "Inspección enviada correctamente. Las fotos se han guardado."
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error upload inspección: {type(e).__name__}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))


@api_router.get("/inspections", response_model=List[Inspection])
async def get_inspections(
    vehicle_id: Optional[str] = None,
    center: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 100,
    skip: int = 0,
    _=Depends(require_admin),
):
    query = {"deleted": {"$ne": True}}
    if vehicle_id:
        query["vehicle_id"] = vehicle_id
    if center and center != "Todos":
        query["center"] = center
    if date_from:
        query.setdefault("created_at", {})["$gte"] = date_from
    if date_to:
        query.setdefault("created_at", {})["$lte"] = date_to + "T23:59:59"
    limit = max(1, min(limit, 500))
    skip = max(0, skip)
    inspections = await db.inspections.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).to_list(limit)
    return inspections


@api_router.get("/inspections/vehicle/{vehicle_id}")
async def get_vehicle_inspections(vehicle_id: str, user: dict = Depends(require_any_auth)):
    if user["role"] == "driver":
        assigned = await db.vehicles.find_one({"id": vehicle_id, "current_driver_id": user["sub"]})
        if not assigned:
            raise HTTPException(status_code=403, detail="Acceso denegado")

    inspections = await db.inspections.find(
        {"deleted": {"$ne": True}, "vehicle_id": vehicle_id}, {"_id": 0}
    ).sort("created_at", -1).to_list(50)

    # Enrich with driver name
    driver_ids = list({i["driver_id"] for i in inspections if i.get("driver_id")})
    if driver_ids:
        drivers = await db.drivers.find({"id": {"$in": driver_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(200)
        name_map = {d["id"]: d["name"] for d in drivers}
        for insp in inspections:
            if insp.get("driver_id"):
                insp["driver_name"] = name_map.get(insp["driver_id"], "")

    return inspections


@api_router.get("/vehicles/{vehicle_id}/damage-ledger")
async def get_vehicle_damage_ledger(vehicle_id: str, _admin: dict = Depends(require_admin)):
    """Ledger de daños del vehículo (gemelo digital 3D): daños abiertos
    (con backfill perezoso si aún no existe) + historial de reparados."""
    open_entries = await _get_vehicle_ledger(vehicle_id)
    repaired = await db.vehicle_damage_ledger.find(
        {"vehicle_id": vehicle_id, "status": "repaired"}, {"_id": 0}
    ).sort("updated_at", -1).to_list(100)
    return {"open": open_entries, "repaired": repaired}


# =========================================================================
# GEMELO DIGITAL — CATÁLOGO DE MODELOS + RESOLVER (marca/modelo → malla 3D)
# =========================================================================
# Cada entrada lleva la CONFIG DE CARROCERÍA real del modelo (dimensiones,
# silueta) para que el modelo provisional se parezca a ESE vehículo y no a una
# furgoneta genérica, y un `glb_url` para el modelo 3D real cuando lo tengamos
# (se sirve desde nuestro CDN/R2). Mientras glb_url es None → provisional.
# body: L=largo H=alto W=ancho (m), cab=proporción cabina, roofDrop=caída de
# techo cabina, nose=morro. Fuente: fichas técnicas de cada modelo.
VEHICLE_MODEL_CATALOG = {
    "mercedes_sprinter": {"name": "Mercedes-Benz Sprinter", "match": ["sprinter"],
        "body": {"L": 5.93, "H": 2.62, "W": 2.02, "cab": 0.30, "roofDrop": 0.16, "nose": 0.10}, "glb_url": None},
    "vw_crafter": {"name": "Volkswagen Crafter", "match": ["crafter"],
        "body": {"L": 5.99, "H": 2.59, "W": 2.04, "cab": 0.30, "roofDrop": 0.16, "nose": 0.11}, "glb_url": None},
    "man_tge": {"name": "MAN TGE", "match": ["tge"],
        "body": {"L": 5.98, "H": 2.59, "W": 2.04, "cab": 0.30, "roofDrop": 0.16, "nose": 0.11}, "glb_url": None},
    "ford_transit": {"name": "Ford Transit", "match": ["transit"],
        "body": {"L": 5.98, "H": 2.55, "W": 2.06, "cab": 0.32, "roofDrop": 0.20, "nose": 0.16}, "glb_url": None},
    "ford_transit_custom": {"name": "Ford Transit Custom", "match": ["transit custom", "custom", "tourneo custom"],
        "body": {"L": 5.34, "H": 1.98, "W": 1.99, "cab": 0.36, "roofDrop": 0.14, "nose": 0.22}, "glb_url": None},
    "renault_master": {"name": "Renault Master", "match": ["master"],
        "body": {"L": 6.20, "H": 2.50, "W": 2.07, "cab": 0.31, "roofDrop": 0.18, "nose": 0.17}, "glb_url": None},
    "renault_trafic": {"name": "Renault Trafic", "match": ["trafic", "traffic"],
        "body": {"L": 5.48, "H": 1.97, "W": 1.96, "cab": 0.36, "roofDrop": 0.13, "nose": 0.22}, "glb_url": None},
    "opel_vivaro": {"name": "Opel Vivaro", "match": ["vivaro"],
        "body": {"L": 5.30, "H": 1.93, "W": 1.92, "cab": 0.37, "roofDrop": 0.13, "nose": 0.22}, "glb_url": None},
    "opel_movano": {"name": "Opel Movano", "match": ["movano"],
        "body": {"L": 6.20, "H": 2.50, "W": 2.07, "cab": 0.31, "roofDrop": 0.18, "nose": 0.17}, "glb_url": None},
    "fiat_ducato": {"name": "Fiat Ducato", "match": ["ducato"],
        "body": {"L": 5.99, "H": 2.52, "W": 2.05, "cab": 0.30, "roofDrop": 0.15, "nose": 0.14}, "glb_url": None},
    "peugeot_boxer": {"name": "Peugeot Boxer", "match": ["boxer"],
        "body": {"L": 5.99, "H": 2.52, "W": 2.05, "cab": 0.30, "roofDrop": 0.15, "nose": 0.14}, "glb_url": None},
    "citroen_jumper": {"name": "Citroën Jumper", "match": ["jumper", "relay"],
        "body": {"L": 5.99, "H": 2.52, "W": 2.05, "cab": 0.30, "roofDrop": 0.15, "nose": 0.14}, "glb_url": None},
    "peugeot_expert": {"name": "Peugeot Expert", "match": ["expert"],
        "body": {"L": 5.31, "H": 1.94, "W": 1.92, "cab": 0.37, "roofDrop": 0.13, "nose": 0.22}, "glb_url": None},
    "citroen_jumpy": {"name": "Citroën Jumpy", "match": ["jumpy", "dispatch"],
        "body": {"L": 5.31, "H": 1.94, "W": 1.92, "cab": 0.37, "roofDrop": 0.13, "nose": 0.22}, "glb_url": None},
    "toyota_proace": {"name": "Toyota Proace", "match": ["proace"],
        "body": {"L": 5.31, "H": 1.94, "W": 1.92, "cab": 0.37, "roofDrop": 0.13, "nose": 0.22},
        "glb_url": "/models/toyota_proace.glb"},
    "iveco_daily": {"name": "Iveco Daily", "match": ["daily"],
        "body": {"L": 6.00, "H": 2.60, "W": 2.00, "cab": 0.30, "roofDrop": 0.16, "nose": 0.13}, "glb_url": None},
    "vw_transporter": {"name": "Volkswagen Transporter", "match": ["transporter", "t6", "t5", "caravelle"],
        "body": {"L": 5.30, "H": 1.99, "W": 1.90, "cab": 0.38, "roofDrop": 0.12, "nose": 0.20}, "glb_url": None},
    "nissan_primastar": {"name": "Nissan Primastar", "match": ["primastar", "nv300"],
        "body": {"L": 5.48, "H": 1.97, "W": 1.96, "cab": 0.36, "roofDrop": 0.13, "nose": 0.22}, "glb_url": None},
    "mercedes_vito": {"name": "Mercedes-Benz Vito", "match": ["vito", "viano"],
        "body": {"L": 5.14, "H": 1.91, "W": 1.93, "cab": 0.38, "roofDrop": 0.12, "nose": 0.20}, "glb_url": None},
    "vw_caddy": {"name": "Volkswagen Caddy", "match": ["caddy"],
        "body": {"L": 4.50, "H": 1.83, "W": 1.79, "cab": 0.44, "roofDrop": 0.10, "nose": 0.26}, "glb_url": None},
    "citroen_berlingo": {"name": "Citroën Berlingo", "match": ["berlingo", "partner", "combo", "doblo", "kangoo", "rifter"],
        "body": {"L": 4.40, "H": 1.80, "W": 1.85, "cab": 0.45, "roofDrop": 0.10, "nose": 0.24}, "glb_url": None},
}


def _resolve_vehicle_model(brand: str = "", model: str = "") -> dict:
    """Marca+modelo (texto libre) → mejor entrada del catálogo. El match más
    específico (más largo) gana, para que 'Transit Custom' no caiga en 'Transit'."""
    s = f"{brand} {model}".lower().strip()
    best_key, best_len = None, 0
    for key, ent in VEHICLE_MODEL_CATALOG.items():
        for kw in ent["match"]:
            if kw in s and len(kw) > best_len:
                best_key, best_len = key, len(kw)
    if not best_key:
        return {"key": None, "name": None, "body": None, "glb_url": None, "provisional": True}
    ent = VEHICLE_MODEL_CATALOG[best_key]
    return {"key": best_key, "name": ent["name"], "body": ent["body"],
            "glb_url": ent.get("glb_url"), "provisional": ent.get("glb_url") is None}


@api_router.get("/vehicle-models/resolve")
async def resolve_vehicle_model(brand: str = "", model: str = "", year: Optional[int] = None,
                                version: str = "", _admin: dict = Depends(require_admin)):
    """VehicleModelResolver: devuelve la config de carrocería del modelo y, si
    existe, la URL del GLB real. year/version reservados para futuras variantes."""
    return _resolve_vehicle_model(brand, model)


@api_router.post("/vehicles/{vehicle_id}/identify-model")
async def identify_vehicle_model(vehicle_id: str, _admin: dict = Depends(require_admin)):
    """La IA (Gemini) identifica marca/modelo/carrocería/color mirando las fotos
    de la última inspección del vehículo. Se guarda en vehicle.ai_model."""
    v = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not v:
        raise HTTPException(status_code=404, detail="Vehículo no encontrado")
    insp = await db.inspections.find_one(
        {"vehicle_id": vehicle_id, "deleted": {"$ne": True}, "photos.0": {"$exists": True}},
        {"_id": 0, "photos": 1}, sort=[("created_at", -1)])
    photos = (insp or {}).get("photos") or []
    if not photos:
        raise HTTPException(status_code=400, detail="Este vehículo aún no tiene fotos de inspección para identificar el modelo")

    import httpx as _httpx
    imgs = []
    async with _httpx.AsyncClient(timeout=20) as _c:
        for url in photos[:4]:
            try:
                r = await _c.get(url)
                if r.status_code == 200 and r.content:
                    imgs.append(r.content)
            except Exception:
                pass
    if not imgs:
        raise HTTPException(status_code=502, detail="No se pudieron descargar las fotos para el análisis")

    prompt = (
        "Eres un perito de vehículos comerciales. Observa estas fotos de una furgoneta "
        "de reparto y deduce el MODELO EXACTO. Devuelve SOLO un JSON con: "
        '{"brand": marca (ej. Mercedes-Benz, Ford, Renault, Fiat, Peugeot, Citroën, '
        'Volkswagen, Iveco, Opel, Toyota, Nissan), "model": modelo (ej. Sprinter, '
        'Transit, Transit Custom, Master, Trafic, Ducato, Boxer, Jumper, Daily, '
        'Crafter, Vivaro, Proace, Expert, Transporter, Vito, Caddy, Berlingo), '
        '"body_type": carrocería (ej. L2H2, furgón medio, furgón largo, combi), '
        '"color": color dominante en español, "confidence": 0.0-1.0}. '
        "Si no estás seguro del modelo exacto, da tu mejor estimación y baja la confidence. "
        "No inventes: si solo ves parte del vehículo, usa lo visible."
    )
    try:
        from google.genai import types as genai_types
        client = _gemini_client_plantilla()
        model_name = os.environ.get("GEMINI_MODEL", "gemini-2.5-flash")
        contents = [prompt] + [genai_types.Part.from_bytes(data=b, mime_type="image/jpeg") for b in imgs]
        cfg = genai_types.GenerateContentConfig(
            response_mime_type="application/json",
            thinking_config=genai_types.ThinkingConfig(thinking_budget=0))
        loop = asyncio.get_event_loop()
        resp = await asyncio.wait_for(
            loop.run_in_executor(_executor, lambda: client.models.generate_content(
                model=model_name, contents=contents, config=cfg)),
            timeout=45.0)
        parsed = json.loads((resp.text or "{}").strip())
    except Exception as e:
        logger.warning(f"identify-model {vehicle_id}: {e}")
        raise HTTPException(status_code=502, detail="La IA no pudo identificar el modelo ahora mismo, inténtalo de nuevo")

    ai_model = {
        "brand": (parsed.get("brand") or "").strip(),
        "model": (parsed.get("model") or "").strip(),
        "body_type": (parsed.get("body_type") or "").strip(),
        "color": (parsed.get("color") or "").strip(),
        "confidence": float(parsed.get("confidence") or 0.0),
        "identified_at": datetime.now(timezone.utc).isoformat(),
        "source": "gemini_photos",
    }
    resolved = _resolve_vehicle_model(ai_model["brand"], ai_model["model"])
    await db.vehicles.update_one({"id": vehicle_id}, {"$set": {"ai_model": ai_model}})
    return {"ai_model": ai_model, "resolved": resolved}


@api_router.get("/inspections/review-queue")
async def get_review_queue(center: Optional[str] = None, _=Depends(require_admin)):
    """Cola de inspecciones pendientes de revisar (reviewed != true), enriquecidas
    con matrícula y nombre del conductor. Más recientes primero."""
    query = {"reviewed": {"$ne": True}, "deleted": {"$ne": True}}
    insps = await db.inspections.find(query, {"_id": 0}).sort("created_at", -1).to_list(200)

    # Enriquecer con matrícula y conductor (lookups en lote)
    veh_ids = list({i.get("vehicle_id") for i in insps if i.get("vehicle_id")})
    drv_ids = list({i.get("driver_id") for i in insps if i.get("driver_id")})
    vehicles = {v["id"]: v for v in await db.vehicles.find(
        {"id": {"$in": veh_ids}}, {"_id": 0, "id": 1, "license_plate": 1, "brand": 1, "model": 1, "center": 1}
    ).to_list(500)}
    drivers = {dr["id"]: dr for dr in await db.drivers.find(
        {"id": {"$in": drv_ids}}, {"_id": 0, "id": 1, "name": 1, "center": 1}
    ).to_list(500)}

    out = []
    for i in insps:
        v = vehicles.get(i.get("vehicle_id"), {})
        dr = drivers.get(i.get("driver_id"), {})
        i_center = _normalize_center_code(dr.get("center") or v.get("center") or "")
        if center and center != "Todos" and i_center != center:
            continue
        analysis = i.get("analysis") or {}
        # Verificación de matrícula: la leída en las fotos vs la del vehículo
        detected_plate = (analysis.get("detected_plate") or "").replace(" ", "").replace("-", "").upper()
        real_plate = (v.get("license_plate") or "").replace(" ", "").replace("-", "").upper()
        plate_mismatch = bool(detected_plate and real_plate and detected_plate != real_plate)
        raw_new = analysis.get("new_damages") or []
        new_damages = [{
            "part": nd.get("part", ""),
            "severity": nd.get("severity", ""),
            "description": nd.get("description", ""),
            "location_hint": nd.get("location_hint", ""),
            "photo_index": nd.get("photo_index"),
            "box_2d": nd.get("box_2d"),
        } for nd in raw_new if isinstance(nd, dict)]
        out.append({
            "id": i.get("id"),
            "created_at": i.get("created_at"),
            "vehicle_id": i.get("vehicle_id"),
            "license_plate": v.get("license_plate", "—"),
            "vehicle_label": f"{v.get('brand', '')} {v.get('model', '')}".strip(),
            "driver_name": dr.get("name", "—"),
            "center": i_center,
            "photos": i.get("photos", []),
            "analysis_status": i.get("analysis_status"),
            "severity": analysis.get("severity") or "sin_analisis",
            "total_damages_count": analysis.get("total_damages_count") or 0,
            "executive_summary": analysis.get("executive_summary") or "",
            "new_damages": new_damages,
            "new_damages_count": len(new_damages),
            "has_reference": bool(i.get("reference_photos")),
            "annotated_photos": i.get("annotated_photos") or [],
            "detected_plate": analysis.get("detected_plate") or "",
            "plate_mismatch": plate_mismatch,
            "image_quality_warnings": list(analysis.get("image_quality_warnings") or [])[:5],
            "dirt_level": analysis.get("dirt_level"),
            "fraud_warnings": list(analysis.get("fraud_warnings") or [])[:5],
        })
    return {"queue": out, "total": len(out)}


@api_router.get("/inspections/{inspection_id}", response_model=Inspection)
async def get_inspection(inspection_id: str, _=Depends(require_admin)):
    insp = await db.inspections.find_one({"id": inspection_id}, {"_id": 0})
    if not insp:
        raise HTTPException(status_code=404, detail="Inspección no encontrada")
    return insp


@api_router.delete("/inspections/{inspection_id}")
async def delete_inspection(inspection_id: str, user: dict = Depends(get_current_user)):
    """Soft-delete: la inspección desaparece de la app pero la evidencia se conserva
    (fotos, análisis, fecha) junto con quién la borró y cuándo. Crítico para disputas."""
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores")
    result = await db.inspections.update_one(
        {"id": inspection_id, "deleted": {"$ne": True}},
        {"$set": {
            "deleted": True,
            "deleted_at": datetime.now(timezone.utc).isoformat(),
            "deleted_by": user.get("name", "?"),
        }}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Inspección no encontrada")
    logger.info(f"Inspección {inspection_id} eliminada (soft) por {user.get('name')}")
    return {"success": True, "message": "Inspección eliminada"}


@api_router.post("/inspections/{inspection_id}/damage-feedback")
async def damage_feedback(inspection_id: str, data: dict, user: dict = Depends(get_current_user)):
    """Validación humana de un daño detectado por la IA: ✓ correcto / ✗ falso.
    Cada veredicto se guarda como ejemplo de entrenamiento (dataset propio) con
    una copia completa del daño + la foto donde está — independiente de la
    inspección original, listo para entrenar el modelo propio."""
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores")
    verdict = data.get("verdict")
    if verdict not in ("correct", "wrong", "corrected"):
        raise HTTPException(status_code=400, detail="verdict debe ser 'correct', 'wrong' o 'corrected'")
    damage_index = data.get("damage_index")
    scope = data.get("scope", "new")  # 'new' = new_damages, 'all' = damages

    insp = await db.inspections.find_one({"id": inspection_id}, {"_id": 0})
    if not insp:
        raise HTTPException(status_code=404, detail="Inspección no encontrada")
    analysis = insp.get("analysis") or {}
    pool = analysis.get("new_damages" if scope == "new" else "damages") or []
    if damage_index is None or damage_index < 0 or damage_index >= len(pool):
        raise HTTPException(status_code=404, detail="Daño no encontrado")
    dmg = pool[damage_index] if isinstance(pool[damage_index], dict) else {}

    # Foto concreta donde la IA localizó el daño
    photos = insp.get("photos") or []
    pi = dmg.get("photo_index")
    photo_url = photos[pi - 1] if (isinstance(pi, int) and 1 <= pi <= len(photos)) else (photos[0] if photos else None)

    sample = {
        "id": str(uuid.uuid4()),
        "inspection_id": inspection_id,
        "vehicle_id": insp.get("vehicle_id"),
        "scope": scope,
        "damage_index": damage_index,
        "damage": dmg,                      # copia completa: part, severity, box_2d, photo_index…
        "photo_url": photo_url,
        "all_photo_urls": photos,
        "verdict": verdict,                 # correct = acertó · wrong = falso positivo · corrected = caja corregida a mano
        "corrected_box": data.get("corrected_box"),   # [ymin,xmin,ymax,xmax] 0-1000 dibujada por el humano
        "reviewed_by": user.get("name", "?"),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "model_version": "gemini-2.5-flash",
    }
    # Upsert: si re-valida el mismo daño, se actualiza el veredicto
    await db.ai_feedback.update_one(
        {"inspection_id": inspection_id, "scope": scope, "damage_index": damage_index},
        {"$set": sample}, upsert=True
    )

    # ── Sincronizar el LEDGER del vehículo con el veredicto humano ──
    try:
        _p = _canon_panel(dmg.get("part") or "")
        _vid = insp.get("vehicle_id")
        if _p and _vid:
            if verdict in ("correct", "corrected"):
                # Confirmado por humano → entrada consolidada (fuente humana)
                await db.vehicle_damage_ledger.update_one(
                    {"vehicle_id": _vid, "panel": _p, "status": "open"},
                    {"$set": {"part": dmg.get("part"), "source": "human",
                              "severity": _norm_sev(dmg.get("severity")),
                              "rank": _SEV_RANK.get(_norm_sev(dmg.get("severity")), 1),
                              "updated_at": datetime.now(timezone.utc).isoformat()},
                     "$setOnInsert": {"vehicle_id": _vid, "panel": _p, "status": "open",
                                      "first_seen": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                                      "first_seen_inspection": inspection_id}},
                    upsert=True)
            elif verdict == "wrong":
                # El humano dice que era un fantasma → fuera del ledger las
                # entradas creadas por la IA en ese panel (las humanas quedan)
                await db.vehicle_damage_ledger.delete_many(
                    {"vehicle_id": _vid, "panel": _p, "status": "open", "source": "ai"})
    except Exception as _le:
        logger.debug(f"ledger feedback: {_le}")

    total = await db.ai_feedback.count_documents({})
    return {"success": True, "dataset_size": total}


@api_router.post("/ai-feedback")
async def ai_feedback_simple(data: dict, user: dict = Depends(get_current_user)):
    """Endpoint simplificado para botones ✅/❌/✏️ del panel IAPeritaje.
    Llama internamente a save_feedback de ai_learning."""
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores")
    verdict = data.get("verdict")
    if verdict not in ("correct", "wrong", "corrected"):
        raise HTTPException(status_code=400, detail="verdict inválido")
    inspection_id = data.get("inspection_id")
    damage_index = data.get("damage_index")
    if not inspection_id or damage_index is None:
        raise HTTPException(status_code=400, detail="inspection_id y damage_index requeridos")

    insp = await db.inspections.find_one({"id": inspection_id}, {"_id": 0})
    if not insp:
        raise HTTPException(status_code=404, detail="Inspección no encontrada")
    analysis = insp.get("analysis") or {}
    pool = analysis.get("damages") or []
    dmg = pool[damage_index] if (0 <= damage_index < len(pool)) else {}

    photos = insp.get("photos") or []
    pi = dmg.get("photo_index")
    photo_url = photos[pi - 1] if (isinstance(pi, int) and 1 <= pi <= len(photos)) else (photos[0] if photos else None)

    await _save_ai_feedback(
        db=db,
        inspection_id=inspection_id,
        damage_index=damage_index,
        damage=dmg,
        photo_url=photo_url or "",
        verdict=verdict,
        corrected_box=data.get("corrected_box"),
        corrected_polygon=data.get("corrected_polygon_points"),
        reviewed_by=user.get("name", "?"),
    )
    total = await db.ai_feedback.count_documents({})
    return {"success": True, "dataset_size": total}


@api_router.post("/inspections/{inspection_id}/missed-damage")
async def missed_damage(inspection_id: str, data: dict, user: dict = Depends(get_current_user)):
    """Daño REAL que la IA no detectó, marcado y dibujado por el humano.
    El ejemplo más valioso del dataset: enseña a la IA lo que se le escapa
    (tulipas rotas, daños bajos, etc.)."""
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores")
    box = data.get("box_2d")
    part = (data.get("part") or "").strip()[:80]
    if not part:
        raise HTTPException(status_code=400, detail="Indica la pieza (ej: tulipa trasera)")
    if not (isinstance(box, list) and len(box) == 4):
        raise HTTPException(status_code=400, detail="Dibuja la caja del daño en la foto")
    photo_index = int(data.get("photo_index") or 1)

    insp = await db.inspections.find_one({"id": inspection_id}, {"_id": 0})
    if not insp:
        raise HTTPException(status_code=404, detail="Inspección no encontrada")
    photos = insp.get("photos") or []
    photo_url = photos[photo_index - 1] if 1 <= photo_index <= len(photos) else (photos[0] if photos else None)

    sample = {
        "id": str(uuid.uuid4()),
        "inspection_id": inspection_id,
        "vehicle_id": insp.get("vehicle_id"),
        "scope": "missed",
        "damage": {
            "part": part,
            "severity": data.get("severity") or "leve",
            "description": (data.get("description") or "").strip()[:300],
            "box_2d": [int(b) for b in box],
            "photo_index": photo_index,
        },
        "photo_url": photo_url,
        "all_photo_urls": photos,
        "verdict": "missed",               # la IA NO lo vio — falso negativo
        "reviewed_by": user.get("name", "?"),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "model_version": "gemini-2.5-flash",
    }
    await db.ai_feedback.insert_one(sample)
    total = await db.ai_feedback.count_documents({})
    return {"success": True, "dataset_size": total}


@api_router.get("/ai-dataset/export")
async def ai_dataset_export(_=Depends(require_admin)):
    """Exporta el dataset de feedback humano (✓/✗/corregido/no-visto) en formato
    COCO-like: la materia prima para entrenar el detector propio (Fase 2) junto
    a datasets públicos como CarDD. Coordenadas 0-1000 [ymin,xmin,ymax,xmax]."""
    docs = await db.ai_feedback.find({}, {"_id": 0}).to_list(20000)
    images: dict = {}
    annotations = []
    for i, f in enumerate(docs):
        url = f.get("photo_url")
        if not url:
            continue
        if url not in images:
            images[url] = {"id": len(images) + 1, "file_name": url}
        dmg = f.get("damage") or {}
        annotations.append({
            "id": i + 1,
            "image_id": images[url]["id"],
            "verdict": f.get("verdict"),          # correct | wrong (hard negative) | corrected | missed
            "part": dmg.get("part"),
            "severity": dmg.get("severity"),
            "location_hint": dmg.get("location_hint"),
            "box_2d": f.get("corrected_box") or dmg.get("box_2d"),
            "polygon_points": f.get("corrected_polygon_points") or dmg.get("polygon_points"),
            "inspection_id": f.get("inspection_id"),
            "reviewed_by": f.get("reviewed_by"),
            "created_at": f.get("created_at"),
        })
    by_verdict: dict = {}
    for a in annotations:
        by_verdict[a["verdict"]] = by_verdict.get(a["verdict"], 0) + 1
    return {
        "format": "flotadsp-coco-v1",
        "coords": "0-1000, box_2d=[ymin,xmin,ymax,xmax], polygon=[[y,x],...]",
        "counts": {"images": len(images), "annotations": len(annotations), "by_verdict": by_verdict},
        "images": list(images.values()),
        "annotations": annotations,
    }


@api_router.get("/ai-dataset/stats")
async def ai_dataset_stats(_=Depends(require_admin)):
    """Progreso del dataset de entrenamiento de la IA propia."""
    total = await db.ai_feedback.count_documents({})
    correct = await db.ai_feedback.count_documents({"verdict": "correct"})
    wrong = await db.ai_feedback.count_documents({"verdict": "wrong"})
    return {"total": total, "correct": correct, "wrong": wrong,
            "precision_ia": round(correct / total * 100, 1) if total else None,
            "goal": 3000,
            "progress_pct": round(total / 3000 * 100, 1)}


@api_router.get("/ai/export-dataset")
async def ai_export_dataset(_=Depends(require_superadmin)):
    """Exporta las correcciones como dataset ENTRENABLE (un solo objetivo: 'daño').
    YOLO localiza el daño (caja) y Gemini lo describe → con pocos ejemplos mejora.
    Cada muestra: foto + caja real marcada por un humano."""
    # Positivos = la IA acertó, la corregiste a mano, o se le escapó (el oro)
    cursor = db.ai_feedback.find(
        {"verdict": {"$in": ["correct", "corrected", "missed"]}}, {"_id": 0})
    samples = []
    async for f in cursor:
        dmg = f.get("damage") or {}
        box = f.get("corrected_box") if f.get("verdict") == "corrected" else dmg.get("box_2d")
        url = f.get("photo_url")
        if not (url and isinstance(box, list) and len(box) == 4):
            continue
        samples.append({
            "image_url": url,
            "box_2d": [int(v) for v in box],   # [ymin,xmin,ymax,xmax] 0-1000
            "label": "damage",
            "part": dmg.get("part"),
            "verdict": f.get("verdict"),
        })
    return {"classes": ["damage"], "n": len(samples), "samples": samples,
            "formato_box": "[ymin,xmin,ymax,xmax] normalizado 0-1000",
            "listo_para_entrenar": len(samples) >= 50}


@api_router.post("/inspections/{inspection_id}/mark-reviewed")
async def mark_inspection_reviewed(inspection_id: str, _=Depends(require_admin)):
    """Marca una inspección como revisada — desaparece de la cola de revisión rápida."""
    result = await db.inspections.update_one(
        {"id": inspection_id},
        {"$set": {"reviewed": True, "reviewed_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Inspección no encontrada")
    return {"success": True}


@app.on_event("startup")
async def start_analysis_recovery():
    """Auto-recuperación de análisis perdidos: si el servidor se reinicia a mitad
    de un análisis (queda 'pending' colgado) o Gemini falló, se reintenta solo.
    Máximo 3 reintentos automáticos por inspección, 5 por ciclo, cada 10 min."""
    # Al arrancar, dar OTRA ronda de reintentos a las fallidas recientes: si el
    # reinicio viene de arreglar la config de Gemini (billing/clave caída), las
    # que agotaron sus 3 intentos contra un Gemini roto deben curarse solas.
    try:
        _reset = await db.inspections.update_many(
            {"analysis_status": {"$in": ["error", "gemini_failed", "gemini_timeout"]},
             "deleted": {"$ne": True}, "auto_retries": {"$gte": 3},
             "created_at": {"$gt": (datetime.now(timezone.utc) - timedelta(days=14)).isoformat()}},
            {"$set": {"auto_retries": 0}},
        )
        if _reset.modified_count:
            logger.info(f"Auto-recuperación: {_reset.modified_count} inspecciones fallidas reencoladas tras reinicio")
    except Exception as _rr:
        logger.warning(f"Reset de reintentos al arrancar: {_rr}")

    async def _recovery_loop():
        await asyncio.sleep(60)  # dejar arrancar el servidor con calma
        _chain = ["gemini-2.5-flash", "gemini-flash-latest",
                  "gemini-flash-lite-latest", "gemini-2.5-flash-lite"]
        while True:
            try:
                # Toda la cadena con el día agotado (free tier): reintentar solo
                # quemaría los auto_retries de cada inspección contra un muro.
                if all(_is_daily_exhausted(m) for m in _chain):
                    logger.info("Auto-recuperación en pausa: cuota diaria de Gemini agotada en toda la cadena")
                    await asyncio.sleep(1800)
                    continue
                now = datetime.now(timezone.utc)
                cutoff_stuck = (now - timedelta(minutes=15)).isoformat()
                # 14 días: una caída larga (p. ej. créditos de Gemini agotados
                # durante días) no debe dejar inspecciones huérfanas sin análisis.
                cutoff_recent = (now - timedelta(days=14)).isoformat()
                stuck = await db.inspections.find({
                    "$and": [
                        {"$or": [
                            {"analysis_status": "pending", "created_at": {"$lt": cutoff_stuck}},
                            {"analysis_status": {"$in": ["error", "gemini_failed", "gemini_timeout"]},
                             "created_at": {"$gt": cutoff_recent}},
                        ]},
                        {"auto_retries": {"$not": {"$gte": 3}}},
                        {"photos": {"$exists": True, "$ne": []}},
                        {"deleted": {"$ne": True}},
                    ]
                }, {"_id": 0, "id": 1, "analysis_status": 1}).sort("created_at", -1).to_list(5)

                for insp in stuck:
                    iid = insp["id"]
                    await db.inspections.update_one({"id": iid}, {"$inc": {"auto_retries": 1}})
                    logger.info(f"Auto-recuperación: reintentando análisis de {iid} (estaba {insp.get('analysis_status')})")
                    try:
                        await reanalyze_inspection(iid, None)
                    except Exception as _re:
                        logger.warning(f"Auto-recuperación {iid} falló: {_re}")
                    await asyncio.sleep(10)  # no saturar Gemini
            except Exception as e:
                logger.error(f"Recovery loop error: {e}")
            await asyncio.sleep(600)
    asyncio.create_task(_recovery_loop())


@api_router.get("/inspections/{inspection_id}/debug-segment")
async def debug_segment(inspection_id: str, photo_index: int = 0, damage_index: int = 0,
                        _=Depends(require_admin)):
    """Devuelve imagen compuesta con las 6 fases del pipeline de segmentación OpenCV.
    photo_index: 0-based. damage_index: índice del daño en analysis.damages."""
    from fastapi.responses import Response as _FR
    insp = await db.inspections.find_one({"id": inspection_id}, {"_id": 0})
    if not insp:
        raise HTTPException(status_code=404, detail="Inspección no encontrada")
    photos = insp.get("photos") or []
    if photo_index >= len(photos):
        raise HTTPException(status_code=400, detail="photo_index fuera de rango")

    analysis = insp.get("analysis") or {}
    damages = analysis.get("damages") or analysis.get("new_damages") or []
    if damage_index >= len(damages):
        raise HTTPException(status_code=400, detail="damage_index fuera de rango")

    dmg = damages[damage_index]
    box_2d = dmg.get("box_2d")
    if not box_2d or len(box_2d) != 4 or not any(v > 0 for v in box_2d):
        raise HTTPException(status_code=400, detail="Daño sin box_2d válido")

    img_bytes = await _fetch_photo_bytes(photos[photo_index])
    if not img_bytes:
        raise HTTPException(status_code=400, detail="No se pudo cargar la foto")

    loop = asyncio.get_running_loop()
    poly, debug_imgs = await loop.run_in_executor(
        _executor, lambda: _segment_damage_opencv(img_bytes, box_2d, debug=True)
    )

    if not debug_imgs:
        raise HTTPException(status_code=422, detail="La segmentación no produjo resultado")

    # Construir imagen compuesta: original | edges | shadows | reflection | candidate | mask
    import cv2 as _cv2
    import numpy as _np
    from PIL import Image as _PILImg
    import io as _io

    labels = ['original', 'bordes', 'sombras', 'reflexión', 'candidato', 'máscara']
    frames_raw = [
        debug_imgs['crop'],
        debug_imgs['edges'],
        debug_imgs['shadows'],
        debug_imgs['reflection'],
        debug_imgs['candidate'],
        debug_imgs['mask'],
    ]

    # Convertir todos a uint8 BGR para cv2 (float32 → normalizar, grises → BGR)
    frames = []
    for f in frames_raw:
        if f.dtype != _np.uint8:
            f = _cv2.normalize(f, None, 0, 255, _cv2.NORM_MINMAX).astype(_np.uint8)
        if len(f.shape) == 2:
            frames.append(_cv2.cvtColor(f, _cv2.COLOR_GRAY2BGR))
        else:
            frames.append(f)

    # Redimensionar todos al mismo alto
    H_target = 300
    resized = []
    for f in frames:
        h, w = f.shape[:2]
        new_w = int(w * H_target / h)
        resized.append(_cv2.resize(f, (new_w, H_target)))

    # Añadir etiqueta encima de cada frame
    labeled = []
    for i, (f, lbl) in enumerate(zip(resized, labels)):
        bar = _np.zeros((28, f.shape[1], 3), dtype=_np.uint8)
        _cv2.putText(bar, lbl, (4, 20), _cv2.FONT_HERSHEY_SIMPLEX, 0.55, (255, 255, 255), 1)
        labeled.append(_np.vstack([bar, f]))

    composite = _np.hstack(labeled)

    # Si hay polígono, dibujarlo sobre el crop en la primera columna (frame original)
    if poly and len(poly) >= 4:
        orig_with_poly = labeled[0].copy()
        pts_px = _np.array([
            [int(pt[1] / 1000 * resized[0].shape[1]),
             int(pt[0] / 1000 * resized[0].shape[0]) + 28]
            for pt in poly
        ], dtype=_np.int32)
        _cv2.polylines(orig_with_poly, [pts_px], True, (0, 80, 255), 2)
        composite[:, :labeled[0].shape[1]] = orig_with_poly

    # Codificar como JPEG
    ok, buf = _cv2.imencode('.jpg', composite, [_cv2.IMWRITE_JPEG_QUALITY, 88])
    if not ok:
        raise HTTPException(status_code=500, detail="Error codificando imagen")

    return _FR(content=buf.tobytes(), media_type="image/jpeg",
               headers={"Content-Disposition": f'inline; filename="debug_{inspection_id[:8]}_p{photo_index}_d{damage_index}.jpg"'})


@api_router.post("/inspections/{inspection_id}/reanalyze")
async def reanalyze_inspection(inspection_id: str, silent: bool = False, _=Depends(require_admin)):
    """Relanza el análisis IA de una inspección usando las fotos ya guardadas.
    silent=true suprime la notificación de Telegram (útil para pruebas)."""
    insp = await db.inspections.find_one({"id": inspection_id}, {"_id": 0})
    if not insp:
        raise HTTPException(status_code=404, detail="Inspección no encontrada")

    photo_urls = insp.get("photos") or []
    if not photo_urls:
        raise HTTPException(status_code=400, detail="La inspección no tiene fotos guardadas")

    # Descargar las fotos desde R2/disco
    photo_bytes = await load_reference_images(photo_urls)
    if not photo_bytes:
        raise HTTPException(status_code=502, detail="No se pudieron descargar las fotos guardadas")

    photos_base64 = [base64.b64encode(b).decode("utf-8") for b in photo_bytes]

    # Solo carrocería al análisis (photo_roles existe desde 2026-07; en
    # inspecciones antiguas sin roles se analizan todas, como siempre).
    _roles = insp.get("photo_roles") or []
    if len(_roles) == len(photos_base64):
        _aidx = [i for i, r in enumerate(_roles) if r == "angle"] or list(range(len(photos_base64)))
    else:
        _aidx = list(range(len(photos_base64)))
    analysis_b64 = [photos_base64[i] for i in _aidx]
    _idx_map = {k + 1: _aidx[k] + 1 for k in range(len(_aidx))}

    # Fotos de referencia (estado anterior) si las había
    ref_urls = insp.get("reference_photos") or []
    ref_bytes_list = await load_reference_images(ref_urls) if ref_urls else None

    # Marcar como pendiente mientras se reanaliza
    await db.inspections.update_one(
        {"id": inspection_id},
        {"$set": {"analysis_status": "pending", "analysis_error": None}}
    )

    # CV primero (igual que en la subida): guía el prompt y luego el snap
    cv_dets_full = {}
    cv_text = ""
    if AI_SERVICE_URL:
        try:
            _cv_lines = []
            for _k, _full_i in enumerate(_aidx, start=1):
                if _full_i < len(photo_bytes):
                    dets = await _call_ai_service_detect(inspection_id, _full_i, photo_bytes[_full_i])
                    if dets:
                        cv_dets_full[_full_i + 1] = dets
                        for det in dets:
                            _pnl = f" en {det.panel}" if getattr(det, "panel", None) else ""
                            _cv_lines.append(f"Foto {_k}: {det.label}{_pnl} (confianza {det.confidence:.2f})")
            if _cv_lines:
                cv_text = (
                    "\n=== DETECCIONES DEL DETECTOR CV PROPIO (geometría verificada) ===\n"
                    + "\n".join(_cv_lines) +
                    "\nEstas detecciones vienen de un modelo entrenado con miles de daños "
                    "reales de carrocería; la pieza está determinada por intersección "
                    "geométrica, no por estimación. Úsalas como BASE del análisis: "
                    "confirma y describe con precisión cada una que veas en la foto. "
                    "NO reportes daños adicionales salvo evidencia clara e inequívoca; "
                    "si no ves alguna de estas detecciones, inclúyela con confidence < 0.5."
                )
        except Exception as _cve:
            logger.warning(f"CV pre-reanálisis: {_cve}")

    _known_txt = await _known_damages_prompt(insp.get("vehicle_id"), exclude_inspection_id=inspection_id)
    analysis, analysis_status, analysis_error = await analyze_images_with_gemini(
        analysis_b64, ref_bytes_list if ref_bytes_list else None, db=db,
        known_damages_text=_known_txt,
        cv_detections_text=cv_text,
    )
    if analysis:
        _remap_photo_indexes(analysis, _idx_map)
    if analysis_status == "ok" and analysis:
        await _apply_vehicle_memory(insp.get("vehicle_id"), analysis, inspection_id=inspection_id)

    # Snap de cajas al detector CV (reutiliza las detecciones ya pedidas)
    if cv_dets_full and analysis and analysis_status == "ok" and (analysis.damages or analysis.new_damages):
        try:
            _s1 = _snap_damage_boxes_to_yolo(analysis.damages or [], cv_dets_full)
            _s2 = _snap_damage_boxes_to_yolo(analysis.new_damages or [], cv_dets_full)
            if _s1 or _s2:
                logger.info(f"Snap YOLO (reanálisis): {_s1}+{_s2} cajas corregidas insp={inspection_id[:8]}")
        except Exception as _se:
            logger.warning(f"Snap YOLO reanálisis: {_se}")

    await db.inspections.update_one(
        {"id": inspection_id},
        {"$set": {"analysis": serialize_doc(analysis.model_dump()) if analysis else None,
                  "analysis_status": analysis_status,
                  "analysis_error": analysis_error,
                  "analyzed_at": datetime.now(timezone.utc).isoformat()}}
    )

    # ─ Regenerar fotos anotadas ─
    if analysis and analysis_status == "ok" and analysis.damages:
        try:
            ann_urls = await generate_annotated_photos(
                inspection_id, photo_urls, analysis.damages
            )
            if any(u for u in ann_urls):
                await db.inspections.update_one(
                    {"id": inspection_id},
                    {"$set": {"annotated_photos": ann_urls}}
                )
        except Exception as _ae:
            logger.warning(f"Error generando anotaciones en reanálisis: {_ae}")

    # Telegram si el reanálisis revela daños graves/críticos (omitir si silent=True)
    if not silent and analysis and analysis_status == "ok":
        sev = (analysis.severity or "").lower()
        if sev in ("grave", "critico"):
            try:
                vehicle_id = insp.get("vehicle_id", "")
                driver_id = insp.get("driver_id", "")
                veh_doc = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0, "license_plate": 1, "center": 1})
                plate = veh_doc.get("license_plate", vehicle_id) if veh_doc else vehicle_id
                drv_doc = await db.drivers.find_one({"id": driver_id}, {"_id": 0, "name": 1, "center": 1}) if driver_id else None
                drv_name = drv_doc.get("name", "Desconocido") if drv_doc else "Desconocido"
                center = (drv_doc.get("center", "") if drv_doc else "") or (veh_doc.get("center", "") if veh_doc else "")
                await send_telegram_damage_alert(
                    plate=plate, driver_name=drv_name, analysis=analysis,
                    photo_urls=photo_urls, inspection_id=inspection_id, center=center
                )
            except Exception as _te:
                logger.warning(f"Telegram tras reanálisis falló (no crítico): {_te}")

    logger.info(f"Reanálisis {inspection_id}: status={analysis_status}, severity={analysis.severity if analysis else 'n/a'}")
    return {
        "success": analysis_status == "ok",
        "analysis_status": analysis_status,
        "analysis_error": analysis_error,
        "analysis": serialize_doc(analysis.model_dump()) if analysis else None,
    }


# =========================
# PERITAJE TÉCNICO FIRMADO (S1) — cadena de custodia con hash encadenado
# =========================
# Cada firma calcula content_hash = SHA-256(prev_hash + payload canónico de la inspección).
# El prev_hash es el hash de la firma INMEDIATAMENTE anterior de la MISMA organización.
# Si alguien manipula o elimina una firma intermedia, la cadena se rompe (detectable).
# NOTA LEGAL: es evidencia técnica con cadena de custodia, NO firma electrónica avanzada eIDAS.

_GENESIS_HASH = "0" * 64


def _canonical_inspection_payload(insp: dict, prev_hash: str) -> str:
    """Serialización determinista de la inspección para hashear.
    NO usamos json.dumps con sort_keys directamente porque queremos un subconjunto fijo
    para que cambios futuros en otros campos no rompan la cadena."""
    payload = {
        "prev_hash": prev_hash,
        "inspection_id": insp.get("id"),
        "vehicle_id": insp.get("vehicle_id"),
        "driver_id": insp.get("driver_id"),
        "created_at": insp.get("created_at"),
        "photos": list(insp.get("photos") or []),  # URLs en R2; URL inmutable = foto inmutable
        "plate_text": (insp.get("plate_text") or ""),
        "analysis_summary": {
            "severity": (insp.get("analysis") or {}).get("severity"),
            "total_damages_count": (insp.get("analysis") or {}).get("total_damages_count"),
            "executive_summary": (insp.get("analysis") or {}).get("executive_summary"),
        },
    }
    return json.dumps(payload, sort_keys=True, ensure_ascii=False, default=str)


async def _last_signature_hash() -> str:
    """Último hash de la cadena (esta org). _GENESIS_HASH si la cadena está vacía."""
    last = await db.forensic_signatures.find_one({}, sort=[("signed_at", -1)])
    return (last or {}).get("content_hash", _GENESIS_HASH)


class SignInspectionRequest(BaseModel):
    signature_text: str   # declaración aceptada por el firmante (texto libre, queda inmortalizada)


@api_router.post("/inspections/{inspection_id}/sign")
async def sign_inspection(inspection_id: str, data: SignInspectionRequest, request: Request,
                          user: dict = Depends(require_any_auth)):
    """Firma una inspección con cadena de custodia hash. Idempotente: 409 si ya firmada.
    Auth: el conductor asignado al vehículo de esa inspección, o cualquier admin de la org."""
    await _require_plan_feature(user, "forensics")
    # Rate-limit: 3 firmas/min por usuario (anti-rebote).
    _rl_public_action(f"sign:{user.get('sub')}", max_count=3, window_s=60,
                      detail="Demasiadas firmas seguidas. Espera un minuto.")

    insp = await db.inspections.find_one({"id": inspection_id, "deleted": {"$ne": True}}, {"_id": 0})
    if not insp:
        raise HTTPException(status_code=404, detail="Inspección no encontrada")
    if insp.get("forensic_signed"):
        raise HTTPException(status_code=409,
                            detail="Esta inspección ya está firmada (cadena de custodia cerrada)")

    # Autorización: admin puede firmar cualquiera. Conductor solo la suya.
    if user.get("role") == "driver":
        if insp.get("driver_id") and insp["driver_id"] != user.get("sub"):
            # Si la inspección no tiene driver_id, comprobamos por current_driver del vehículo.
            v = await db.vehicles.find_one({"id": insp.get("vehicle_id")}, {"_id": 0, "current_driver_id": 1})
            if not v or v.get("current_driver_id") != user.get("sub"):
                raise HTTPException(status_code=403, detail="No puedes firmar inspecciones de otro conductor")

    # Calcular hashes con reintento simple ante race condition de prev_hash.
    import hashlib as _hl
    now = datetime.now(timezone.utc).isoformat()
    sig_doc = None
    for attempt in range(3):
        prev_hash = await _last_signature_hash()
        payload = _canonical_inspection_payload(insp, prev_hash)
        content_hash = _hl.sha256(payload.encode("utf-8")).hexdigest()
        try:
            sig_doc = {
                "id": str(uuid.uuid4()),
                "inspection_id": inspection_id,
                "revision": 1,
                "prev_hash": prev_hash,
                "content_hash": content_hash,
                "payload_canonical": payload,           # útil para reverificar sin recalcular
                "signed_by_user_id": user.get("sub"),
                "signed_by_user_role": user.get("role"),
                "signed_by_name": user.get("name", ""),
                "signed_at": now,
                "client_ip": _rl_key_ip(request),
                "user_agent": (request.headers.get("user-agent") or "")[:300],
                "signature_text": (data.signature_text or "").strip()[:1000],
            }
            await db.forensic_signatures.insert_one(sig_doc)
            break
        except DuplicateKeyError:
            # Otra firma se coló en el medio: recalcula y reintenta.
            if attempt == 2:
                raise HTTPException(status_code=503, detail="Conflicto al firmar. Reintenta.")
            await asyncio.sleep(0.05)

    # Marca la inspección como firmada (no rompemos compatibilidad: campos nuevos opcionales).
    await db.inspections.update_one(
        {"id": inspection_id},
        {"$set": {"forensic_signed": True, "forensic_hash": sig_doc["content_hash"],
                  "forensic_signed_at": now, "forensic_signed_by": sig_doc["signed_by_name"]}}
    )
    # Índice global hash → tenant (para que /verify/{hash} público pueda localizar la firma sin auth).
    # Aditivo: si falla (DuplicateKey, race), no rompe la firma local.
    try:
        await global_db.forensic_index.insert_one({
            "content_hash": sig_doc["content_hash"],
            "org_id": user.get("org_id"),
            "db_name": _tenant_db_name(await get_org(user.get("org_id"))),
            "inspection_id": inspection_id,
            "signed_at": now,
        })
    except DuplicateKeyError:
        logger.warning(f"forensic_index duplicado para hash={sig_doc['content_hash'][:12]}")
    logger.info(f"Inspección {inspection_id} firmada por {user.get('name')} hash={sig_doc['content_hash'][:12]}…")
    return {
        "ok": True,
        "hash": sig_doc["content_hash"],
        "prev_hash": sig_doc["prev_hash"],
        "signed_at": now,
        "signed_by_name": sig_doc["signed_by_name"],
    }


@api_router.get("/inspections/{inspection_id}/forensic")
async def get_forensic_status(inspection_id: str, _=Depends(require_any_auth)):
    """Estado de firma de una inspección. Sin secretos: NO devuelve payload completo."""
    # FIX (bug S1.A): incluir 'id' en projection para que find_one no devuelva {} en docs
    # sin campos forensic_*. Sin esto, /forensic devolvía 404 para inspecciones no firmadas.
    insp = await db.inspections.find_one(
        {"id": inspection_id}, {"_id": 0, "id": 1, "forensic_signed": 1, "forensic_hash": 1,
                                "forensic_signed_at": 1, "forensic_signed_by": 1}
    )
    if not insp:
        raise HTTPException(status_code=404, detail="Inspección no encontrada")
    if not insp.get("forensic_signed"):
        return {"signed": False}
    sig = await db.forensic_signatures.find_one(
        {"inspection_id": inspection_id}, {"_id": 0, "prev_hash": 1, "content_hash": 1,
                                            "signed_at": 1, "signed_by_name": 1, "signature_text": 1,
                                            "revision": 1}
    )
    return {
        "signed": True,
        "hash": insp.get("forensic_hash"),
        "prev_hash": (sig or {}).get("prev_hash"),
        "signed_at": insp.get("forensic_signed_at"),
        "signed_by_name": insp.get("forensic_signed_by"),
        "signature_text": (sig or {}).get("signature_text"),
        "revision": (sig or {}).get("revision", 1),
    }


# --- Generación del PDF forense (reportlab + qrcode) ---

_PORTAL_BASE_FRONT = os.environ.get("PUBLIC_BASE_URL_FRONT", "https://flotadsp.com").rstrip("/")


def _mask_plate(plate: str) -> str:
    """Oculta parcialmente una matrícula para mostrarla en el verificador público.
    1234ABC -> 1234A** ; corto -> tal cual."""
    if not plate:
        return "—"
    p = re.sub(r"\s+", "", plate.strip().upper())
    if len(p) < 5:
        return p
    keep = max(4, len(p) - 2)
    return p[:keep] + "*" * (len(p) - keep)


async def _fetch_photo_bytes(url: str, timeout: int = 8):
    """Descarga una foto de R2/CDN. None si falla. Usado solo para embedirla en PDF."""
    if not url:
        return None
    try:
        async with _aiohttp.ClientSession() as s:
            async with s.get(url, timeout=_aiohttp.ClientTimeout(total=timeout)) as r:
                if r.status == 200:
                    return await r.read()
    except Exception as e:
        logger.warning(f"forensic PDF: foto inaccesible {url}: {e}")
    return None


# Umbral de confianza mínimo para DIBUJAR caja en el PDF (compromiso de veracidad).
# Daños bajo el umbral aparecen listados como "sugeridos" pero NO se marcan visualmente.
_FORENSIC_CONFIDENCE_MIN = 0.5
# Mínimo área de caja (en % del área de la imagen) para considerarla un daño real.
_FORENSIC_MIN_BOX_AREA_PCT = 0.5  # 0.5% mínimo

_SEVERITY_COLORS = {
    "leve":     (250, 204, 21),   # amarillo
    "moderado": (251, 146, 60),   # naranja
    "grave":    (239, 68, 68),    # rojo
    "critico":  (220, 38, 38),    # rojo intenso
    "sin_danos": (34, 197, 94),   # verde (no se dibuja normalmente)
}


def _is_box_valid(box_2d) -> bool:
    """Valida que box_2d sea una caja real (no placeholder, dentro de rango, con área mínima)."""
    if not isinstance(box_2d, (list, tuple)) or len(box_2d) != 4:
        return False
    try:
        y1, x1, y2, x2 = [float(v) for v in box_2d]
    except Exception:
        return False
    if y1 + x1 + y2 + x2 == 0:                # placeholder [0,0,0,0]
        return False
    if not (0 <= y1 < y2 <= 1000 and 0 <= x1 < x2 <= 1000):
        return False
    area_pct = ((y2 - y1) * (x2 - x1)) / 10000.0  # 1000*1000/10000 = 100%
    if area_pct < _FORENSIC_MIN_BOX_AREA_PCT:
        return False
    return True


def _annotate_photo_with_damages(photo_bytes: bytes, damages_for_photo: list) -> bytes:
    """Dibuja cajas sobre los daños VERÍDICOS de una foto.
    Solo dibuja: confidence ≥ 0.5, box_2d válido, severity ≠ sin_danos.
    Si la imagen falla → devuelve los bytes originales (fallback seguro)."""
    if not photo_bytes:
        return photo_bytes
    try:
        from PIL import Image as PILImage, ImageDraw, ImageFont
        import io as _io
        img = PILImage.open(_io.BytesIO(photo_bytes))
        if img.mode != "RGB":
            img = img.convert("RGB")
        W, H = img.size
        draw = ImageDraw.Draw(img)
        # Grosor proporcional a la imagen (mínimo 3 px).
        thick = max(3, int(min(W, H) * 0.006))
        try:
            font = ImageFont.truetype("DejaVuSans-Bold.ttf", max(14, int(min(W, H) * 0.022)))
        except Exception:
            font = ImageFont.load_default()

        drawn = 0
        for d in (damages_for_photo or []):
            if (d.get("severity") in (None, "sin_danos", "sin_analisis")):
                continue
            if float(d.get("confidence", 0)) < _FORENSIC_CONFIDENCE_MIN:
                continue
            box = d.get("box_2d")
            if not _is_box_valid(box):
                continue
            y1, x1, y2, x2 = box
            # Mapear 0-1000 a píxeles.
            left, top = int(x1 / 1000.0 * W), int(y1 / 1000.0 * H)
            right, bottom = int(x2 / 1000.0 * W), int(y2 / 1000.0 * H)
            color = _SEVERITY_COLORS.get(d.get("severity"), (239, 68, 68))
            verified = bool(d.get("_verified"))
            if verified:
                # Sólido + grueso → daño cross-validado con YOLO/SAM.
                draw.rectangle([left, top, right, bottom], outline=color, width=thick)
            else:
                # Discontinuo → solo Gemini (caja aproximada, sin confirmación geométrica).
                seg = max(8, thick * 3)
                for sx in range(left, right, seg * 2):
                    draw.line([(sx, top), (min(sx + seg, right), top)], fill=color, width=thick)
                    draw.line([(sx, bottom), (min(sx + seg, right), bottom)], fill=color, width=thick)
                for sy in range(top, bottom, seg * 2):
                    draw.line([(left, sy), (left, min(sy + seg, bottom))], fill=color, width=thick)
                    draw.line([(right, sy), (right, min(sy + seg, bottom))], fill=color, width=thick)
            # Etiqueta arriba del rectángulo: parte + confidence (+ ~ si no verificada)
            tag = "" if verified else "~"
            label = f"{tag}{(d.get('part') or 'daño')[:22]}  {int(float(d.get('confidence', 0)) * 100)}%"
            try:
                tb = draw.textbbox((0, 0), label, font=font)
                tw, th = tb[2] - tb[0], tb[3] - tb[1]
            except Exception:
                tw, th = (len(label) * 8, 14)
            ly = max(0, top - th - 6)
            draw.rectangle([left, ly, left + tw + 8, ly + th + 4], fill=color)
            draw.text((left + 4, ly), label, fill=(0, 0, 0), font=font)
            drawn += 1

        buf = _io.BytesIO()
        # JPEG para no inflar PDF; calidad alta para mantener detalle del daño.
        img.save(buf, format="JPEG", quality=85, optimize=True)
        return buf.getvalue()
    except Exception as e:
        logger.warning(f"forensic PDF: anotación de foto falló: {e}")
        return photo_bytes


def _filter_damages_for_peritaje(damages: list):
    """Devuelve (incluidos, sugeridos): incluidos pasan los filtros estrictos."""
    incluidos, sugeridos = [], []
    for d in (damages or []):
        sev = d.get("severity")
        if sev in (None, "sin_danos", "sin_analisis"):
            continue
        conf = float(d.get("confidence", 0))
        if conf >= _FORENSIC_CONFIDENCE_MIN and _is_box_valid(d.get("box_2d")):
            incluidos.append(d)
        else:
            sugeridos.append(d)
    return incluidos, sugeridos


# --- ENSEMBLE GEMINI + YOLO+SAM (precisión geométrica del bounding box) ---

def _bbox_iou(a, b) -> float:
    """IoU entre 2 cajas [y1,x1,y2,x2] (cualquier escala). 0 si no se solapan."""
    try:
        ay1, ax1, ay2, ax2 = float(a[0]), float(a[1]), float(a[2]), float(a[3])
        by1, bx1, by2, bx2 = float(b[0]), float(b[1]), float(b[2]), float(b[3])
    except Exception:
        return 0.0
    iy1, ix1 = max(ay1, by1), max(ax1, bx1)
    iy2, ix2 = min(ay2, by2), min(ax2, bx2)
    iw, ih = max(0.0, ix2 - ix1), max(0.0, iy2 - iy1)
    inter = iw * ih
    if inter <= 0:
        return 0.0
    area_a = max(0.0, ay2 - ay1) * max(0.0, ax2 - ax1)
    area_b = max(0.0, by2 - by1) * max(0.0, bx2 - bx1)
    union = area_a + area_b - inter
    return inter / union if union > 0 else 0.0


def _nms_damages(damages: list, iou_thresh: float = 0.7) -> list:
    """Non-Max Suppression: si 2 daños solapan IoU>thresh, queda el de mayor confidence."""
    if not damages:
        return []
    sorted_d = sorted(damages, key=lambda d: float(d.get("confidence", 0)), reverse=True)
    keep, suppressed = [], set()
    for i, d in enumerate(sorted_d):
        if i in suppressed:
            continue
        keep.append(d)
        for j in range(i + 1, len(sorted_d)):
            if j in suppressed:
                continue
            if _bbox_iou(d.get("box_2d") or [0, 0, 0, 0],
                        sorted_d[j].get("box_2d") or [0, 0, 0, 0]) > iou_thresh:
                suppressed.add(j)
    return keep


async def _refine_damage_boxes_with_yolo_sam(inspection_id: str, photo_index_0based: int,
                                              gemini_damages_for_photo: list) -> list:
    """Cruza cajas Gemini con detecciones YOLO+SAM ya almacenadas en inspection_ai_results.
    Cuando hay coincidencia (IoU≥0.3) → caja YOLO/SAM (geométricamente precisa)
                                       + label semántica Gemini (qué es el daño).
    Sin coincidencia → mantiene Gemini SOLO si confidence ≥ 0.7 (compromiso de veracidad).
    Sin detecciones YOLO disponibles → degrada elegantemente a Gemini sin cambios."""
    ai_doc = await db.inspection_ai_results.find_one(
        {"inspection_id": inspection_id, "photo_index": photo_index_0based}, {"_id": 0}
    )
    yolo_dets = (ai_doc or {}).get("detections") or []
    # Filtra detecciones YOLO/SAM con cajas válidas y confianza decente.
    yolo_dets = [y for y in yolo_dets
                 if y.get("box_2d") and _is_box_valid(y["box_2d"])
                 and float(y.get("confidence", 0)) >= 0.35]

    # Filtro geométrico estricto: cajas > 15% del área = "toda la pieza" → no es daño puntual.
    # Aplicable a Gemini y a YOLO. Un daño real raramente supera el 15% de la foto.
    MAX_AREA_PCT = 15.0
    refined = []
    for g in gemini_damages_for_photo:
        g_box = g.get("box_2d")
        if not _is_box_valid(g_box):
            continue
        y1, x1, y2, x2 = g_box
        g_area = ((y2 - y1) * (x2 - x1)) / 10000.0
        if g_area > MAX_AREA_PCT:
            continue  # Gemini marcó la pieza entera, no el daño.
        # Cross-validation con YOLO/SAM.
        best_iou, best_y = 0.0, None
        for y in yolo_dets:
            iou = _bbox_iou(g_box, y["box_2d"])
            if iou > best_iou:
                best_iou, best_y = iou, y
        if best_iou >= 0.3 and best_y is not None:
            yy1, yx1, yy2, yx2 = best_y["box_2d"]
            y_area = ((yy2 - yy1) * (yx2 - yx1)) / 10000.0
            if y_area <= MAX_AREA_PCT:
                # Cross-validado: caja YOLO precisa + label Gemini. Borde SÓLIDO.
                refined.append({**g, "box_2d": list(best_y["box_2d"]),
                                "_box_source": "yolo+sam", "_iou": round(best_iou, 2),
                                "_verified": True})
                continue
        # Gemini sin confirmación pero con caja pequeña razonable. Borde DISCONTINUO.
        refined.append({**g, "_box_source": "gemini_only", "_verified": False})
    return _nms_damages(refined, iou_thresh=0.7)


def _generate_qr_png(data: str) -> bytes:
    """QR como PNG bytes (para embed en PDF). Importado lazy."""
    import qrcode as _qr
    import io as _io
    q = _qr.QRCode(border=2, box_size=4)
    q.add_data(data)
    q.make(fit=True)
    img = q.make_image()
    buf = _io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


def _legend_pill(color_hex, label):
    """Genera una mini-tabla 'cuadrito de color + label' para la leyenda."""
    from reportlab.platypus import Table, TableStyle, Paragraph
    from reportlab.lib import colors as _c
    from reportlab.lib.styles import getSampleStyleSheet
    st = getSampleStyleSheet()
    return Table([["", Paragraph(label, st["BodyText"])]],
                 colWidths=[4*1.5, None],
                 style=TableStyle([
                     ("BACKGROUND", (0, 0), (0, 0), color_hex),
                     ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                     ("LEFTPADDING", (0, 0), (-1, -1), 2),
                 ]))


def _build_forensic_pdf(insp: dict, sig: dict, vehicle: dict, driver_name: str,
                        photo_bytes_list: list, verify_url: str,
                        damages_incluidos: list = None, damages_sugeridos: list = None) -> bytes:
    """Construye el PDF de peritaje técnico. Síncrono, llamado vía run_in_executor.
    photo_bytes_list contiene fotos YA anotadas con cajas (donde aplique)."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Image, Table,
                                    TableStyle, PageBreak)
    from reportlab.lib.enums import TA_LEFT
    import io as _io

    damages_incluidos = damages_incluidos or []
    damages_sugeridos = damages_sugeridos or []

    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=15*mm, leftMargin=15*mm,
                            topMargin=15*mm, bottomMargin=15*mm,
                            title=f"Peritaje técnico {insp.get('id','')[:8]}",
                            author="FlotaDSP")
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("H1", parent=styles["Heading1"], fontSize=18, textColor=colors.HexColor("#0b1220"))
    h2 = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=12, textColor=colors.HexColor("#0ea5e9"),
                        spaceBefore=8, spaceAfter=4)
    body = ParagraphStyle("body", parent=styles["BodyText"], fontSize=9.5, leading=13, alignment=TA_LEFT)
    small = ParagraphStyle("small", parent=styles["BodyText"], fontSize=8, textColor=colors.HexColor("#6b7280"))
    mono = ParagraphStyle("mono", parent=styles["BodyText"], fontSize=8.5, fontName="Courier",
                          textColor=colors.HexColor("#0b1220"))

    story = []

    # Cabecera
    story.append(Paragraph("Peritaje técnico de inspección", h1))
    story.append(Paragraph("Cadena de custodia con hash inmutable · FlotaDSP", small))
    story.append(Spacer(1, 5*mm))

    # Bloque datos
    analysis = (insp.get("analysis") or {})
    rows = [
        ["Matrícula", vehicle.get("license_plate", "—"), "Centro", vehicle.get("center", "—")],
        ["Marca/Modelo", f"{vehicle.get('brand','—')} {vehicle.get('model','')}".strip(), "VIN", vehicle.get("vin", "—")],
        ["Conductor", driver_name or "—", "Inspección ID", (insp.get("id","")[:8] + "…")],
        ["Fecha inspección", (insp.get("created_at") or "")[:19].replace("T", " "), "Severidad", str(analysis.get("severity", "—"))],
        ["Daños incluidos", str(len(damages_incluidos)), "Coste estimado", f"{analysis.get('total_estimated_cost', 0)} €"],
    ]
    t = Table(rows, colWidths=[28*mm, 60*mm, 28*mm, 64*mm])
    t.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#6b7280")),
        ("TEXTCOLOR", (2, 0), (2, -1), colors.HexColor("#6b7280")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LINEBELOW", (0, 0), (-1, -1), 0.4, colors.HexColor("#e5e7eb")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(t)
    story.append(Spacer(1, 5*mm))

    # Leyenda de colores
    if damages_incluidos:
        legend_data = [[
            "Leyenda:",
            _legend_pill(colors.HexColor("#facc15"), "Leve"),
            _legend_pill(colors.HexColor("#fb923c"), "Moderado"),
            _legend_pill(colors.HexColor("#ef4444"), "Grave"),
            _legend_pill(colors.HexColor("#dc2626"), "Crítico"),
        ]]
        leg = Table(legend_data, colWidths=[18*mm, 30*mm, 35*mm, 30*mm, 30*mm])
        leg.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("TEXTCOLOR", (0, 0), (0, 0), colors.HexColor("#6b7280")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        story.append(leg)
        story.append(Spacer(1, 3*mm))

    # Resumen ejecutivo si hay
    if analysis.get("executive_summary"):
        story.append(Paragraph("Resumen del peritaje IA", h2))
        story.append(Paragraph((analysis.get("executive_summary") or "")[:1500], body))

    # Fotos con cajas
    photos_valid = [b for b in photo_bytes_list if b]
    if photos_valid:
        story.append(Paragraph("Evidencia fotográfica con daños marcados", h2))
        msg = (f"{len(photos_valid)} fotos analizadas. "
               f"<b>Borde sólido</b> = daño confirmado por dos modelos (Gemini+YOLO/SAM). "
               f"<b>Borde discontinuo</b> = detectado por Gemini sin confirmación geométrica (caja aproximada, marcado con ~). "
               f"El hash de cada imagen está sellado en la cadena de custodia.")
        story.append(Paragraph(msg, small))
        story.append(Spacer(1, 3*mm))
        thumbs = []
        for b in photos_valid:
            try:
                img = Image(_io.BytesIO(b), width=85*mm, height=64*mm, kind="proportional")
                thumbs.append(img)
            except Exception as e:
                logger.warning(f"forensic PDF: foto inválida: {e}")
        # Tabla 2 columnas
        if thumbs:
            grid = []
            for i in range(0, len(thumbs), 2):
                row = [thumbs[i], thumbs[i+1] if i + 1 < len(thumbs) else ""]
                grid.append(row)
            tg = Table(grid, colWidths=[90*mm, 90*mm])
            tg.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 2),
                ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(tg)

    # Tabla de daños incluidos en el peritaje
    if damages_incluidos:
        story.append(Spacer(1, 4*mm))
        story.append(Paragraph(f"Daños incluidos en el peritaje ({len(damages_incluidos)})", h2))
        story.append(Paragraph("Cada daño cumple los criterios de veracidad: confianza ≥ 50%, caja delimitada, severidad clasificada.", small))
        story.append(Spacer(1, 2*mm))
        tdata = [["Pieza", "Severidad", "Confianza", "Foto", "Coste est."]]
        for d in damages_incluidos[:30]:
            tdata.append([
                (d.get("part") or "—")[:38],
                (d.get("severity") or "—"),
                f"{int(float(d.get('confidence', 0)) * 100)}%",
                str(d.get("photo_index") or "—"),
                f"{d.get('estimated_cost', 0)} €",
            ])
        tdmg = Table(tdata, colWidths=[68*mm, 24*mm, 22*mm, 14*mm, 22*mm])
        tdmg.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#0b1220")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("LINEBELOW", (0, 0), (-1, -1), 0.3, colors.HexColor("#e5e7eb")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(tdmg)

    # Daños sugeridos (NO incluidos por baja confianza)
    if damages_sugeridos:
        story.append(Spacer(1, 4*mm))
        story.append(Paragraph(f"Sugerencias no incluidas ({len(damages_sugeridos)})", h2))
        story.append(Paragraph("La IA detectó estos posibles daños pero NO cumplen el umbral de veracidad para formar parte del peritaje. Revisión humana recomendada.", small))
        story.append(Spacer(1, 2*mm))
        tdata2 = [["Pieza", "Severidad", "Confianza", "Motivo de exclusión"]]
        for d in damages_sugeridos[:20]:
            conf = float(d.get("confidence", 0))
            box_ok = _is_box_valid(d.get("box_2d"))
            motivo = []
            if d.get("_motivo"): motivo.append(d["_motivo"])
            if conf < _FORENSIC_CONFIDENCE_MIN: motivo.append(f"confianza {int(conf*100)}%<50%")
            if not box_ok: motivo.append("caja inválida")
            tdata2.append([
                (d.get("part") or "—")[:38],
                (d.get("severity") or "—"),
                f"{int(conf * 100)}%",
                ", ".join(motivo) or "—",
            ])
        tsug = Table(tdata2, colWidths=[60*mm, 24*mm, 22*mm, 64*mm])
        tsug.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#6b7280")),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#fef3c7")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#854d0e")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("LINEBELOW", (0, 0), (-1, -1), 0.3, colors.HexColor("#fde68a")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(tsug)

    # Sección de firma + cadena
    story.append(PageBreak())
    story.append(Paragraph("Firma electrónica y cadena de custodia", h2))
    sig_rows = [
        ["Firmado por", sig.get("signed_by_name", "—")],
        ["Fecha de firma", (sig.get("signed_at") or "")[:19].replace("T", " ")],
        ["IP del firmante", sig.get("client_ip", "—")],
        ["Hash de esta inspección", Paragraph(sig.get("content_hash", "—"), mono)],
        ["Hash anterior en la cadena", Paragraph(sig.get("prev_hash", "—"), mono)],
        ["Revisión", str(sig.get("revision", 1))],
    ]
    st = Table(sig_rows, colWidths=[55*mm, 119*mm])
    st.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#6b7280")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LINEBELOW", (0, 0), (-1, -1), 0.4, colors.HexColor("#e5e7eb")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(st)
    story.append(Spacer(1, 4*mm))

    # Declaración firmada
    if sig.get("signature_text"):
        story.append(Paragraph("Declaración aceptada por el firmante", h2))
        story.append(Paragraph(f"«{sig.get('signature_text','')}»", body))
    story.append(Spacer(1, 6*mm))

    # QR al verificador
    try:
        qr_png = _generate_qr_png(verify_url)
        qr_img = Image(_io.BytesIO(qr_png), width=32*mm, height=32*mm)
        qr_tbl = Table([[qr_img,
                         Paragraph(f"<b>Verifica este peritaje</b><br/>{verify_url}<br/><br/>"
                                   f"Escanea el QR o introduce el hash en el verificador público.", body)]],
                       colWidths=[36*mm, 138*mm])
        qr_tbl.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
        story.append(qr_tbl)
    except Exception as e:
        logger.warning(f"forensic PDF: QR falló: {e}")

    # Disclaimer legal honesto
    story.append(Spacer(1, 8*mm))
    story.append(Paragraph(
        "<b>Aviso:</b> este documento es evidencia técnica con cadena de custodia hash, NO una firma "
        "electrónica avanzada conforme al reglamento eIDAS. FlotaDSP no presta servicios de asesoría "
        "jurídica. La validez probatoria del documento dependerá de la valoración que haga el "
        "destinatario o, en su caso, la autoridad competente.", small))

    doc.build(story)
    return buf.getvalue()


def _build_damage_report_pdf(insp: dict, vehicle: dict, driver_name: str,
                             dmg: dict, annotated_photo: Optional[bytes]) -> bytes:
    """Parte de daño individual (1-2 páginas): foto anotada + datos clave + hash.
    Pensado para enviarse tal cual al renting o al seguro."""
    import io as _io
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Image as RLImage,
                                    Table, TableStyle)

    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=16 * mm, bottomMargin=14 * mm,
                            leftMargin=16 * mm, rightMargin=16 * mm,
                            title="Parte de daño — FlotaDSP")
    ss = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=ss["Title"], fontSize=17, spaceAfter=2)
    small = ParagraphStyle("small", parent=ss["Normal"], fontSize=8.5, textColor=colors.grey)
    body = ParagraphStyle("body", parent=ss["Normal"], fontSize=10.5, leading=15)

    sev = (dmg.get("severity") or "").capitalize()
    fecha = (insp.get("created_at") or "")[:16].replace("T", " ")
    story = [
        Paragraph("PARTE DE DAÑO DE VEHÍCULO", h1),
        Paragraph("Documento generado por FlotaDSP — registro fotográfico con sello temporal", small),
        Spacer(1, 6 * mm),
    ]

    rows = [
        ["Matrícula", vehicle.get("license_plate") or "—",
         "Marca / Modelo", f"{vehicle.get('brand') or ''} {vehicle.get('model') or ''}".strip() or "—"],
        ["Fecha de detección", fecha or "—", "Conductor asignado", driver_name or "—"],
        ["Pieza afectada", dmg.get("part") or "—", "Severidad", sev or "—"],
        ["Coste estimado reparación", f"{float(dmg.get('estimated_cost') or 0):,.0f} €".replace(",", "."),
         "Centro", vehicle.get("center") or "—"],
    ]
    t = Table(rows, colWidths=[38 * mm, 52 * mm, 42 * mm, 46 * mm])
    t.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#555555")),
        ("TEXTCOLOR", (2, 0), (2, -1), colors.HexColor("#555555")),
        ("LINEBELOW", (0, 0), (-1, -2), 0.4, colors.HexColor("#dddddd")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(t)
    story.append(Spacer(1, 5 * mm))

    if dmg.get("description"):
        story.append(Paragraph(f"<b>Descripción técnica:</b> {dmg['description'][:600]}", body))
        story.append(Spacer(1, 4 * mm))

    if annotated_photo:
        try:
            from PIL import Image as PILImage
            pil = PILImage.open(_io.BytesIO(annotated_photo))
            iw, ih = pil.size
            max_w = 178 * mm
            max_h = 150 * mm
            ratio = min(max_w / iw, max_h / ih)
            story.append(RLImage(_io.BytesIO(annotated_photo), width=iw * ratio, height=ih * ratio))
            story.append(Spacer(1, 2 * mm))
            story.append(Paragraph("Fotografía original de la inspección con el daño señalizado automáticamente.", small))
        except Exception as _ie:
            logger.warning(f"parte PDF: imagen no embebible: {_ie}")

    story.append(Spacer(1, 6 * mm))
    fh = insp.get("forensic_hash") or ""
    integrity = (f"Integridad: las fotos originales están selladas con hash SHA-256 ({fh[:24]}…) "
                 f"en el momento de la subida." if fh else
                 "Las fotos originales se conservan sin alteración en el sistema.")
    story.append(Paragraph(
        f"{integrity} Inspección {insp.get('id', '')} · Generado el "
        f"{datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} · flotadsp.com", small))

    doc.build(story)
    return buf.getvalue()


@api_router.get("/inspections/{inspection_id}/damage-report")
async def damage_report_pdf(inspection_id: str, damage_index: int, scope: str = "new",
                            _=Depends(require_admin)):
    """PARTE DE DAÑO EN UN CLIC: PDF de un daño concreto (foto anotada + datos +
    coste + sello), listo para enviar al renting o al seguro."""
    insp = await db.inspections.find_one({"id": inspection_id}, {"_id": 0})
    if not insp:
        raise HTTPException(status_code=404, detail="Inspección no encontrada")
    analysis = insp.get("analysis") or {}
    pool = analysis.get("new_damages" if scope == "new" else "damages") or []
    if not (0 <= damage_index < len(pool)):
        raise HTTPException(status_code=404, detail="Daño no encontrado")
    dmg = pool[damage_index] if isinstance(pool[damage_index], dict) else {}

    vehicle = await db.vehicles.find_one({"id": insp.get("vehicle_id")}, {"_id": 0}) or {}
    driver_name = ""
    if insp.get("driver_id"):
        drv = await db.drivers.find_one({"id": insp["driver_id"]}, {"_id": 0, "name": 1})
        driver_name = (drv or {}).get("name", "")

    # Foto donde está el daño, anotada con su caja
    photos = insp.get("photos") or []
    pi = dmg.get("photo_index")
    photo_url = photos[pi - 1] if (isinstance(pi, int) and 1 <= pi <= len(photos)) else (photos[0] if photos else None)
    annotated = None
    if photo_url:
        raw = await _fetch_photo_bytes(photo_url, timeout=10)
        if raw:
            annotated = _annotate_photo_with_damages(raw, [dmg])

    loop = asyncio.get_running_loop()
    pdf_bytes = await loop.run_in_executor(
        None, _build_damage_report_pdf, insp, vehicle, driver_name, dmg, annotated)

    from fastapi.responses import Response as _Resp
    plate = (vehicle.get("license_plate") or "vehiculo").replace(" ", "")
    return _Resp(
        content=pdf_bytes, media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="parte-dano-{plate}-{inspection_id[:8]}.pdf"'},
    )


@api_router.get("/inspections/{inspection_id}/forensic-pdf")
async def forensic_pdf(inspection_id: str, _=Depends(require_admin)):
    """Genera y devuelve el PDF de peritaje técnico de una inspección firmada."""
    insp = await db.inspections.find_one({"id": inspection_id, "deleted": {"$ne": True}}, {"_id": 0})
    if not insp:
        raise HTTPException(status_code=404, detail="Inspección no encontrada")
    if not insp.get("forensic_signed"):
        raise HTTPException(status_code=409, detail="Esta inspección aún no está firmada. Fírmala antes de generar el peritaje.")
    sig = await db.forensic_signatures.find_one({"inspection_id": inspection_id}, {"_id": 0})
    if not sig:
        raise HTTPException(status_code=500, detail="Firma marcada pero no encontrada (estado inconsistente).")
    vehicle = await db.vehicles.find_one({"id": insp.get("vehicle_id")}, {"_id": 0}) or {}
    driver_name = "—"
    did = insp.get("driver_id") or vehicle.get("current_driver_id")
    if did:
        d = await db.drivers.find_one({"id": did}, {"_id": 0, "name": 1})
        if d:
            driver_name = d.get("name", "—")
    photos = (insp.get("photos") or [])[:6]
    raw_bytes_list = await asyncio.gather(*[_fetch_photo_bytes(u) for u in photos]) if photos else []

    # Filtrar daños incluidos vs sugeridos (umbral de veracidad inicial sobre Gemini).
    all_damages = ((insp.get("analysis") or {}).get("damages") or []) + \
                  ((insp.get("analysis") or {}).get("new_damages") or [])
    # Dedupe por (part, photo_index, hash de box) para no duplicar entre damages y new_damages.
    seen, dedup = set(), []
    for d in all_damages:
        k = (d.get("part"), d.get("photo_index"), tuple(d.get("box_2d") or []))
        if k in seen: continue
        seen.add(k); dedup.append(d)
    gemini_incluidos, damages_sugeridos = _filter_damages_for_peritaje(dedup)

    # ENSEMBLE STRICT: cruzar Gemini con YOLO+SAM por foto. Solo dibujamos cajas confirmadas.
    refined_per_photo = []
    incluidos_keys = set()
    for i in range(len(raw_bytes_list)):
        gem_for_photo = [d for d in gemini_incluidos if (d.get("photo_index") or 1) == (i + 1)]
        rfn = await _refine_damage_boxes_with_yolo_sam(inspection_id, i, gem_for_photo)
        refined_per_photo.append(rfn)
        for d in rfn:
            incluidos_keys.add((d.get("part"), d.get("photo_index")))
    damages_incluidos = [d for sub in refined_per_photo for d in sub]
    # Los Gemini que NO pasaron el filtro estricto pasan a "sugeridos" (transparencia total).
    for g in gemini_incluidos:
        if (g.get("part"), g.get("photo_index")) not in incluidos_keys:
            damages_sugeridos.append({**g, "_motivo": "no confirmado por YOLO/SAM"})

    # Anotar cada foto con sus daños refinados.
    annotated_bytes_list = []
    for i, raw in enumerate(raw_bytes_list):
        if not raw:
            annotated_bytes_list.append(None)
            continue
        annotated_bytes_list.append(_annotate_photo_with_damages(raw, refined_per_photo[i]))

    verify_url = f"{_PORTAL_BASE_FRONT}/verify/{sig['content_hash']}"
    # Generación PDF en hilo (reportlab es síncrono y CPU-bound).
    loop = asyncio.get_event_loop()
    pdf_bytes = await loop.run_in_executor(
        None, _build_forensic_pdf, insp, sig, vehicle, driver_name, annotated_bytes_list, verify_url,
        damages_incluidos, damages_sugeridos
    )
    fn_plate = re.sub(r"[^A-Za-z0-9]", "", (vehicle.get("license_plate") or "INSP"))
    fn_date = (sig.get("signed_at") or "")[:10]
    filename = f"peritaje-{fn_plate}-{fn_date}.pdf"
    from starlette.responses import Response
    return Response(
        content=pdf_bytes, media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api_router.get("/verify/{content_hash}")
async def verify_hash_public(content_hash: str, request: Request):
    """Verificador PÚBLICO (sin auth) de un hash de peritaje.
    Devuelve solo info mínima no sensible. Rate-limited para evitar scraping enumerativo."""
    ip = _rl_key_ip(request)
    _rl_public_action(f"verify-ip-min:{ip}", max_count=20, window_s=60,
                      detail="Demasiadas verificaciones. Espera un minuto.")
    _rl_public_action(f"verify-ip-hour:{ip}", max_count=200, window_s=3600,
                      detail="Demasiadas verificaciones esta hora.")
    h = (content_hash or "").strip().lower()
    if len(h) != 64 or not re.match(r"^[0-9a-f]+$", h):
        raise HTTPException(status_code=400, detail="Hash inválido (debe ser SHA-256 hex de 64 caracteres).")

    # Localiza el tenant donde vive este hash (índice global creado al firmar).
    idx = await global_db.forensic_index.find_one({"content_hash": h}, {"_id": 0})
    if not idx:
        return {"valid": False, "error": "Hash no encontrado. Puede que no exista o que se firmara antes de junio 2026."}

    # Carga el cliente de Mongo del tenant correcto.
    tenant_db = client[idx["db_name"]]
    sig = await tenant_db.forensic_signatures.find_one({"content_hash": h}, {"_id": 0})
    if not sig:
        return {"valid": False, "error": "Inconsistencia interna. Contacta con soporte."}
    insp = await tenant_db.inspections.find_one({"id": sig.get("inspection_id")},
                                                 {"_id": 0, "vehicle_id": 1, "created_at": 1, "id": 1}) or {}
    vehicle = await tenant_db.vehicles.find_one({"id": insp.get("vehicle_id")},
                                                 {"_id": 0, "license_plate": 1}) or {}
    next_sig = await tenant_db.forensic_signatures.find_one(
        {"signed_at": {"$gt": sig.get("signed_at", "")}}, {"_id": 0, "content_hash": 1}
    )

    return {
        "valid": True,
        "hash": h,
        "prev_hash": sig.get("prev_hash"),
        "signed_at": sig.get("signed_at"),
        "signed_by_name": sig.get("signed_by_name", ""),
        "inspection_date": insp.get("created_at"),
        "vehicle_plate_masked": _mask_plate(vehicle.get("license_plate", "")),
        "has_next_in_chain": bool(next_sig),
        "disclaimer": "Evidencia técnica con cadena de custodia hash. No constituye firma electrónica avanzada eIDAS.",
    }


# =========================
# CHECKLIST OPERATIVO POR CENTRO/TURNO (dispatcher)
# =========================
# 1 doc por (center, date, shift) en db.daily_checklists. Items con done/done_by/done_at.
# Permisos: el admin ve solo los centros de su allowed_centers (o todos si no esta seteado).

_DEFAULT_CHECKLIST_ITEMS_MANANA = [
    "Confirmar asignaciones conductor-furgoneta",
    "Validar plantilla con Cortex",
    "Briefing de mañana completado",
    "Verificar combustible flota",
    "Comprobar estado de móviles entregados",
    "Confirmar rutas cargadas en sistema",
]

_DEFAULT_CHECKLIST_ITEMS_TARDE = [
    "Recepción de furgonetas del turno de mañana",
    "Revisión de paquetes dañados / incidencias del día",
    "Confirmar asignaciones conductor-furgoneta (tarde)",
    "Validar plantilla Cortex turno tarde",
    "Briefing de tarde completado",
    "Cierre y documentación de turno",
]

def _default_items_for_shift(shift: str) -> list:
    src = _DEFAULT_CHECKLIST_ITEMS_TARDE if shift == "tarde" else _DEFAULT_CHECKLIST_ITEMS_MANANA
    return [{"id": str(uuid.uuid4()), "text": t, "done": False, "done_by": None, "done_at": None} for t in src]


async def _template_items_for(center: str, shift: str) -> list:
    """Items iniciales del día para UN centro: su plantilla propia si la tiene,
    y si no, la genérica. Así cada centro tiene sus tareas recurrentes SUYAS
    (antes todos los centros clonaban la misma lista fija cada día)."""
    tpl = await db.checklist_templates.find_one(
        {"center": center, "shift": shift}, {"_id": 0, "texts": 1})
    texts = (tpl or {}).get("texts")
    if not texts:
        return _default_items_for_shift(shift)
    return [{"id": str(uuid.uuid4()), "text": t, "done": False, "done_by": None, "done_at": None}
            for t in texts if (t or "").strip()]


@api_router.post("/checklist/template")
async def save_checklist_template(data: dict = Body(...), user: dict = Depends(require_admin)):
    """Guarda la lista actual como plantilla recurrente DE ESE CENTRO y turno:
    a partir de mañana, sus checklists nacen con estas tareas."""
    center = data.get("center")
    shift = data.get("shift")
    if not center or shift not in ("manana", "tarde"):
        raise HTTPException(400, "Faltan center/shift válidos")
    if not _user_can_see_center(user, center):
        raise HTTPException(403, "No tienes acceso a este centro")
    texts = [(it.get("text") or "").strip()[:300] if isinstance(it, dict) else str(it).strip()[:300]
             for it in (data.get("items") or [])]
    texts = [t for t in texts if t][:50]
    await db.checklist_templates.update_one(
        {"center": center, "shift": shift},
        {"$set": {"center": center, "shift": shift, "texts": texts,
                  "updated_by": user.get("name"),
                  "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True,
    )
    return {"ok": True, "count": len(texts)}


def _user_can_see_center(user: dict, center: str) -> bool:
    """True si el usuario admin puede ver ese centro. Super-admin/owner ve todos."""
    if user.get("sa") or user.get("account_type") == "owner":
        return True
    ac = user.get("allowed_centers")
    if ac is None:
        return True   # sin restricción: ve todos los de la org
    return (center or "") in ac


@api_router.get("/checklist")
async def get_checklist(center: str, date: Optional[str] = None,
                        user: dict = Depends(require_admin)):
    """Devuelve los 2 turnos (mañana, tarde) de un centro y día. Crea con defaults si no existe."""
    if not center:
        raise HTTPException(400, "Centro requerido")
    if not _user_can_see_center(user, center):
        raise HTTPException(403, "No tienes acceso a este centro")
    date = date or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    docs = await db.daily_checklists.find({"center": center, "date": date}, {"_id": 0}).to_list(2)
    by_shift = {d["shift"]: d for d in docs}
    result = {}
    now = datetime.now(timezone.utc).isoformat()
    for shift in ("manana", "tarde"):
        if shift in by_shift:
            result[shift] = by_shift[shift]
            continue
        # No existe: crear y PERSISTIR con la plantilla DEL CENTRO (IDs estables).
        new_doc = {
            "id": str(uuid.uuid4()), "center": center, "date": date, "shift": shift,
            "items": await _template_items_for(center, shift),
            "created_at": now, "updated_at": now,
        }
        try:
            await db.daily_checklists.insert_one(new_doc)
        except DuplicateKeyError:
            # Race: alguien lo creó entre el find y el insert. Recargamos.
            existing = await db.daily_checklists.find_one(
                {"center": center, "date": date, "shift": shift}, {"_id": 0})
            if existing:
                result[shift] = existing
                continue
        result[shift] = new_doc
    return result


@api_router.put("/checklist")
async def upsert_checklist(data: dict = Body(...), user: dict = Depends(require_admin)):
    """Crea/actualiza la lista de items de un (center,date,shift). El cliente envía items completos."""
    center = data.get("center")
    date = data.get("date") or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    shift = data.get("shift")
    if not center or shift not in ("manana", "tarde"):
        raise HTTPException(400, "Faltan center/shift válidos")
    if not _user_can_see_center(user, center):
        raise HTTPException(403, "No tienes acceso a este centro")
    raw_items = data.get("items") or []
    items = []
    for it in raw_items:
        if not isinstance(it, dict):
            continue
        text = (it.get("text") or "").strip()
        if not text:
            continue
        items.append({
            "id": it.get("id") or str(uuid.uuid4()),
            "text": text[:300],
            "done": bool(it.get("done")),
            "done_by": it.get("done_by"),
            "done_at": it.get("done_at"),
        })
    # Detectar tareas NUEVAS (para avisar por push a los del centro)
    _prev = await db.daily_checklists.find_one(
        {"center": center, "date": date, "shift": shift}, {"_id": 0, "items.id": 1})
    _prev_ids = {i.get("id") for i in (_prev or {}).get("items", [])}
    _new_items = [i for i in items if i["id"] not in _prev_ids] if _prev else []

    now = datetime.now(timezone.utc).isoformat()
    await db.daily_checklists.update_one(
        {"center": center, "date": date, "shift": shift},
        {"$set": {"items": items, "updated_at": now, "updated_by": user.get("name")},
         "$setOnInsert": {"id": str(uuid.uuid4()), "center": center, "date": date, "shift": shift,
                          "created_at": now}},
        upsert=True,
    )
    # Push a los coordinadores del centro por cada tarea nueva (máx 3 avisos)
    for _ni in _new_items[:3]:
        await push_center_event(
            center, f"📝 Nueva tarea · {center}", _ni.get("text", ""),
            url="/panel/checklist-operativo", exclude_id=user.get("sub"))
    return {"ok": True, "count": len(items), "new_items": len(_new_items)}


@api_router.delete("/chat/{center}/{message_id}")
async def chat_delete_message(center: str, message_id: str, user: dict = Depends(require_admin)):
    """Borra un mensaje del chat del centro. El autor puede borrar los suyos;
    el super-admin/owner puede borrar cualquiera."""
    if not await _chat_room_can_access(user, center):
        raise HTTPException(403, "No tienes acceso a este chat")
    msg = await db.chat_messages.find_one({"id": message_id, "center": center}, {"_id": 0})
    if not msg:
        raise HTTPException(404, "Mensaje no encontrado")
    is_owner_like = user.get("sa") or user.get("account_type") == "owner"
    if msg.get("author_id") != user.get("sub") and not is_owner_like:
        raise HTTPException(403, "Solo puedes borrar tus propios mensajes")
    await db.chat_messages.delete_one({"id": message_id, "center": center})
    return {"ok": True}


@api_router.post("/checklist/toggle")
async def toggle_checklist_item(data: dict = Body(...), user: dict = Depends(require_admin)):
    """Marca/desmarca un item concreto. Más rápido que reenviar la lista entera."""
    center = data.get("center")
    date = data.get("date") or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    shift = data.get("shift")
    item_id = data.get("item_id")
    done = bool(data.get("done"))
    if not center or shift not in ("manana", "tarde") or not item_id:
        raise HTTPException(400, "Faltan center/shift/item_id")
    if not _user_can_see_center(user, center):
        raise HTTPException(403, "No tienes acceso a este centro")
    now = datetime.now(timezone.utc).isoformat()
    # 1) Asegurar que el doc existe (con defaults si no).
    doc = await db.daily_checklists.find_one({"center": center, "date": date, "shift": shift}, {"_id": 0})
    if not doc:
        doc = {
            "id": str(uuid.uuid4()), "center": center, "date": date, "shift": shift,
            "items": _default_items_for_shift(shift),
            "created_at": now, "updated_at": now,
        }
        await db.daily_checklists.insert_one(doc)
    # 2) Toggle item por id.
    items = doc.get("items", [])
    found = False
    for it in items:
        if it.get("id") == item_id:
            it["done"] = done
            it["done_by"] = user.get("name") if done else None
            it["done_at"] = now if done else None
            found = True
            break
    if not found:
        raise HTTPException(404, "Item no encontrado")
    await db.daily_checklists.update_one(
        {"center": center, "date": date, "shift": shift},
        {"$set": {"items": items, "updated_at": now}}
    )
    return {"ok": True}


# =========================
# CHAT INTERNO POR CENTRO (entre admins del mismo centro)
# =========================
# 1 sala = 1 centro. Cada org tiene N salas (una por centro de la org).
# Solo ven la sala los admins con allowed_centers compatible.
# Polling cada N s desde el cliente (no hay WS — KISS).

async def _chat_room_can_access(user: dict, center: str) -> bool:
    return _user_can_see_center(user, center)


# =========================
# WEB PUSH — notificaciones al móvil de los coordinadores del panel (PWA)
# =========================
VAPID_PRIVATE_KEY = os.environ.get("VAPID_PRIVATE_KEY", "")
VAPID_PUBLIC_KEY = os.environ.get("VAPID_PUBLIC_KEY", "")
VAPID_SUBJECT = os.environ.get("VAPID_SUBJECT", "mailto:hola@flotadsp.com")
_push_enabled = bool(VAPID_PRIVATE_KEY and VAPID_PUBLIC_KEY)
if _push_enabled:
    try:
        from pywebpush import webpush as _webpush, WebPushException as _WebPushException
    except Exception as _pe:
        logger.warning(f"pywebpush no disponible, push desactivado: {_pe}")
        _push_enabled = False


@api_router.get("/push/vapid-key")
async def push_vapid_key(_=Depends(require_any_auth)):
    """Clave pública VAPID para que el navegador se suscriba (admins Y conductores)."""
    return {"key": VAPID_PUBLIC_KEY, "enabled": _push_enabled}


@api_router.post("/push/subscribe")
async def push_subscribe(data: dict, user: dict = Depends(require_any_auth)):
    """Guarda la suscripción push del dispositivo (coordinador o conductor)."""
    sub = data.get("subscription") or {}
    if not sub.get("endpoint"):
        raise HTTPException(400, "Suscripción inválida")
    await global_db.push_subscriptions.update_one(
        {"endpoint": sub["endpoint"]},
        {"$set": {
            "endpoint": sub["endpoint"],
            "subscription": sub,
            "user_id": user.get("sub"),
            "role": user.get("role"),
            "org_id": user.get("org_id"),
            "db_name": user.get("db_name"),
            "updated_at": datetime.now(timezone.utc).isoformat(),
        }},
        upsert=True,
    )
    return {"ok": True}


@api_router.post("/push/unsubscribe")
async def push_unsubscribe(data: dict, user: dict = Depends(require_any_auth)):
    ep = (data.get("endpoint") or "").strip()
    if ep:
        await global_db.push_subscriptions.delete_one({"endpoint": ep})
    return {"ok": True}


def _do_webpush(sub, payload_json):
    _webpush(subscription_info=sub, data=payload_json,
             vapid_private_key=VAPID_PRIVATE_KEY,
             vapid_claims={"sub": VAPID_SUBJECT}, timeout=10)


async def send_web_push_to_users(user_ids: list, title: str, body: str, url: str = "/panel"):
    """Envía push a los dispositivos de esos usuarios. Fire-and-forget: tolera
    fallos y borra las suscripciones caducadas (404/410). No bloquea el event loop."""
    if not (_push_enabled and user_ids):
        return
    try:
        subs = await global_db.push_subscriptions.find(
            {"user_id": {"$in": list(set(user_ids))}}, {"_id": 0}
        ).to_list(1000)
    except Exception:
        return
    if not subs:
        return
    payload = json.dumps({"title": (title or "")[:80], "body": (body or "")[:180], "url": url})
    loop = asyncio.get_running_loop()
    for s in subs:
        sub = s.get("subscription")
        if not sub:
            continue
        try:
            await loop.run_in_executor(_executor, _do_webpush, sub, payload)
        except _WebPushException as e:
            code = getattr(getattr(e, "response", None), "status_code", None)
            if code in (404, 410):
                await global_db.push_subscriptions.delete_one({"endpoint": s.get("endpoint")})
            else:
                logger.debug(f"web push {code}")
        except Exception as ex:
            logger.debug(f"web push error: {ex}")


async def _push_org_id() -> Optional[str]:
    """org_id de la organización del contexto actual (por su BD de tenant)."""
    try:
        org = await global_db.organizations.find_one(
            {"db_name": _current_db_name.get()}, {"_id": 0, "id": 1})
        return org.get("id") if org else None
    except Exception:
        return None


async def _center_recipient_ids(org_id, center, exclude_id=None) -> list:
    """IDs de admins de la org que pueden ver ese centro (destinatarios de push)."""
    if not org_id:
        return []
    try:
        users = await global_db.admin_users.find(
            {"org_id": org_id}, {"_id": 0, "id": 1, "allowed_centers": 1}).to_list(500)
    except Exception:
        return []
    out = []
    for u in users:
        if exclude_id and u.get("id") == exclude_id:
            continue
        ac = u.get("allowed_centers")
        if ac is None or (center or "") in ac:
            out.append(u.get("id"))
    return out


async def push_center_event(center, title, body, url, exclude_id=None):
    """Resuelve destinatarios del centro y lanza el push en segundo plano."""
    if not _push_enabled:
        return
    try:
        org_id = await _push_org_id()
        ids = await _center_recipient_ids(org_id, center, exclude_id=exclude_id)
        if ids:
            asyncio.create_task(send_web_push_to_users(ids, title, body, url))
    except Exception as e:
        logger.debug(f"push_center_event: {e}")


@api_router.get("/chat/{center}")
async def chat_get(center: str, since: Optional[str] = None,
                   limit: int = 100, user: dict = Depends(require_admin)):
    """Devuelve los últimos N mensajes de la sala del centro (o desde 'since' ISO si se pasa)."""
    await _require_plan_feature(user, "chat")
    if not await _chat_room_can_access(user, center):
        raise HTTPException(403, "No tienes acceso a este chat")
    q = {"center": center}
    if since:
        q["created_at"] = {"$gt": since}
    limit = max(1, min(limit, 200))
    msgs = await db.chat_messages.find(q, {"_id": 0}).sort("created_at", -1).to_list(limit)
    msgs.reverse()  # cronológico ascendente para el cliente
    return {"messages": msgs}


@api_router.post("/chat/{center}")
async def chat_post(center: str, data: dict = Body(...), user: dict = Depends(require_admin)):
    """Envía un mensaje a la sala del centro."""
    await _require_plan_feature(user, "chat")
    if not await _chat_room_can_access(user, center):
        raise HTTPException(403, "No tienes acceso a este chat")
    text = (data.get("text") or "").strip()
    if not text:
        raise HTTPException(400, "Mensaje vacío")
    if len(text) > 2000:
        raise HTTPException(400, "Mensaje demasiado largo (máx 2000)")
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "id": str(uuid.uuid4()),
        "center": center,
        "text": text[:2000],
        "author_id": user.get("sub"),
        "author_name": user.get("name") or user.get("username") or "—",
        "created_at": now,
        "pinned_to_checklist": False,
    }
    await db.chat_messages.insert_one(doc)
    doc.pop("_id", None)  # insert_one añade _id (ObjectId) que FastAPI no serializa
    # Push al móvil de los demás miembros del centro
    await push_center_event(
        center, f"💬 {doc['author_name']} · {center}", text,
        url="/panel/chat", exclude_id=user.get("sub"))
    return {"ok": True, "message": doc}


@api_router.post("/chat/{center}/{message_id}/to-checklist")
async def chat_pin_to_checklist(center: str, message_id: str,
                                 data: dict = Body(default={}),
                                 user: dict = Depends(require_admin)):
    """Convierte un mensaje de chat en un item de la checklist del turno actual.
    Marca el mensaje como 'pinned_to_checklist' para feedback visual."""
    if not await _chat_room_can_access(user, center):
        raise HTTPException(403, "No tienes acceso a este chat")
    msg = await db.chat_messages.find_one({"id": message_id, "center": center}, {"_id": 0})
    if not msg:
        raise HTTPException(404, "Mensaje no encontrado")
    # Decidir turno: el cliente puede pasarlo, si no, basado en hora UTC.
    shift = data.get("shift")
    if shift not in ("manana", "tarde"):
        h = datetime.now(timezone.utc).hour
        shift = "manana" if h < 14 else "tarde"
    date = data.get("date") or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    now = datetime.now(timezone.utc).isoformat()
    # Asegurar doc checklist + añadir item.
    doc = await db.daily_checklists.find_one({"center": center, "date": date, "shift": shift}, {"_id": 0})
    if not doc:
        doc = {
            "id": str(uuid.uuid4()), "center": center, "date": date, "shift": shift,
            "items": _default_items_for_shift(shift),
            "created_at": now, "updated_at": now,
        }
        await db.daily_checklists.insert_one(doc)
    items = doc.get("items", [])
    items.append({
        "id": str(uuid.uuid4()),
        "text": (msg["text"][:280] + ("…" if len(msg["text"]) > 280 else "")),
        "done": False, "done_by": None, "done_at": None,
        "from_chat_id": message_id, "from_chat_author": msg.get("author_name"),
    })
    await db.daily_checklists.update_one(
        {"center": center, "date": date, "shift": shift},
        {"$set": {"items": items, "updated_at": now}}
    )
    await db.chat_messages.update_one({"id": message_id}, {"$set": {"pinned_to_checklist": True, "pinned_shift": shift, "pinned_date": date}})
    return {"ok": True, "shift": shift, "date": date}


# =========================
# AI FRAUD GUARD (S3) — detecta intentos de engaño del conductor
# =========================
# 3 heurísticas:
#  1) EXIF DateTimeOriginal MUY anterior al upload → "foto antigua"
#  2) Perceptual hash coincide con inspección anterior reciente del mismo vehículo → "reusa foto"
#  3) plate_text del análisis ≠ matrícula del vehículo asignado → "matrícula no coincide"
# Score 0-100. Notifica Telegram si ≥85.

_FRAUD_OLD_PHOTO_HOURS = 24      # foto con EXIF >24h antes del upload → flag
_FRAUD_PHASH_DISTANCE = 8        # distancia Hamming <8 = misma foto (de 64 bits)
_FRAUD_PHASH_LOOKBACK_DAYS = 30  # ventana para buscar foto reusada


def _normalize_plate(p: str) -> str:
    """Quita espacios, guiones y pasa a mayúsculas para comparar matrículas.
    (Definición única: antes había un duplicado más abajo que la pisaba.)"""
    return "".join(ch for ch in (p or "").upper() if ch.isalnum())


def _exif_datetime(img_bytes: bytes):
    """Lee EXIF DateTimeOriginal de la foto. None si no hay."""
    try:
        from PIL import Image as PILImage
        import io as _io
        img = PILImage.open(_io.BytesIO(img_bytes))
        exif = img._getexif() if hasattr(img, "_getexif") else None
        if not exif:
            return None
        # tag 36867 = DateTimeOriginal
        raw = exif.get(36867) or exif.get(306)
        if not raw:
            return None
        return datetime.strptime(raw, "%Y:%m:%d %H:%M:%S")
    except Exception:
        return None


def _phash_bytes(img_bytes: bytes) -> str:
    """Perceptual hash (pHash) como string hex. None si falla."""
    try:
        import imagehash
        from PIL import Image as PILImage
        import io as _io
        img = PILImage.open(_io.BytesIO(img_bytes))
        return str(imagehash.phash(img))
    except Exception as e:
        logger.warning(f"phash failed: {e}")
        return None


def _phash_distance(h1: str, h2: str) -> int:
    """Distancia Hamming entre 2 phash hex (64 bits)."""
    try:
        import imagehash
        return imagehash.hex_to_hash(h1) - imagehash.hex_to_hash(h2)
    except Exception:
        return 999


async def _calculate_fraud_score(inspection_id: str) -> dict:
    """Calcula fraud_score y guarda en la inspección. Devuelve {score, reasons}."""
    insp = await db.inspections.find_one({"id": inspection_id}, {"_id": 0})
    if not insp:
        return {"score": 0, "reasons": []}
    vehicle = await db.vehicles.find_one({"id": insp.get("vehicle_id")}, {"_id": 0}) or {}
    reasons = []
    score = 0
    now = datetime.now(timezone.utc)

    # 3) Plate check (rápido, sin descargar fotos)
    plate_text = (insp.get("analysis") or {}).get("plate_text") or insp.get("plate_text") or ""
    expected_plate = vehicle.get("license_plate") or ""
    if plate_text and expected_plate:
        if _normalize_plate(plate_text) != _normalize_plate(expected_plate):
            reasons.append({
                "type": "plate_mismatch",
                "detail": f"IA leyó '{plate_text}' pero el vehículo asignado es '{expected_plate}'",
                "weight": 40,
            })
            score += 40

    # 1+2) EXIF + pHash sobre la primera foto (suficiente para detección rápida)
    photos = (insp.get("photos") or [])[:3]   # primeras 3 fotos máximo
    if photos:
        photo_bytes = await _fetch_photo_bytes(photos[0])
        if photo_bytes:
            # EXIF: comparar con created_at del upload
            exif_dt = _exif_datetime(photo_bytes)
            if exif_dt:
                # Asumimos zona local (no hay TZ en EXIF); margen amplio.
                try:
                    upload_dt = datetime.fromisoformat((insp.get("created_at") or "").replace("Z", "+00:00"))
                    delta_h = abs((upload_dt.replace(tzinfo=None) - exif_dt).total_seconds()) / 3600.0
                    if delta_h > _FRAUD_OLD_PHOTO_HOURS:
                        reasons.append({
                            "type": "old_photo",
                            "detail": f"Foto tomada hace {delta_h:.0f}h (EXIF {exif_dt.isoformat()}), subida posterior",
                            "weight": 35,
                        })
                        score += 35
                except Exception:
                    pass

            # pHash: comparar con inspecciones anteriores del mismo vehículo (últimos 30 días)
            this_phash = _phash_bytes(photo_bytes)
            if this_phash:
                cutoff = (now - timedelta(days=_FRAUD_PHASH_LOOKBACK_DAYS)).isoformat()
                prev = await db.inspections.find(
                    {"vehicle_id": insp.get("vehicle_id"),
                     "id": {"$ne": inspection_id},
                     "deleted": {"$ne": True},
                     "created_at": {"$gte": cutoff},
                     "first_phash": {"$exists": True}},
                    {"_id": 0, "id": 1, "first_phash": 1, "created_at": 1}
                ).limit(50).to_list(50)
                for p in prev:
                    d = _phash_distance(this_phash, p.get("first_phash", ""))
                    if d < _FRAUD_PHASH_DISTANCE:
                        reasons.append({
                            "type": "reused_photo",
                            "detail": f"Foto casi idéntica (distancia pHash {d}) a inspección anterior {p.get('id','')[:8]} del {(p.get('created_at') or '')[:10]}",
                            "weight": 50,
                        })
                        score += 50
                        break
                # Guarda el phash para futuras comparaciones (en TODAS las inspecciones, no solo fraud).
                await db.inspections.update_one({"id": inspection_id}, {"$set": {"first_phash": this_phash}})

    score = min(100, score)
    await db.inspections.update_one(
        {"id": inspection_id},
        {"$set": {"fraud_score": score, "fraud_reasons": reasons,
                  "fraud_checked_at": now.isoformat()}}
    )
    logger.info(f"fraud check insp={inspection_id[:8]} score={score} reasons={[r['type'] for r in reasons]}")

    # Notificación Telegram si alto.
    if score >= 85:
        try:
            cfg = await db.telegram_config.find_one({}, {"_id": 0})
            if cfg and cfg.get("enabled") and cfg.get("bot_token"):
                txt = (f"🚨 <b>POSIBLE FRAUDE</b>\n\nInspección <code>{inspection_id[:8]}</code> · "
                       f"vehículo {expected_plate} · score {score}/100\n\n" +
                       "\n".join(f"• {r['detail']}" for r in reasons))
                async with _aiohttp.ClientSession() as s:
                    for cid in cfg.get("chat_ids", []):
                        if cid.strip():
                            await s.post(f"https://api.telegram.org/bot{cfg['bot_token']}/sendMessage",
                                         json={"chat_id": cid, "text": txt, "parse_mode": "HTML"},
                                         timeout=_aiohttp.ClientTimeout(total=8))
        except Exception as e:
            logger.warning(f"Telegram fraud alert failed: {e}")
    return {"score": score, "reasons": reasons}


@api_router.post("/inspections/{inspection_id}/recheck-fraud")
async def recheck_fraud(inspection_id: str, _=Depends(require_admin)):
    """Fuerza recálculo del fraud_score (útil tras ediciones)."""
    return await _calculate_fraud_score(inspection_id)


# =========================
# ALERTS — solo admin
# =========================

@api_router.get("/alerts")
async def get_alerts(unread_only: bool = False, _=Depends(require_admin)):
    query = {"read": False} if unread_only else {}
    alerts = await db.alerts.find(query, {"_id": 0}).sort("created_at", -1).to_list(200)
    return alerts


@api_router.patch("/alerts/{alert_id}/read")
async def mark_alert_read(alert_id: str, _=Depends(require_admin)):
    result = await db.alerts.update_one({"id": alert_id}, {"$set": {"read": True}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Alerta no encontrada")
    return {"success": True}


# =========================
# HEALTH — público
# =========================

@api_router.get("/r2-test")
async def r2_test(_=Depends(require_admin)):
    """Diagnóstico: intenta subir un pixel de prueba a R2 y devuelve el resultado detallado."""
    r2 = get_r2()
    if not r2:
        return {
            "r2_client": False,
            "error": "Cliente R2 no inicializado",
            "R2_ENDPOINT": R2_ENDPOINT or "vacío",
            "R2_BUCKET": R2_BUCKET,
            "R2_ACCESS_KEY_set": bool(R2_ACCESS_KEY),
            "R2_SECRET_KEY_set": bool(R2_SECRET_KEY),
            "R2_PUBLIC_URL": R2_PUBLIC_URL or "vacío",
        }
    try:
        loop = asyncio.get_running_loop()
        test_bytes = b'\xff\xd8\xff\xe0\x00\x10JFIF\x00\x01\x01\x00\x00\x01\x00\x01\x00\x00'
        test_key = "_test_connectivity.jpg"
        await loop.run_in_executor(_executor, lambda: r2.put_object(
            Bucket=R2_BUCKET, Key=test_key, Body=test_bytes,
            ContentType="image/jpeg"
        ))
        test_url = f"{R2_PUBLIC_URL}/{test_key}"
        return {"r2_client": True, "upload": "ok", "test_url": test_url,
                "bucket": R2_BUCKET, "endpoint": R2_ENDPOINT}
    except Exception as e:
        return {"r2_client": True, "upload": "FAILED",
                "error": f"{type(e).__name__}: {e}",
                "bucket": R2_BUCKET, "endpoint": R2_ENDPOINT}


@api_router.get("/health")
async def health():
    try:
        await db.command("ping")
        mongo_ok = True
    except Exception:
        mongo_ok = False

    r2_ok = get_r2() is not None

    return {
        "status": "ok",
        "version": "5.3.4",
        "gemini_configured": bool(os.environ.get("GEMINI_API_KEY")) or os.environ.get("USE_VERTEX_AI","").lower() in ("1","true","yes"),
        "gemini_mode": "vertex_ai" if os.environ.get("USE_VERTEX_AI","").lower() in ("1","true","yes") else "ai_studio",
        "gemini_model": os.environ.get("GEMINI_MODEL", "gemini-2.5-flash"),
        "gemini_role": "text_analysis_only",
        "ai_service_configured": bool(AI_SERVICE_URL),
        "ai_service_url": AI_SERVICE_URL or "not configured — using location_hint fallback",
        "detection_mode": "yolo11+sam2" if AI_SERVICE_URL else "location_hint_deterministic",
        "public_base_url": PUBLIC_BASE_URL or "no configurada",
        "r2_configured": r2_ok,
        "r2_bucket": R2_BUCKET if r2_ok else "no configurado",
        "r2_public_url": R2_PUBLIC_URL or "no configurada",
        "mongo_connected": mongo_ok,
        "upload_dir_exists": UPLOAD_DIR.exists(),
        "secret_key_set": bool(os.environ.get("SECRET_KEY")),
    }


# =========================
# CORS + MIDDLEWARE + ROUTERS
# =========================

@app.middleware("http")
async def security_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    response.headers["Permissions-Policy"] = "geolocation=(), microphone=()"
    return response


# CORS: por defecto SOLO los dominios oficiales de FlotaDSP (lista blanca).
# En desarrollo se puede ampliar con CORS_ORIGINS="https://flotadsp.com,http://localhost:5175,..."
# Para abrir a todo (NO recomendado en producción): CORS_ORIGINS="*"
_DEFAULT_CORS = ",".join([
    "https://flotadsp.com",
    "https://www.flotadsp.com",
    "https://app.flotadsp.com",
    "https://flotadsp-v2.pages.dev",
    "https://test.flotadsp-v2.pages.dev",
])
cors_origins_raw = os.environ.get("CORS_ORIGINS", _DEFAULT_CORS)
cors_origins = [o.strip() for o in cors_origins_raw.split(",") if o.strip()]
# Si alguien explícitamente quiere abrirlo, usar wildcard. Si no, credenciales activas.
use_credentials = cors_origins != ["*"]
logger.info(f"CORS allow_origins={cors_origins} credentials={use_credentials}")

# IMPORTANTE: middleware ANTES de montar rutas estáticas
app.add_middleware(
    CORSMiddleware,
    allow_origins=cors_origins,
    # Cualquier preview deploy de Cloudflare Pages del proyecto flotadsp-v2
    # (<hash>.flotadsp-v2.pages.dev) y el dev server local de Vite (verificación).
    allow_origin_regex=r"https://([a-z0-9-]+\.)?flotadsp-v2\.pages\.dev|http://localhost:517[0-9]",
    allow_credentials=use_credentials,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["X-Boxes-Found", "X-Legend", "X-Diag", "X-Img-Bytes"],
)

# NOTA: app.include_router() se llama al FINAL del archivo, despues de definir
# TODOS los endpoints, para que las rutas anadidas se registren correctamente.

# Montar uploads locales DESPUÉS del middleware (solo fallback sin R2)
app.mount("/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="uploads")


@app.get("/")
async def home():
    return {"status": "FlotaDSP backend running", "version": "5.3.4"}


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
    _executor.shutdown(wait=False)


if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8080))
    uvicorn.run("server:app", host="0.0.0.0", port=port, reload=False)


# =============================================================
# ENDPOINTS ADICIONALES — añadidos para completar funcionalidad
# =============================================================

import aiohttp as _aiohttp

# =========================
# TELEGRAM
# =========================

class TelegramConfig(BaseModel):
    bot_token: str
    chat_ids: List[str]
    enabled: bool = True
    notify_critical: bool = True
    notify_new_damage: bool = True
    notify_inspections: bool = False


@api_router.post("/telegram/config")
async def save_telegram_config(config: TelegramConfig, _=Depends(require_admin)):
    doc = serialize_doc(config.model_dump())
    doc["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.telegram_config.replace_one({}, doc, upsert=True)
    return {"success": True, "message": "Configuración de Telegram guardada"}


@api_router.get("/telegram/config")
async def get_telegram_config(_=Depends(require_admin)):
    config = await db.telegram_config.find_one({}, {"_id": 0})
    if not config:
        return {"bot_token": "", "chat_ids": [""], "enabled": False,
                "notify_critical": True, "notify_new_damage": True, "notify_inspections": False}
    return config


@api_router.post("/telegram/test")
async def test_telegram(_=Depends(require_admin)):
    config = await db.telegram_config.find_one({}, {"_id": 0})
    if not config or not config.get("bot_token"):
        raise HTTPException(status_code=400, detail="Telegram no configurado")

    token = config["bot_token"]
    chat_ids = config.get("chat_ids", [])

    results = []
    async with _aiohttp.ClientSession() as session:
        for chat_id in chat_ids:
            if not chat_id.strip():
                continue
            url = f"https://api.telegram.org/bot{token}/sendMessage"
            payload = {"chat_id": chat_id, "text": "✅ FlotaDSP — Test de conexión exitoso", "parse_mode": "HTML"}
            try:
                async with session.post(url, json=payload, timeout=_aiohttp.ClientTimeout(total=10)) as resp:
                    data = await resp.json()
                    results.append({"chat_id": chat_id, "ok": data.get("ok", False)})
            except Exception as e:
                results.append({"chat_id": chat_id, "ok": False, "error": str(e)})

    all_ok = all(r["ok"] for r in results)
    return {"success": all_ok, "results": results}


async def send_telegram_alert(title: str, message: str, severity: str = "critico"):
    """Envía alerta a Telegram. Llamar cuando hay daños graves."""
    try:
        config = await db.telegram_config.find_one({}, {"_id": 0})
        if not config or not config.get("enabled") or not config.get("bot_token"):
            return
        if severity in ["critico", "grave"] and not config.get("notify_critical"):
            return

        token = config["bot_token"]
        emoji = "🚨" if severity == "critico" else "⚠️"
        text = f"{emoji} <b>{title}</b>\n\n{message}"

        async with _aiohttp.ClientSession() as session:
            for chat_id in config.get("chat_ids", []):
                if not chat_id.strip():
                    continue
                url = f"https://api.telegram.org/bot{token}/sendMessage"
                await session.post(url, json={"chat_id": chat_id, "text": text, "parse_mode": "HTML"},
                                   timeout=_aiohttp.ClientTimeout(total=5))
    except Exception as e:
        logger.warning(f"Error enviando Telegram: {e}")


# =========================
# MONITORIZACIÓN DE ERRORES — el panel/portal reporta errores JS y el backend
# avisa por Telegram de los 500 no controlados. Dedupe 1h por mensaje.
# =========================

_err_alerted: dict = {}          # hash del error → timestamp del último aviso
_ERR_ALERT_WINDOW_S = 3600       # backend: máx 1 aviso por error único y hora
_ERR_ALERT_WINDOW_FRONT_S = 6 * 3600  # frontend: máx 1 aviso por error único cada 6h

# Ruido que NO merece Telegram (se loguea y ya): chunks viejos tras un deploy,
# cortes de red de móvil, errores opacos de extensiones, desconexiones de cliente…
_ERR_NOISE_PATTERNS = (
    "dynamically imported module", "importing a module script failed", "chunkloaderror",
    "failed to fetch", "networkerror", "load failed", "network error",
    "script error", "resizeobserver",
    # backend: desconexiones y cancelaciones normales, no son fallos reales
    "clientdisconnect", "connection reset", "broken pipe", "cancellederror",
    "connectionreset", "peer closed", "response ended prematurely",
)

# Los avisos de "error backend" por Telegram eran ruido inútil ("no sé ni a qué
# se refiere"): se apagan por defecto (todo queda en el log). Para reactivarlos:
# fly secrets set BACKEND_ERROR_ALERTS=1
def _backend_alerts_on() -> bool:
    return os.environ.get("BACKEND_ERROR_ALERTS", "0").lower() in ("1", "true", "yes")


async def _notify_error_once(kind: str, message: str, extra: str = ""):
    """Log + Telegram con dedupe. kind: 'frontend' | 'backend'."""
    import hashlib as _hl
    logger.error(f"[{kind}] {message} {extra}"[:2000])
    # Ruido conocido (cualquier origen): queda en el log, no molesta por Telegram.
    if any(p in (message or "").lower() for p in _ERR_NOISE_PATTERNS):
        return
    # Errores de backend: silenciados por defecto (solo log). Reactivables por env.
    if kind == "backend" and not _backend_alerts_on():
        return
    key = _hl.sha256(f"{kind}:{message[:300]}".encode()).hexdigest()
    now_ts = datetime.now(timezone.utc).timestamp()
    window = _ERR_ALERT_WINDOW_FRONT_S if kind == "frontend" else _ERR_ALERT_WINDOW_S
    if now_ts - _err_alerted.get(key, 0) < window:
        return
    _err_alerted[key] = now_ts
    # Limpieza para que el dict no crezca sin límite
    if len(_err_alerted) > 500:
        cutoff = now_ts - _ERR_ALERT_WINDOW_S
        for k in [k for k, t in _err_alerted.items() if t < cutoff]:
            _err_alerted.pop(k, None)
    emoji = "🖥️" if kind == "frontend" else "🔥"
    await send_telegram_alert(
        f"{emoji} Error en {'el panel/portal' if kind == 'frontend' else 'el backend'}",
        f"{message[:500]}\n{extra[:400]}".strip(),
        severity="critico",
    )


@api_router.post("/client-error")
async def report_client_error(data: dict, request: Request):
    """Recibe errores JS del frontend (window.onerror). Público pero rate-limited."""
    _rl_public_action(f"ce:{_rl_key_ip(request)}", max_count=10, window_s=600,
                      detail="Demasiados reportes")
    message = str(data.get("message") or "")[:500]
    if not message:
        return {"success": True}
    stack = str(data.get("stack") or "")[:800]
    url = str(data.get("url") or "")[:200]
    ua = (request.headers.get("user-agent") or "")[:120]
    await _notify_error_once("frontend", message, f"URL: {url}\nStack: {stack[:300]}\nUA: {ua}")
    return {"success": True}


@app.exception_handler(Exception)
async def unhandled_exception_handler(request: Request, exc: Exception):
    """Los 500 no controlados se loguean con traza y avisan por Telegram (dedupe 1h)."""
    import traceback as _tb
    trace = "".join(_tb.format_exception(type(exc), exc, exc.__traceback__))[-1500:]
    try:
        await _notify_error_once("backend", f"{type(exc).__name__}: {exc}",
                                 f"{request.method} {request.url.path}\n{trace[-600:]}")
    except Exception:
        logger.error(f"Error notificando excepción: {exc}")
    return JSONResponse(status_code=500, content={"detail": "Error interno del servidor"})


def _severity_emoji(sev: str) -> str:
    return {"critico": "🔴", "grave": "🟠", "moderado": "🟡", "leve": "🟢"}.get(sev, "⚪")


async def send_telegram_damage_alert(plate, driver_name, analysis, photo_urls, inspection_id, center=None):
    """Envía a Telegram una alerta de daños con formato detallado: matrícula, conductor,
    daños NUEVOS y enlaces a las fotos. Además dispara push a los coordinadores del centro."""
    # Push al móvil (independiente de Telegram: aunque Telegram no esté configurado)
    try:
        _sev = (getattr(analysis, "severity", None)
                or (analysis.get("severity") if isinstance(analysis, dict) else "") or "")
        await push_center_event(
            center or "", f"🚨 Daño {_sev} · {center or ''}".strip(),
            f"{plate}: nueva inspección con daño, revísala", url="/panel/revision")
    except Exception as _pe:
        logger.debug(f"push daño: {_pe}")
    try:
        config = await db.telegram_config.find_one({}, {"_id": 0})
        if not config or not config.get("enabled") or not config.get("bot_token"):
            return
        if not config.get("notify_critical", True):
            return

        token = config["bot_token"]
        sev = analysis.severity
        head_emoji = "🚨" if sev == "critico" else "⚠️"

        # Daños nuevos (si la IA los distingue); si no, usar todos
        nuevos = analysis.new_damages if analysis.new_damages else analysis.damages

        # Extraer código de centro corto (ej: "AMZL OGA5 SANTIAGO XPT" → "OGA5")
        center_display = ""
        if center:
            import re as _re_tg
            m = _re_tg.search(r'\b(OGA\d|DGA\d|[A-Z]{2,4}\d)\b', center.upper())
            center_display = m.group(1) if m else center[:6]

        lines = []
        center_tag = f" [{center_display}]" if center_display else ""
        lines.append(f"{head_emoji} <b>PERITAJE — DAÑOS DETECTADOS{center_tag}</b>")
        lines.append("")
        if center_display:
            lines.append(f"🏢 <b>Centro:</b> {center_display}")
        lines.append(f"🚗 <b>Matrícula:</b> {plate}")
        lines.append(f"👤 <b>Conductor:</b> {driver_name}")
        lines.append(f"{_severity_emoji(sev)} <b>Severidad:</b> {sev.upper()}")
        lines.append(f"💶 <b>Coste estimado:</b> {analysis.total_estimated_cost:.0f} €")

        # Aviso anti-fraude si la matrícula leída no coincide
        detected = (analysis.detected_plate or "").replace(" ", "").upper()
        plate_norm = (plate or "").replace(" ", "").upper()
        if detected and plate_norm and detected != plate_norm:
            lines.append("")
            lines.append(f"⛔ <b>ALERTA FRAUDE:</b> la matrícula leída en la foto ({analysis.detected_plate}) NO coincide con la del vehículo ({plate}).")
        if analysis.fraud_warnings:
            lines.append("")
            for fw in analysis.fraud_warnings[:3]:
                lines.append(f"⛔ {fw}")

        # Daños nuevos
        lines.append("")
        if nuevos:
            lines.append(f"<b>🆕 DAÑOS NUEVOS ({len(nuevos)}):</b>")
            for i, d in enumerate(nuevos[:10], 1):
                loc = f" [{d.location_hint}]" if getattr(d, "location_hint", "") else ""
                lines.append(f"{i}. <b>{d.part}</b>{loc} — {d.severity}")
                if d.description:
                    lines.append(f"   {d.description}")
        else:
            lines.append("Sin daños nuevos respecto al peritaje anterior.")

        # Enlaces a las fotos — etiquetadas por zona (orden del portal conductor)
        zonas = ["Frontal", "Trasera", "Lateral izquierdo", "Lateral derecho",
                 "Foto 5", "Foto 6", "Foto 7", "Foto 8"]
        if photo_urls:
            lines.append("")
            lines.append("<b>📸 Fotos por zona:</b>")
            for i, u in enumerate(photo_urls[:8]):
                etiqueta = zonas[i] if i < len(zonas) else f"Foto {i+1}"
                lines.append(f"• {etiqueta}: <a href=\"{u}\">ver</a>")

        text = "\n".join(lines)

        async with _aiohttp.ClientSession() as session:
            for chat_id in config.get("chat_ids", []):
                if not chat_id.strip():
                    continue
                url = f"https://api.telegram.org/bot{token}/sendMessage"
                await session.post(
                    url,
                    json={"chat_id": chat_id, "text": text, "parse_mode": "HTML",
                          "disable_web_page_preview": True},
                    timeout=_aiohttp.ClientTimeout(total=8)
                )
    except Exception as e:
        logger.warning(f"Error enviando alerta detallada Telegram: {e}")


async def send_daily_inspection_summary():
    """Resumen nocturno por centro: qué furgonetas del cuadrante de hoy tienen
    inspección y cuáles no. Un mensaje de Telegram por centro con cuadrante."""
    try:
        config = await db.telegram_config.find_one({}, {"_id": 0})
        if not config or not config.get("enabled") or not config.get("bot_token"):
            logger.info("Resumen diario: Telegram no configurado — omitido")
            return
        token = config["bot_token"]
        chat_ids = [c for c in config.get("chat_ids", []) if c.strip()]
        if not chat_ids:
            return

        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        assignments = await db.daily_assignments.find({"date": today}, {"_id": 0}).to_list(20)
        if not assignments:
            logger.info("Resumen diario: sin cuadrante hoy — nada que enviar")
            return

        # Fecha bonita en español
        dias = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
        now_dt = datetime.now(timezone.utc)
        fecha_str = f"{dias[now_dt.weekday()]} {now_dt.strftime('%d/%m/%Y')}"

        # Inspecciones de hoy (una consulta, agrupadas por vehículo)
        insps_today = await db.inspections.find(
            {"deleted": {"$ne": True}, "created_at": {"$regex": f"^{today}"}},
            {"_id": 0, "vehicle_id": 1, "analysis": 1}
        ).to_list(1000)
        insp_by_vehicle = {}
        for i in insps_today:
            vid = i.get("vehicle_id")
            if vid:
                insp_by_vehicle[vid] = i

        for assignment in assignments:
            center = assignment.get("center", "—")
            slots = [s for s in assignment.get("slots", []) if s.get("vehicle_id")]
            if not slots:
                continue

            total = len(slots)
            done = []
            missing = []
            sev_counts = {"sin": 0, "leve": 0, "grave": 0}
            for slot in slots:
                vid = slot["vehicle_id"]
                insp = insp_by_vehicle.get(vid)
                if insp:
                    done.append(slot)
                    sev = ((insp.get("analysis") or {}).get("severity") or "").lower()
                    if sev in ("grave", "critico", "crítico"):
                        sev_counts["grave"] += 1
                    elif sev in ("leve", "moderado"):
                        sev_counts["leve"] += 1
                    else:
                        sev_counts["sin"] += 1
                else:
                    missing.append(slot)

            pct = round(len(done) / total * 100) if total else 0
            lines = []
            if not missing:
                lines.append(f"✅ <b>RESUMEN DIARIO — {center}</b>")
                lines.append(f"📅 {fecha_str}")
                lines.append("")
                lines.append("🎉 <b>¡Todas las inspecciones completadas!</b>")
                lines.append("")
                lines.append(f"🚐 Furgonetas asignadas: {total}")
                lines.append(f"📸 Inspecciones recibidas: {len(done)} (100%)")
                lines.append("")
                lines.append(f"🟢 Sin daños: {sev_counts['sin']}")
                lines.append(f"🟡 Leves: {sev_counts['leve']}")
                lines.append(f"🔴 Graves/críticos: {sev_counts['grave']}")
                lines.append("")
                lines.append("💪 Día redondo. Sin pendientes.")
            else:
                lines.append(f"⚠️ <b>RESUMEN DIARIO — {center}</b>")
                lines.append(f"📅 {fecha_str}")
                lines.append("")
                lines.append(f"❌ <b>Faltan {len(missing)} inspecciones:</b>")
                for slot in missing[:15]:
                    drv = slot.get("driver_name") or "Sin conductor asignado"
                    lines.append(f"• {drv} — {slot.get('vehicle_plate', '—')}")
                lines.append("")
                lines.append(f"🚐 Asignadas: {total} · 📸 Recibidas: {len(done)} ({pct}%)")
                lines.append("")
                lines.append(f"🟢 Sin daños: {sev_counts['sin']} · 🟡 Leves: {sev_counts['leve']} · 🔴 Graves: {sev_counts['grave']}")

            text = "\n".join(lines)
            async with _aiohttp.ClientSession() as session:
                for chat_id in chat_ids:
                    try:
                        await session.post(
                            f"https://api.telegram.org/bot{token}/sendMessage",
                            json={"chat_id": chat_id, "text": text, "parse_mode": "HTML",
                                  "disable_web_page_preview": True},
                            timeout=_aiohttp.ClientTimeout(total=8)
                        )
                    except Exception as _se:
                        logger.warning(f"Resumen diario: error enviando a {chat_id}: {_se}")
            logger.info(f"Resumen diario enviado: {center} — {len(done)}/{total} inspecciones")
    except Exception as e:
        logger.error(f"Error en resumen diario: {e}", exc_info=True)


async def backup_database_to_r2() -> dict:
    """Exporta TODAS las colecciones a un JSON.gz y lo sube a R2 (backups/).
    Conserva los últimos 14 backups. El cluster Atlas es M0 (sin backups nativos):
    esto es la red de seguridad contra corrupción o borrado accidental."""
    import gzip as _gzip
    r2 = get_r2()
    if not r2:
        return {"success": False, "error": "R2 no configurado"}
    try:
        dump = {}
        total_docs = 0
        names = await db.list_collection_names()
        for cname in names:
            docs = await db[cname].find({}, {"_id": 0}).to_list(100000)
            dump[cname] = docs
            total_docs += len(docs)

        raw = json.dumps(dump, default=str, ensure_ascii=False).encode("utf-8")
        compressed = _gzip.compress(raw, compresslevel=6)
        stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d_%H%M")
        key = f"backups/flotadsp_{stamp}.json.gz"

        loop = asyncio.get_running_loop()
        await loop.run_in_executor(_executor, lambda: r2.put_object(
            Bucket=R2_BUCKET, Key=key, Body=compressed,
            ContentType="application/gzip"
        ))

        # Rotación: conservar solo los 14 más recientes
        deleted = 0
        try:
            listing = await loop.run_in_executor(_executor, lambda: r2.list_objects_v2(
                Bucket=R2_BUCKET, Prefix="backups/"
            ))
            keys = sorted([o["Key"] for o in listing.get("Contents", [])], reverse=True)
            for old_key in keys[14:]:
                await loop.run_in_executor(_executor, lambda k=old_key: r2.delete_object(Bucket=R2_BUCKET, Key=k))
                deleted += 1
        except Exception as _rot:
            logger.warning(f"Rotación de backups falló (no crítico): {_rot}")

        size_mb = round(len(compressed) / 1024 / 1024, 2)
        logger.info(f"Backup BD OK: {key} — {total_docs} docs, {size_mb}MB, {len(names)} colecciones, {deleted} antiguos borrados")
        return {"success": True, "key": key, "collections": len(names),
                "documents": total_docs, "size_mb": size_mb}
    except Exception as e:
        logger.error(f"Backup BD FALLÓ: {e}", exc_info=True)
        return {"success": False, "error": str(e)}


@api_router.post("/admin/backup-now")
async def trigger_backup(_=Depends(require_admin)):
    """Lanza un backup manual de la base de datos a R2."""
    return await backup_database_to_r2()


@app.on_event("startup")
async def start_backup_scheduler():
    """Backup automático diario a las 04:00 hora española."""
    async def _backup_loop():
        from zoneinfo import ZoneInfo
        madrid = ZoneInfo("Europe/Madrid")
        while True:
            try:
                now = datetime.now(madrid)
                target = now.replace(hour=4, minute=0, second=0, microsecond=0)
                if now >= target:
                    target = target + timedelta(days=1)
                await asyncio.sleep((target - now).total_seconds())
                result = await backup_database_to_r2()
                if not result.get("success"):
                    # Avisar por Telegram si el backup falla — esto SÍ es crítico
                    try:
                        config = await db.telegram_config.find_one({}, {"_id": 0})
                        if config and config.get("enabled") and config.get("bot_token"):
                            async with _aiohttp.ClientSession() as session:
                                for chat_id in config.get("chat_ids", []):
                                    if chat_id.strip():
                                        await session.post(
                                            f"https://api.telegram.org/bot{config['bot_token']}/sendMessage",
                                            json={"chat_id": chat_id,
                                                  "text": f"🚨 EL BACKUP DIARIO DE LA BASE DE DATOS HA FALLADO: {result.get('error')}"},
                                            timeout=_aiohttp.ClientTimeout(total=8))
                    except Exception:
                        pass
                await asyncio.sleep(70)
            except Exception as e:
                logger.error(f"Backup scheduler: {e}")
                await asyncio.sleep(600)
    asyncio.create_task(_backup_loop())


async def send_weekly_summary():
    """Informe de gerencia de la semana (lunes 08:00): cumplimiento por centro,
    € de daños nuevos, top 3 del scoring y furgonetas con más daños nuevos."""
    try:
        config = await db.telegram_config.find_one({}, {"_id": 0})
        if not config or not config.get("enabled") or not config.get("bot_token"):
            return {"success": False, "error": "Telegram no configurado"}
        token = config["bot_token"]
        chat_ids = [c for c in config.get("chat_ids", []) if c.strip()]
        if not chat_ids:
            return {"success": False, "error": "sin chats"}

        now = datetime.now(timezone.utc)
        week_ago = (now - timedelta(days=7)).isoformat()
        week_start_d = (now - timedelta(days=7)).strftime("%d/%m")
        week_end_d = now.strftime("%d/%m")

        # Inspecciones de la semana
        insps = await db.inspections.find(
            {"deleted": {"$ne": True}, "created_at": {"$gte": week_ago}},
            {"_id": 0, "vehicle_id": 1, "analysis": 1, "analysis_status": 1, "reference_photos": 1}
        ).to_list(3000)

        # Cumplimiento: slots asignados vs inspecciones, por los cuadrantes de la semana
        assignments = await db.daily_assignments.find(
            {"date": {"$gte": (now - timedelta(days=7)).strftime("%Y-%m-%d")}},
            {"_id": 0, "center": 1, "slots": 1, "date": 1}
        ).to_list(100)
        total_slots = sum(len([sl for sl in a.get("slots", []) if sl.get("driver_id")]) for a in assignments)

        # € de daños nuevos de la semana + furgonetas con más daños nuevos
        total_eur = 0.0
        dam_by_vehicle = {}
        for i in insps:
            a = i.get("analysis") or {}
            if i.get("analysis_status") != "ok":
                continue
            if i.get("reference_photos"):
                nds = [nd for nd in (a.get("new_damages") or []) if isinstance(nd, dict)]
                total_eur += sum((nd.get("estimated_cost") or 0) for nd in nds)
                if nds:
                    dam_by_vehicle[i.get("vehicle_id")] = dam_by_vehicle.get(i.get("vehicle_id"), 0) + len(nds)

        # Matrículas de las furgonetas problemáticas
        worst = sorted(dam_by_vehicle.items(), key=lambda x: -x[1])[:3]
        plates = {}
        if worst:
            vids = [w[0] for w in worst]
            async for v in db.vehicles.find({"id": {"$in": vids}}, {"_id": 0, "id": 1, "license_plate": 1}):
                plates[v["id"]] = v.get("license_plate", "?")

        # Top 3 del scoring del mes en curso
        top3_txt = ""
        try:
            sc = await get_driver_scoring(month=now.month, year=now.year, _=None)
            ranked = [x for x in sc.get("scores", []) if not x.get("insufficient")][:3]
            medals = ["🥇", "🥈", "🥉"]
            top3_txt = "\n".join(f"{medals[i]} {r['name']} — {r['total']} pts" for i, r in enumerate(ranked))
        except Exception as _se:
            logger.warning(f"Weekly: scoring falló: {_se}")

        lines = [
            f"📊 <b>RESUMEN SEMANAL</b> ({week_start_d} – {week_end_d})",
            "",
            f"📸 Inspecciones: <b>{len(insps)}</b>" + (f" de {total_slots} asignaciones ({round(len(insps)/total_slots*100)}%)" if total_slots else ""),
            f"💶 Daños nuevos detectados: <b>{round(total_eur):,} €</b>".replace(",", "."),
        ]
        if worst:
            lines.append("")
            lines.append("🚨 <b>Furgonetas con más daños nuevos:</b>")
            for vid, cnt in worst:
                lines.append(f"  • {plates.get(vid, '?')} — {cnt} daño(s)")
        if top3_txt:
            lines.append("")
            lines.append("🏆 <b>Top scoring del mes:</b>")
            lines.append(top3_txt)
        lines.append("")
        lines.append("¡Buena semana! 💪")

        text = "\n".join(lines)
        async with _aiohttp.ClientSession() as session:
            for cid in chat_ids:
                await session.post(
                    f"https://api.telegram.org/bot{token}/sendMessage",
                    json={"chat_id": cid, "text": text, "parse_mode": "HTML"},
                    timeout=_aiohttp.ClientTimeout(total=8))
        logger.info("Resumen semanal enviado")
        return {"success": True}
    except Exception as e:
        logger.error(f"Resumen semanal: {e}", exc_info=True)
        return {"success": False, "error": str(e)}


@api_router.post("/telegram/send-weekly-summary")
async def trigger_weekly_summary(_=Depends(require_admin)):
    """Dispara el resumen semanal a mano (para probar sin esperar al lunes)."""
    return await send_weekly_summary()


@app.on_event("startup")
async def start_weekly_summary_scheduler():
    """Resumen semanal cada lunes a las 08:00 hora española."""
    async def _loop():
        from zoneinfo import ZoneInfo
        madrid = ZoneInfo("Europe/Madrid")
        while True:
            try:
                now = datetime.now(madrid)
                days_ahead = (0 - now.weekday()) % 7  # 0 = lunes
                target = (now + timedelta(days=days_ahead)).replace(hour=8, minute=0, second=0, microsecond=0)
                if target <= now:
                    target = target + timedelta(days=7)
                await asyncio.sleep((target - now).total_seconds())
                await send_weekly_summary()
                await asyncio.sleep(70)
            except Exception as e:
                logger.error(f"Weekly scheduler: {e}")
                await asyncio.sleep(600)
    asyncio.create_task(_loop())


# =========================
# DIGEST SEMANAL POR EMAIL A LOS DSPs (retención) — lunes 07:30
# =========================

async def send_weekly_email_digest():
    """Email semanal a cada organización con email: ITVs próximas, inspecciones
    y daños nuevos de la semana. Dedupe por semana ISO en global_db.app_meta."""
    if os.environ.get("WEEKLY_DIGEST_DISABLED", "").lower() in ("1", "true"):
        return {"success": False, "error": "deshabilitado por env"}
    week_key = datetime.now(timezone.utc).strftime("%G-W%V")
    meta = await global_db.app_meta.find_one({"_id": "weekly_email_digest"})
    if meta and meta.get("last_week") == week_key:
        return {"success": False, "error": f"ya enviado esta semana ({week_key})"}

    since = (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()
    orgs = await global_db.organizations.find(
        {"email": {"$nin": [None, ""]}, "status": {"$nin": ["suspended", "deleted"]}},
        {"_id": 0, "id": 1, "name": 1, "email": 1, "db_name": 1}
    ).to_list(500)

    sent = 0
    for org in orgs:
        try:
            set_current_org_db(org.get("db_name"))
            n_vehicles = await db.vehicles.count_documents({"status": {"$ne": "deleted"}})
            if n_vehicles == 0:
                continue  # sin flota aún: un digest vacío no aporta
            n_taller = await db.vehicles.count_documents({"status": "taller"})
            n_insp = await db.inspections.count_documents(
                {"deleted": {"$ne": True}, "created_at": {"$gte": since}})
            n_damages = await db.inspections.count_documents(
                {"deleted": {"$ne": True}, "created_at": {"$gte": since},
                 "analysis.new_damages.0": {"$exists": True}})

            hoy = datetime.now(timezone.utc).strftime("%Y-%m-%d")
            limite = (datetime.now(timezone.utc) + timedelta(days=30)).strftime("%Y-%m-%d")
            itvs = await db.vehicles.find(
                {"status": {"$ne": "deleted"}, "itv_date": {"$gte": hoy, "$lte": limite}},
                {"_id": 0, "license_plate": 1, "itv_date": 1}
            ).sort("itv_date", 1).to_list(5)

            itv_html = "".join(
                f"<div style='font-size:13px;color:#cbd3e0;margin-bottom:4px'>🛡 <b>{v.get('license_plate','?')}</b> — ITV el {v.get('itv_date','')}</div>"
                for v in itvs
            ) or "<div style='font-size:13px;color:#34d399'>✓ Ninguna ITV caduca en los próximos 30 días</div>"

            html = f"""
<div style="font-family:Inter,sans-serif;max-width:520px;margin:0 auto;background:#0b0d10;color:#eef1f6;border-radius:16px;overflow:hidden">
  <div style="background:linear-gradient(135deg,#0ea5e9,#0369a1);padding:26px 28px;text-align:center">
    <div style="font-size:26px;margin-bottom:6px">⚡</div>
    <h1 style="margin:0;font-size:20px;font-weight:900;color:#fff">Tu semana en FlotaDSP</h1>
    <p style="margin:6px 0 0;color:rgba(255,255,255,.8);font-size:13px">{org.get('name','')}</p>
  </div>
  <div style="padding:26px 28px">
    <div style="display:flex;gap:10px;margin-bottom:20px">
      <div style="flex:1;background:#13161b;border:1px solid rgba(255,255,255,.08);border-radius:12px;padding:14px;text-align:center">
        <div style="font-size:22px;font-weight:900;color:#0ea5e9">{n_insp}</div>
        <div style="font-size:11px;color:#64748b">inspecciones</div>
      </div>
      <div style="flex:1;background:#13161b;border:1px solid rgba(255,255,255,.08);border-radius:12px;padding:14px;text-align:center">
        <div style="font-size:22px;font-weight:900;color:{'#f87171' if n_damages else '#34d399'}">{n_damages}</div>
        <div style="font-size:11px;color:#64748b">con daños nuevos</div>
      </div>
      <div style="flex:1;background:#13161b;border:1px solid rgba(255,255,255,.08);border-radius:12px;padding:14px;text-align:center">
        <div style="font-size:22px;font-weight:900;color:{'#fb923c' if n_taller else '#34d399'}">{n_taller}</div>
        <div style="font-size:11px;color:#64748b">en taller</div>
      </div>
    </div>
    <div style="background:#13161b;border:1px solid rgba(255,255,255,.08);border-radius:12px;padding:16px 18px;margin-bottom:22px">
      <div style="font-size:11px;color:#64748b;font-weight:700;text-transform:uppercase;letter-spacing:.06em;margin-bottom:10px">ITV próximas (30 días)</div>
      {itv_html}
    </div>
    <a href="https://flotadsp.com/panel" style="display:block;text-align:center;background:linear-gradient(135deg,#0ea5e9,#0369a1);color:#fff;text-decoration:none;padding:13px 24px;border-radius:10px;font-weight:800;font-size:14px">
      Abrir el panel →
    </a>
    <p style="margin:12px 0 0;color:#475569;font-size:11px;text-align:center;line-height:1.5">
      🇬🇧 Your week in FlotaDSP: inspections, new damage and upcoming MOT dates above. Open the panel for details.
    </p>
  </div>
</div>"""
            ok = await _send_resend_email(
                org["email"],
                f"Tu semana en FlotaDSP: {n_insp} inspecciones, {n_damages} con daños nuevos",
                html,
            )
            if ok:
                sent += 1
        except Exception as _oe:
            logger.warning(f"Digest semanal: error con org {org.get('id')}: {_oe}")
    set_current_org_db(None)

    await global_db.app_meta.update_one(
        {"_id": "weekly_email_digest"},
        {"$set": {"last_week": week_key, "sent": sent, "at": datetime.now(timezone.utc).isoformat()}},
        upsert=True,
    )
    logger.info(f"Digest semanal por email: {sent} orgs")
    return {"success": True, "sent": sent, "week": week_key}


@api_router.post("/admin/send-weekly-digest")
async def trigger_weekly_digest(_=Depends(require_superadmin)):
    """Dispara el digest por email a mano (para probar). Dedupe semanal incluido."""
    return await send_weekly_email_digest()


async def _send_inspection_reminders():
    """Push a cada conductor con furgoneta asignada HOY que aún no subió inspección."""
    if not _push_enabled:
        return
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    orgs = await global_db.organizations.find(
        {"status": {"$nin": ["suspended", "deleted"]}},
        {"_id": 0, "id": 1, "db_name": 1}).to_list(500)
    total = 0
    for org in orgs:
        try:
            set_current_org_db(org.get("db_name"))
            assignments = await db.daily_assignments.find(
                {"date": today}, {"_id": 0, "slots": 1}).to_list(50)
            pending = {}
            for a in assignments:
                for s in a.get("slots", []):
                    if s.get("driver_id") and s.get("vehicle_id"):
                        pending[s["driver_id"]] = s.get("vehicle_plate") or ""
            if not pending:
                continue
            done = await db.inspections.find(
                {"deleted": {"$ne": True}, "created_at": {"$regex": f"^{today}"},
                 "driver_id": {"$in": list(pending)}},
                {"_id": 0, "driver_id": 1}).to_list(2000)
            for i in done:
                pending.pop(i.get("driver_id"), None)
            for did, plate in pending.items():
                asyncio.create_task(send_web_push_to_users(
                    [did], "📸 Inspección pendiente",
                    f"Recuerda hacer las fotos de la furgoneta {plate} antes de salir.",
                    url="/conductor"))
                total += 1
        except Exception as _oe:
            logger.debug(f"reminder org {org.get('id')}: {_oe}")
    set_current_org_db(None)
    if total:
        logger.info(f"Recordatorios de inspección push enviados: {total}")


@app.on_event("startup")
async def start_inspection_reminder_scheduler():
    """Recordatorio push a conductores a las 08:15 hora española: quien tiene
    furgoneta asignada hoy y no ha subido inspección, recibe aviso en el móvil."""
    async def _loop():
        from zoneinfo import ZoneInfo
        madrid = ZoneInfo("Europe/Madrid")
        while True:
            try:
                now = datetime.now(madrid)
                target = now.replace(hour=8, minute=15, second=0, microsecond=0)
                if target <= now:
                    target += timedelta(days=1)
                await asyncio.sleep((target - now).total_seconds())
                await _send_inspection_reminders()
                await asyncio.sleep(70)
            except Exception as e:
                logger.error(f"Reminder scheduler: {e}")
                await asyncio.sleep(600)
    asyncio.create_task(_loop())


@app.on_event("startup")
async def start_weekly_email_digest_scheduler():
    """Digest por email a los DSPs cada lunes a las 07:30 hora española."""
    async def _loop():
        from zoneinfo import ZoneInfo
        madrid = ZoneInfo("Europe/Madrid")
        while True:
            try:
                now = datetime.now(madrid)
                days_ahead = (0 - now.weekday()) % 7  # 0 = lunes
                target = (now + timedelta(days=days_ahead)).replace(hour=7, minute=30, second=0, microsecond=0)
                if target <= now:
                    target = target + timedelta(days=7)
                await asyncio.sleep((target - now).total_seconds())
                await send_weekly_email_digest()
                await asyncio.sleep(70)
            except Exception as e:
                logger.error(f"Digest email scheduler: {e}")
                await asyncio.sleep(600)
    asyncio.create_task(_loop())


@app.on_event("startup")
async def start_daily_summary_scheduler():
    """Programa el resumen diario de inspecciones a las 22:00 hora española."""
    async def _scheduler():
        from zoneinfo import ZoneInfo
        madrid = ZoneInfo("Europe/Madrid")
        while True:
            try:
                now = datetime.now(madrid)
                target = now.replace(hour=22, minute=0, second=0, microsecond=0)
                if now >= target:
                    target = target + timedelta(days=1)
                wait_s = (target - now).total_seconds()
                logger.info(f"Resumen diario programado para {target.isoformat()} (en {wait_s/3600:.1f}h)")
                await asyncio.sleep(wait_s)
                await send_daily_inspection_summary()
                # Pequeño margen para no disparar dos veces en el mismo minuto
                await asyncio.sleep(70)
            except Exception as e:
                logger.error(f"Scheduler resumen diario: {e}")
                await asyncio.sleep(300)
    asyncio.create_task(_scheduler())


@api_router.post("/telegram/send-daily-summary")
async def trigger_daily_summary(_=Depends(require_admin)):
    """Dispara manualmente el resumen diario (para probar sin esperar a las 22:00)."""
    await send_daily_inspection_summary()
    return {"success": True, "message": "Resumen enviado (si hay cuadrante y Telegram configurado)"}


# =========================
# INCIDENCIAS
# =========================

class Incident(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    vehicle_id: str
    driver_id: Optional[str] = None
    created_by_name: Optional[str] = None   # nombre del admin/conductor que abrió la incidencia
    title: str = ""
    description: str
    severity: str = "leve"
    status: str = "open"
    photos: List[str] = []
    notes: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    resolved_at: Optional[datetime] = None


class IncidentCreate(BaseModel):
    vehicle_id: str
    driver_id: Optional[str] = None
    title: str = ""
    description: str
    severity: str = "leve"
    photos: List[str] = []
    notes: str = ""


async def _auto_incident_on_workshop(vehicle_id: str, prev_status, new_status):
    """Al marcar un vehículo 'en taller', crea (o reabre) su incidencia automáticamente
    para que siempre quede en el histórico."""
    if new_status != "taller" or prev_status == "taller":
        return
    v = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0, "license_plate": 1})
    plate = (v or {}).get("license_plate", vehicle_id)
    titulo = f"Vehículo en taller — {plate}"
    # ¿Ya hay una incidencia de taller para este vehículo? → reabrirla en vez de duplicar
    existing = await db.incidents.find_one({"vehicle_id": vehicle_id, "title": titulo})
    now_iso = datetime.now(timezone.utc).isoformat()
    if existing:
        if existing.get("status") in ("resolved", "closed"):
            await db.incidents.update_one(
                {"id": existing["id"]},
                {"$set": {"status": "open", "resolved_at": None, "reopened_at": now_iso},
                 "$push": {"history": {"date": now_iso, "event": "Reabierta: vuelve a entrar en taller"}}}
            )
            logger.info(f"Incidencia de taller reabierta para {plate}")
        return
    inc = Incident(
        vehicle_id=vehicle_id,
        title=titulo,
        description=f"Entrada en taller registrada automáticamente el {now_iso[:10]}.",
        severity="media",
        status="open",
    )
    doc = serialize_doc(inc.model_dump())
    doc["auto_created"] = True
    await db.incidents.insert_one(doc)
    logger.info(f"Incidencia de taller creada automáticamente para {plate}")


@api_router.get("/incidents")
async def get_incidents(vehicle_id: Optional[str] = None, user: dict = Depends(require_any_auth)):
    query = {}
    if vehicle_id:
        query["vehicle_id"] = vehicle_id
    if user["role"] == "driver":
        query["driver_id"] = user["sub"]
    incidents = await db.incidents.find(query, {"_id": 0}).sort("created_at", -1).to_list(200)
    return incidents


@api_router.post("/incidents")
async def create_incident(data: IncidentCreate, user: dict = Depends(require_any_auth)):
    incident = Incident(**data.model_dump())
    if user["role"] == "driver":
        incident.driver_id = user["sub"]
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        assigned_today = await db.daily_assignments.find_one({
            "date": today,
            "slots": {"$elemMatch": {"driver_id": user["sub"], "vehicle_id": incident.vehicle_id}}
        }, {"_id": 0, "id": 1})
        inspected_today = None
        if not assigned_today:
            inspected_today = await db.inspections.find_one({
                "vehicle_id": incident.vehicle_id,
                "driver_id": user["sub"],
                "created_at": {"$regex": f"^{today}"}
            }, {"_id": 0, "id": 1})
        if not assigned_today and not inspected_today:
            raise HTTPException(status_code=403, detail="Solo puedes crear incidencias en tu vehículo asignado")
    incident.created_by_name = user.get("name") or user.get("username") or "Admin"
    doc = serialize_doc(incident.model_dump())
    await db.incidents.insert_one(doc)
    logger.info(f"Incidencia creada: {incident.id} — vehículo {incident.vehicle_id}")
    # Push a los coordinadores del centro de ese vehículo
    try:
        veh = await db.vehicles.find_one({"id": incident.vehicle_id}, {"_id": 0, "center": 1, "license_plate": 1})
        _center = (veh or {}).get("center") or ""
        _plate = (veh or {}).get("license_plate") or ""
        await push_center_event(
            _center, f"⚠️ Incidencia · {_center}".strip(),
            f"{_plate}: {incident.title or incident.description[:60]}",
            url="/panel/incidencias", exclude_id=user.get("sub"))
    except Exception as _pe:
        logger.debug(f"push incidencia: {_pe}")
    return serialize_doc(incident.model_dump())


@api_router.put("/incidents/{incident_id}/reopen")
async def reopen_incident(incident_id: str, _=Depends(require_admin)):
    """Reabre una incidencia cerrada (sin necesidad de crear otra)."""
    result = await db.incidents.update_one(
        {"id": incident_id},
        {"$set": {"status": "open", "resolved_at": None,
                  "reopened_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Incidencia no encontrada")
    return {"success": True}


@api_router.put("/incidents/{incident_id}/resolve")
async def resolve_incident(incident_id: str, _=Depends(require_admin)):
    result = await db.incidents.update_one(
        {"id": incident_id},
        {"$set": {"status": "resolved", "resolved_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Incidencia no encontrada")
    return {"success": True}


@api_router.patch("/incidents/{incident_id}")
async def update_incident(incident_id: str, request: Request, _=Depends(require_admin)):
    """Actualiza campos editables de una incidencia."""
    _ALLOWED = {"title", "description", "severity", "notes", "status"}
    data = await request.json()
    data = {k: v for k, v in data.items() if k in _ALLOWED}
    if not data:
        raise HTTPException(status_code=400, detail="Sin campos válidos")
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    if data.get("status") == "resolved" and "resolved_at" not in data:
        data["resolved_at"] = data["updated_at"]
    elif data.get("status") == "open":
        data["resolved_at"] = None
    result = await db.incidents.update_one({"id": incident_id}, {"$set": data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Incidencia no encontrada")
    return {"success": True}


@api_router.delete("/incidents/{incident_id}")
async def delete_incident(incident_id: str, _=Depends(require_admin)):
    """Elimina permanentemente una incidencia."""
    result = await db.incidents.delete_one({"id": incident_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Incidencia no encontrada")
    return {"success": True}


# =========================
# CONTACTOS — directorio interno de empleados
# =========================

class ContactCreate(BaseModel):
    name: str = Field(min_length=1, max_length=120)
    role: Optional[str] = ""          # cargo / rol
    phone: Optional[str] = ""
    email: Optional[str] = ""
    center: Optional[str] = ""
    notes: Optional[str] = ""

class ContactUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    center: Optional[str] = None
    notes: Optional[str] = None


@api_router.get("/contacts")
async def get_contacts(_=Depends(require_admin)):
    contacts = await db.contacts.find({}, {"_id": 0}).sort("name", 1).to_list(500)
    return contacts


@api_router.post("/contacts")
async def create_contact(data: ContactCreate, _=Depends(require_admin)):
    doc = {"id": str(uuid.uuid4()), "created_at": datetime.now(timezone.utc).isoformat(), **data.model_dump()}
    await db.contacts.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api_router.patch("/contacts/{contact_id}")
async def update_contact(contact_id: str, data: ContactUpdate, _=Depends(require_admin)):
    updates = {k: v for k, v in data.model_dump().items() if v is not None}
    if not updates:
        raise HTTPException(status_code=400, detail="Sin cambios")
    result = await db.contacts.update_one({"id": contact_id}, {"$set": updates})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Contacto no encontrado")
    return {"success": True}


@api_router.delete("/contacts/{contact_id}")
async def delete_contact(contact_id: str, _=Depends(require_admin)):
    result = await db.contacts.delete_one({"id": contact_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Contacto no encontrado")
    return {"success": True}


# =========================
# TALLERES — CRUD + sugerencias por daño (v5.1)
# =========================

def _normalize_center_code(center_raw: Optional[str]) -> str:
    """Convierte 'AMZL OGA5 SANTIAGO XPT' → 'OGA5'. Devuelve '' si no reconocido."""
    if not center_raw or not isinstance(center_raw, str):
        return ""
    up = center_raw.upper()
    for code in ("OGA5", "DGA1", "DGA2"):
        if code in up:
            return code
    return ""


def _damage_category(damage: dict) -> str:
    """Infiere la categoría de trabajo necesaria a partir del daño.
    Devuelve: 'lunas' | 'mecanica' | 'chapa' (default)."""
    haystack = " ".join([
        str(damage.get("part") or ""),
        str(damage.get("description") or ""),
        str(damage.get("repair_suggestion") or ""),
    ]).lower()
    if any(k in haystack for k in ("luna", "cristal", "parabrisas", "vidrio")):
        return "lunas"
    if any(k in haystack for k in ("motor", "transmis", "embrague", "freno", "suspen", "dirección", "direccion")):
        return "mecanica"
    if any(k in haystack for k in ("neumát", "neumat", "rueda", "llanta")):
        return "neumaticos"
    return "chapa"


def _provider_matches(workshop: dict, provider: str) -> bool:
    """¿El taller tiene convenio con este proveedor?"""
    if not provider:
        return False
    convenios = workshop.get("convenios") or []
    if "*" in convenios:
        return True
    pup = provider.upper()
    for c in convenios:
        cup = str(c).upper()
        if cup and (cup in pup or pup in cup):
            return True
    return False


@api_router.get("/workshops", response_model=List[Workshop])
async def list_workshops(
    center: Optional[str] = None,
    category: Optional[str] = None,
    provider: Optional[str] = None,
    only_active: bool = True,
    _=Depends(require_admin),
):
    query = {}
    if only_active:
        query["active"] = {"$ne": False}
    if center:
        query["center"] = center
    if category:
        query["categories"] = category
    workshops = await db.workshops.find(query, {"_id": 0}).to_list(500)
    # Filtro por proveedor en Python (más flexible que regex en arrays)
    if provider:
        workshops = [w for w in workshops if _provider_matches(w, provider)]
    for w in workshops:
        for k in ("created_at", "updated_at"):
            if isinstance(w.get(k), str):
                try:
                    w[k] = datetime.fromisoformat(w[k])
                except Exception:
                    pass
    return workshops


@api_router.get("/workshops/nearby")
async def workshops_nearby(
    lat: float,
    lng: float,
    provider: Optional[str] = None,
    max_km: float = 80,
    limit: int = 15,
    category: Optional[str] = None,
    current_user: dict = Depends(require_any_auth),
):
    """
    Devuelve talleres ordenados por distancia al punto (lat, lng).
    Accesible para cualquier usuario autenticado (no solo admin).
    Incluye info de asistencia en carretera del proveedor.
    """
    query = {"active": {"$ne": False}}
    if category:
        query["categories"] = category

    workshops = await db.workshops.find(query, {"_id": 0}).to_list(500)

    if provider:
        workshops = [w for w in workshops if _provider_matches(w, provider)]

    # Calcular distancia y filtrar
    results = []
    for w in workshops:
        if w.get("latitude") is None or w.get("longitude") is None:
            continue
        dist = _haversine_km((lat, lng), (w["latitude"], w["longitude"])) or 9999
        if dist <= max_km:
            w["distance_km"] = round(dist, 1)
            results.append(w)

    results.sort(key=lambda x: x["distance_km"])
    results = results[:limit]

    # Inyectar info de asistencia en carretera
    roadside = _provider_roadside(provider or "")
    return {
        "workshops": results,
        "roadside": roadside,
        "provider": provider,
        "user_lat": lat,
        "user_lng": lng,
    }


@api_router.post("/workshops", response_model=Workshop)
async def create_workshop(data: WorkshopCreate, _=Depends(require_admin)):
    w = Workshop(**data.model_dump())
    doc = serialize_doc(w.model_dump())
    await db.workshops.insert_one(doc)
    logger.info(f"Taller creado: {w.name} ({w.id})")
    return w


@api_router.get("/workshops/{workshop_id}", response_model=Workshop)
async def get_workshop(workshop_id: str, _=Depends(require_admin)):
    w = await db.workshops.find_one({"id": workshop_id}, {"_id": 0})
    if not w:
        raise HTTPException(status_code=404, detail="Taller no encontrado")
    for k in ("created_at", "updated_at"):
        if isinstance(w.get(k), str):
            try:
                w[k] = datetime.fromisoformat(w[k])
            except Exception:
                pass
    return w


@api_router.put("/workshops/{workshop_id}")
async def update_workshop(workshop_id: str, data: dict, _=Depends(require_admin)):
    data.pop("_id", None)
    data.pop("id", None)
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.workshops.update_one({"id": workshop_id}, {"$set": data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Taller no encontrado")
    return {"success": True}


@api_router.delete("/workshops/{workshop_id}")
async def delete_workshop(workshop_id: str, _=Depends(require_admin)):
    # Soft delete: marca inactive en lugar de borrar (preserva históricos)
    result = await db.workshops.update_one(
        {"id": workshop_id},
        {"$set": {"active": False, "updated_at": datetime.now(timezone.utc).isoformat()}},
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Taller no encontrado")
    return {"success": True}


# --- Actualización de un daño individual (coste real, taller asignado, estado) ---

@api_router.patch("/inspections/{inspection_id}/damages/{damage_index}")
async def update_damage(
    inspection_id: str,
    damage_index: int,
    data: dict,
    _=Depends(require_admin),
):
    """Actualiza UN daño dentro de una inspección por índice de array.
    Campos editables: actual_cost, workshop_id, repair_status, repair_notes."""
    insp = await db.inspections.find_one({"id": inspection_id})
    if not insp:
        raise HTTPException(status_code=404, detail="Inspección no encontrada")

    analysis = insp.get("analysis") or {}
    damages = analysis.get("damages") or []
    if damage_index < 0 or damage_index >= len(damages):
        raise HTTPException(status_code=404, detail="Daño no encontrado en esa inspección")

    allowed = {"actual_cost", "workshop_id", "repair_status", "repair_notes"}
    now_iso = datetime.now(timezone.utc).isoformat()

    current = damages[damage_index] if isinstance(damages[damage_index], dict) else {}
    for k, v in (data or {}).items():
        if k in allowed:
            current[k] = v

    # Marcas de tiempo automáticas
    if "workshop_id" in data and data.get("workshop_id"):
        current.setdefault("assigned_at", now_iso)
        if current.get("repair_status") in (None, "", "pending"):
            current["repair_status"] = "assigned"
    if data.get("repair_status") == "done":
        current["completed_at"] = now_iso
        # Reparado → el panel queda LIMPIO en el ledger: un golpe futuro ahí
        # vuelve a contar como daño nuevo (lo justo para scoring y €).
        try:
            _p = _canon_panel(current.get("part") or "")
            if _p:
                await db.vehicle_damage_ledger.update_many(
                    {"vehicle_id": insp.get("vehicle_id"), "panel": _p, "status": "open"},
                    {"$set": {"status": "repaired", "repaired_at": now_iso}})
                logger.info(f"[Ledger] panel '{_p}' de {insp.get('vehicle_id')} limpiado por reparación")
        except Exception as _le:
            logger.debug(f"ledger repair: {_le}")

    damages[damage_index] = current

    # Recalcular total de coste real (suma de actual_cost de los daños que lo tengan)
    actual_total = sum(
        float(d.get("actual_cost") or 0)
        for d in damages
        if isinstance(d, dict) and d.get("actual_cost") is not None
    )

    update = {"analysis.damages": damages}
    if actual_total > 0:
        update["analysis.total_actual_cost"] = round(actual_total, 2)

    await db.inspections.update_one({"id": inspection_id}, {"$set": update})

    return {
        "success": True,
        "damage": current,
        "total_actual_cost": round(actual_total, 2),
    }


@api_router.get("/inspections/{inspection_id}/damages/{damage_index}/suggested-workshops")
async def suggest_workshops_for_damage(
    inspection_id: str,
    damage_index: int,
    _=Depends(require_admin),
):
    """Devuelve talleres sugeridos para un daño concreto, filtrados ESTRICTAMENTE
    por el centro logístico de la furgoneta (no mezcla talleres de otros centros),
    y ordenados por score.

    Filtro previo: solo talleres del mismo centro que la furgoneta, MÁS los
    universales (convenios contiene "*", típico Carglass). El resto se descarta
    aunque tenga convenio con el proveedor — un taller en Vigo no le sirve a
    una furgoneta en Santiago.

    Lógica de score:
      - +100 si la furgoneta es Kinto Y el taller es oficial Toyota
      - +80  si el taller tiene convenio explícito con el proveedor de la furgoneta
      - +60  si el taller es 'universal' (convenios incluye '*', p.ej. Carglass)
      - +50  si la categoría del taller coincide con el tipo de daño (lunas/chapa/mecánica)
      - +5   por cada 0.1 sobre 4.0 en rating (max +50)
    """
    insp = await db.inspections.find_one({"id": inspection_id})
    if not insp:
        raise HTTPException(status_code=404, detail="Inspección no encontrada")

    analysis = insp.get("analysis") or {}
    damages = analysis.get("damages") or []
    if damage_index < 0 or damage_index >= len(damages):
        raise HTTPException(status_code=404, detail="Daño no encontrado en esa inspección")

    damage = damages[damage_index] if isinstance(damages[damage_index], dict) else {}
    category = _damage_category(damage)

    vehicle = await db.vehicles.find_one({"id": insp.get("vehicle_id", "")})
    provider = (vehicle or {}).get("provider", "") or ""
    center_code = _normalize_center_code((vehicle or {}).get("center"))
    is_kinto = "KINTO" in provider.upper()

    # Filtro ESTRICTO por centro: cada taller (incluso cadenas como Carglass) está
    # dado de alta con su centro — un taller de Vigo nunca sirve para Santiago.
    # Sin centro de furgoneta → no filtramos (mejor mostrar todo que nada).
    query = {"active": {"$ne": False}}
    if center_code:
        query["center"] = center_code

    workshops = await db.workshops.find(query, {"_id": 0}).to_list(500)
    # Respaldo: si el centro no tiene ningún taller, mostrar todos marcados como de otro centro
    other_center_fallback = False
    if center_code and not workshops:
        workshops = await db.workshops.find({"active": {"$ne": False}}, {"_id": 0}).to_list(500)
        other_center_fallback = True

    scored = []
    for w in workshops:
        score = 0
        reasons = []
        cats = w.get("categories") or []
        convenios = w.get("convenios") or []

        if is_kinto and ("oficial_toyota" in cats or w.get("is_official")):
            score += 100
            reasons.append("oficial Toyota (Kinto)")
        elif _provider_matches(w, provider):
            if "*" in convenios:
                score += 60
                reasons.append("universal (vía seguro)")
            else:
                score += 80
                reasons.append(f"convenio con {provider}")
        elif "*" in convenios:
            # Universal sin coincidencia explícita de proveedor: igual sirve (Carglass para lunas)
            score += 40
            reasons.append("trabaja con todos")

        if category in cats:
            score += 50
            reasons.append(f"especialista en {category}")
        elif category == "chapa" and "pintura" in cats:
            score += 30
            reasons.append("hace pintura")

        rating = w.get("rating")
        if isinstance(rating, (int, float)) and rating >= 4.0:
            bonus = min(50, int((rating - 4.0) * 50))
            score += bonus
            if bonus > 0:
                reasons.append(f"{rating}★")

        if score > 0:
            w_copy = dict(w)
            w_copy["_match_score"] = score
            w_copy["_match_reasons"] = reasons
            scored.append(w_copy)

    scored.sort(key=lambda x: -x.get("_match_score", 0))

    # Marcar talleres de otro centro cuando se usó el respaldo
    if other_center_fallback:
        for w in scored:
            w["_other_center"] = True
            w["_match_reasons"] = [f"⚠️ taller de {w.get('center', 'otro centro')}"] + (w.get("_match_reasons") or [])

    return {
        "damage_index": damage_index,
        "damage_category": category,
        "vehicle_provider": provider,
        "vehicle_center": center_code,
        "is_kinto": is_kinto,
        "other_center_fallback": other_center_fallback,
        "provider_network": _provider_network_for(provider),
        "workshops": scored[:15],
        "total_matched": len(scored),
    }


# =========================
# DELETE vehículos y conductores
# =========================

@api_router.delete("/vehicles/{vehicle_id}")
async def delete_vehicle(vehicle_id: str, _=Depends(require_admin)):
    result = await db.vehicles.update_one(
        {"id": vehicle_id},
        {"$set": {"status": "deleted", "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Vehículo no encontrado")
    return {"success": True}


@api_router.delete("/drivers/{driver_id}")
async def delete_driver(driver_id: str, _=Depends(require_admin)):
    result = await db.drivers.update_one(
        {"id": driver_id},
        {"$set": {"active": False}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Conductor no encontrado")
    return {"success": True}


@api_router.post("/drivers/{driver_id}/photo")
async def upload_driver_photo(driver_id: str, file: UploadFile = File(...), _=Depends(require_admin)):
    """Sube foto de perfil del conductor a R2 y guarda la URL en MongoDB."""
    driver = await db.drivers.find_one({"id": driver_id})
    if not driver:
        raise HTTPException(status_code=404, detail="Conductor no encontrado")
    content = await file.read()
    if len(content) > 8 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="La foto no puede superar 8 MB")
    try:
        photo_url, _ = await process_and_save_image(content, f"driver-{driver_id}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error procesando imagen: {e}")
    await db.drivers.update_one({"id": driver_id}, {"$set": {"photo_url": photo_url}})
    return {"success": True, "photo_url": photo_url}


# =========================
# PUT (alias de PATCH) para vehículos y conductores
# =========================

@api_router.put("/vehicles/{vehicle_id}")
async def put_vehicle(vehicle_id: str, data: dict, _=Depends(require_admin)):
    data.pop("_id", None)
    data.pop("id", None)
    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    prev = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0, "status": 1})
    result = await db.vehicles.update_one({"id": vehicle_id}, {"$set": data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Vehículo no encontrado")
    if "status" in data:
        try:
            await _auto_incident_on_workshop(vehicle_id, (prev or {}).get("status"), data.get("status"))
        except Exception as _ai:
            logger.warning(f"Auto-incidencia taller: {_ai}")
    return {"success": True}


@api_router.put("/drivers/{driver_id}")
async def put_driver(driver_id: str, data: dict, _=Depends(require_admin)):
    data.pop("_id", None)
    data.pop("id", None)
    result = await db.drivers.update_one({"id": driver_id}, {"$set": data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Conductor no encontrado")
    return {"success": True}


# =========================
# MARCAR TODAS LAS ALERTAS COMO LEÍDAS
# =========================

@api_router.put("/alerts/read-all")
async def mark_all_alerts_read(_=Depends(require_admin)):
    result = await db.alerts.update_many({}, {"$set": {"read": True}})
    return {"success": True, "updated": result.modified_count}


# =========================
# PDF DE INSPECCIONES
# =========================



# =========================
# IMPORTACIÓN MASIVA VEHÍCULOS
# =========================

def _load_workbook_reparado(content):
    """Repara un .xlsx con styles.xml corrupto (típico del Excel de Amazon Fleet)
    y lo recarga. Reemplaza la hoja de estilos por una mínima válida."""
    import io as _io, zipfile, openpyxl
    src = _io.BytesIO(content)
    out = _io.BytesIO()
    # Generamos MUCHOS estilos (300) porque las celdas del Excel de Amazon referencian
    # índices de estilo altos; si solo ponemos uno, falla con "list index out of range".
    N = 300
    xfs = '<xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>' * N
    styles_minimo = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<numFmts count="0"/>'
        '<fonts count="1"><font><sz val="11"/><name val="Calibri"/></font></fonts>'
        '<fills count="2"><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill></fills>'
        '<borders count="1"><border/></borders>'
        '<cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>'
        f'<cellXfs count="{N}">{xfs}</cellXfs>'
        '<cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>'
        '</styleSheet>'
    )
    with zipfile.ZipFile(src, "r") as zin:
        names = zin.namelist()
        with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
            for n in names:
                data = zin.read(n)
                if n == "xl/styles.xml":
                    data = styles_minimo.encode("utf-8")
                zout.writestr(n, data)
    out.seek(0)
    return openpyxl.load_workbook(out, data_only=False)


@api_router.post("/import/vehicles")
async def import_vehicles(
    file: UploadFile = File(...),
    center_filter: Optional[str] = Form(None),
    _=Depends(require_admin)
):
    content = await file.read()
    try:
        import openpyxl
        wb = None
        # Intento 1: carga normal con data_only
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        except Exception as e1:
            # Intento 2: sin data_only (evita leer la hoja de estilos calculada)
            try:
                wb = openpyxl.load_workbook(io.BytesIO(content), data_only=False)
            except Exception as e2:
                # Intento 3: reparar el XML de estilos corrupto del Excel de Amazon
                wb = _load_workbook_reparado(content)
        if wb is None:
            raise HTTPException(status_code=400, detail="No se pudo leer el Excel")
        # Elegir la hoja con MÁS filas (no siempre es la activa)
        hojas_validas = [w for w in wb.worksheets if (w.max_row or 0) > 0]
        if not hojas_validas:
            raise HTTPException(status_code=400, detail="El Excel no tiene hojas con datos")
        ws = max(hojas_validas, key=lambda w: w.max_row or 0)
        # Forzar recálculo de dimensiones por si vienen mal declaradas
        try:
            ws.reset_dimensions()
        except Exception:
            pass
        # Leer todas las filas como valores
        all_rows = list(ws.iter_rows(values_only=True))
        # Filtrar filas completamente vacías al principio
        all_rows = [r for r in all_rows if r is not None]
        if not all_rows or len(all_rows) < 1:
            raise HTTPException(status_code=400, detail="El Excel no tiene filas legibles")
        headers = [str(c).strip().lower() if c is not None else "" for c in all_rows[0]]
        if not any(headers):
            raise HTTPException(status_code=400, detail="No se detectaron cabeceras en el Excel")
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        logger.error(f"IMPORT-ERR lectura Excel: {e}\n{traceback.format_exc()}")
        raise HTTPException(status_code=400, detail=f"Error leyendo Excel: {e}")

    import unicodedata
    def _strip_accents(txt):
        return "".join(c for c in unicodedata.normalize("NFD", txt) if unicodedata.category(c) != "Mn")
    # Normalizar cabeceras: minúsculas y sin tildes
    headers_norm = [_strip_accents(h) for h in headers]

    def col(row, *names):
        for name in names:
            key = _strip_accents(name.lower())
            if key in headers_norm:
                idx = headers_norm.index(key)
                v = row[idx] if idx < len(row) else None
                return str(v).strip() if v is not None else ""
        return ""

    imported = 0
    updated = 0
    skipped = 0
    errors = []

    # Precargar todas las furgonetas indexadas por matrícula normalizada (sin espacios)
    _all_vehicles = await db.vehicles.find({}).to_list(10000)
    _plate_map = {}
    for _v in _all_vehicles:
        _key = (_v.get("license_plate","") or "").replace(" ","").upper()
        if _key:
            _plate_map[_key] = _v

    _diag_count = 0
    _diag_examples = []
    for row in all_rows[1:]:
        if not any(c is not None and str(c).strip() for c in row):
            continue
        plate = col(row, "matricula", "license_plate", "matrícula").upper()
        if not plate:
            skipped += 1
            continue
        try:
            center = col(row, "centro", "center")
            if center_filter and center != center_filter:
                skipped += 1
                continue
            # Buscar usando el mapa normalizado precargado
            plate_nospace = plate.replace(" ", "").upper()
            existing = _plate_map.get(plate_nospace)
            if len(_diag_examples) < 8:
                _diag_examples.append({"excel_plate": plate, "normalizada": plate_nospace, "encontrada_en_bd": bool(existing)})
            if existing:
                plate = existing.get("license_plate", plate)

            # Construir solo los campos que vienen con valor en el Excel
            campos = {}
            field_map = [
                (["marca","brand","vehículo","vehiculo"], "brand"),
                (["modelo","model"], "model"),
                (["color"], "color"),
                (["centro","center"], "center"),
                (["proveedor","provider"], "provider"),
                (["tipo","vehicle_type"], "vehicle_type"),
            ]
            for names, field in field_map:
                val = col(row, *names)
                if val:
                    campos[field] = val
            km_val = col(row, "kilómetros", "kilometros", "km", "mileage")
            if km_val:
                try:
                    campos["mileage"] = int(float(str(km_val).replace(".","").replace(",","").replace(" ","")))
                except Exception:
                    pass
            # Fechas de ITV y renting (formato Excel DD/MM/YYYY -> guardamos tal cual y normalizada ISO)
            def _parse_fecha(val):
                if not val:
                    return None
                txt = str(val).strip()[:10]
                for sep in ["/", "-"]:
                    if sep in txt:
                        parts = txt.split(sep)
                        if len(parts) == 3:
                            try:
                                if len(parts[0]) == 4:  # YYYY-MM-DD
                                    y, m, d = parts
                                else:  # DD/MM/YYYY
                                    d, m, y = parts
                                return f"{int(y):04d}-{int(m):02d}-{int(d):02d}"
                            except Exception:
                                return None
                return None
            itv = _parse_fecha(col(row, "i.t.v.", "itv", "fecha itv"))
            if itv:
                campos["itv_date"] = itv
            vto_renting = _parse_fecha(col(row, "vto. cont.renting.", "vto cont renting", "vto renting", "vencimiento renting"))
            if vto_renting:
                campos["renting_end_date"] = vto_renting
            baja_renting = _parse_fecha(col(row, "fecha baja renting", "baja renting"))
            if baja_renting:
                campos["renting_baja_date"] = baja_renting
            yr = col(row, "año", "year")
            if yr:
                try: campos["year"] = int(yr)
                except Exception: pass

            if existing:
                if campos:
                    campos["updated_at"] = datetime.now(timezone.utc)
                    await db.vehicles.update_one({"license_plate": plate}, {"$set": campos})
                    updated += 1
                else:
                    skipped += 1
                continue

            # Furgoneta NO está en tu flota: la omitimos (no creamos las 600 ajenas)
            skipped += 1
        except Exception as row_err:
            errors.append(f"{plate}: {str(row_err)[:80]}")
            skipped += 1
            continue

    return {
        "imported": imported,
        "updated": updated,
        "skipped": skipped,
        "errors": errors[:10],
        "diag_total_filas": len(all_rows),
        "diag_furgonetas_bd": len(_plate_map),
        "diag_ejemplos_excel": _diag_examples,
        "message": f"{imported} creados, {updated} actualizados, {skipped} omitidos"
    }


# =========================
# BULK VIN y FIX CENTERS
# =========================

@api_router.post("/vehicles/fix-centers")
async def fix_centers(_=Depends(require_admin)):
    """Normaliza los nombres de centros."""
    center_map = {"OGA5": "OGA5", "DGA1": "DGA1", "DGA2": "DGA2"}
    updated = 0
    async for v in db.vehicles.find({}, {"_id": 0, "id": 1, "center": 1}):
        c = v.get("center", "")
        normalized = center_map.get(c.upper().strip(), c)
        if normalized != c:
            await db.vehicles.update_one({"id": v["id"]}, {"$set": {"center": normalized}})
            updated += 1
    return {"success": True, "updated": updated}


# =========================
# REPROCESAR INSPECCIONES FALLIDAS
# =========================

@api_router.post("/inspections/reanalyze-failed")
async def reanalyze_all_failed(_=Depends(require_admin)):
    """Reencola TODAS las inspecciones cuyo analisis fallo: resetea sus contadores
    de reintentos para que el ciclo de auto-recuperacion (cada 10 min, 5 por ciclo)
    las vaya reanalizando solo. Util tras recuperar cuota/creditos de Gemini."""
    failed = await db.inspections.find(
        {"deleted": {"$ne": True}, "analysis_status": {"$ne": "ok"},
         "photos": {"$exists": True, "$ne": []}}, {"_id": 0, "id": 1}
    ).to_list(500)
    ids = [f["id"] for f in failed]
    if ids:
        await db.inspections.update_many(
            {"id": {"$in": ids}}, {"$set": {"auto_retries": 0}})

    return {
        "found": len(ids),
        "message": f"{len(ids)} inspecciones reencoladas. El reproceso automatico "
                   f"las ira analizando (5 cada 10 min, ~{max(1, len(ids) // 30)}h).",
        "inspection_ids": ids,
    }


# =========================
# STATS DASHBOARD
# =========================

# ── Valoración de daños tipo taller: por PANEL, no por rayón ──
# La IA etiqueta la zona con texto libre (cientos de variantes). La
# normalizamos a ~16 paneles de carrocería y aplicamos un baremo por
# panel × gravedad (mano de obra + pintura incluidas, mercado ES 2026).
# Un panel se paga UNA vez: pintar una puerta cubre todos sus rayones.
_PANEL_BAREMO = {
    "paragolpes":  {"leve": 90,  "moderado": 200, "grave": 350, "critico": 500},
    "puerta":      {"leve": 150, "moderado": 300, "grave": 500, "critico": 850},
    "porton":      {"leve": 170, "moderado": 340, "grave": 580, "critico": 1000},
    "aleta":       {"leve": 140, "moderado": 280, "grave": 480, "critico": 750},
    "lateral":     {"leve": 180, "moderado": 380, "grave": 650, "critico": 1300},
    "faldon":      {"leve": 90,  "moderado": 180, "grave": 320, "critico": 500},
    "paso_rueda":  {"leve": 60,  "moderado": 130, "grave": 220, "critico": 350},
    "capo":        {"leve": 150, "moderado": 320, "grave": 550, "critico": 900},
    "techo":       {"leve": 200, "moderado": 420, "grave": 700, "critico": 1400},
    "retrovisor":  {"leve": 60,  "moderado": 120, "grave": 180, "critico": 220},
    "optica":      {"leve": 90,  "moderado": 160, "grave": 240, "critico": 320},
    "parabrisas":  {"leve": 80,  "moderado": 250, "grave": 350, "critico": 450},
    "rueda":       {"leve": 40,  "moderado": 90,  "grave": 150, "critico": 220},
    "rejilla":     {"leve": 50,  "moderado": 110, "grave": 180, "critico": 280},
    "moldura":     {"leve": 40,  "moderado": 90,  "grave": 140, "critico": 200},
    "menor":       {"leve": 40,  "moderado": 80,  "grave": 140, "critico": 220},
    "otros":       {"leve": 100, "moderado": 250, "grave": 450, "critico": 800},
}

_SEV_RANK = {"leve": 1, "moderado": 2, "grave": 3, "critico": 4}


def _norm_sev(sev):
    s = (sev or "").strip().lower()
    if "crit" in s or "críti" in s:
        return "critico"
    if "grav" in s:
        return "grave"
    if "moder" in s:
        return "moderado"
    return "leve"


def _canon_panel(part):
    """Normaliza la zona de texto libre a un panel de carrocería. None = no
    cuenta para reparación (suciedad, mecánica, interior — no es chapa)."""
    s = (part or "").lower()

    def has(*ws):
        return any(w in s for w in ws)

    # no es carrocería / no fiable desde foto → no se valora
    if has("sucied", "limpieza"):
        return None
    if has("motor", "freno", "tpms", "fluido", "electrón", "electron", "cuadro",
           "salpicadero", "instrument", "interior", "mantenimiento", "neumát", "neumat"):
        return None
    if has("parabrisas", "luna"):
        return "parabrisas"
    if has("retrovisor", "espejo"):
        return "retrovisor"
    if has("faro", "piloto", "óptica", "optica", "intermitente", "luz trasera", "luz delantera"):
        return "optica"
    if has("llanta", "rueda", "tapacubo"):
        return "rueda"
    if has("rejilla", "calandra", "parrilla"):
        return "rejilla"
    if has("paragolpes", "parachoques"):
        return "paragolpes"
    if has("portón", "porton"):
        return "porton"
    if has("puerta"):
        return "puerta"
    if has("aleta", "guardabarros"):
        return "aleta"
    if has("faldón", "faldon", "umbral", "estrib", "moldura inferior", "panel inferior",
           "panel basculante", "rocker"):
        return "faldon"
    if has("paso de rueda"):
        return "paso_rueda"
    if has("capó", "capo"):
        return "capo"
    if has("techo"):
        return "techo"
    if has("moldura", "embellecedor"):
        return "moldura"
    if has("panel lateral", "lateral", "panel trasero", "pilar"):
        return "lateral"
    if has("matrícula", "matricula", "tapa de combustible", "tapa del depós", "depósito",
           "maneta", "cerradura", "mecanismo"):
        return "menor"
    if has("general", "completa", "completo", "exterior", "carrocería", "carroceria", "paneles"):
        return None  # descripciones globales vagas: no inventar coste
    return "otros"


def _vehicle_panel_cost(new_damages):
    """Coste de reparación de un vehículo agrupando por panel (peor gravedad
    de cada panel; usa actual_cost si el admin lo metió)."""
    panels = {}  # panel -> {"rank":, "sev":, "actual":}
    for nd in (new_damages or []):
        if not isinstance(nd, dict):
            continue
        panel = _canon_panel(nd.get("part") or nd.get("zone") or nd.get("location"))
        if not panel:
            continue
        sev = _norm_sev(nd.get("severity"))
        rank = _SEV_RANK[sev]
        cur = panels.get(panel)
        actual = nd.get("actual_cost") or 0
        if cur is None:
            panels[panel] = {"rank": rank, "sev": sev, "actual": actual}
        else:
            if rank > cur["rank"]:
                cur["rank"], cur["sev"] = rank, sev
            cur["actual"] = max(cur["actual"], actual)
    total = 0
    for panel, p in panels.items():
        if p["actual"] > 0:
            total += p["actual"]
        else:
            total += _PANEL_BAREMO.get(panel, _PANEL_BAREMO["otros"]).get(p["sev"], 0)
    return total, panels


async def _known_damaged_panels(vehicle_id, before_iso=None, exclude_id=None):
    """Paneles que YA estaban dañados en inspecciones anteriores de la furgo.
    Un panel ya conocido no puede ser 'daño nuevo' otra vez."""
    q = {"vehicle_id": vehicle_id, "deleted": {"$ne": True}, "analysis_status": "ok"}
    if before_iso:
        q["created_at"] = {"$lt": before_iso}
    if exclude_id:
        q["id"] = {"$ne": exclude_id}
    panels = set()
    async for pi in db.inspections.find(q, {"_id": 0, "analysis.damages": 1}):
        for d in ((pi.get("analysis") or {}).get("damages") or []):
            if isinstance(d, dict):
                p = _canon_panel(d.get("part") or d.get("zone") or d.get("location"))
                if p:
                    panels.add(p)
    return panels


@api_router.post("/admin/backfill-new-damages")
async def backfill_new_damages(_=Depends(require_admin)):
    """Reprocesa el histórico: por cada furgo, en orden cronológico, deja como
    'daño nuevo' SOLO la primera vez que aparece cada panel. Idempotente."""
    vehicles = await db.inspections.distinct(
        "vehicle_id", {"deleted": {"$ne": True}, "analysis_status": "ok"})
    corregidas = 0
    for veh in vehicles:
        insps = await db.inspections.find(
            {"vehicle_id": veh, "deleted": {"$ne": True}, "analysis_status": "ok"},
            {"_id": 0, "id": 1, "created_at": 1, "analysis.damages": 1, "analysis.new_damages": 1}
        ).sort("created_at", 1).to_list(20000)
        known = set()
        for ins in insps:
            a = ins.get("analysis") or {}
            nds = a.get("new_damages") or []
            kept = [nd for nd in nds if isinstance(nd, dict)
                    and _canon_panel(nd.get("part") or nd.get("zone") or nd.get("location")) not in known]
            panel_total = float(_vehicle_panel_cost(a.get("damages") or [])[0])
            update = {"analysis.total_estimated_cost": panel_total}
            if len(kept) != len(nds):
                update["analysis.new_damages"] = kept
                update["analysis.new_damages_count"] = len(kept)
            await db.inspections.update_one({"id": ins["id"]}, {"$set": update})
            corregidas += 1
            for d in (a.get("damages") or []):
                if isinstance(d, dict):
                    p = _canon_panel(d.get("part") or d.get("zone") or d.get("location"))
                    if p:
                        known.add(p)
    return {"success": True, "vehiculos": len(vehicles), "inspecciones_corregidas": corregidas}


@api_router.get("/stats/attention")
async def stats_attention(_=Depends(require_admin)):
    """Panel 'qué necesita mi atención HOY': pendientes de revisar, incidentes
    abiertos, inspecciones que faltan según cuadrante, análisis fallidos y
    € de daños del mes vs mes anterior."""
    now = datetime.now(timezone.utc)
    today = now.strftime("%Y-%m-%d")

    # 1. Pendientes de revisar
    pending_review = await db.inspections.count_documents({
        "reviewed": {"$ne": True}, "deleted": {"$ne": True}
    })

    # 2. Incidentes abiertos
    open_incidents = await db.incidents.count_documents({"status": {"$nin": ["closed", "resolved", "cerrada", "resuelta"]}})

    # 3. Cuadrante de hoy: faltan por inspeccionar
    assignments = await db.daily_assignments.find({"date": today}, {"_id": 0, "slots": 1}).to_list(20)
    slots = [s for a in assignments for s in a.get("slots", []) if s.get("vehicle_id")]
    inspected_vids = set()
    if slots:
        async for i in db.inspections.find(
            {"deleted": {"$ne": True}, "created_at": {"$regex": f"^{today}"}},
            {"_id": 0, "vehicle_id": 1}
        ):
            inspected_vids.add(i.get("vehicle_id"))
    missing_today = [
        {"plate": s.get("vehicle_plate", "—"), "driver": s.get("driver_name", "—")}
        for s in slots if s["vehicle_id"] not in inspected_vids
    ]

    # 4. Análisis fallidos (últimas 48h, recuperables)
    cutoff = (now - timedelta(hours=48)).isoformat()
    failed_analyses = await db.inspections.count_documents({
        "deleted": {"$ne": True},
        "analysis_status": {"$nin": ["ok", None]},
        "created_at": {"$gt": cutoff},
    })

    # 5. € de daños nuevos: mes actual vs anterior
    import calendar as _cal
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    prev_month_end = month_start - timedelta(seconds=1)
    prev_month_start = prev_month_end.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    async def _month_cost(start_iso, end_iso):
        # Coste realista tipo taller:
        #  1) SOLO inspecciones con foto de referencia (sin baseline no se sabe
        #     qué es nuevo; lo demás es "lo que ya estaba").
        #  2) Se juntan los daños nuevos del mes por vehículo y se valora por
        #     PANEL (peor gravedad de cada panel, pagado una vez).
        #  3) Usa el coste real (actual_cost) si lo metió el admin.
        veh_damages = {}
        async for i in db.inspections.find(
            {"deleted": {"$ne": True}, "analysis_status": "ok",
             "reference_photos": {"$exists": True, "$ne": []},
             "created_at": {"$gte": start_iso, "$lte": end_iso}},
            {"_id": 0, "vehicle_id": 1, "analysis.new_damages": 1}
        ):
            a = i.get("analysis") or {}
            veh = i.get("vehicle_id") or "?"
            veh_damages.setdefault(veh, []).extend(a.get("new_damages") or [])
        total = 0
        for veh, dmgs in veh_damages.items():
            cost, _ = _vehicle_panel_cost(dmgs)
            total += cost
        return round(total)

    cost_this_month = await _month_cost(month_start.isoformat(), now.isoformat())
    cost_prev_month = await _month_cost(prev_month_start.isoformat(), prev_month_end.isoformat())

    return {
        "pending_review": pending_review,
        "open_incidents": open_incidents,
        "assigned_today": len(slots),
        "missing_today": missing_today[:10],
        "missing_today_count": len(missing_today),
        "failed_analyses": failed_analyses,
        "cost_this_month": cost_this_month,
        "cost_prev_month": cost_prev_month,
    }


@api_router.get("/stats/damage-costs")
async def get_damage_costs(center: Optional[str] = None, _=Depends(require_admin)):
    """€ estimados de daños NUEVOS: mes actual vs anterior + top conductores.
    Justo: los daños marcados ✗ (falso positivo) en Revisión Rápida no cuentan,
    y los 'sugeridos' sin confirmación humana tampoco."""
    now = datetime.now(timezone.utc)
    month_start = datetime(now.year, now.month, 1, tzinfo=timezone.utc)
    prev_start = (month_start - timedelta(days=1)).replace(day=1)

    query = {"deleted": {"$ne": True}, "analysis_status": "ok",
             "created_at": {"$gte": prev_start.isoformat()}}
    if center and center != "Todos":
        query["center"] = center
    insps = await db.inspections.find(
        query, {"_id": 0, "id": 1, "driver_id": 1, "created_at": 1, "analysis.new_damages": 1}
    ).to_list(5000)

    # Falsos positivos marcados por humanos → excluidos del cómputo
    wrong = await db.ai_feedback.find(
        {"verdict": "wrong"}, {"_id": 0, "inspection_id": 1, "damage.part": 1}
    ).to_list(5000)
    wrong_parts: dict = {}
    for f in wrong:
        p = ((f.get("damage") or {}).get("part") or "").strip().lower()
        if p:
            wrong_parts.setdefault(f.get("inspection_id"), set()).add(p)

    cur = {"eur": 0.0, "count": 0}
    prev = {"eur": 0.0, "count": 0}
    by_driver: dict = {}
    m_iso = month_start.isoformat()
    for i in insps:
        nd = ((i.get("analysis") or {}).get("new_damages")) or []
        bad = wrong_parts.get(i.get("id"), set())
        eur, cnt = 0.0, 0
        for d in nd:
            if not isinstance(d, dict):
                continue
            if d.get("confirmed") is False:
                continue
            if (d.get("part") or "").strip().lower() in bad:
                continue
            eur += float(d.get("estimated_cost") or 0)
            cnt += 1
        if cnt == 0:
            continue
        is_current = i.get("created_at", "") >= m_iso
        bucket = cur if is_current else prev
        bucket["eur"] += eur
        bucket["count"] += cnt
        if is_current and i.get("driver_id"):
            e = by_driver.setdefault(i["driver_id"], {"eur": 0.0, "count": 0})
            e["eur"] += eur
            e["count"] += cnt

    top = sorted(by_driver.items(), key=lambda kv: -kv[1]["eur"])[:5]
    names = {}
    if top:
        ds = await db.drivers.find(
            {"id": {"$in": [k for k, _ in top]}}, {"_id": 0, "id": 1, "name": 1}
        ).to_list(10)
        names = {d["id"]: d.get("name", "?") for d in ds}

    return {
        "month_eur": round(cur["eur"]), "month_count": cur["count"],
        "prev_month_eur": round(prev["eur"]), "prev_month_count": prev["count"],
        "top_drivers": [
            {"driver_id": k, "name": names.get(k, "?"), "eur": round(v["eur"]), "count": v["count"]}
            for k, v in top
        ],
    }


@api_router.get("/stats/dashboard")
async def stats_dashboard(_=Depends(require_admin)):
    """Devuelve los contadores y metricas del dashboard principal."""
    total_vehicles = await db.vehicles.count_documents({"status": {"$ne": "deleted"}})
    vehicles_in_workshop = await db.vehicles.count_documents({"status": "taller"})
    total_drivers = await db.drivers.count_documents({"active": {"$ne": False}})
    total_inspections = await db.inspections.count_documents({})
    unread_alerts = await db.alerts.count_documents({"read": False})
    open_incidents = await db.incidents.count_documents({"status": "open"})

    # Desglose de severidad de inspecciones
    severity_counts = {"sin_danos": 0, "leve": 0, "moderado": 0, "grave": 0, "critico": 0}
    async for insp in db.inspections.find({"deleted": {"$ne": True}}, {"_id": 0, "analysis.severity": 1}):
        sev = (insp.get("analysis") or {}).get("severity", "sin_danos")
        if sev in severity_counts:
            severity_counts[sev] += 1

    # Actividad ultimos 7 dias (inspecciones por dia)
    from collections import defaultdict
    daily = defaultdict(lambda: {"inspecciones": 0, "danos": 0})
    cutoff = datetime.now(timezone.utc) - timedelta(days=7)
    async for insp in db.inspections.find({"deleted": {"$ne": True}}, {"_id": 0, "created_at": 1, "analysis.severity": 1}):
        ca = insp.get("created_at")
        if isinstance(ca, str):
            try:
                ca = datetime.fromisoformat(ca)
            except Exception:
                continue
        if not ca:
            continue
        if ca.tzinfo is None:
            ca = ca.replace(tzinfo=timezone.utc)
        if ca >= cutoff:
            key = ca.strftime("%Y-%m-%d")
            daily[key]["inspecciones"] += 1
            sev = (insp.get("analysis") or {}).get("severity", "sin_danos")
            if sev in ["grave", "critico", "moderado"]:
                daily[key]["danos"] += 1

    return {
        "total_vehicles": total_vehicles,
        "vehicles_in_workshop": vehicles_in_workshop,
        "total_drivers": total_drivers,
        "total_inspections": total_inspections,
        "unread_alerts": unread_alerts,
        "open_incidents": open_incidents,
        "severity_breakdown": severity_counts,
        "weekly_activity": dict(daily),
    }

# =========================
# ASIGNACION CONDUCTOR <-> VEHICULO
# =========================

class AssignDriverRequest(BaseModel):
    driver_id: Optional[str] = None  # None = desasignar


@api_router.put("/vehicles/{vehicle_id}/assign-driver")
async def assign_driver(vehicle_id: str, req: AssignDriverRequest, _=Depends(require_admin)):
    """Asigna (o desasigna) un conductor a un vehiculo. Registra el cambio en el historial."""
    vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehiculo no encontrado")

    prev = vehicle.get("current_driver_id")

    if req.driver_id:
        driver = await db.drivers.find_one({"id": req.driver_id}, {"_id": 0})
        if not driver:
            raise HTTPException(status_code=404, detail="Conductor no encontrado")

    await db.vehicles.update_one(
        {"id": vehicle_id},
        {"$set": {"current_driver_id": req.driver_id,
                  "updated_at": datetime.now(timezone.utc).isoformat()}}
    )

    # Registrar en historial de asignaciones (para deteccion de responsabilidad)
    await db.driver_assignments.insert_one({
        "id": str(uuid.uuid4()),
        "vehicle_id": vehicle_id,
        "driver_id": req.driver_id,
        "previous_driver_id": prev,
        "assigned_at": datetime.now(timezone.utc).isoformat(),
    })

    return {"success": True, "vehicle_id": vehicle_id, "driver_id": req.driver_id}


@api_router.get("/vehicles/{vehicle_id}/driver")
async def get_vehicle_driver(vehicle_id: str, _=Depends(require_admin)):
    """Devuelve el conductor asignado actualmente a un vehiculo."""
    vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0, "current_driver_id": 1})
    if not vehicle:
        raise HTTPException(status_code=404, detail="Vehiculo no encontrado")
    did = vehicle.get("current_driver_id")
    if not did:
        return {"driver": None}
    driver = await db.drivers.find_one({"id": did}, {"_id": 0})
    return {"driver": driver}


# =========================
# DETECCION DE RESPONSABILIDAD DE DANOS
# =========================

@api_router.get("/inspections/{inspection_id}/responsibility")
async def damage_responsibility(inspection_id: str, _=Depends(require_admin)):
    """Determina que conductor tenia asignado el vehiculo cuando se hizo esta inspeccion con danos."""
    insp = await db.inspections.find_one({"id": inspection_id}, {"_id": 0})
    if not insp:
        raise HTTPException(status_code=404, detail="Inspeccion no encontrada")

    vehicle_id = insp.get("vehicle_id")
    insp_date = insp.get("created_at")

    # Conductor asignado en el momento de la inspeccion
    driver_id = insp.get("driver_id")
    if not driver_id:
        v = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0, "current_driver_id": 1})
        driver_id = v.get("current_driver_id") if v else None

    driver = None
    if driver_id:
        driver = await db.drivers.find_one({"id": driver_id}, {"_id": 0})

    analysis = insp.get("analysis") or {}
    return {
        "inspection_id": inspection_id,
        "vehicle_id": vehicle_id,
        "inspection_date": insp_date,
        "driver": driver,
        "severity": analysis.get("severity", "sin_danos"),
        "damages_count": analysis.get("total_damages_count", 0),
        "note": "Conductor asignado en el momento de la inspeccion. Para danos NUEVOS, compara con la inspeccion anterior."
    }


# =========================
# SCORE DE CONDUCTOR
# =========================

@api_router.get("/drivers/{driver_id}/score")
async def driver_score(driver_id: str, _=Depends(require_admin)):
    """Calcula una puntuacion del conductor segun sus inspecciones y danos."""
    driver = await db.drivers.find_one({"id": driver_id}, {"_id": 0})
    if not driver:
        raise HTTPException(status_code=404, detail="Conductor no encontrado")

    inspections = await db.inspections.find({"deleted": {"$ne": True}, "driver_id": driver_id}, {"_id": 0}).to_list(500)
    total = len(inspections)
    con_danos = 0
    graves = 0
    for i in inspections:
        sev = (i.get("analysis") or {}).get("severity", "sin_danos")
        if sev not in ["sin_danos", "sin_analisis"]:
            con_danos += 1
        if sev in ["grave", "critico"]:
            graves += 1

    # Score base 100, penaliza danos graves
    score = 100
    if total > 0:
        score -= int((con_danos / total) * 30)
        score -= int((graves / total) * 40)
    score = max(0, min(100, score))

    nivel = "Excelente" if score >= 85 else "Bueno" if score >= 65 else "Regular" if score >= 40 else "Necesita atencion"

    return {
        "driver_id": driver_id,
        "driver_name": driver.get("name", ""),
        "total_inspections": total,
        "inspections_with_damage": con_danos,
        "severe_damages": graves,
        "score": score,
        "level": nivel,
    }


@api_router.get("/drivers/ranking")
async def drivers_ranking(_=Depends(require_admin)):
    """Ranking de conductores por puntuacion."""
    drivers = await db.drivers.find({"active": {"$ne": False}}, {"_id": 0}).to_list(200)
    ranking = []
    for d in drivers:
        inspections = await db.inspections.find({"deleted": {"$ne": True}, "driver_id": d["id"]}, {"_id": 0, "analysis.severity": 1}).to_list(500)
        total = len(inspections)
        con_danos = sum(1 for i in inspections if (i.get("analysis") or {}).get("severity") not in ["sin_danos", "sin_analisis", None])
        graves = sum(1 for i in inspections if (i.get("analysis") or {}).get("severity") in ["grave", "critico"])
        score = 100
        if total > 0:
            score -= int((con_danos / total) * 30) + int((graves / total) * 40)
        score = max(0, min(100, score))
        ranking.append({"driver_id": d["id"], "name": d.get("name", ""),
                        "center": d.get("center", ""), "score": score,
                        "total_inspections": total})
    ranking.sort(key=lambda x: -x["score"])
    return ranking


# =========================
# ASIGNACIÓN DIARIA DE FURGONETAS POR CENTRO
# =========================

@api_router.get("/assignments/daily")
async def get_daily_assignments(
    date: Optional[str] = None, center: Optional[str] = None, user: dict = Depends(require_admin)
):
    await _require_plan_feature(user, "assignments")
    """Obtiene las asignaciones del día para un centro."""
    if not date:
        date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    elif not re.match(r'^\d{4}-\d{2}-\d{2}$', date):
        raise HTTPException(400, "Formato de fecha inválido (YYYY-MM-DD)")
    query = {"date": date}
    if center and center != "Todos":
        query["center"] = center
    docs = await db.daily_assignments.find(query, {"_id": 0}).to_list(100)
    # Enrich: check if each assigned vehicle has an inspection today
    for doc in docs:
        for slot in doc.get("slots", []):
            vid = slot.get("vehicle_id")
            if vid:
                insp = await db.inspections.find_one(
                    {"deleted": {"$ne": True}, "vehicle_id": vid, "created_at": {"$regex": f"^{date}"}},
                    {"_id": 0, "id": 1, "analysis": 1}
                )
                slot["has_inspection"] = insp is not None
                if insp:
                    slot["inspection_severity"] = (insp.get("analysis") or {}).get("severity")
    return docs


@api_router.put("/assignments/daily")
async def upsert_daily_assignment(request: Request, user: dict = Depends(require_admin)):
    """Crea o actualiza la asignación diaria de un centro."""
    await _require_plan_feature(user, "assignments")
    try:
        data = await request.json()
    except Exception as e:
        logger.error(f"Cuadrante: body inválido: {e}")
        raise HTTPException(status_code=400, detail="Body JSON inválido")

    date = data.get("date") or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    center = data.get("center")
    if not center:
        raise HTTPException(status_code=400, detail="Centro requerido")

    # Limpia los slots: guarda solo los campos núcleo (descarta has_inspection,
    # inspection_severity y cualquier campo calculado que venga del GET).
    raw_slots = data.get("slots", []) or []
    slots = []
    for s in raw_slots:
        if not isinstance(s, dict):
            continue
        slots.append({
            "vehicle_id": s.get("vehicle_id", ""),
            "vehicle_plate": s.get("vehicle_plate", ""),
            "driver_id": s.get("driver_id", "") or "",
            "driver_name": s.get("driver_name", "") or "",
        })

    now = datetime.now(timezone.utc).isoformat()
    try:
        existing = await db.daily_assignments.find_one(
            {"date": date, "center": center}, {"_id": 0}
        )
        if existing:
            await db.daily_assignments.update_one(
                {"date": date, "center": center},
                {"$set": {"slots": slots, "updated_at": now}}
            )
            return {"success": True, "action": "updated", "id": existing.get("id")}
        else:
            doc_id = str(uuid.uuid4())
            doc = {
                "id": doc_id, "date": date, "center": center,
                "slots": slots, "created_at": now, "updated_at": now
            }
            await db.daily_assignments.insert_one(doc)
            return {"success": True, "action": "created", "id": doc_id}
    except Exception as e:
        logger.error(f"Cuadrante: error guardando en BD: {e}")
        raise HTTPException(status_code=500, detail=f"Error BD: {str(e)}")


def _norm_plate(p):
    """Normaliza matrícula: quita espacios/guiones y pasa a mayúsculas."""
    return "".join(c for c in (p or "").upper() if c.isalnum())


def _norm_name_words(name):
    """Extrae palabras del nombre normalizadas (sin acentos, minúsculas)."""
    import unicodedata
    s = unicodedata.normalize("NFD", (name or "").lower())
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.replace(",", " ")
    return set(w for w in s.split() if len(w) > 1)


@api_router.post("/assignments/import-text")
async def import_roster_text(
    data: dict,
    user: dict = Depends(require_admin),
):
    await _require_plan_feature(user, "assignments")
    """Importa el roster pegado desde la plataforma de Amazon, EN CUALQUIER FORMATO.

    Estrategia en cascada:
      1. Gemini estructura el texto crudo en pares matricula->conductor (entiende
         cualquier orden de columnas, saltos raros y formatos cambiantes).
      2. Si Gemini no responde: deteccion automatica de columnas (busca la columna
         que parece matricula y la que parece nombre, sin posiciones fijas).
      3. Matching contra la BD insensible a acentos y con desempate por apellidos.
    """
    raw_text = data.get("text", "")
    date = data.get("date") or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    center = data.get("center", "OGA5")

    if not raw_text.strip():
        raise HTTPException(status_code=400, detail="Texto vacío")
    raw_text = raw_text[:15000]  # límite de seguridad (textos más largos ralentizan a Gemini)

    PLATE_RE = re.compile(r'\b(\d{4})\s*-?\s*([A-Z]{3})\b', re.IGNORECASE)

    # 1) Intento con Gemini: entiende cualquier formato
    parsed = []
    parse_method = "gemini"
    try:
        from google import genai as genai_sdk
        from google.genai import types as genai_types

        use_vertex = os.environ.get("USE_VERTEX_AI", "").lower() in ("1", "true", "yes")
        if use_vertex:
            from google.oauth2 import service_account
            import json as _json
            import base64 as _b64
            sa_clean = os.environ.get("GCP_SERVICE_ACCOUNT_JSON", "").strip()
            if sa_clean and not sa_clean.startswith("{"):
                sa_clean = _b64.b64decode(sa_clean).decode("utf-8")
            credentials = service_account.Credentials.from_service_account_info(
                _json.loads(sa_clean), scopes=["https://www.googleapis.com/auth/cloud-platform"]
            ) if sa_clean else None
            client = genai_sdk.Client(
                vertexai=True, project=os.environ.get("GCP_PROJECT", ""),
                location=os.environ.get("GCP_LOCATION", "us-central1"), credentials=credentials
            )
        else:
            client = genai_sdk.Client(api_key=os.environ.get("GEMINI_API_KEY", ""))

        prompt = (
            "Este texto es un roster/cuadrante de reparto copiado de una plataforma "
            "(formato impredecible: columnas tabuladas, líneas sueltas, etc.). "
            "Extrae TODAS las parejas de matrícula española (4 dígitos + 3 letras) y "
            "nombre del conductor asignado a esa matrícula.\n"
            "Reglas:\n"
            "- Una pareja por matrícula. Si una matrícula no tiene conductor (vacío, 'SIN ASIGNAR', '-'), omítela.\n"
            "- El nombre puede venir como 'Apellidos, Nombre' o 'Nombre Apellidos' — devuélvelo tal cual aparece.\n"
            "- Ignora cabeceras, totales, horas, rutas y cualquier otra cosa.\n"
            "Responde SOLO este JSON sin markdown: "
            '{"pairs": [{"plate": "1234 ABC", "driver": "nombre tal cual"}]}\n\n'
            "TEXTO:\n" + raw_text
        )
        gen_config = genai_types.GenerateContentConfig(temperature=0.0, response_mime_type="application/json")
        loop = asyncio.get_running_loop()
        async with _gemini_sem:
            response = await asyncio.wait_for(
                loop.run_in_executor(
                    _executor,
                    lambda: client.models.generate_content(
                        model=os.environ.get("GEMINI_MODEL", "gemini-2.5-flash"),
                        contents=[prompt], config=gen_config
                    )
                ),
                timeout=20.0
            )
        gdata = json.loads(_strip_markdown_json(response.text or "{}"))
        for p in gdata.get("pairs", []):
            plate = (p.get("plate") or "").strip()
            drv = (p.get("driver") or "").strip()
            if PLATE_RE.search(plate.upper()) and drv and "sin asign" not in drv.lower():
                parsed.append({"plate": plate, "driver": drv})
    except Exception as _ge:
        logger.warning(f"Import roster: Gemini falló ({_ge}), usando detección automática")
        parsed = []

    # 2) Respaldo: detección automática de columnas (sin posiciones fijas)
    if not parsed:
        parse_method = "auto-columns"
        for line in raw_text.splitlines():
            cols = [c.strip() for c in re.split(r'\t|\s{3,}', line)]
            if len(cols) < 2:
                continue
            plate_val, plate_idx = None, None
            for idx, c in enumerate(cols):
                m = PLATE_RE.search(c.upper())
                if m:
                    plate_val = f"{m.group(1)} {m.group(2).upper()}"
                    plate_idx = idx
                    break
            if not plate_val:
                continue
            best_name, best_len = "", 0
            for idx, c in enumerate(cols):
                if idx == plate_idx:
                    continue
                letters = len(re.sub(r'[^A-Za-zÁÉÍÓÚÜÑáéíóúüñ ]', '', c))
                if letters >= 6 and letters > best_len and not PLATE_RE.search(c.upper()):
                    if re.search(r'\d{2}/\d{2}|AMZL|^ETT', c, re.IGNORECASE):
                        continue
                    best_name, best_len = c, letters
            if best_name and "sin asign" not in best_name.lower():
                parsed.append({"plate": plate_val, "driver": best_name})

    if not parsed:
        raise HTTPException(status_code=400, detail="No se encontraron parejas matrícula+conductor en el texto. Copia la tabla completa del roster.")

    # 3) Matching contra la BD: insensible a acentos, desempate por apellidos
    import unicodedata as _ud

    def _fold(s):
        s = _ud.normalize("NFD", (s or "").lower())
        return "".join(c for c in s if _ud.category(c) != "Mn")

    def _norm_plate_local(p):
        return re.sub(r'[^A-Z0-9]', '', (p or "").upper())

    def _norm_words(name):
        return set(w for w in re.sub(r'[^a-z ]', ' ', _fold(name)).split() if len(w) > 1)

    vehicles = await db.vehicles.find(
        {"center": {"$regex": re.escape(center[:4]), "$options": "i"}, "status": {"$ne": "deleted"}},
        {"_id": 0}
    ).to_list(500)
    drivers = await db.drivers.find({"center": {"$regex": re.escape(center[:4]), "$options": "i"}}, {"_id": 0}).to_list(500)
    if len(drivers) < 3:
        drivers = await db.drivers.find({}, {"_id": 0}).to_list(500)

    vehicle_map = {_norm_plate_local(v.get("license_plate", "")): v for v in vehicles}
    driver_words = [(_d, _norm_words(_d.get("name", ""))) for _d in drivers]

    slots = []
    matched = 0
    unmatched_plate = []
    unmatched_driver = []
    ambiguous = []

    for row in parsed:
        pnorm = _norm_plate_local(row["plate"])
        veh = vehicle_map.get(pnorm)
        if not veh:
            unmatched_plate.append(row["plate"])
            continue
        words = _norm_words(row["driver"])
        scored = []
        for d, dw in driver_words:
            inter = words & dw
            if not inter:
                continue
            score = sum(2 if len(w) >= 5 else 1 for w in inter)
            scored.append((score, d))
        scored.sort(key=lambda x: -x[0])
        best = scored[0] if scored else None
        second = scored[1] if len(scored) > 1 else None

        ok = bool(best) and (best[0] >= 3 or (best[0] >= 2 and (not second or second[0] < best[0])))
        if ok and second and second[0] == best[0]:
            ok = False
            ambiguous.append(row["driver"])

        if ok:
            slots.append({
                "vehicle_id": veh.get("id"), "vehicle_plate": veh.get("license_plate"),
                "driver_id": best[1].get("id"), "driver_name": best[1].get("name"),
            })
            matched += 1
        else:
            slots.append({
                "vehicle_id": veh.get("id"), "vehicle_plate": veh.get("license_plate"),
                "driver_id": "", "driver_name": "",
            })
            if row["driver"] not in ambiguous:
                unmatched_driver.append(row["driver"])

    plates_in_roster = {_norm_plate_local(s["vehicle_plate"]) for s in slots}
    for v in vehicles:
        if v.get("status") in ("inactive", "deleted"):
            continue
        if _norm_plate_local(v.get("license_plate", "")) not in plates_in_roster:
            slots.append({
                "vehicle_id": v.get("id"), "vehicle_plate": v.get("license_plate"),
                "driver_id": "", "driver_name": "",
            })

    logger.info(f"Import roster ({parse_method}): {matched}/{len(parsed)} emparejados, "
                f"{len(unmatched_plate)} matriculas desconocidas, {len(unmatched_driver)} sin match")
    return {
        "date": date, "center": center, "slots": slots,
        "matched": matched, "total_roster": len(parsed),
        "parse_method": parse_method,
        "unmatched_plate": unmatched_plate, "unmatched_driver": unmatched_driver,
        "ambiguous": ambiguous,
    }


@api_router.post("/assignments/import-image")
async def import_roster_image(
    file: UploadFile = File(...),
    date: str = Form(None),
    center: str = Form(...),
    user: dict = Depends(require_admin),
):
    """Lee una captura del roster de Amazon con Gemini Vision, extrae los pares
    matrícula↔conductor y los cruza con la BD para rellenar el cuadrante."""
    await _require_plan_feature(user, "assignments")
    if not date:
        date = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    img_bytes = await file.read()
    if not img_bytes:
        raise HTTPException(status_code=400, detail="Imagen vacía")

    # --- Gemini Vision: extraer roster ---
    gemini_key = os.environ.get("GEMINI_API_KEY", "")
    use_vertex = os.environ.get("USE_VERTEX_AI", "").lower() in ("1", "true", "yes")
    try:
        from google import genai as genai_sdk
        from google.genai import types as genai_types
        model_name = os.environ.get("GEMINI_MODEL", "gemini-2.5-flash")
        if use_vertex:
            vertex_project = os.environ.get("GCP_PROJECT", "")
            vertex_location = os.environ.get("GCP_LOCATION", "us-central1")
            sa_json = os.environ.get("GCP_SERVICE_ACCOUNT_JSON", "")
            if sa_json:
                from google.oauth2 import service_account
                import json as _json, base64 as _b64
                sa_clean = sa_json.strip()
                if not sa_clean.startswith("{"):
                    sa_clean = _b64.b64decode(sa_clean).decode("utf-8")
                creds_info = _json.loads(sa_clean)
                credentials = service_account.Credentials.from_service_account_info(
                    creds_info, scopes=["https://www.googleapis.com/auth/cloud-platform"]
                )
                client = genai_sdk.Client(vertexai=True, project=vertex_project,
                                          location=vertex_location, credentials=credentials)
            else:
                client = genai_sdk.Client(vertexai=True, project=vertex_project, location=vertex_location)
        else:
            client = genai_sdk.Client(api_key=gemini_key)

        prompt = (
            "Esta es una captura de un roster diario de reparto de Amazon. "
            "Extrae TODAS las filas que tengan una matrícula de furgoneta y un nombre de conductor. "
            "Las matrículas españolas son 4 números + 3 letras (ej: '9883 NFX'). "
            "Los nombres vienen como 'Apellidos, Nombre' (ej: 'Botana Vilar, Victor Manuel'). "
            "Ignora filas con 'SIN ASIGN' o sin matrícula. "
            "Devuelve SOLO un JSON array, sin texto adicional, con este formato exacto: "
            '[{"plate":"9883 NFX","driver":"Botana Vilar, Victor Manuel"}, ...]'
        )
        # Detectar tipo MIME real para aceptar JPEG y PNG
        _img_mime = "image/png" if img_bytes[:4] == b'\x89PNG' else "image/jpeg"
        contents = [prompt, genai_types.Part.from_bytes(data=img_bytes, mime_type=_img_mime)]
        gen_config = genai_types.GenerateContentConfig(temperature=0.1, response_mime_type="application/json")
        import json as _json2
        import time as _time
        loop = asyncio.get_running_loop()

        # Reintentos con fallback de modelo si hay 429 RESOURCE_EXHAUSTED
        fallback_models = [model_name, "gemini-flash-latest", "gemini-flash-lite-latest"]
        last_err = None
        response = None
        for _attempt_model in fallback_models:
            for _retry in range(3):
                try:
                    response = await asyncio.wait_for(
                        loop.run_in_executor(_executor, lambda m=_attempt_model: client.models.generate_content(
                            model=m, contents=contents, config=gen_config)),
                        timeout=90.0
                    )
                    last_err = None
                    break  # éxito
                except Exception as _e:
                    last_err = _e
                    err_str = str(_e).lower()
                    if "429" in err_str or "resource_exhausted" in err_str or "quota" in err_str:
                        logger.warning(f"Cuadrante import: 429 en {_attempt_model} intento {_retry+1}, reintentando...")
                        await asyncio.sleep(3 * (_retry + 1))  # 3s, 6s, 9s
                    else:
                        break  # error no recuperable, probar siguiente modelo
            if response is not None:
                break
            logger.warning(f"Cuadrante import: modelo {_attempt_model} agotado, probando siguiente...")

        if response is None:
            raise last_err or Exception("Todos los modelos fallaron")

        raw = response.text.strip()
        # quitar fences markdown si los hubiera
        if raw.startswith("```"):
            raw = raw.split("```")[1].replace("json", "", 1).strip() if "```" in raw else raw
        parsed = _json2.loads(raw)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Cuadrante import: error Gemini: {e}")
        raise HTTPException(status_code=500, detail=f"Error leyendo la imagen: {str(e)}")

    # --- Cruzar con vehículos y conductores de la BD ---
    # Usar regex igual que en el resto del API (coincidencia flexible de centros)
    vehicles = await db.vehicles.find(
        {"center": {"$regex": re.escape(center), "$options": "i"}, "status": {"$ne": "deleted"}},
        {"_id": 0}
    ).to_list(500)
    drivers = await db.drivers.find({"center": {"$regex": re.escape(center), "$options": "i"}}, {"_id": 0}).to_list(500)
    veh_by_plate = {_norm_plate(v.get("license_plate")): v for v in vehicles}
    driver_words = [(d, _norm_name_words(d.get("name"))) for d in drivers]

    slots = []
    matched, unmatched_plate, unmatched_driver = 0, [], []
    for row in parsed:
        plate = row.get("plate", "")
        dname = row.get("driver", "")
        np = _norm_plate(plate)
        veh = veh_by_plate.get(np)
        if not veh:
            unmatched_plate.append(plate)
            continue
        # match conductor por solapamiento de palabras
        words = _norm_name_words(dname)
        best, best_score = None, 0
        for d, dw in driver_words:
            score = len(words & dw)
            if score > best_score:
                best, best_score = d, score
        if best and best_score >= 2:
            slots.append({
                "vehicle_id": veh.get("id"), "vehicle_plate": veh.get("license_plate"),
                "driver_id": best.get("id"), "driver_name": best.get("name"),
            })
            matched += 1
        else:
            slots.append({
                "vehicle_id": veh.get("id"), "vehicle_plate": veh.get("license_plate"),
                "driver_id": "", "driver_name": "",
            })
            unmatched_driver.append(dname)

    # añadir furgonetas del centro que NO salían en el roster (en blanco)
    plates_in_roster = {_norm_plate(s["vehicle_plate"]) for s in slots}
    for v in vehicles:
        if v.get("status") in ("inactive", "deleted"):
            continue
        if _norm_plate(v.get("license_plate")) not in plates_in_roster:
            slots.append({
                "vehicle_id": v.get("id"), "vehicle_plate": v.get("license_plate"),
                "driver_id": "", "driver_name": "",
            })

    return {
        "date": date, "center": center, "slots": slots,
        "matched": matched, "total_roster": len(parsed),
        "unmatched_plate": unmatched_plate, "unmatched_driver": unmatched_driver,
    }


# =========================
# INFORME DE FLOTA (PDF para Amazon)
# =========================

@api_router.get("/reports/fleet-pdf")
async def fleet_report_pdf(_=Depends(require_admin)):
    """Genera un informe PDF del estado de la flota, listo para enviar al coordinador."""
    total_vehicles = await db.vehicles.count_documents({"status": {"$ne": "deleted"}})
    total_inspections = await db.inspections.count_documents({})
    alerts = await db.alerts.find({}, {"_id": 0}).sort("created_at", -1).to_list(100)

    # Desglose severidad
    sev_counts = {"leve": 0, "moderado": 0, "grave": 0, "critico": 0}
    total_cost = 0.0
    async for insp in db.inspections.find({"deleted": {"$ne": True}}, {"_id": 0, "analysis": 1}):
        a = insp.get("analysis") or {}
        sev = a.get("severity")
        if sev in sev_counts:
            sev_counts[sev] += 1
        total_cost += float(a.get("total_estimated_cost", 0) or 0)

    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story = []
    story.append(Paragraph("<b>FlotaDSP - Informe de Estado de Flota</b>", styles["Title"]))
    story.append(Paragraph(datetime.now(timezone.utc).strftime("Generado: %d/%m/%Y"), styles["Normal"]))
    story.append(Spacer(1, 0.6*cm))

    resumen = [
        ["Metrica", "Valor"],
        ["Vehiculos totales", str(total_vehicles)],
        ["Inspecciones realizadas", str(total_inspections)],
        ["Danos leves", str(sev_counts["leve"])],
        ["Danos moderados", str(sev_counts["moderado"])],
        ["Danos graves", str(sev_counts["grave"])],
        ["Danos criticos", str(sev_counts["critico"])],
        ["Coste estimado total", f"{total_cost:.2f} EUR"],
    ]
    t = Table(resumen, colWidths=[8*cm, 8*cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#EA580C")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 10),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.whitesmoke, colors.white]),
        ("GRID", (0,0), (-1,-1), 0.5, colors.lightgrey),
        ("PADDING", (0,0), (-1,-1), 8),
    ]))
    story.append(t)
    doc.build(story)
    buf.seek(0)

    from fastapi.responses import StreamingResponse
    fname = "informe_flota_" + datetime.now(timezone.utc).strftime("%Y%m%d") + ".pdf"
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f"attachment; filename={fname}"})


# =============================================================
# REGISTRO DE ROUTERS — AL FINAL, tras definir todos los endpoints
# =============================================================

# =========================
# SUBIDA MASIVA POR CARPETAS (matrícula = nombre de carpeta)
# =========================

async def _process_single_inspection(vehicle_id, driver_id, photo_urls, photos_base64):
    """Procesa una inspección: referencia + Gemini + guardar + alerta Telegram.
    Devuelve (analysis, status). Reutilizada por subida individual y masiva."""
    # Fotos de referencia de la última inspección OK
    ref_results = await db.inspections.find(
        {"deleted": {"$ne": True}, "vehicle_id": vehicle_id, "analysis": {"$ne": None}, "analysis_status": "ok"},
        {"_id": 0, "photos": 1}
    ).sort("created_at", -1).to_list(1)
    last_insp = ref_results[0] if ref_results else None
    ref_photo_urls = last_insp.get("photos", [])[:4] if last_insp else []
    ref_bytes_list = await load_reference_images(ref_photo_urls) if ref_photo_urls else []

    _known_txt = await _known_damages_prompt(vehicle_id)
    analysis, analysis_status, analysis_error = await analyze_images_with_gemini(
        photos_base64, ref_bytes_list if ref_bytes_list else None, db=db,
        known_damages_text=_known_txt,
    )
    if analysis_status == "ok" and analysis:
        await _apply_vehicle_memory(vehicle_id, analysis)

    inspection = Inspection(
        vehicle_id=vehicle_id, driver_id=driver_id, photos=photo_urls,
        reference_photos=(ref_photo_urls[:2] if ref_photo_urls else []),
        analysis=analysis, analysis_status=analysis_status, analysis_error=analysis_error,
        notes="Subida masiva", analyzed_at=datetime.now(timezone.utc)
    )
    await db.inspections.insert_one(serialize_doc(inspection.model_dump()))

    # YOLO en background también para subidas masivas
    asyncio.create_task(_run_yolo_for_inspection(inspection.id, photo_urls))

    if analysis_status == "ok" and (
        analysis.severity in ["grave", "critico"] or analysis.critical_damages_count > 0
    ):
        vehicle = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
        plate = vehicle.get("license_plate") if vehicle else vehicle_id
        alert = Alert(
            vehicle_id=vehicle_id, inspection_id=inspection.id,
            title=f"Daños detectados en {plate}",
            description=analysis.executive_summary,
            severity="critical" if analysis.severity == "critico" else "high"
        )
        await db.alerts.insert_one(serialize_doc(alert.model_dump()))
        driver_name = "Sin asignar"
        did = driver_id or (vehicle.get("current_driver_id") if vehicle else None)
        if did:
            drv = await db.drivers.find_one({"id": did}, {"_id": 0, "name": 1})
            if drv:
                driver_name = drv.get("name", "Sin asignar")
        await send_telegram_damage_alert(
            plate=plate, driver_name=driver_name, analysis=analysis,
            photo_urls=photo_urls, inspection_id=inspection.id
        )
    return analysis, analysis_status


@api_router.post("/inspections/batch-upload")
async def batch_upload_inspections(
    plates: List[str] = Form(...),
    files: List[UploadFile] = File(...),
    file_plates: List[str] = Form(...),
    _=Depends(require_admin)
):
    """Subida masiva. Cada archivo viene con su matrícula (file_plates[i] corresponde a files[i]).
    Agrupa por matrícula, resuelve el vehículo y crea un peritaje por furgoneta."""
    if len(files) != len(file_plates):
        raise HTTPException(status_code=400, detail="Desajuste entre archivos y matrículas")

    # Agrupar archivos por matrícula normalizada
    grupos = {}
    for f, p in zip(files, file_plates):
        key = _normalize_plate(p)
        grupos.setdefault(key, {"raw": p, "files": []})["files"].append(f)

    # Mapa de vehículos: matrícula normalizada -> vehículo
    vehiculos = await db.vehicles.find({}, {"_id": 0, "id": 1, "license_plate": 1, "current_driver_id": 1}).to_list(5000)
    mapa = {_normalize_plate(v.get("license_plate", "")): v for v in vehiculos}

    resultados = []
    for key, grupo in grupos.items():
        raw_plate = grupo["raw"]
        vehiculo = mapa.get(key)
        if not vehiculo:
            resultados.append({"plate": raw_plate, "ok": False,
                               "error": "No existe ninguna furgoneta con esa matrícula"})
            continue
        try:
            photo_urls, photos_base64 = [], []
            for file in grupo["files"][:20]:
                content = await file.read()
                if not content:
                    continue
                validate_image_content(content)
                photo_url, processed = await process_and_save_image(content, vehiculo["id"])
                photo_urls.append(photo_url)
                photos_base64.append(base64.b64encode(processed).decode("utf-8"))
            if not photo_urls:
                resultados.append({"plate": raw_plate, "ok": False, "error": "Sin fotos válidas"})
                continue
            analysis, status = await _process_single_inspection(
                vehiculo["id"], vehiculo.get("current_driver_id"), photo_urls, photos_base64
            )
            resultados.append({
                "plate": raw_plate, "ok": status == "ok",
                "severity": analysis.severity if status == "ok" else None,
                "damages": analysis.total_damages_count if status == "ok" else 0,
                "photos": len(photo_urls),
                "error": None if status == "ok" else status
            })
        except Exception as e:
            logger.warning(f"Batch: error con {raw_plate}: {e}")
            resultados.append({"plate": raw_plate, "ok": False, "error": str(e)[:200]})

    ok_count = sum(1 for r in resultados if r["ok"])
    return {"success": True, "total_grupos": len(grupos), "procesados_ok": ok_count,
            "resultados": resultados}



# =========================
# BOLSAS, ACEITE Y KILOMETRAJE
# =========================

@api_router.post("/vehicles/{vehicle_id}/bags/set")
async def set_bags(vehicle_id: str, data: dict, _=Depends(require_admin)):
    """Fija el stock de bolsas de una furgoneta."""
    cantidad = int(data.get("bags", 0))
    v = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not v:
        raise HTTPException(status_code=404, detail="Furgoneta no encontrada")
    entry = {"date": datetime.now(timezone.utc).isoformat(), "change": cantidad - v.get("bags_remaining", 0),
             "note": data.get("note", "Ajuste de stock"), "remaining_after": cantidad}
    await db.vehicles.update_one({"id": vehicle_id},
        {"$set": {"bags_remaining": cantidad, "updated_at": datetime.now(timezone.utc)},
         "$push": {"bags_history": entry}})
    return {"success": True, "bags_remaining": cantidad}


@api_router.post("/vehicles/{vehicle_id}/bags/consume")
async def consume_bags(vehicle_id: str, data: dict, _=Depends(require_admin)):
    """Descuenta bolsas gastadas y registra cuándo."""
    gastadas = int(data.get("used", 1))
    v = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not v:
        raise HTTPException(status_code=404, detail="Furgoneta no encontrada")
    restantes = max(0, v.get("bags_remaining", 0) - gastadas)
    entry = {"date": datetime.now(timezone.utc).isoformat(), "change": -gastadas,
             "note": data.get("note", f"Gastadas {gastadas} bolsas"), "remaining_after": restantes}
    await db.vehicles.update_one({"id": vehicle_id},
        {"$set": {"bags_remaining": restantes, "updated_at": datetime.now(timezone.utc)},
         "$push": {"bags_history": entry}})
    # Aviso si quedan pocas
    if restantes <= 5:
        await db.alerts.insert_one(serialize_doc(Alert(
            vehicle_id=vehicle_id, inspection_id="",
            title=f"Quedan {restantes} bolsas en {v.get('license_plate','')}",
            description=f"Stock bajo de bolsas. Repón pronto.",
            severity="medium").model_dump()))
    return {"success": True, "bags_remaining": restantes}


@api_router.post("/vehicles/{vehicle_id}/oil/change")
async def register_oil_change(vehicle_id: str, data: dict, _=Depends(require_admin)):
    """Registra un cambio de aceite: km actuales + fecha + intervalo."""
    km = int(data.get("km", 0))
    fecha = data.get("date") or datetime.now(timezone.utc).date().isoformat()
    intervalo = int(data.get("interval_km", 15000))
    aviso_antes = int(data.get("warning_before_km", 2500))
    v = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not v:
        raise HTTPException(status_code=404, detail="Furgoneta no encontrada")
    await db.vehicles.update_one({"id": vehicle_id}, {"$set": {
        "oil_last_change_km": km, "oil_last_change_date": fecha,
        "oil_interval_km": intervalo, "oil_warning_before_km": aviso_antes,
        "mileage": max(km, v.get("mileage") or 0), "updated_at": datetime.now(timezone.utc)}})
    return {"success": True, "next_change_km": km + intervalo,
            "warning_at_km": km + intervalo - aviso_antes}


# Mantenimientos por km adicionales (mismo patrón que el aceite)
_MAINT_KINDS = {
    "ruedas": {"label": "Cambio de ruedas", "default_interval": 40000, "default_warning": 3000},
    "pastillas": {"label": "Pastillas de freno", "default_interval": 30000, "default_warning": 3000},
}


@api_router.post("/vehicles/{vehicle_id}/maintenance/{kind}/change")
async def register_maintenance_change(vehicle_id: str, kind: str, data: dict, _=Depends(require_admin)):
    """Registra un mantenimiento por km (ruedas, pastillas): igual que el aceite."""
    spec = _MAINT_KINDS.get(kind)
    if not spec:
        raise HTTPException(status_code=400, detail=f"Tipo desconocido. Válidos: {list(_MAINT_KINDS)}")
    km = int(data.get("km", 0))
    fecha = data.get("date") or datetime.now(timezone.utc).date().isoformat()
    intervalo = int(data.get("interval_km", spec["default_interval"]))
    aviso_antes = int(data.get("warning_before_km", spec["default_warning"]))
    v = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not v:
        raise HTTPException(status_code=404, detail="Furgoneta no encontrada")
    await db.vehicles.update_one({"id": vehicle_id}, {"$set": {
        f"{kind}_last_change_km": km, f"{kind}_last_change_date": fecha,
        f"{kind}_interval_km": intervalo, f"{kind}_warning_before_km": aviso_antes,
        "mileage": max(km, v.get("mileage") or 0), "updated_at": datetime.now(timezone.utc)}})
    return {"success": True, "kind": kind, "label": spec["label"],
            "next_change_km": km + intervalo, "warning_at_km": km + intervalo - aviso_antes}


# NOTA: la subida de foto de conductor vive en POST /drivers/{driver_id}/photo
# (definida más arriba). Aquí había un duplicado inalcanzable que se eliminó.

# =========================
# STICKERS DE MERY — persistentes en servidor
# =========================

@api_router.get("/mery/stickers")
async def get_mery_stickers(_=Depends(require_admin)):
    doc = await db.app_meta.find_one({"_id": "mery_stickers"})
    return {"stickers": (doc or {}).get("data", {})}


@api_router.put("/mery/stickers")
async def set_mery_sticker(data: dict, _=Depends(require_admin)):
    key = (data.get("key") or "").strip()[:120].replace(".", "_").replace("$", "_")
    value = data.get("value")
    if not key:
        raise HTTPException(status_code=400, detail="key requerida")
    if value:
        await db.app_meta.update_one(
            {"_id": "mery_stickers"},
            {"$set": {f"data.{key}": str(value)[:120]}},
            upsert=True
        )
    else:
        await db.app_meta.update_one(
            {"_id": "mery_stickers"},
            {"$unset": {f"data.{key}": ""}}
        )
    return {"success": True}


@api_router.post("/inspections/validate-photo")
async def validate_inspection_photo(
    vehicle_id: str = Form(...),
    expected_zone: str = Form(...),
    file: UploadFile = File(...),
    user: dict = Depends(require_any_auth),
):
    """Valida UNA foto en el momento de capturarla: zona correcta, nitidez y matrícula.
    FAIL-OPEN: si la IA no responde, la foto se acepta (nunca bloquear por fallo técnico)."""
    ZONES = {
        "frontal": "el FRONTAL (morro, parabrisas, matrícula delantera)",
        "trasera": "la TRASERA (portón/puertas traseras, matrícula trasera)",
        "lateral_izq": "el LATERAL IZQUIERDO completo",
        "lateral_izquierdo": "el LATERAL IZQUIERDO completo",
        "lateral_der": "el LATERAL DERECHO completo",
        "lateral_derecho": "el LATERAL DERECHO completo",
    }
    zone_desc = ZONES.get(expected_zone.lower().strip())
    if not zone_desc:
        return {"valid": True, "checked": False, "reason": "zona desconocida — no validada"}

    v = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0, "license_plate": 1})
    real_plate = ((v or {}).get("license_plate") or "").replace(" ", "").replace("-", "").upper()

    try:
        content = await file.read()
        img = Image.open(io.BytesIO(content)).convert("RGB")
        img.thumbnail((1024, 1024))
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=75)
        img_bytes = buf.getvalue()

        from google import genai as genai_sdk
        from google.genai import types as genai_types

        use_vertex = os.environ.get("USE_VERTEX_AI", "").lower() in ("1", "true", "yes")
        if use_vertex:
            from google.oauth2 import service_account
            import json as _json
            import base64 as _b64
            sa_clean = os.environ.get("GCP_SERVICE_ACCOUNT_JSON", "").strip()
            if sa_clean and not sa_clean.startswith("{"):
                sa_clean = _b64.b64decode(sa_clean).decode("utf-8")
            credentials = service_account.Credentials.from_service_account_info(
                _json.loads(sa_clean), scopes=["https://www.googleapis.com/auth/cloud-platform"]
            ) if sa_clean else None
            client = genai_sdk.Client(
                vertexai=True, project=os.environ.get("GCP_PROJECT", ""),
                location=os.environ.get("GCP_LOCATION", "us-central1"), credentials=credentials
            )
        else:
            client = genai_sdk.Client(api_key=os.environ.get("GEMINI_API_KEY", ""))

        prompt = (
            f"Foto de inspección de una furgoneta de reparto. Debe mostrar {zone_desc} del vehículo.\n"
            "Evalúa y responde SOLO este JSON sin markdown:\n"
            '{"zone_ok": <true si la foto muestra esa zona del vehículo>, '
            '"zone_seen": "<frontal|trasera|lateral_izquierdo|lateral_derecho|otra|no_es_vehiculo>", '
            '"too_blurry": <true SOLO si está tan borrosa/oscura que no permite inspeccionar daños>, '
            '"plate": "<matrícula legible en la foto o vacío>", '
            '"is_vehicle": <true si se ve una furgoneta/vehículo>}\n'
            "Sé tolerante con ángulos imperfectos: zone_ok=true si la zona pedida es claramente la protagonista."
        )
        contents = [prompt, genai_types.Part.from_bytes(data=img_bytes, mime_type="image/jpeg")]
        gen_config = genai_types.GenerateContentConfig(temperature=0.0, response_mime_type="application/json")
        loop = asyncio.get_running_loop()
        async with _gemini_sem:
            response = await asyncio.wait_for(
                loop.run_in_executor(
                    _executor,
                    lambda: client.models.generate_content(
                        model=os.environ.get("GEMINI_MODEL", "gemini-2.5-flash"),
                        contents=contents, config=gen_config
                    )
                ),
                timeout=25.0
            )
        data = json.loads(_strip_markdown_json(response.text or "{}"))

        if not data.get("is_vehicle", True):
            return {"valid": False, "checked": True, "reason": "En la foto no se ve ninguna furgoneta. Haz la foto al vehículo."}
        if data.get("too_blurry"):
            return {"valid": False, "checked": True, "reason": "La foto está demasiado borrosa u oscura. Repítela con más luz y el móvil quieto."}
        if not data.get("zone_ok", True):
            seen = data.get("zone_seen", "otra zona")
            return {"valid": False, "checked": True,
                    "reason": f"Esta casilla es para {zone_desc.split('(')[0].strip()}, pero la foto muestra: {seen}. Haz la foto de la zona correcta."}
        # Matrícula: solo exigible en frontal/trasera y solo si se lee alguna
        detected = (data.get("plate") or "").replace(" ", "").replace("-", "").upper()
        if detected and real_plate and detected != real_plate and expected_zone.lower() in ("frontal", "trasera"):
            return {"valid": False, "checked": True,
                    "reason": f"La matrícula de la foto ({data.get('plate')}) NO es la del vehículo seleccionado ({(v or {}).get('license_plate','')}). ¿Estás fotografiando la furgoneta correcta?"}
        return {"valid": True, "checked": True, "detected_plate": data.get("plate", "")}
    except Exception as e:
        logger.warning(f"validate-photo fail-open ({expected_zone}): {e}")
        return {"valid": True, "checked": False, "reason": "validación no disponible — foto aceptada"}


@api_router.post("/vehicles/{vehicle_id}/odometer-photo")
async def read_odometer_photo(vehicle_id: str, file: UploadFile = File(...), user: dict = Depends(require_any_auth)):
    """Lee los km del cuentakilómetros con Gemini a partir de una foto del salpicadero.
    Devuelve el número leído. NO actualiza el kilometraje — el cliente confirma después."""
    v = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0, "license_plate": 1, "mileage": 1})
    if not v:
        raise HTTPException(status_code=404, detail="Furgoneta no encontrada")
    content = await file.read()
    if not content or len(content) > 15 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Imagen inválida o demasiado grande")
    try:
        # Comprimir para acelerar
        img = Image.open(io.BytesIO(content)).convert("RGB")
        img.thumbnail((1280, 1280))
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=80)
        img_bytes = buf.getvalue()

        from google import genai as genai_sdk
        from google.genai import types as genai_types

        use_vertex = os.environ.get("USE_VERTEX_AI", "").lower() in ("1", "true", "yes")
        if use_vertex:
            from google.oauth2 import service_account
            import json as _json
            import base64 as _b64
            sa_clean = os.environ.get("GCP_SERVICE_ACCOUNT_JSON", "").strip()
            if sa_clean and not sa_clean.startswith("{"):
                sa_clean = _b64.b64decode(sa_clean).decode("utf-8")
            credentials = service_account.Credentials.from_service_account_info(
                _json.loads(sa_clean), scopes=["https://www.googleapis.com/auth/cloud-platform"]
            ) if sa_clean else None
            client = genai_sdk.Client(
                vertexai=True, project=os.environ.get("GCP_PROJECT", ""),
                location=os.environ.get("GCP_LOCATION", "us-central1"), credentials=credentials
            )
        else:
            client = genai_sdk.Client(api_key=os.environ.get("GEMINI_API_KEY", ""))

        prompt = (
            "Esta es una foto del salpicadero/cuadro de instrumentos de una furgoneta. "
            "Lee el ODÓMETRO (kilometraje total acumulado, normalmente 5-6 dígitos, NO el parcial 'trip'). "
            "Responde SOLO este JSON sin markdown: "
            '{"km": <número entero o null si no es legible>, "confidence": <0.0-1.0>, "legible": <true|false>}'
        )
        contents = [prompt, genai_types.Part.from_bytes(data=img_bytes, mime_type="image/jpeg")]
        gen_config = genai_types.GenerateContentConfig(temperature=0.0, response_mime_type="application/json")
        loop = asyncio.get_running_loop()
        async with _gemini_sem:
            response = await asyncio.wait_for(
                loop.run_in_executor(
                    _executor,
                    lambda: client.models.generate_content(
                        model=os.environ.get("GEMINI_MODEL", "gemini-2.5-flash"),
                        contents=contents, config=gen_config
                    )
                ),
                timeout=45.0
            )
        raw = _strip_markdown_json(response.text or "")
        data = json.loads(raw)
        km_read = data.get("km")
        legible = bool(data.get("legible", False)) and km_read is not None
        current = v.get("mileage") or 0
        warning = None
        if legible and km_read is not None and current and km_read < current:
            warning = f"El km leído ({km_read:,}) es MENOR que el registrado ({current:,}). Revisa la foto."
        logger.info(f"Odómetro {v.get('license_plate','')}: leído={km_read} legible={legible}")
        return {
            "success": legible,
            "km": km_read,
            "confidence": float(data.get("confidence", 0)),
            "current_mileage": current,
            "warning": warning,
        }
    except asyncio.TimeoutError:
        raise HTTPException(status_code=504, detail="La lectura tardó demasiado. Inténtalo de nuevo.")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error leyendo odómetro: {e}")
        raise HTTPException(status_code=502, detail="No se pudo leer el cuentakilómetros. Introduce el km a mano.")


@api_router.post("/vehicles/{vehicle_id}/mileage")
async def update_mileage(vehicle_id: str, data: dict, user: dict = Depends(require_any_auth)):
    """Actualiza el kilometraje. Conductores solo su vehículo. Genera aviso de aceite si toca."""
    km = int(data.get("km", 0))
    v = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not v:
        raise HTTPException(status_code=404, detail="Furgoneta no encontrada")
    if user["role"] == "driver" and v.get("current_driver_id") != user["sub"]:
        # Los conductores rotan por cuadrante diario: aceptar si HOY tiene
        # asignada esta furgoneta, o si acaba de inspeccionarla hoy.
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        assigned_today = await db.daily_assignments.find_one({
            "date": today,
            "slots": {"$elemMatch": {"driver_id": user["sub"], "vehicle_id": vehicle_id}}
        }, {"_id": 0, "id": 1})
        inspected_today = None
        if not assigned_today:
            inspected_today = await db.inspections.find_one({
                "vehicle_id": vehicle_id, "driver_id": user["sub"],
                "created_at": {"$regex": f"^{today}"}
            }, {"_id": 0, "id": 1})
        if not assigned_today and not inspected_today:
            raise HTTPException(status_code=403, detail="Solo tu vehículo asignado")
    # No permitir km menores que los actuales (salvo admin)
    if user["role"] != "admin" and km < (v.get("mileage") or 0):
        raise HTTPException(status_code=400, detail="Los km no pueden ser menores que los actuales")

    km_entry = {"date": datetime.now(timezone.utc).isoformat(), "km": km,
                "source": user.get("role", "?")}
    await db.vehicles.update_one({"id": vehicle_id},
        {"$set": {"mileage": km, "updated_at": datetime.now(timezone.utc)},
         "$push": {"mileage_history": km_entry}})

    # ¿Toca avisar de aceite?
    oil_km = v.get("oil_last_change_km")
    oil_alert = None
    if oil_km is not None:
        intervalo = v.get("oil_interval_km", 15000)
        aviso_antes = v.get("oil_warning_before_km", 2500)
        recorridos = km - oil_km
        restantes = intervalo - recorridos
        if restantes <= aviso_antes:
            if restantes <= 0:
                titulo = f"CAMBIO DE ACEITE VENCIDO en {v.get('license_plate','')}"
                desc = f"Se ha pasado {abs(restantes)} km del cambio de aceite."
                sev = "high"
            else:
                titulo = f"Cambio de aceite próximo en {v.get('license_plate','')}"
                desc = f"Quedan {restantes} km para el cambio de aceite (intervalo {intervalo} km)."
                sev = "medium"
            # Evitar duplicar el aviso si ya hay uno reciente sin leer
            existe = await db.alerts.find_one({"vehicle_id": vehicle_id, "title": titulo, "read": {"$ne": True}})
            if not existe:
                await db.alerts.insert_one(serialize_doc(Alert(
                    vehicle_id=vehicle_id, inspection_id="", title=titulo,
                    description=desc, severity=sev).model_dump()))
                oil_alert = {"remaining_km": restantes, "message": desc}
                # Notificar a Telegram también
                try:
                    config = await db.telegram_config.find_one({}, {"_id": 0})
                    if config and config.get("enabled") and config.get("bot_token"):
                        async with _aiohttp.ClientSession() as session:
                            for cid in config.get("chat_ids", []):
                                if cid.strip():
                                    await session.post(
                                        f"https://api.telegram.org/bot{config['bot_token']}/sendMessage",
                                        json={"chat_id": cid, "text": f"🛢️ <b>{titulo}</b>\n{desc}", "parse_mode": "HTML"},
                                        timeout=_aiohttp.ClientTimeout(total=8))
                except Exception as e:
                    logger.warning(f"Telegram aceite: {e}")

    # ¿Toca avisar de ruedas o pastillas? (mismo patrón que el aceite)
    for _mk, _mspec in (("ruedas", {"label": "CAMBIO DE RUEDAS", "emoji": "🛞", "interval": 40000, "warn": 3000}),
                        ("pastillas", {"label": "PASTILLAS DE FRENO", "emoji": "🛑", "interval": 30000, "warn": 3000})):
        last_km = v.get(f"{_mk}_last_change_km")
        if last_km is None:
            continue
        _interval = v.get(f"{_mk}_interval_km", _mspec["interval"])
        _warn = v.get(f"{_mk}_warning_before_km", _mspec["warn"])
        _rest = _interval - (km - last_km)
        if _rest <= _warn:
            if _rest <= 0:
                _tit = f"{_mspec['label']} VENCIDO en {v.get('license_plate','')}"
                _desc = f"Se ha pasado {abs(_rest)} km del cambio."
                _sev = "high"
            else:
                _tit = f"{_mspec['label'].capitalize()} próximo en {v.get('license_plate','')}"
                _desc = f"Quedan {_rest} km (intervalo {_interval} km)."
                _sev = "medium"
            _existe = await db.alerts.find_one({"vehicle_id": vehicle_id, "title": _tit, "read": {"$ne": True}})
            if not _existe:
                await db.alerts.insert_one(serialize_doc(Alert(
                    vehicle_id=vehicle_id, inspection_id="", title=_tit,
                    description=_desc, severity=_sev).model_dump()))
                try:
                    config = await db.telegram_config.find_one({}, {"_id": 0})
                    if config and config.get("enabled") and config.get("bot_token"):
                        async with _aiohttp.ClientSession() as session:
                            for cid in config.get("chat_ids", []):
                                if cid.strip():
                                    await session.post(
                                        f"https://api.telegram.org/bot{config['bot_token']}/sendMessage",
                                        json={"chat_id": cid, "text": f"{_mspec['emoji']} <b>{_tit}</b>\n{_desc}", "parse_mode": "HTML"},
                                        timeout=_aiohttp.ClientTimeout(total=8))
                except Exception as e:
                    logger.warning(f"Telegram {_mk}: {e}")

    return {"success": True, "mileage": km, "oil_alert": oil_alert}


@api_router.get("/alerts/itv")
async def get_itv_alerts(_=Depends(require_admin)):
    """Lista furgonetas con ITV próxima a caducar (30 días) o caducada."""
    from datetime import date as _date
    hoy = _date.today()
    vehicles = await db.vehicles.find({"status": {"$ne": "deleted"}}, {"_id": 0}).to_list(2000)
    result = []
    for v in vehicles:
        itv = v.get("itv_date")
        if not itv:
            continue
        try:
            y, m, d = [int(x) for x in itv.split("-")]
            fecha = _date(y, m, d)
        except Exception:
            continue
        dias = (fecha - hoy).days
        if dias <= 30:  # caducada o caduca en <=30 días
            result.append({
                "vehicle_id": v.get("id"), "license_plate": v.get("license_plate"),
                "brand": v.get("brand"), "model": v.get("model"), "center": v.get("center"),
                "itv_date": itv, "days_left": dias,
                "status": "caducada" if dias < 0 else ("urgente" if dias <= 7 else "proxima"),
            })
    result.sort(key=lambda x: x["days_left"])
    return result


@api_router.get("/alerts/renting")
async def get_renting_alerts(_=Depends(require_admin)):
    """Lista furgonetas con vencimiento de renting próximo (30 días) o vencido."""
    from datetime import date as _date
    hoy = _date.today()
    vehicles = await db.vehicles.find({"status": {"$ne": "deleted"}}, {"_id": 0}).to_list(2000)
    result = []
    for v in vehicles:
        rent = v.get("renting_end_date")
        if not rent:
            continue
        try:
            y, m, d = [int(x) for x in rent.split("-")]
            fecha = _date(y, m, d)
        except Exception:
            continue
        dias = (fecha - hoy).days
        if dias <= 30:
            result.append({
                "vehicle_id": v.get("id"), "license_plate": v.get("license_plate"),
                "brand": v.get("brand"), "model": v.get("model"), "center": v.get("center"),
                "provider": v.get("provider"), "renting_end_date": rent, "days_left": dias,
                "status": "vencido" if dias < 0 else ("urgente" if dias <= 7 else "proximo"),
            })
    result.sort(key=lambda x: x["days_left"])
    return result


@api_router.get("/vehicles/{vehicle_id}/history")
async def get_vehicle_history(vehicle_id: str, _=Depends(require_admin)):
    """Devuelve el histórico de km y bolsas de una furgoneta, listo para graficar."""
    v = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not v:
        raise HTTPException(status_code=404, detail="Furgoneta no encontrada")
    # Historial de km (ordenado por fecha)
    mh = v.get("mileage_history", [])
    km_points = [{"date": e.get("date", "")[:10], "km": e.get("km", 0)} for e in mh if e.get("km")]
    # Historial de bolsas: reconstruir el valor restante tras cada movimiento
    bh = v.get("bags_history", [])
    bags_points = []
    for e in bh:
        bags_points.append({
            "date": e.get("date", "")[:10],
            "remaining": e.get("remaining_after", e.get("remaining", 0)),
            "change": e.get("change", 0),
        })
    return {
        "license_plate": v.get("license_plate"),
        "current_mileage": v.get("mileage"),
        "current_bags": v.get("bags_remaining", 0),
        "km_history": km_points,
        "bags_history": bags_points,
    }


def _build_maint_item(v: dict, kind: str, default_interval: int, default_warn: int) -> Optional[dict]:
    """Construye el objeto de estado de un ítem de mantenimiento a partir del documento de vehículo."""
    last_km = v.get(f"{kind}_last_change_km")
    if last_km is None:
        return None
    km_actual = v.get("mileage") or last_km
    interval = v.get(f"{kind}_interval_km", default_interval)
    warn = v.get(f"{kind}_warning_before_km", default_warn)
    recorridos = km_actual - last_km
    restantes = interval - recorridos
    return {
        "last_change_km": last_km,
        "last_change_date": v.get(f"{kind}_last_change_date"),
        "interval_km": interval,
        "warning_before_km": warn,
        "km_until_change": restantes,
        "next_change_at_km": last_km + interval,
        "overdue": restantes <= 0,
        "warning": 0 < restantes <= warn,
    }


@api_router.get("/vehicles/{vehicle_id}/maintenance")
async def get_maintenance_info(vehicle_id: str, _=Depends(require_admin)):
    """Devuelve info de mantenimiento completa de una furgoneta (aceite, ruedas, pastillas).
    Incluye PREDICCIÓN: km/día reales (mileage_history, últimos 60 días) → días
    estimados hasta cada cambio, para planificar citas de taller con antelación."""
    v = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
    if not v:
        raise HTTPException(status_code=404, detail="Furgoneta no encontrada")

    # km/día reales: pendiente entre el primer y último registro recientes.
    # Mínimo 2 registros con 7+ días de separación para no extrapolar ruido.
    km_per_day = None
    hist = sorted(
        (h for h in (v.get("mileage_history") or [])
         if isinstance(h, dict) and h.get("date") and isinstance(h.get("km"), (int, float))),
        key=lambda h: str(h["date"]),
    )
    if len(hist) >= 2:
        cutoff = (datetime.now(timezone.utc) - timedelta(days=60)).strftime("%Y-%m-%d")
        recent = [h for h in hist if str(h["date"])[:10] >= cutoff] or hist[-5:]
        if len(recent) >= 2:
            try:
                d0 = datetime.fromisoformat(str(recent[0]["date"])[:10])
                d1 = datetime.fromisoformat(str(recent[-1]["date"])[:10])
                span = (d1 - d0).days
                dk = recent[-1]["km"] - recent[0]["km"]
                if span >= 7 and dk > 0:
                    km_per_day = round(dk / span, 1)
            except Exception:
                km_per_day = None

    def _with_estimate(item):
        if item and km_per_day and isinstance(item.get("km_until_change"), (int, float)):
            item["days_left_estimate"] = max(0, round(item["km_until_change"] / km_per_day))
        return item

    return {
        "mileage": v.get("mileage"),
        "bags_remaining": v.get("bags_remaining", 0),
        "bags_history": v.get("bags_history", [])[-10:],
        "provider": v.get("provider"),
        "km_per_day": km_per_day,
        "oil":       _with_estimate(_build_maint_item(v, "oil",       15000, 2500)),
        "ruedas":    _with_estimate(_build_maint_item(v, "ruedas",    40000, 3000)),
        "pastillas": _with_estimate(_build_maint_item(v, "pastillas", 30000, 3000)),
    }


@api_router.get("/alerts/maintenance")
async def get_maintenance_alerts(_=Depends(require_admin)):
    """Devuelve todas las furgonetas con mantenimiento vencido o próximo."""
    vehicles = await db.vehicles.find(
        {"status": {"$ne": "baja"}},
        {"_id": 0, "id": 1, "license_plate": 1, "brand": 1, "model": 1, "center": 1,
         "mileage": 1,
         "oil_last_change_km": 1, "oil_last_change_date": 1, "oil_interval_km": 1, "oil_warning_before_km": 1,
         "ruedas_last_change_km": 1, "ruedas_last_change_date": 1, "ruedas_interval_km": 1, "ruedas_warning_before_km": 1,
         "pastillas_last_change_km": 1, "pastillas_last_change_date": 1, "pastillas_interval_km": 1, "pastillas_warning_before_km": 1}
    ).to_list(500)

    alerts = []
    for v in vehicles:
        for kind, default_i, default_w, label in [
            ("oil", 15000, 2500, "Aceite"),
            ("ruedas", 40000, 3000, "Ruedas"),
            ("pastillas", 30000, 3000, "Pastillas de freno"),
        ]:
            item = _build_maint_item(v, kind, default_i, default_w)
            if item and (item["overdue"] or item["warning"]):
                alerts.append({
                    "vehicle_id": v["id"],
                    "license_plate": v.get("license_plate"),
                    "brand": v.get("brand"),
                    "center": v.get("center"),
                    "mileage": v.get("mileage"),
                    "kind": kind,
                    "label": label,
                    **item,
                })
    alerts.sort(key=lambda a: a["km_until_change"])
    return alerts



@api_router.post("/import/diagnose")
async def diagnose_excel(file: UploadFile = File(...), _=Depends(require_admin)):
    """DIAGNÓSTICO TEMPORAL: analiza el Excel y devuelve qué hay en cada columna clave."""
    content = await file.read()
    import openpyxl, unicodedata
    def _sa(t): return "".join(c for c in unicodedata.normalize("NFD",t) if unicodedata.category(c)!="Mn")

    out = {}
    # Probar AMBOS modos de carga
    for modo in ["data_only_true", "data_only_false"]:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=(modo=="data_only_true"))
        # Listar todas las hojas
        out["hojas"] = wb.sheetnames
        ws = wb.active
        out["hoja_activa"] = ws.title
        headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        headers_norm = [_sa(h) for h in headers]
        # Índices de proveedor, matrícula, kilómetros
        def idx(name):
            k=_sa(name.lower())
            return headers_norm.index(k) if k in headers_norm else -1
        i_prov = idx("proveedor")
        i_mat = idx("matricula")
        # Leer las primeras 5 filas de datos, columna proveedor
        muestras = []
        cnt = 0
        for row in ws.iter_rows(min_row=2):
            if not any(c.value for c in row): continue
            if cnt >= 5: break
            mat = row[i_mat].value if i_mat>=0 else None
            prov_cell = row[i_prov] if i_prov>=0 else None
            muestras.append({
                "matricula": str(mat),
                "proveedor_valor": str(prov_cell.value) if prov_cell else "SIN_COLUMNA",
                "proveedor_tipo": type(prov_cell.value).__name__ if prov_cell else "N/A",
                "proveedor_data_type": prov_cell.data_type if prov_cell else "N/A"
            })
            cnt += 1
        out[modo] = {
            "indice_proveedor": i_prov,
            "indice_matricula": i_mat,
            "total_headers": len(headers),
            "muestras": muestras
        }
    return out



# =========================
# INFORME PDF DE PERITAJE
# =========================


async def _cargar_foto_bytes(foto):
    """Carga una foto (URL http o base64) y devuelve sus bytes."""
    import base64 as _b64, aiohttp as _aio
    if not foto:
        return None
    try:
        if foto.startswith("http"):
            async with _aio.ClientSession(headers={"User-Agent":"FlotaDSP/1.0"}) as ss:
                async with ss.get(foto, timeout=_aio.ClientTimeout(total=15)) as r:
                    if r.status == 200:
                        return await r.read()
        else:
            b64 = foto.split(",",1)[1] if foto.startswith("data:") else foto
            return _b64.b64decode(b64)
    except Exception as e:
        logger.warning(f"cargar_foto error: {e}")
    return None


def _damages_to_detections(damages: list) -> List[AIDetection]:
    """Maps stored analysis.damages → AIDetection using location_hint zone boxes.
    Shifts overlapping boxes in the same zone by a small offset for readability."""
    detections: List[AIDetection] = []
    zone_count: dict = {}
    for d in damages:
        if isinstance(d, dict):
            loc = d.get("location_hint", "otra") or "otra"
            label = d.get("part", "daño")
            severity = d.get("severity", "leve")
            conf = float(d.get("confidence", 0.7))
        else:
            loc = getattr(d, "location_hint", "otra") or "otra"
            label = getattr(d, "part", "daño")
            severity = getattr(d, "severity", "leve")
            conf = float(getattr(d, "confidence", 0.7))

        base = list(_LOCATION_BOX.get(loc, _LOCATION_BOX["otra"]))
        n = zone_count.get(loc, 0)
        zone_count[loc] = n + 1
        shift = n * 55
        box = [
            min(base[0] + shift,      900),
            min(base[1] + shift // 2, 900),
            min(base[2] + shift,      950),
            min(base[3] + shift // 2, 970),
        ]
        detections.append(AIDetection(
            label=label, severity=severity, box_2d=box,
            confidence=conf, source="location_hint",
        ))
    return detections


def _merge_yolo_with_gemini(
    yolo_detections: List[AIDetection],
    gemini_damages: list,
    photo_index: int = 0,
) -> List[AIDetection]:
    """
    Enhances YOLO detections with Gemini damage labels.

    YOLO gives precise bounding boxes but generic class names (Crack, Dent...).
    Gemini gives professional labels (Paragolpes delantero, Puerta del conductor...)
    with location_hint zones.

    Strategy: for each YOLO box, find the Gemini damage whose location_hint zone
    best overlaps with the box position, then replace the YOLO label with the
    Gemini professional label. Falls back to YOLO label if no match.
    """
    if not gemini_damages or not yolo_detections:
        return yolo_detections

    # Zone → normalized center x range (0-1000)
    _zone_x = {
        "frontal":           (0,   500),
        "trasera":           (500, 1000),
        "lateral_izquierdo": (0,   500),
        "lateral_derecho":   (500, 1000),
        "techo":             (200, 800),
        "otra":              (0,   1000),
    }
    # Zone → normalized center y range (0-1000)
    _zone_y = {
        "frontal":           (0,   1000),
        "trasera":           (0,   1000),
        "lateral_izquierdo": (200, 900),
        "lateral_derecho":   (200, 900),
        "techo":             (0,   400),
        "otra":              (0,   1000),
    }

    # Build list of Gemini damages with zone info — filter to photo zone hints
    # For photo_index: 0=frontal, 1=trasera, 2=lateral_izq, 3=lateral_der (heuristic)
    _photo_zone_hints = {
        0: ["frontal", "otra"],
        1: ["trasera", "otra"],
        2: ["lateral_izquierdo", "otra"],
        3: ["lateral_derecho", "otra"],
    }
    relevant_hints = _photo_zone_hints.get(photo_index, ["otra"])

    gemini_list = []
    for d in gemini_damages:
        if isinstance(d, dict):
            part = d.get("part", "")
            loc = d.get("location_hint", "otra") or "otra"
            sev = d.get("severity", "leve")
            desc = d.get("description", "")
        else:
            part = getattr(d, "part", "")
            loc = getattr(d, "location_hint", "otra") or "otra"
            sev = getattr(d, "severity", "leve")
            desc = getattr(d, "description", "")
        gemini_list.append({"part": part, "loc": loc, "sev": sev, "desc": desc, "used": False})

    merged = []
    for det in yolo_detections:
        box = det.box_2d  # [ymin, xmin, ymax, xmax] 0-1000
        cx = (box[1] + box[3]) / 2
        cy = (box[0] + box[2]) / 2

        best_idx = None
        best_score = -1
        for i, g in enumerate(gemini_list):
            if g["used"]:
                continue
            loc = g["loc"]
            xr = _zone_x.get(loc, (0, 1000))
            yr = _zone_y.get(loc, (0, 1000))
            # Score: how well the box center falls within this zone
            x_overlap = max(0, min(cx, xr[1]) - max(cx - 1, xr[0]))
            y_overlap = max(0, min(cy, yr[1]) - max(cy - 1, yr[0]))
            in_zone = (xr[0] <= cx <= xr[1]) and (yr[0] <= cy <= yr[1])
            zone_match = loc in relevant_hints
            score = (2 if in_zone else 0) + (1 if zone_match else 0)
            if score > best_score:
                best_score = score
                best_idx = i

        if best_idx is not None and best_score > 0:
            g = gemini_list[best_idx]
            g["used"] = True
            # Build professional label: "Parte — tipo daño"
            sev_label = {"leve": "leve", "moderado": "moderado",
                         "grave": "grave", "critico": "crítico"}.get(g["sev"], g["sev"])
            label = g["part"] if g["part"] else det.label
            merged.append(AIDetection(
                label=label,
                severity=g["sev"],
                box_2d=det.box_2d,
                confidence=det.confidence,
                source="yolo+gemini",
            ))
        else:
            merged.append(det)

    return merged


async def _run_yolo_for_inspection(inspection_id: str, photo_urls: List[str]) -> None:
    """Background task: runs YOLO detection on every photo of a new inspection
    and stores results in inspection_ai_results. Called after upload completes."""
    if not AI_SERVICE_URL:
        logger.info(f"AI service not configured — skipping YOLO for {inspection_id[:8]}")
        return
    for idx, url in enumerate(photo_urls):
        try:
            img_bytes = await _cargar_foto_bytes(url)
            if not img_bytes:
                continue
            detections = await _call_ai_service_detect(inspection_id, idx, img_bytes)
            if detections is None:
                continue
            # Enrich YOLO boxes with Gemini professional labels
            insp_doc = await db.inspections.find_one({"id": inspection_id}, {"_id": 0, "analysis": 1})
            gemini_damages = ((insp_doc or {}).get("analysis") or {}).get("damages", [])
            if gemini_damages:
                detections = _merge_yolo_with_gemini(detections, gemini_damages, idx)
            result = InspectionAIResult(
                inspection_id=inspection_id,
                photo_index=idx,
                detections=detections,
                source="yolo+gemini" if gemini_damages else "yolo",
                updated_at=datetime.now(timezone.utc),
            )
            await db.inspection_ai_results.update_one(
                {"inspection_id": inspection_id, "photo_index": idx},
                {"$set": serialize_doc(result.model_dump())},
                upsert=True,
            )
            logger.info(f"YOLO auto-detect: insp={inspection_id[:8]} photo={idx} → {len(detections)} detecciones")
        except Exception as e:
            logger.warning(f"YOLO auto-detect error insp={inspection_id[:8]} photo={idx}: {e}")


async def _call_ai_service_detect(
    inspection_id: str, photo_index: int, img_bytes: bytes
) -> Optional[List[AIDetection]]:
    """Calls the external GPU microservice (YOLO11+SAM2). Returns None if unavailable."""
    if not AI_SERVICE_URL:
        return None
    try:
        import aiohttp as _aio
        payload = {
            "inspection_id": inspection_id,
            "photo_index": photo_index,
            "image_b64": base64.b64encode(img_bytes).decode(),
        }
        async with _aio.ClientSession() as session:
            async with session.post(
                f"{AI_SERVICE_URL}/detect",
                json=payload,
                timeout=_aio.ClientTimeout(total=30),
            ) as resp:
                if resp.status == 200:
                    data = await resp.json()
                    return [AIDetection(**det) for det in data.get("detections", [])]
                logger.warning(f"AI service HTTP {resp.status}")
                return None
    except Exception as e:
        logger.warning(f"AI service unavailable: {e}")
        return None


@api_router.post("/ai/detect/{inspection_id}")
async def ai_detect_inspection(
    inspection_id: str, photo_index: int = 0, _=Depends(require_admin)
):
    """Runs AI detection on an inspection photo.
    Uses the GPU microservice (YOLO11+SAM2) if AI_SERVICE_URL is set,
    otherwise falls back to deterministic location_hint mapping from stored analysis."""
    insp = await db.inspections.find_one({"id": inspection_id}, {"_id": 0})
    if not insp:
        raise HTTPException(status_code=404, detail="Inspección no encontrada")
    photos = insp.get("photos", [])
    if not photos or photo_index >= len(photos):
        raise HTTPException(status_code=400, detail="Foto no encontrada")

    img_bytes = await _cargar_foto_bytes(photos[photo_index])
    if not img_bytes:
        raise HTTPException(status_code=400, detail="No se pudo cargar la foto")

    detections = await _call_ai_service_detect(inspection_id, photo_index, img_bytes)
    source = "yolo"
    gemini_damages = (insp.get("analysis") or {}).get("damages", [])
    if detections is None:
        detections = _damages_to_detections(gemini_damages)
        source = "location_hint"
    elif gemini_damages:
        # Enrich YOLO boxes with Gemini professional labels
        detections = _merge_yolo_with_gemini(detections, gemini_damages, photo_index)
        source = "yolo+gemini"

    result = InspectionAIResult(
        inspection_id=inspection_id,
        photo_index=photo_index,
        detections=detections,
        source=source,
        updated_at=datetime.now(timezone.utc),
    )
    await db.inspection_ai_results.update_one(
        {"inspection_id": inspection_id, "photo_index": photo_index},
        {"$set": serialize_doc(result.model_dump())},
        upsert=True,
    )
    return {
        "inspection_id": inspection_id,
        "photo_index": photo_index,
        "count": len(detections),
        "source": source,
        "detections": [d.model_dump() for d in detections],
    }


@api_router.get("/ai/status/{inspection_id}")
async def ai_detection_status(inspection_id: str, _=Depends(require_admin)):
    """Returns stored AI detection results for all photos of an inspection."""
    results = await db.inspection_ai_results.find(
        {"inspection_id": inspection_id}, {"_id": 0}
    ).sort("photo_index", 1).to_list(20)
    return {
        "inspection_id": inspection_id,
        "photos_processed": len(results),
        "ai_service_configured": bool(AI_SERVICE_URL),
        "detection_mode": "yolo11+sam2" if AI_SERVICE_URL else "location_hint",
        "results": results,
    }


async def _detectar_cajas_danos(img_bytes):
    """Pide a Gemini las cajas de daños sobre una foto. Devuelve lista [{label, box_2d, severity}]."""
    try:
        from google import genai as genai_sdk
        from google.genai import types as genai_types
        use_vertex = os.environ.get("USE_VERTEX_AI","").lower() in ("1","true","yes")
        model_name = os.environ.get("GEMINI_MODEL","gemini-2.5-flash")
        if use_vertex:
            from google.oauth2 import service_account
            import json as _json, base64 as _b64c
            sa_clean = os.environ.get("GCP_SERVICE_ACCOUNT_JSON","").strip()
            if sa_clean and not sa_clean.startswith("{"):
                try:
                    sa_clean = _b64c.b64decode(sa_clean).decode("utf-8")
                except Exception:
                    pass
            creds = service_account.Credentials.from_service_account_info(_json.loads(sa_clean), scopes=["https://www.googleapis.com/auth/cloud-platform"])
            client = genai_sdk.Client(vertexai=True, project=os.environ.get("GCP_PROJECT",""), location=os.environ.get("GCP_LOCATION","us-central1"), credentials=creds)
        else:
            client = genai_sdk.Client(api_key=os.environ.get("GEMINI_API_KEY",""))
        prompt = (
            "Detecta los daños visibles en la carrocería de esta furgoneta (golpes, arañazos, "
            "abolladuras, roturas, piezas rotas). Devuelve SOLO un JSON array, sin markdown. "
            'Cada elemento: {"label": "descripción corta", "severity": "leve|moderado|grave|critico", "box_2d": [ymin, xmin, ymax, xmax]}. '
            "box_2d normalizado de 0 a 1000 (0=arriba/izquierda). Si no hay daños, []. SOLO el JSON."
        )
        # Prompt PRIMERO, imagen después (como la función de análisis que funciona)
        contents = [prompt, genai_types.Part.from_bytes(data=img_bytes, mime_type="image/jpeg")]
        cfg = genai_types.GenerateContentConfig(temperature=0.1, max_output_tokens=4096)
        loop = asyncio.get_running_loop()
        resp = await asyncio.wait_for(loop.run_in_executor(_executor, lambda: client.models.generate_content(model=model_name, contents=contents, config=cfg)), timeout=60.0)
        import json as _json2, re as _re2
        # Extraer texto de forma defensiva (resp.text puede venir vacío en Vertex)
        raw = ""
        try:
            raw = (resp.text or "").strip()
        except Exception:
            raw = ""
        if not raw:
            try:
                for cand in (resp.candidates or []):
                    for part in (cand.content.parts or []):
                        if getattr(part, "text", None):
                            raw += part.text
                raw = raw.strip()
            except Exception:
                pass
        logger.info(f"detectar_cajas raw (len={len(raw)}): {raw[:150]}")
        globals()["_ultimo_raw_gemini"] = f"len={len(raw)} | {raw[:200]}"
        # Extraer el array JSON [...] esté donde esté (dentro de markdown o no)
        m = _re2.search(r"\[.*\]", raw, _re2.DOTALL)
        json_str = m.group(0) if m else "[]"
        try:
            data = _json2.loads(json_str)
        except Exception as _je:
            logger.error(f"detectar_cajas parseo falló: {_je} | json_str={json_str[:100]}")
            data = []
        return data if isinstance(data, list) else []
    except Exception as e:
        logger.error(f"detectar_cajas error: {type(e).__name__}: {e}")
        return []


def _dibujar_numeros(img_bytes, boxes, start_num=0):
    """Dibuja daños con relleno semitransparente + etiqueta flotante estilo IA.
    Estilo: área de daño coloreada semitransparente, esquinas reticle, badge numerado
    y etiqueta oscura flotante con tipo de daño, severidad y confianza.
    start_num: offset de numeración para que el grid multi-foto numere de forma
    continua (foto 1: 1,2,3 · foto 2: 4,5,6...) y coincida 1:1 con la leyenda."""
    from PIL import Image, ImageDraw, ImageFont
    import io as _io

    img = Image.open(_io.BytesIO(img_bytes)).convert("RGBA")
    W, H = img.size

    # Overlay for semitransparent fills (painted once, alpha-composited)
    overlay = Image.new("RGBA", (W, H), (0, 0, 0, 0))
    drw = ImageDraw.Draw(overlay)

    # Font setup
    fsize_badge = max(18, W // 36)
    fsize_label = max(13, W // 54)
    _font_paths = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
        "/usr/share/fonts/truetype/freefont/FreeSansBold.ttf",
    ]
    _font_reg_paths = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
        "/usr/share/fonts/truetype/freefont/FreeSans.ttf",
    ]
    def _load_font(paths, size):
        for p in paths:
            try:
                return ImageFont.truetype(p, size)
            except Exception:
                pass
        return ImageFont.load_default()

    font_badge = _load_font(_font_paths, fsize_badge)
    font_label = _load_font(_font_reg_paths, fsize_label)

    # Severity → (R, G, B)
    _SEV_RGB = {
        "critico":  (220, 38,  38),
        "grave":    (234, 179,  8),
        "moderado": (249, 115, 22),
        "leve":     (34,  197, 94),
    }
    _SEV_ES = {
        "critico": "Crítico", "grave": "Grave",
        "moderado": "Moderado", "leve": "Leve",
    }

    # ── Pass 1: draw fills & borders onto overlay ─────────────────
    for b in boxes:
        box = b.get("box_2d") or b.get("box") or []
        if len(box) != 4:
            continue
        ymin, xmin, ymax, xmax = box
        x1 = int(xmin / 1000 * W); y1 = int(ymin / 1000 * H)
        x2 = int(xmax / 1000 * W); y2 = int(ymax / 1000 * H)
        sev = b.get("severity", "moderado")
        rgb = _SEV_RGB.get(sev, (249, 115, 22))
        r, g, b_c = rgb

        # Semitransparent fill
        drw.rectangle([x1, y1, x2, y2], fill=(r, g, b_c, 48))

        # Solid border
        bw = max(2, W // 280)
        drw.rectangle([x1, y1, x2, y2], outline=(r, g, b_c, 210), width=bw)

        # Corner reticle lines (targeting style)
        cl = max(10, min(28, (x2 - x1) // 5, (y2 - y1) // 5))
        lw = max(2, W // 220)
        for sx, sy, dx, dy in [
            (x1, y1,  1,  1), (x2, y1, -1,  1),
            (x1, y2,  1, -1), (x2, y2, -1, -1),
        ]:
            drw.line([(sx, sy + dy * cl), (sx, sy), (sx + dx * cl, sy)],
                     fill=(r, g, b_c, 255), width=lw + 1)

        # Badge circle (top-left of box)
        br = fsize_badge
        bx = max(0, x1 - br // 2)
        by = max(0, y1 - br)
        drw.ellipse([bx, by, bx + br * 2, by + br * 2], fill=(r, g, b_c, 230),
                    outline=(255, 255, 255, 160), width=max(1, bw - 1))

        # Label background
        label_text = b.get("label", "Daño")
        sev_str = _SEV_ES.get(sev, sev.capitalize())
        conf = b.get("confidence")
        conf_str = f"  {int(float(conf) * 100)}%" if conf is not None else ""
        full_label = f"{label_text}   {sev_str}{conf_str}"

        try:
            tb = ImageDraw.Draw(Image.new("RGBA", (1, 1))).textbbox((0, 0), full_label, font=font_label)
            tw, th = tb[2] - tb[0], tb[3] - tb[1]
        except Exception:
            tw, th = int(len(full_label) * fsize_label * 0.6), fsize_label + 4

        pad = 7
        lbl_w = tw + pad * 2
        lbl_h = th + pad * 2
        lbl_x = max(0, min(x1, W - lbl_w))
        lbl_y = y2 + 5 if y2 + lbl_h + 5 < H else max(0, y1 - lbl_h - 5)

        # Dark rounded pill background
        drw.rounded_rectangle(
            [lbl_x, lbl_y, lbl_x + lbl_w, lbl_y + lbl_h],
            radius=5, fill=(10, 10, 18, 195), outline=(r, g, b_c, 120), width=1,
        )

    # Composite overlay over image
    merged = Image.alpha_composite(img, overlay).convert("RGB")
    draw_text = ImageDraw.Draw(merged)

    # ── Pass 2: draw text ─────────────────────────────────────────
    n = start_num
    for b in boxes:
        box = b.get("box_2d") or b.get("box") or []
        if len(box) != 4:
            continue
        n += 1
        ymin, xmin, ymax, xmax = box
        x1 = int(xmin / 1000 * W); y1 = int(ymin / 1000 * H)
        x2 = int(xmax / 1000 * W); y2 = int(ymax / 1000 * H)
        sev = b.get("severity", "moderado")
        rgb = _SEV_RGB.get(sev, (249, 115, 22))
        r, g, b_c = rgb

        # Badge number
        br = fsize_badge
        bx = max(0, x1 - br // 2)
        by = max(0, y1 - br)
        num = str(n)
        try:
            nb = draw_text.textbbox((0, 0), num, font=font_badge)
            nw, nh = nb[2] - nb[0], nb[3] - nb[1]
        except Exception:
            nw, nh = br, br
        draw_text.text(
            (bx + br - nw // 2, by + br - nh // 2),
            num, fill=(255, 255, 255), font=font_badge,
        )

        # Label text
        label_text = b.get("label", "Daño")
        sev_str = _SEV_ES.get(sev, sev.capitalize())
        conf = b.get("confidence")
        conf_str = f"  {int(float(conf) * 100)}%" if conf is not None else ""
        full_label = f"{label_text}   {sev_str}{conf_str}"

        try:
            tb = draw_text.textbbox((0, 0), full_label, font=font_label)
            tw, th = tb[2] - tb[0], tb[3] - tb[1]
        except Exception:
            tw, th = int(len(full_label) * fsize_label * 0.6), fsize_label + 4

        pad = 7
        lbl_w = tw + pad * 2
        lbl_h = th + pad * 2
        lbl_x = max(0, min(x1, W - lbl_w))
        lbl_y = y2 + 5 if y2 + lbl_h + 5 < H else max(0, y1 - lbl_h - 5)

        draw_text.text((lbl_x + pad, lbl_y + pad), full_label,
                       fill=(255, 255, 255), font=font_label)

    out = _io.BytesIO()
    merged.save(out, format="JPEG", quality=88)
    return out.getvalue(), n




@api_router.get("/inspections/{inspection_id}/annotated")
async def inspection_annotated(inspection_id: str, photo_index: int = 0, _=Depends(require_admin)):
    """Returns a collage of ALL inspection photos with numbered damage boxes.
    When photo_index=0 (default), returns all photos in a 2-column grid.
    When photo_index>0, returns that specific photo only (backward compat).
    Priority: stored YOLO+Gemini detections → location_hint → legacy Gemini
    """
    from fastapi.responses import Response
    import json as _json
    from PIL import Image as _PILImage
    import io as _io

    insp = await db.inspections.find_one({"id": inspection_id}, {"_id": 0})
    if not insp:
        raise HTTPException(status_code=404, detail="Inspección no encontrada")
    photos = insp.get("photos", [])
    if not photos:
        raise HTTPException(status_code=400, detail="Sin fotos en esta inspección")

    # If specific photo requested (not default), return just that one
    single_mode = photo_index > 0
    photo_indices = [photo_index] if single_mode else list(range(len(photos)))

    async def _get_annotated_bytes(idx, start_num=0):
        if idx >= len(photos):
            return None, []
        img_bytes = await _cargar_foto_bytes(photos[idx])
        if not img_bytes:
            return None, []
        boxes = []
        ai_result = await db.inspection_ai_results.find_one(
            {"inspection_id": inspection_id, "photo_index": idx}, {"_id": 0}
        )
        if ai_result and ai_result.get("detections"):
            for det in ai_result["detections"]:
                if len(det.get("box_2d", [])) == 4:
                    boxes.append(det)
        if not boxes:
            damages = (insp.get("analysis") or {}).get("damages", [])
            if damages:
                for det in _damages_to_detections(damages):
                    boxes.append({"label": det.label, "severity": det.severity, "box_2d": det.box_2d})
        if not boxes:
            boxes = await _detectar_cajas_danos(img_bytes)
        annotated_bytes, _ = _dibujar_numeros(img_bytes, boxes, start_num=start_num)
        return annotated_bytes, boxes

    all_leyenda = []
    box_counter = 0

    if single_mode:
        annotated, boxes = await _get_annotated_bytes(photo_index, start_num=0)
        if not annotated:
            raise HTTPException(status_code=400, detail="No se pudo cargar la foto")
        for b in boxes:
            if len((b.get("box_2d") or b.get("box") or [])) == 4:
                box_counter += 1
                all_leyenda.append({"n": box_counter, "label": b.get("label",""), "severity": b.get("severity",""), "confidence": b.get("confidence")})
        final_bytes = annotated
    else:
        # Build collage: 2 columns grid
        annotated_images = []
        zone_labels = ["Frontal", "Trasera", "Lateral izq.", "Lateral der."]
        for idx in photo_indices:
            # start_num = nº de daños ya contados → numeración continua que
            # coincide 1:1 con la leyenda (foto 1: 1,2,3 · foto 2: 4,5...).
            ann_bytes, boxes = await _get_annotated_bytes(idx, start_num=box_counter)
            if ann_bytes:
                annotated_images.append((ann_bytes, boxes, idx))
                for b in boxes:
                    if len((b.get("box_2d") or b.get("box") or [])) == 4:
                        box_counter += 1
                        label = b.get("label","")
                        zone = zone_labels[idx] if idx < len(zone_labels) else f"Foto {idx+1}"
                        all_leyenda.append({"n": box_counter, "label": f"[{zone}] {label}", "severity": b.get("severity",""), "confidence": b.get("confidence")})

        if not annotated_images:
            raise HTTPException(status_code=400, detail="No se pudieron cargar las fotos")

        # Compose 2-column grid
        pil_imgs = []
        for ann_bytes, _, _ in annotated_images:
            pil_imgs.append(_PILImage.open(_io.BytesIO(ann_bytes)).convert("RGB"))

        cols = 2
        rows = (len(pil_imgs) + 1) // 2
        cell_w = max(img.width for img in pil_imgs)
        cell_h = max(img.height for img in pil_imgs)
        # Scale to reasonable size
        max_cell = 800
        if cell_w > max_cell:
            scale = max_cell / cell_w
            cell_w = int(cell_w * scale)
            cell_h = int(cell_h * scale)
        grid = _PILImage.new("RGB", (cell_w * cols, cell_h * rows), (20, 20, 20))
        for i, img in enumerate(pil_imgs):
            img_resized = img.resize((cell_w, cell_h), _PILImage.LANCZOS)
            col = i % cols
            row = i // cols
            grid.paste(img_resized, (col * cell_w, row * cell_h))

        buf = _io.BytesIO()
        grid.save(buf, format="JPEG", quality=82)
        final_bytes = buf.getvalue()

    # v5.1: enriquecer la leyenda con damage_index para que el frontend pueda
    # mapear cada zona numerada a su daño correspondiente en analysis.damages.
    # Estrategia: matchear por label (case-insensitive). Si el label viene como
    # "[Frontal] Paragolpes delantero", quitar el prefijo de zona antes de matchear.
    try:
        analysis_damages = (insp.get("analysis") or {}).get("damages", []) or []
        label_to_idx: dict = {}
        for di, dmg in enumerate(analysis_damages):
            if isinstance(dmg, dict):
                key = (dmg.get("part") or "").lower().strip()
                if key and key not in label_to_idx:
                    label_to_idx[key] = di
        for item in all_leyenda:
            raw_label = item.get("label", "")
            clean = raw_label
            if "] " in clean:
                clean = clean.split("] ", 1)[1]
            item["damage_index"] = label_to_idx.get(clean.lower().strip())
    except Exception as _e:
        logger.warning(f"No se pudo enriquecer leyenda con damage_index: {_e}")

    return Response(
        content=final_bytes, media_type="image/jpeg",
        headers={
            "X-Boxes-Found": str(box_counter),
            "X-Detection-Source": "yolo+gemini",
            "X-Img-Bytes": str(len(final_bytes)),
            "X-Legend": base64.b64encode(_json.dumps(all_leyenda, ensure_ascii=False).encode()).decode(),
        },
    )

    boxes = []
    detection_source = "none"

    # 1. Stored AI detections (YOLO or location_hint from previous /ai/detect call)
    ai_result = await db.inspection_ai_results.find_one(
        {"inspection_id": inspection_id, "photo_index": photo_index}, {"_id": 0}
    )
    if ai_result and ai_result.get("detections"):
        for det in ai_result["detections"]:
            if len(det.get("box_2d", [])) == 4:
                boxes.append(det)
        detection_source = ai_result.get("source", "stored")

    # 2. location_hint fallback from stored analysis.damages
    if not boxes:
        damages = (insp.get("analysis") or {}).get("damages", [])
        if damages:
            detections = _damages_to_detections(damages)
            for det in detections:
                boxes.append({
                    "label": det.label, "severity": det.severity,
                    "box_2d": det.box_2d,
                })
            detection_source = "location_hint"

    # 3. Legacy Gemini call — only when absolutely no damage data available
    if not boxes:
        boxes = await _detectar_cajas_danos(img_bytes)
        detection_source = "gemini_legacy"

    annotated, n = _dibujar_numeros(img_bytes, boxes)
    leyenda = []
    idx = 0
    for b in boxes:
        if len((b.get("box_2d") or b.get("box") or [])) == 4:
            idx += 1
            leyenda.append({"n": idx, "label": b.get("label", ""), "severity": b.get("severity", "")})

    return Response(
        content=annotated, media_type="image/jpeg",
        headers={
            "X-Boxes-Found": str(n),
            "X-Detection-Source": detection_source,
            "X-Img-Bytes": str(len(img_bytes)),
            "X-Legend": base64.b64encode(_json.dumps(leyenda, ensure_ascii=False).encode()).decode(),
        },
    )


@api_router.get("/inspections/{inspection_id}/pdf")
async def inspection_pdf(inspection_id: str, boxes: int = 0, _=Depends(require_admin)):
    """Genera un PDF profesional del peritaje de una inspección."""
    from fastapi.responses import Response
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.utils import ImageReader
    import io as _io

    insp = await db.inspections.find_one({"id": inspection_id}, {"_id": 0})
    if not insp:
        raise HTTPException(status_code=404, detail="Inspección no encontrada")

    vehicle = await db.vehicles.find_one({"id": insp.get("vehicle_id")}, {"_id": 0}) or {}
    plate = vehicle.get("license_plate", insp.get("vehicle_id", "?"))
    driver_name = "Sin asignar"
    did = insp.get("driver_id")
    if did:
        drv = await db.drivers.find_one({"id": did}, {"_id": 0, "name": 1})
        if drv:
            driver_name = drv.get("name", "Sin asignar")

    a = insp.get("analysis") or {}
    severity = a.get("severity", "sin_danos")
    damages = a.get("damages", [])
    total_cost = a.get("total_estimated_cost", 0)
    summary = a.get("executive_summary", "")
    detected_plate = a.get("detected_plate", "")
    fecha = str(insp.get("created_at", ""))[:10]

    # Descargar fotos (máx 4)
    raw_photos = insp.get("photos", [])[:4]
    photo_bytes = []
    import base64 as _b64
    import aiohttp as _aio
    _diag = []
    async with _aio.ClientSession(headers={"User-Agent": "FlotaDSP-PDF/1.0"}) as _sess:
        for p in raw_photos:
            try:
                if not p:
                    continue
                if p.startswith("http"):
                    async with _sess.get(p, timeout=_aio.ClientTimeout(total=15)) as _r:
                        if _r.status == 200:
                            _data = await _r.read()
                            photo_bytes.append(_data)
                            _diag.append(f"url:200:{len(_data)}b")
                        else:
                            _diag.append(f"url:{_r.status}")
                elif p.startswith("/uploads/"):
                    _fp = ROOT_DIR / p.lstrip("/")
                    if _fp.exists():
                        photo_bytes.append(_fp.read_bytes())
                        _diag.append("local:ok")
                    else:
                        _diag.append("local:missing")
                else:
                    b64str = p.split(",", 1)[1] if p.startswith("data:") else p
                    photo_bytes.append(_b64.b64decode(b64str))
                    _diag.append("b64:ok")
            except Exception as _pe:
                _diag.append(f"err:{str(_pe)[:40]}")
    logger.info(f"PDF-DIAG inspeccion={inspection_id[:8]} fotos_entrada={len(raw_photos)} bytes_listos={len(photo_bytes)} detalle={_diag}")

    # Colores por severidad
    sev_colors = {
        "critico": colors.HexColor("#A32D2D"), "grave": colors.HexColor("#BA7517"),
        "moderado": colors.HexColor("#BA7517"), "leve": colors.HexColor("#639922"),
        "sin_danos": colors.HexColor("#639922"),
    }
    sev_color = sev_colors.get(severity, colors.HexColor("#5F5E5A"))

    buf = _io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=A4)
    W, H = A4

    # --- Cabecera ---
    c.setFillColor(colors.HexColor("#15181d"))
    c.rect(0, H-28*mm, W, 28*mm, fill=1, stroke=0)
    c.setFillColor(colors.HexColor("#ff6b1a"))
    c.roundRect(15*mm, H-22*mm, 10*mm, 10*mm, 2*mm, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 16)
    c.drawString(28*mm, H-18*mm, "FlotaDSP")
    c.setFont("Helvetica", 9)
    c.setFillColor(colors.HexColor("#9aa0aa"))
    c.drawRightString(W-15*mm, H-16*mm, "Informe de peritaje IA")

    y = H - 38*mm
    # --- Datos vehículo ---
    c.setFillColor(colors.HexColor("#15181d"))
    c.setFont("Helvetica-Bold", 20)
    c.drawString(15*mm, y, plate)
    c.setFont("Helvetica", 10)
    c.setFillColor(colors.HexColor("#666666"))
    c.drawString(15*mm, y-6*mm, f"{vehicle.get('brand','')} {vehicle.get('model','')} - {vehicle.get('center','')}")
    c.drawRightString(W-15*mm, y, f"Fecha: {fecha}")
    c.drawRightString(W-15*mm, y-6*mm, f"Conductor: {driver_name}")

    y -= 16*mm
    # --- Resumen (3 cajas) ---
    box_w = (W - 30*mm - 10*mm) / 3
    boxes = [("SEVERIDAD", severity.upper(), sev_color),
             ("DAÑOS", str(len(damages)), colors.HexColor("#2C2C2A")),
             ("COSTE ESTIMADO", f"{total_cost:.0f} EUR", colors.HexColor("#2C2C2A"))]
    bx = 15*mm
    for label, val, col in boxes:
        c.setFillColor(colors.HexColor("#F1EFE8"))
        c.roundRect(bx, y-14*mm, box_w, 14*mm, 2*mm, fill=1, stroke=0)
        c.setFillColor(colors.HexColor("#888888"))
        c.setFont("Helvetica", 7)
        c.drawCentredString(bx+box_w/2, y-5*mm, label)
        c.setFillColor(col)
        c.setFont("Helvetica-Bold", 12)
        c.drawCentredString(bx+box_w/2, y-11*mm, val)
        bx += box_w + 5*mm

    y -= 22*mm
    # --- Resumen ejecutivo ---
    if summary:
        c.setFillColor(colors.HexColor("#888888"))
        c.setFont("Helvetica", 8)
        c.drawString(15*mm, y, "RESUMEN")
        y -= 5*mm
        c.setFillColor(colors.HexColor("#333333"))
        c.setFont("Helvetica", 9)
        # Word-wrap del resumen
        import textwrap
        for line in textwrap.wrap(summary, 105)[:4]:
            c.drawString(15*mm, y, line)
            y -= 4.5*mm
        y -= 3*mm

    # --- Fotos ---
    if photo_bytes:
        c.setFillColor(colors.HexColor("#888888"))
        c.setFont("Helvetica", 8)
        c.drawString(15*mm, y, "FOTOGRAFIAS")
        y -= 4*mm
        zonas = ["Frontal", "Trasera", "Lat. izq", "Lat. der"]
        img_w = (W - 30*mm - 9*mm) / 4
        img_h = img_w * 0.75
        ix = 15*mm
        for idx, pb in enumerate(photo_bytes[:4]):
            try:
                # Normalizar la imagen con Pillow (convierte a RGB JPEG limpio que reportlab acepta siempre)
                pil_img = Image.open(_io.BytesIO(pb))
                if pil_img.mode not in ("RGB", "L"):
                    pil_img = pil_img.convert("RGB")
                norm_buf = _io.BytesIO()
                pil_img.save(norm_buf, format="JPEG", quality=70)
                norm_buf.seek(0)
                img = ImageReader(norm_buf)
                c.drawImage(img, ix, y-img_h, img_w, img_h, preserveAspectRatio=True, anchor='c')
                c.setFillColor(colors.HexColor("#999999"))
                c.setFont("Helvetica", 6)
                c.drawCentredString(ix+img_w/2, y-img_h-3*mm, zonas[idx] if idx < len(zonas) else f"Foto {idx+1}")
            except Exception as _imgerr:
                logger.warning(f"PDF-DIAG error insertando foto {idx}: {_imgerr}")
            ix += img_w + 3*mm
        y -= (img_h + 8*mm)

    # --- Foto con daños numerados + leyenda (solo si se pide con ?boxes=1, porque la IA tarda ~15s) ---
    if photo_bytes and boxes:
        try:
            boxes = await _detectar_cajas_danos(photo_bytes[0])
            if boxes:
                annotated, n_boxes = _dibujar_numeros(photo_bytes[0], boxes)
                if y < 90*mm:
                    c.showPage(); y = H - 25*mm
                c.setFillColor(colors.HexColor("#888888"))
                c.setFont("Helvetica", 8)
                c.drawString(15*mm, y, "LOCALIZACION DE DAÑOS (IA)")
                y -= 5*mm
                # Imagen anotada (más grande, centrada)
                aimg = ImageReader(_io.BytesIO(annotated))
                aw = 110*mm
                ah = aw * 0.75
                c.drawImage(aimg, 15*mm, y-ah, aw, ah, preserveAspectRatio=True, anchor='nw')
                # Leyenda al lado derecho de la imagen
                lx = 15*mm + aw + 6*mm
                ly = y - 4*mm
                sev_hex = {"critico":"#A32D2D","grave":"#BA7517","moderado":"#BA7517","leve":"#639922"}
                idx_n = 0
                for b in boxes:
                    if len((b.get("box_2d") or b.get("box") or [])) != 4:
                        continue
                    idx_n += 1
                    c.setFillColor(colors.HexColor(sev_hex.get(b.get("severity",""), "#555555")))
                    c.circle(lx+2*mm, ly-1.2*mm, 2.2*mm, fill=1, stroke=0)
                    c.setFillColor(colors.white)
                    c.setFont("Helvetica-Bold", 7)
                    c.drawCentredString(lx+2*mm, ly-2.2*mm, str(idx_n))
                    c.setFillColor(colors.HexColor("#333333"))
                    c.setFont("Helvetica", 8)
                    lbl = str(b.get("label",""))[:32]
                    c.drawString(lx+6*mm, ly-2*mm, lbl)
                    ly -= 6*mm
                    if ly < y - ah:
                        break
                y -= (ah + 8*mm)
        except Exception as _ae:
            logger.warning(f"PDF foto numerada error: {_ae}")

    # --- Tabla de daños ---
    if damages:
        c.setFillColor(colors.HexColor("#888888"))
        c.setFont("Helvetica", 8)
        c.drawString(15*mm, y, "DAÑOS DETECTADOS")
        y -= 6*mm
        for d in damages[:12]:
            if y < 30*mm:
                c.showPage(); y = H - 25*mm
            c.setFillColor(colors.HexColor("#333333"))
            c.setFont("Helvetica", 9)
            part = str(d.get("part", ""))[:45]
            c.drawString(15*mm, y, part)
            dsev = d.get("severity", "")
            c.setFillColor(sev_colors.get(dsev, colors.HexColor("#666666")))
            c.setFont("Helvetica", 8)
            c.drawString(120*mm, y, dsev)
            c.setFillColor(colors.HexColor("#333333"))
            c.setFont("Helvetica", 9)
            c.drawRightString(W-15*mm, y, f"{d.get('estimated_cost',0):.0f} EUR")
            y -= 4*mm
            c.setStrokeColor(colors.HexColor("#eeeeee"))
            c.line(15*mm, y, W-15*mm, y)
            y -= 4*mm

    # --- Verificación matrícula ---
    y -= 4*mm
    if y < 25*mm:
        c.showPage(); y = H - 25*mm
    det_norm = (detected_plate or "").replace(" ", "").upper()
    plate_norm = (plate or "").replace(" ", "").upper()
    if detected_plate:
        coincide = det_norm == plate_norm
        c.setFillColor(colors.HexColor("#EAF3DE") if coincide else colors.HexColor("#FCEBEB"))
        c.roundRect(15*mm, y-8*mm, W-30*mm, 8*mm, 2*mm, fill=1, stroke=0)
        c.setFillColor(colors.HexColor("#3B6D11") if coincide else colors.HexColor("#A32D2D"))
        c.setFont("Helvetica", 8)
        txt = f"Verificacion matricula: leida '{detected_plate}' - " + ("COINCIDE" if coincide else "NO COINCIDE (posible fraude)")
        c.drawString(18*mm, y-5*mm, txt)
        y -= 12*mm

    # --- Pie ---
    c.setFillColor(colors.HexColor("#aaaaaa"))
    c.setFont("Helvetica", 7)
    c.drawCentredString(W/2, 12*mm, "Generado por FlotaDSP - Peritaje IA - Documento orientativo")

    c.showPage()
    c.save()
    buf.seek(0)
    return Response(content=buf.read(), media_type="application/pdf",
                    headers={"Content-Disposition": f"inline; filename=peritaje_{plate.replace(' ','')}_{fecha}.pdf"})





# =========================
# BÚSQUEDA DE RECAMBIOS — 7 marketplaces hiper-específicos
# =========================

def _slugify_es(text: str) -> str:
    """Convierte texto a slug URL-friendly para AutoDoc/RecambiosCoche."""
    import unicodedata, re
    text = unicodedata.normalize("NFD", text.lower())
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    text = re.sub(r"[^a-z0-9]+", "-", text).strip("-")
    return text


def _autodoc_brand_path(brand: str) -> str:
    """Mapea nombre de marca al slug de AutoDoc.es."""
    _map = {
        "TOYOTA": "toyota", "FORD": "ford", "PEUGEOT": "peugeot",
        "CITROEN": "citroen", "OPEL": "opel", "RENAULT": "renault",
        "VOLKSWAGEN": "volkswagen", "VW": "volkswagen",
        "MERCEDES": "mercedes-benz", "MERCEDES-BENZ": "mercedes-benz",
        "FIAT": "fiat", "IVECO": "iveco", "NISSAN": "nissan",
        "STELLANTIS": "peugeot",
    }
    return _map.get((brand or "").upper().strip(), _slugify_es(brand or "vehiculo"))


def _build_part_search_urls(
    part_name: str,
    brand: str = "",
    model: str = "",
    year: Optional[int] = None,
    vin: Optional[str] = None,
) -> list:
    """
    Construye 7 URLs de marketplace con máxima especificidad.

    Estrategia:
      - Query principal = '{pieza} {marca} {modelo} {año}' (todos los tokens)
      - Amazon: usa category automotive + nodo de recambios para filtrar solo piezas de coche
      - AutoDoc: si hay VIN → lookup directo /car/?vin=; si no → búsqueda filtrada
      - eBay: categoría 131090 (car parts) + si VIN disponible lo añade al query
      - RecambiosCoche / Desguaces: query exacto pieza+marca+modelo
      - Wallapop: categoría motor
      - Google Shopping: locale es + query completo con año
    """
    from urllib.parse import quote_plus, quote

    part_clean  = (part_name or "").strip()
    brand_clean = (brand  or "").strip().upper()
    model_clean = (model  or "").strip().upper()
    year_str    = str(year) if year else ""

    # ── Tokens ordenados para mayor especificidad ────────────────
    tokens_full = [t for t in [part_clean, brand_clean, model_clean, year_str] if t]
    tokens_pbm  = [t for t in [part_clean, brand_clean, model_clean] if t]
    tokens_pb   = [t for t in [part_clean, brand_clean] if t]

    q_full = quote_plus(" ".join(tokens_full))
    q_pbm  = quote_plus(" ".join(tokens_pbm))
    q_pb   = quote_plus(" ".join(tokens_pb))
    q_part = quote_plus(part_clean)

    brand_slug = _autodoc_brand_path(brand_clean)
    model_slug = _slugify_es(model_clean)

    # ── Amazon: nodo 599412031 = Coche y Moto > Recambios en amazon.es ──
    # Filtra SOLO la categoría de recambios, evita artículos de accesorios genéricos.
    amazon_node = "599412031"
    amazon_url = (
        f"https://www.amazon.es/s?k={q_full}"
        f"&i=automotive&bbn={amazon_node}&s=review-rank"
    )
    amazon_label = f"{part_clean} {brand_clean} {model_clean}" + (f" {year_str}" if year_str else "")

    # ── AutoDoc: VIN → lookup directo; si no → búsqueda por pieza+marca+modelo ──
    if vin:
        autodoc_url   = f"https://www.autodoc.es/car/?vin={quote(vin)}"
        autodoc_label = f"{part_clean} (VIN: {vin[:8]}…)"
        autodoc_spec  = "alta"
    else:
        autodoc_url = (
            f"https://www.autodoc.es/busqueda/"
            f"?sSearch={q_pbm}&sMake={quote_plus(brand_slug)}&sModel={quote_plus(model_slug)}"
        )
        autodoc_label = f"{part_clean} — {brand_clean} {model_clean}"
        autodoc_spec  = "alta" if (brand_clean and model_clean) else "media"

    # ── eBay: cat 131090 = Car Parts & Accessories ───────────────
    ebay_q = quote_plus(" ".join([t for t in [part_clean, brand_clean, model_clean, year_str, vin or ""] if t][:5]))
    ebay_url = f"https://www.ebay.es/sch/i.html?_nkw={ebay_q}&_sacat=131090&LH_PrefLoc=3"

    # ── Wallapop: categoría 100 (Motor) ──────────────────────────
    wallapop_url = f"https://es.wallapop.com/app/search?keywords={q_pbm}&category_ids=100&filters_source=quick_filters"

    # ── Desguaces Online ──────────────────────────────────────────
    desguaces_url = f"https://www.desguacesonline.com/buscador?busqueda={q_pbm}"

    # ── RecambiosCoche ────────────────────────────────────────────
    recambios_url = f"https://www.recambioscoche.es/buscar?q={q_pbm}"

    # ── Google Shopping: filtro España, ordenar por relevancia ───
    google_url = f"https://www.google.es/search?tbm=shop&q={q_full}&hl=es&gl=es&tbs=mr:1,merchagg:g116461169"

    return [
        {
            "marketplace": "Amazon",
            "url": amazon_url,
            "label": amazon_label,
            "description": "Recambios nuevos — sección Coche y Moto",
            "specificity": "alta",
            "icon": "📦",
        },
        {
            "marketplace": "AutoDoc",
            "url": autodoc_url,
            "label": autodoc_label,
            "description": "Especialista en recambios — lookup por VIN o modelo",
            "specificity": autodoc_spec,
            "icon": "🔧",
        },
        {
            "marketplace": "RecambiosCoche",
            "url": recambios_url,
            "label": f"{part_clean} — {brand_clean} {model_clean}",
            "description": "Recambios nuevos España",
            "specificity": "alta",
            "icon": "🚗",
        },
        {
            "marketplace": "eBay España",
            "url": ebay_url,
            "label": f"{part_clean} {brand_clean} {model_clean}",
            "description": "Nuevo + segunda mano",
            "specificity": "alta",
            "icon": "🛒",
        },
        {
            "marketplace": "Wallapop",
            "url": wallapop_url,
            "label": f"{part_clean} {brand_clean}",
            "description": "Segunda mano local",
            "specificity": "media",
            "icon": "♻️",
        },
        {
            "marketplace": "Desguaces Online",
            "url": desguaces_url,
            "label": f"{part_clean} {brand_clean} {model_clean}",
            "description": "Piezas de desguace",
            "specificity": "media",
            "icon": "🏭",
        },
        {
            "marketplace": "Google Shopping",
            "url": google_url,
            "label": f"{part_clean} {brand_clean} {model_clean}" + (f" {year_str}" if year_str else ""),
            "description": "Comparador de precios",
            "specificity": "alta",
            "icon": "🔍",
        },
    ]


@api_router.get("/parts/search")
async def search_parts(
    part_name: str,
    vehicle_id: Optional[str] = None,
    vin: Optional[str] = None,
    brand: Optional[str] = None,
    model: Optional[str] = None,
    year: Optional[int] = None,
    _=Depends(require_any_auth),
):
    """
    Devuelve 7 enlaces de marketplace para buscar una pieza concreta.

    Prioridad de especificidad:
      1. Si se pasa vehicle_id → carga VIN, marca, modelo y año del vehículo en BD.
      2. Si se pasan brand/model/year/vin directamente → los usa tal cual.
      3. Si hay VIN → AutoDoc hace lookup directo por VIN (máxima especificidad OEM).

    Todos los enlaces combinan marca + modelo + año + nombre de pieza para
    asegurar que los resultados sean de esa pieza exacta para ese vehículo.
    """
    _vin = vin
    _brand = brand
    _model = model
    _year = year

    if vehicle_id:
        v = await db.vehicles.find_one({"id": vehicle_id}, {"_id": 0})
        if v:
            _vin = _vin or v.get("vin") or v.get("chassis_number")
            _brand = _brand or v.get("brand", "")
            _model = _model or v.get("model", "")
            _year = _year or v.get("year") or v.get("fabrication_year")

    if not part_name or not part_name.strip():
        raise HTTPException(status_code=400, detail="Falta el nombre de la pieza (part_name)")

    links = _build_part_search_urls(
        part_name=part_name,
        brand=_brand or "",
        model=_model or "",
        year=_year,
        vin=_vin,
    )

    return {
        "part_name": part_name,
        "vehicle_info": {
            "brand": _brand,
            "model": _model,
            "year": _year,
            "vin": _vin,
        },
        "links": links,
        "total": len(links),
        "note": "URLs generadas con máxima especificidad. Si hay VIN disponible, AutoDoc hace búsqueda directa OEM.",
    }


# =========================
# LOOKUP PROVEEDOR POR MATRÍCULA
# =========================

@api_router.get("/vehicles/plate/{plate}/provider-info")
async def vehicle_provider_info_by_plate(plate: str, _=Depends(require_any_auth)):
    """
    Devuelve la información de proveedor de renting para una furgoneta por matrícula.

    Incluye:
      - Red de talleres del proveedor (teléfono, proceso, app)
      - Datos del vehículo en BD
      - Talleres concertados con ese proveedor en el mismo centro
    """
    plate_norm = re.sub(r'[^A-Z0-9]', '', plate.upper())
    if not plate_norm or len(plate_norm) > 20:
        raise HTTPException(400, "Matrícula inválida")
    # Buscar en BD por matrícula (normalizada)
    v = await db.vehicles.find_one(
        {"license_plate": {"$regex": re.escape(plate_norm), "$options": "i"}},
        {"_id": 0},
    )
    if not v:
        raise HTTPException(status_code=404, detail=f"Furgoneta con matrícula '{plate}' no encontrada")

    provider = v.get("provider", "")
    network = _provider_network_for(provider)
    center_code = _normalize_center_code(v.get("center"))

    # Talleres concertados con el proveedor en el mismo centro
    workshops_raw = await db.workshops.find(
        {"active": {"$ne": False}}, {"_id": 0}
    ).to_list(500)

    provider_workshops = []
    for w in workshops_raw:
        # Filtrar: mismo centro + convenio con el proveedor (no universales)
        if center_code and w.get("center") != center_code:
            continue
        convs = w.get("convenios", [])
        if "*" in convs:
            continue  # solo mostramos los específicos del proveedor aquí
        if provider and any(
            (provider.upper() in str(c).upper() or str(c).upper() in provider.upper())
            for c in convs
        ):
            provider_workshops.append({
                "id": w.get("id"),
                "name": w.get("name"),
                "address": w.get("address"),
                "city": w.get("city"),
                "phone": w.get("phone"),
                "hours": w.get("hours"),
                "categories": w.get("categories", []),
                "rating": w.get("rating"),
                "maps_url": w.get("maps_url"),
                "notes": w.get("notes"),
            })

    return {
        "license_plate": v.get("license_plate"),
        "vehicle": {
            "brand": v.get("brand"),
            "model": v.get("model"),
            "year": v.get("year"),
            "vin": v.get("vin") or v.get("chassis_number"),
            "center": v.get("center"),
            "provider": provider,
        },
        "provider_network": network,
        "provider_workshops_in_center": provider_workshops,
        "workshops_count": len(provider_workshops),
    }


# =========================
# ALQUILER DE FURGONETAS — directorio por centro (datos reales verificados)
# =========================

# Coordenadas aproximadas de cada centro logístico (para distancia)
_CENTER_COORDS = {
    "OGA5": (42.8782, -8.5448),   # Santiago de Compostela
    "DGA1": (43.3623, -8.4115),   # A Coruña (La Grela)
    "DGA2": (42.2328, -8.7226),   # Vigo
}

_SEED_RENTALS = [
    # ── SANTIAGO CIUDAD (OGA5) — radio ~5 km ──
    {"name": "Iberfurgo Santiago", "center": "OGA5",
     "address": "Rúa de Xoán XXIII, 3, 15703 Santiago de Compostela", "phone": "679954668",
     "email": "santiago@iberfurgo.com",
     "website": "https://www.iberfurgo.com/oficinas/alquiler-furgonetas-santiago-compostela/",
     "latitude": 42.8757, "longitude": -8.5470,
     "logo_domain": "iberfurgo.com", "brand_color": "#E84023",
     "notes": "Furgonetas y camiones. Desde ~30 €/día. Asistencia 24/7."},
    {"name": "OneFurgo Santiago", "center": "OGA5",
     "address": "Rúa de Casas Reais, 20, 15703 Santiago de Compostela", "phone": "900829339",
     "email": "info@onefurgo.com",
     "website": "https://onefurgo.com/red-de-oficinas/alquiler-de-furgonetas-baratas-en-santiago-de-compostela",
     "latitude": 42.8808, "longitude": -8.5453,
     "logo_domain": "onefurgo.com", "brand_color": "#FF6B00",
     "notes": "Furgonetas de carga, pasajeros y carrozadas. Reserva online."},
    {"name": "Hello Rentacar Santiago", "center": "OGA5",
     "address": "Avenida de Lugo, 117, 15703 Santiago de Compostela", "phone": "881972226",
     "email": "",
     "website": "https://www.hellorentacar.es/alquiler-furgonetas/galicia/santiago-compostela/",
     "latitude": 42.8803, "longitude": -8.5338,
     "logo_domain": "hellorentacar.es", "brand_color": "#00B140",
     "notes": "Coches y furgonetas en el centro de Santiago."},
    {"name": "GoRental Santiago", "center": "OGA5",
     "address": "Rúa do Hórreo, 76, 15702 Santiago de Compostela", "phone": "981573993",
     "email": "",
     "website": "http://www.gorental.es/",
     "latitude": 42.8726, "longitude": -8.5442,
     "logo_domain": "gorental.es", "brand_color": "#0073CF",
     "notes": "Alquiler local de vehículos comerciales."},
    {"name": "Europcar Santiago Ciudad", "center": "OGA5",
     "address": "Rúa do Hórreo, 24, 15702 Santiago de Compostela", "phone": "981563668",
     "email": "",
     "website": "https://www.europcar.es/location/spain/santiago-de-compostela/santiago-downtown",
     "latitude": 42.8721, "longitude": -8.5437,
     "logo_domain": "europcar.com", "brand_color": "#009A44",
     "notes": "Oficina en el centro. Coches y furgonetas. L-V 8:00-20:00, S 9:00-13:00."},

    # ── AEROPUERTO SCQ — 14 km del centro, todos dentro del radio de 20 km ──
    {"name": "Hertz — Aeropuerto Santiago", "center": "OGA5",
     "address": "Aeropuerto SCQ, Terminal, 15820 Lavacolla", "phone": "981591323",
     "email": "",
     "website": "https://www.hertz.es/p/alquiler-de-furgonetas/espana/santiago-de-compostela",
     "latitude": 42.8963, "longitude": -8.4151,
     "logo_domain": "hertz.com", "brand_color": "#FFD100",
     "notes": "Mostrador en terminal de llegadas. Coches y furgonetas."},
    {"name": "Europcar — Aeropuerto Santiago", "center": "OGA5",
     "address": "Aeropuerto SCQ, Terminal, 15820 Lavacolla", "phone": "981591861",
     "email": "",
     "website": "https://www.europcar.es/location/spain/santiago-de-compostela/santiago-airport",
     "latitude": 42.8968, "longitude": -8.4147,
     "logo_domain": "europcar.com", "brand_color": "#009A44",
     "notes": "Mostrador en terminal. Recogida y entrega 24h con reserva."},
    {"name": "Sixt — Aeropuerto Santiago", "center": "OGA5",
     "address": "Aeropuerto SCQ, Terminal, 15820 Lavacolla", "phone": "932719000",
     "email": "",
     "website": "https://www.sixt.es/alquiler-de-coches/espana/santiago-de-compostela/aeropuerto-sgo/",
     "latitude": 42.8965, "longitude": -8.4148,
     "logo_domain": "sixt.com", "brand_color": "#FF5F00",
     "notes": "Mostrador en terminal de llegadas. Coches y SUVs."},
    {"name": "Avis — Aeropuerto Santiago", "center": "OGA5",
     "address": "Aeropuerto SCQ, Terminal, 15820 Lavacolla", "phone": "981597648",
     "email": "",
     "website": "https://www.avis.es/es/locations/es/sco",
     "latitude": 42.8961, "longitude": -8.4153,
     "logo_domain": "avis.com", "brand_color": "#C8102E",
     "notes": "Mostrador en terminal. Reserva online con tarifa garantizada."},
    {"name": "Budget — Aeropuerto Santiago", "center": "OGA5",
     "address": "Aeropuerto SCQ, Terminal, 15820 Lavacolla", "phone": "981597648",
     "email": "",
     "website": "https://www.budget.es/brs/carHireSearch.do?selectedOffice=SCQT01",
     "latitude": 42.8962, "longitude": -8.4152,
     "logo_domain": "budget.com", "brand_color": "#E4002B",
     "notes": "Grupo Avis. Tarifas económicas para reservas anticipadas."},
    {"name": "Record Go — Aeropuerto Santiago", "center": "OGA5",
     "address": "Aeropuerto SCQ, P. Exterior, 15820 Lavacolla", "phone": "902500905",
     "email": "",
     "website": "https://www.record-go.com/alquiler-coches/santiago-de-compostela-aeropuerto",
     "latitude": 42.8970, "longitude": -8.4160,
     "logo_domain": "record-go.com", "brand_color": "#E30613",
     "notes": "Parking exterior con lanzadera. Precios muy competitivos."},

    # ── A CORUÑA (DGA1) — ~65 km, fuera del radio Santiago ──
    {"name": "Iberfurgo A Coruña", "center": "DGA1",
     "address": "C/ Gutemberg 38A, P.I. La Grela, 15008 A Coruña", "phone": "698139597",
     "email": "",
     "website": "https://www.iberfurgo.com/oficinas/alquiler-furgonetas-coruna/",
     "latitude": 43.3228, "longitude": -8.4472,
     "logo_domain": "iberfurgo.com", "brand_color": "#E84023",
     "notes": "L-V 8:00-13:30 y 16:00-20:30 · Sáb 9:00-13:00 · Dom/festivos cita previa."},
    {"name": "OneFurgo A Coruña", "center": "DGA1",
     "address": "Carretera Pocomaco, S/N, 15190 A Coruña", "phone": "900829339",
     "email": "info@onefurgo.com",
     "website": "https://onefurgo.com/red-de-oficinas/a-coruna",
     "latitude": 43.3402, "longitude": -8.3793,
     "logo_domain": "onefurgo.com", "brand_color": "#FF6B00",
     "notes": "Furgonetas sin conductor para empresas y particulares. Reserva online."},
    {"name": "Europcar A Coruña", "center": "DGA1",
     "address": "Rúa Federico Tapia, 30, 15005 A Coruña", "phone": "981233397",
     "email": "",
     "website": "https://www.europcar.es/location/spain/a-coruna",
     "latitude": 43.3623, "longitude": -8.4115,
     "logo_domain": "europcar.com", "brand_color": "#009A44",
     "notes": "Oficina en el centro de A Coruña. L-V 8:30-13:30 / 16:00-19:30."},

    # ── VIGO (DGA2) — ~90 km, fuera del radio Santiago ──
    {"name": "OneFurgo Vigo", "center": "DGA2",
     "address": "Camiño Gandariña, 21, Lavadores, 36214 Vigo", "phone": "986933464",
     "email": "info@onefurgo.com",
     "website": "https://onefurgo.com/red-de-oficinas/vigo",
     "latitude": 42.2248, "longitude": -8.7217,
     "logo_domain": "onefurgo.com", "brand_color": "#FF6B00",
     "notes": "Carga, pasajeros y carrozadas. Reserva online."},
    {"name": "Iberfurgo Vigo", "center": "DGA2",
     "address": "Autovía de Madrid, 234 - Nave 4B, 36318 Vigo", "phone": "608096307",
     "email": "",
     "website": "https://www.iberfurgo.com/oficinas/alquiler-furgonetas-vigo/",
     "latitude": 42.1950, "longitude": -8.6600,
     "logo_domain": "iberfurgo.com", "brand_color": "#E84023",
     "notes": "Alquiler por días y renting por meses. Flota nueva. Asistencia 24/7."},
    {"name": "Europcar Vigo", "center": "DGA2",
     "address": "Avda. de Madrid, 7, 36204 Vigo", "phone": "986439282",
     "email": "",
     "website": "https://www.europcar.es/location/spain/vigo",
     "latitude": 42.2328, "longitude": -8.7114,
     "logo_domain": "europcar.com", "brand_color": "#009A44",
     "notes": "Oficina en el centro de Vigo."},
]


@app.on_event("startup")
async def seed_rental_companies():
    """Siembra y actualiza el directorio de empresas de alquiler (upsert por nombre)."""
    try:
        for r in _SEED_RENTALS:
            existing = await db.rental_companies.find_one({"name": r["name"]})
            if existing:
                # Actualiza coordenadas y campos nuevos sin tocar id/created_at
                await db.rental_companies.update_one(
                    {"name": r["name"]},
                    {"$set": {k: v for k, v in r.items() if k not in ("name",)}}
                )
            else:
                doc = dict(r)
                doc["id"] = str(uuid.uuid4())
                doc["active"] = True
                doc["last_check"] = None
                doc["created_at"] = datetime.now(timezone.utc).isoformat()
                await db.rental_companies.insert_one(doc)
        logger.info(f"Seed rentals: {len(_SEED_RENTALS)} empresas sincronizadas")
    except Exception as e:
        logger.error(f"Seed rentals: {e}")


def _haversine_km(c1, c2):
    import math
    if not c1 or not c2:
        return None
    lat1, lon1 = c1
    lat2, lon2 = c2
    R = 6371
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1))*math.cos(math.radians(lat2))*math.sin(dlon/2)**2
    return round(R * 2 * math.atan2(math.sqrt(a), math.sqrt(1-a)), 1)


@api_router.get("/rentals/nearby")
async def rentals_nearby(
    lat: float,
    lng: float,
    max_km: float = 20,
    limit: int = 20,
    current_user: dict = Depends(require_any_auth),
):
    """
    Devuelve casas de alquiler ordenadas por distancia al punto (lat, lng).
    Las que no tienen coordenadas se añaden al final ordenadas por nombre.
    """
    docs = await db.rental_companies.find({"active": {"$ne": False}}, {"_id": 0}).to_list(200)

    with_coords, without_coords = [], []
    for d in docs:
        if d.get("latitude") is not None and d.get("longitude") is not None:
            dist = _haversine_km((lat, lng), (d["latitude"], d["longitude"])) or 9999
            if dist <= max_km:
                d["distance_km"] = round(dist, 1)
                with_coords.append(d)
        else:
            without_coords.append(d)

    with_coords.sort(key=lambda x: x["distance_km"])
    without_coords.sort(key=lambda x: x.get("name", ""))

    return {"rentals": (with_coords + without_coords)[:limit], "user_lat": lat, "user_lng": lng}


@api_router.get("/rentals")
async def list_rentals(center: Optional[str] = None, _=Depends(require_admin)):
    """Empresas de alquiler de furgonetas, opcionalmente filtradas por centro."""
    query = {"active": {"$ne": False}}
    if center and center != "Todos":
        query["center"] = center
    docs = await db.rental_companies.find(query, {"_id": 0}).to_list(200)
    docs.sort(key=lambda d: d.get("name", ""))
    return docs


@api_router.post("/rentals")
async def create_rental(data: dict, _=Depends(require_admin)):
    """Añade una empresa de alquiler manualmente."""
    name = (data.get("name") or "").strip()
    center = (data.get("center") or "").strip()
    if not name or center not in ("OGA5", "DGA1", "DGA2"):
        raise HTTPException(status_code=400, detail="Nombre y centro (OGA5/DGA1/DGA2) requeridos")
    doc = {
        "id": str(uuid.uuid4()),
        "name": name, "center": center,
        "address": (data.get("address") or "").strip(),
        "phone": re.sub(r"[^0-9+]", "", data.get("phone") or ""),
        "email": (data.get("email") or "").strip(),
        "website": (data.get("website") or "").strip(),
        "notes": (data.get("notes") or "").strip(),
        "maps_url": "https://www.google.com/maps/search/?api=1&query=" + ((data.get("address") or name).replace(" ", "+")),
        "active": True, "last_check": None,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.rental_companies.insert_one(doc)
    doc.pop("_id", None)
    return {"success": True, "id": doc["id"]}


@api_router.patch("/rentals/{rental_id}")
async def update_rental(rental_id: str, data: dict, _=Depends(require_admin)):
    """Edita una empresa o registra una verificación de disponibilidad."""
    data.pop("_id", None)
    data.pop("id", None)
    await db.rental_companies.update_one({"id": rental_id}, {"$set": data})
    return {"success": True}


@api_router.post("/rentals/{rental_id}/check")
async def verify_rental_availability(rental_id: str, data: dict, user: dict = Depends(get_current_user)):
    """Registra disponibilidad verificada por el equipo (tras llamar)."""
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores")
    check = {
        "date": datetime.now(timezone.utc).isoformat(),
        "by": user.get("name", "?"),
        "available": data.get("available"),          # nº de furgonetas o texto
        "note": (data.get("note") or "").strip()[:200],
    }
    result = await db.rental_companies.update_one(
        {"id": rental_id}, {"$set": {"last_check": check}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Empresa no encontrada")
    return {"success": True, "last_check": check}


@api_router.delete("/rentals/{rental_id}")
async def delete_rental(rental_id: str, _=Depends(require_admin)):
    await db.rental_companies.update_one({"id": rental_id}, {"$set": {"active": False}})
    return {"success": True}



# =========================
# MÉTRICAS AMAZON — análisis de Scorecard y reports con IA
# =========================

_DSP_ANALYST_PROMPT = """Eres un analista senior de operaciones de un Amazon DSP (Delivery Service Partner) en España, experto en el DSP Scorecard 3.0 y todos los reports semanales que Amazon entrega.

Conoces a fondo estas métricas y sus objetivos:
- SAFETY: Safe Driving Score (Mentor/FICO), Seatbelt-Off Rate, Speeding Event Rate, Distractions Rate, Following Distance Rate, Sign/Signal Violations. (Objetivo: bajos / verde)
- COMPLIANCE: Comprehensive Audit Score (CAS), Working Hours Compliance (WHC), Vehicle Inspection (DVIC pre/post), Comprehensive DVIC.
- QUALITY: Delivery Completion Rate DCR (objetivo ≥98-99%), Delivered Not Received DNR DPMO (objetivo bajo, <1500), Photo-On-Delivery / RTS quality, Contact Compliance (objetivo bajo), Customer Delivery Feedback CDF / Concessions.
- Tier general: Fantastic+ > Fantastic > Great > Fair > Poor > At Risk.

Te paso el contenido de un report real de un DSP. Analízalo a fondo desde la óptica de un JEFE DE TRÁFICO / DISPATCHER que necesita saber QUÉ está mal, QUIÉN falla y QUÉ hacer esta semana.

Responde SOLO con este JSON (sin markdown):
{
  "report_type": "tipo detectado (Scorecard 3.0, Daily Report, POD Quality, CDF/Concessions, DNR Investigations, Contact Compliance, etc.)",
  "period": "semana o fecha del report si aparece",
  "overall_tier": "tier general si es un scorecard, o vacío",
  "headline": "una frase de titular: lo más importante de este report",
  "key_metrics": [
    {"name": "nombre métrica", "value": "valor actual", "target": "objetivo", "status": "good|warning|bad", "trend": "up|down|flat|"}
  ],
  "drivers_at_risk": [
    {
      "name": "nombre del conductor tal cual aparece",
      "metric": "métrica afectada (DNR, POD, Speeding…)",
      "detail": "dato concreto (ej: 3 DNR, 5 speeding events)",
      "what": "QUÉ falla, en una frase clara",
      "why": "POR QUÉ suele pasar esto (causa raíz típica)",
      "work_on": "EN QUÉ debe trabajar concretamente",
      "advice": "CONSEJO accionable para que mejore",
      "how_to_explain": "CÓMO explicárselo al conductor de forma motivadora y sin que se ponga a la defensiva (tono de coaching, frase lista para decirle)"
    }
  ],
  "top_issues": ["problema 1 priorizado", "problema 2", "..."],
  "recommendations": [
    {"priority": "alta|media|baja", "action": "acción concreta y accionable", "for_whom": "dispatcher|conductor concreto|coordinador|flota"}
  ],
  "executive_summary": "3-4 frases para el jefe de tráfico: estado general, qué está en riesgo y qué priorizar."
}

Para drivers_at_risk, eres un COACH de conductores: rellena what/why/work_on/advice/how_to_explain de forma concreta y humana. El campo how_to_explain debe ser una frase que el jefe de tráfico pueda decirle tal cual al conductor, en tono positivo (reconoce lo bueno, señala lo concreto a mejorar, da un objetivo claro). Sé concreto. Si el report no menciona conductores individuales, deja drivers_at_risk vacío. Usa los nombres EXACTOS que aparezcan."""


async def _extract_report_text(content: bytes, filename: str):
    """Devuelve (texto_o_None, pdf_bytes_o_None). Para PDF pasamos los bytes a Gemini;
    para HTML/XLSX/CSV extraemos texto plano."""
    fn = (filename or "").lower()
    if fn.endswith(".pdf"):
        return None, content
    if fn.endswith(".xlsx") or fn.endswith(".xlsm"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            out = []
            for ws in wb.worksheets:
                out.append(f"### Hoja: {ws.title}")
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) for c in row if c is not None]
                    if cells:
                        out.append(" | ".join(cells))
            return "\n".join(out)[:60000], None
        except Exception as e:
            return f"(No se pudo leer el Excel: {e})", None
    # HTML / CSV / TXT → quitar tags
    try:
        txt = content.decode("utf-8", errors="ignore")
    except Exception:
        txt = str(content[:50000])
    if fn.endswith(".html") or "<html" in txt[:500].lower() or "<table" in txt.lower():
        txt = re.sub(r"<script[\s\S]*?</script>", " ", txt, flags=re.I)
        txt = re.sub(r"<style[\s\S]*?</style>", " ", txt, flags=re.I)
        txt = re.sub(r"<[^>]+>", " ", txt)
        txt = re.sub(r"&nbsp;", " ", txt)
        txt = re.sub(r"[ \t]+", " ", txt)
        txt = re.sub(r"\n\s*\n+", "\n", txt)
    return txt[:60000], None


async def _analyze_report_with_gemini(text, pdf_bytes, driver_map=None):
    from google import genai as genai_sdk
    from google.genai import types as genai_types
    use_vertex = os.environ.get("USE_VERTEX_AI", "").lower() in ("1", "true", "yes")
    if use_vertex:
        from google.oauth2 import service_account
        import json as _json, base64 as _b64
        sa = os.environ.get("GCP_SERVICE_ACCOUNT_JSON", "").strip()
        if sa and not sa.startswith("{"):
            sa = _b64.b64decode(sa).decode("utf-8")
        creds = service_account.Credentials.from_service_account_info(
            _json.loads(sa), scopes=["https://www.googleapis.com/auth/cloud-platform"]) if sa else None
        client = genai_sdk.Client(vertexai=True, project=os.environ.get("GCP_PROJECT", ""),
                                  location=os.environ.get("GCP_LOCATION", "us-central1"), credentials=creds)
    else:
        client = genai_sdk.Client(api_key=os.environ.get("GEMINI_API_KEY", ""))

    prompt = _DSP_ANALYST_PROMPT
    if driver_map:
        tabla = "\n".join(f"  {tid} = {nombre}" for tid, nombre in list(driver_map.items())[:300])
        prompt += ("\n\n=== TABLA DE CONDUCTORES (Amazon Transporter ID = Nombre real) ===\n"
                   "Cuando en el report aparezca un ID de conductor de esta tabla, usa SIEMPRE el NOMBRE REAL "
                   "en el campo 'name' (no el ID). Tabla:\n" + tabla)
    contents = [prompt]
    if pdf_bytes:
        contents.append(genai_types.Part.from_bytes(data=pdf_bytes, mime_type="application/pdf"))
    else:
        contents.append("CONTENIDO DEL REPORT:\n\n" + (text or ""))

    cfg = genai_types.GenerateContentConfig(temperature=0.1, response_mime_type="application/json")
    loop = asyncio.get_running_loop()
    async with _gemini_sem:
        resp = await asyncio.wait_for(
            loop.run_in_executor(_executor, lambda: client.models.generate_content(
                model=os.environ.get("GEMINI_MODEL", "gemini-2.5-flash"), contents=contents, config=cfg)),
            timeout=90.0)
    return json.loads(_strip_markdown_json(resp.text or "{}"))


def _parse_routes_excel(content: bytes, driver_info, snapshot_name=None):
    """Parsea el Excel 'Rutas' de Cortex a estructura operativa. driver_map_inv:
    {transporter_id_upper: nombre_real_BD} para mostrar el nombre de la ficha si existe.
    Devuelve dict con kpis + lista de rutas, o None si no es ese formato."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return None
        headers = [str(c or "").strip().lower() for c in rows[0]]

        def col(*names):
            for n in names:
                for i, hd in enumerate(headers):
                    if n in hd:
                        return i
            return None

        ci_code = col("código de ruta", "codigo de ruta", "route code")
        ci_id = col("id del transportista", "transporter")
        ci_name = col("nombre del conductor", "driver name")
        ci_prog = col("progreso de la ruta", "estado del progreso", "route progress")
        ci_service = col("tipo de servicio")
        ci_allstops = col("todas las paradas", "total stops")
        ci_done = col("paradas completadas", "completed")
        ci_notstarted = col("paradas no iniciadas", "not started")
        ci_deliv = col("entregas totales", "total deliveries")
        ci_dep = col("hora de salida")
        # ── Columnas extra del export de Itinerarios/DAs (datos REALES de Amazon) ──
        ci_phone = col("número de teléfono", "numero de telefono", "phone")
        ci_pace = col("avg_pace", "pace_stops_per_hour")           # ritmo real de Amazon
        ci_return = col("regreso previsto a la estación", "regreso previsto")  # cierre real
        ci_battery = col("% de batería", "bateria")
        ci_helper = col("ayudante")
        ci_login = col("inicio de sesión", "inicio de sesion")
        ci_realdep = col("salida real")
        ci_totalpkg = col("total de paquetes")
        if (ci_code is None and ci_id is None) or ci_name is None:
            return None  # no es ninguno de los Excel de rutas

        # Momento del snapshot: preferir la hora del nombre del archivo
        # (ej "Rutas_OGA5_2026-06-13_13_23 (GMT+2).xlsx" → 13:23). Si no, hora actual.
        from zoneinfo import ZoneInfo as _ZI
        now_es = datetime.now(_ZI("Europe/Madrid"))
        snap_h = None
        snap_label = now_es.strftime("%d/%m/%Y %H:%M")
        if snapshot_name:
            fm = re.search(r"(\d{4})-(\d{2})-(\d{2})[_ ](\d{2})[_:](\d{2})", snapshot_name)
            if fm:
                snap_h = int(fm.group(4)) + int(fm.group(5)) / 60.0
                snap_label = f"{fm.group(3)}/{fm.group(2)}/{fm.group(1)} {fm.group(4)}:{fm.group(5)}"
        now_h = snap_h if snap_h is not None else (now_es.hour + now_es.minute / 60.0)
        SHIFT_HOURS = 9.0  # jornada: cierre = hora de salida + 9h

        def _parse_hour(s):
            m = re.match(r"(\d{1,2}):(\d{2})", str(s or ""))
            return int(m.group(1)) + int(m.group(2)) / 60.0 if m else None

        routes = []
        for r in rows[1:]:
            if not r or all(c is None for c in r):
                continue
            tid = (str(r[ci_id]).strip().upper() if ci_id is not None and r[ci_id] else "")
            name = (str(r[ci_name]).strip() if ci_name is not None and r[ci_name] else "")
            info = driver_info.get(tid) if tid else None
            real = info.get("name") if info else None
            # Teléfono: PRIMERO el del Excel (con el que trabaja hoy), luego el de la ficha
            phone_xl = re.sub(r"[^0-9+]", "", str(r[ci_phone])) if ci_phone is not None and r[ci_phone] and str(r[ci_phone]).lower() != "falta" else ""
            phone = phone_xl or (info.get("phone") if info else "") or ""
            allstops = int(r[ci_allstops]) if ci_allstops is not None and isinstance(r[ci_allstops], (int, float)) else 0
            done = int(r[ci_done]) if ci_done is not None and isinstance(r[ci_done], (int, float)) else 0
            prog_raw = (str(r[ci_prog]).strip().upper() if ci_prog is not None and r[ci_prog] else "")
            status = "bad" if "BEHIND" in prog_raw else "good" if ("AHEAD" in prog_raw or "ON TIME" in prog_raw) else "ok"
            pct = round(done / allstops * 100) if allstops else 0
            remaining = max(0, allstops - done)
            dep = str(r[ci_dep]).strip() if ci_dep is not None and r[ci_dep] else ""

            # Datos REALES de Amazon (export Itinerarios) si están disponibles
            amz_pace = None
            if ci_pace is not None and isinstance(r[ci_pace], (int, float)):
                amz_pace = round(float(r[ci_pace]), 1)
            return_str = str(r[ci_return]).strip() if ci_return is not None and r[ci_return] and str(r[ci_return]).lower() != "falta" else ""
            battery = None
            if ci_battery is not None and isinstance(r[ci_battery], (int, float)):
                battery = round(float(r[ci_battery]))
            helper = str(r[ci_helper]).strip() if ci_helper is not None and r[ci_helper] and str(r[ci_helper]).lower() != "falta" else ""
            total_pkg = int(r[ci_totalpkg]) if ci_totalpkg is not None and isinstance(r[ci_totalpkg], (int, float)) else 0

            # ── Predicción de RESCATE EN PARADAS ──
            # Regla fija de la empresa: jornada de 9h desde que SALE (salida real
            # si está, si no la planificada). El cierre de Amazon se ignora.
            real_dep = str(r[ci_realdep]).strip() if ci_realdep is not None and r[ci_realdep] and str(r[ci_realdep]).lower() != "falta" else ""
            dep_h = _parse_hour(real_dep) or _parse_hour(dep)
            cutoff_h = (dep_h + SHIFT_HOURS) if dep_h is not None else None
            cutoff_label = None
            if cutoff_h is not None:
                ch, cm = int(cutoff_h) % 24, int(round((cutoff_h - int(cutoff_h)) * 60))
                if cm == 60:
                    ch, cm = (ch + 1) % 24, 0
                cutoff_label = f"{ch:02d}:{cm:02d}"
            rate = amz_pace  # ritmo real de Amazon (paradas/hora)
            eta = None
            rescue = None          # PARADAS que no le dará tiempo de hacer en sus 9h
            will_finish = None
            if rate is None and dep_h is not None and done > 0 and now_h > dep_h:
                worked = now_h - dep_h
                if worked >= 0.25:
                    rate = round(done / worked, 1)
            if rate and rate > 0:
                hours_left_route = remaining / rate
                eta_h = now_h + hours_left_route
                eh = int(eta_h) % 24
                em = int(round((eta_h - int(eta_h)) * 60))
                if em == 60:
                    eh, em = (eh + 1) % 24, 0
                eta = f"{eh:02d}:{em:02d}"
                if cutoff_h is not None:
                    hours_to_cutoff = max(0, cutoff_h - now_h)
                    can_do = rate * hours_to_cutoff            # paradas que hará antes de su cierre
                    rescue = max(0, round(remaining - can_do)) # paradas de AYUDA necesarias
                    will_finish = eta_h <= cutoff_h

            routes.append({
                "code": str(r[ci_code]).strip() if ci_code is not None and r[ci_code] else "",
                "transporter_id": tid,
                "driver_name": real or name,
                "from_db": bool(real),
                "phone": phone or "",
                "service": str(r[ci_service]).strip() if ci_service is not None and r[ci_service] else "",
                "progress_label": prog_raw,
                "status": status,
                "stops_total": allstops,
                "stops_done": done,
                "stops_remaining": remaining,
                "stops_pct": pct,
                "deliveries": int(r[ci_deliv]) if ci_deliv is not None and isinstance(r[ci_deliv], (int, float)) else 0,
                "departure": dep,
                "cutoff": cutoff_label,
                "return_planned": return_str,
                "rate": rate,
                "rate_source": "amazon" if amz_pace is not None else "estimado",
                "eta": eta,
                "rescue": rescue,
                "will_finish": will_finish,
                "battery": battery,
                "helper": helper,
                "total_packages": total_pkg,
            })

        # ordenar: primero quien más rescate necesita, luego atrasados, luego % bajo
        routes.sort(key=lambda x: (-(x["rescue"] or 0), 0 if x["status"] == "bad" else 1, x["stops_pct"]))

        behind = sum(1 for x in routes if x["status"] == "bad")
        total_stops = sum(x["stops_total"] for x in routes)
        done_stops = sum(x["stops_done"] for x in routes)
        total_rescue = sum((x["rescue"] or 0) for x in routes)
        need_help = [x for x in routes if (x["rescue"] or 0) > 0]
        return {
            "report_type": "Rutas en vivo",
            "period": snap_label,
            "is_routes": True,
            "snapshot_hour": round(now_h, 2),
            "kpis": {
                "total_routes": len(routes),
                "behind": behind,
                "on_track": len(routes) - behind,
                "stops_done": done_stops,
                "stops_total": total_stops,
                "global_pct": round(done_stops / total_stops * 100) if total_stops else 0,
                "total_rescue": total_rescue,
                "routes_need_help": len(need_help),
            },
            "routes": routes,
        }
    except Exception as e:
        logger.warning(f"parse routes excel: {e}")
        return None


@api_router.post("/metrics/upload-routeplan")
async def upload_route_plan(file: UploadFile = File(...), center: str = Form("OGA5"), _=Depends(require_admin)):
    """Sube el archivo de la mañana (sequenced routes / CYCLE) con las paradas reales
    de cada ruta: dirección, código postal, GPS y hora planificada por Amazon.
    Guarda un documento por ruta para luego mostrar el mapa y el progreso geográfico."""
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Archivo vacío")
    fn = (file.filename or "").lower()
    # Lectura universal: .xls (xlrd) y .xlsx (openpyxl) → dict {sheet: [rows]}
    sheet_rows = {}
    try:
        if fn.endswith(".xls"):
            import xlrd
            book = xlrd.open_workbook(file_contents=content)
            for shn in book.sheet_names():
                sh = book.sheet_by_name(shn)
                sheet_rows[shn] = [sh.row_values(i) for i in range(sh.nrows)]
        else:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            for shn in wb.sheetnames:
                sheet_rows[shn] = [list(r) for r in wb[shn].iter_rows(values_only=True)]
    except Exception as _re:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el archivo: {_re}")

    sheets = [s for s in sheet_rows if s.lower().startswith("sequencedroute")]
    if not sheets:
        raise HTTPException(status_code=400, detail="No es el archivo de rutas secuenciadas (CYCLE). Sube el que te llega por la mañana.")
    try:

        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        # Limpiar plan anterior del mismo día/centro
        await db.route_plans.delete_many({"date": today, "center": center})

        saved = 0
        total_stops = 0
        for sn in sheets:
            rows = sheet_rows[sn]
            if len(rows) < 3:
                continue
            code = sn.replace("sequencedRoute_", "").replace("sequencedroute_", "")
            info = str(rows[0][0] or "")
            dist_m = re.search(r"([\d.]+)\s*km", info)
            distance_km = float(dist_m.group(1)) if dist_m else None
            plan_m = re.search(r"Route plan:\s*([^R]+?)(?:Route time|Total|$)", str(rows[0][4] or ""))
            plan_time = plan_m.group(1).strip() if plan_m else ""

            hdr = [str(c or "").strip().lower() for c in rows[1]]

            def hc(*names):
                for n in names:
                    for i, x in enumerate(hdr):
                        if n in x:
                            return i
                return None

            ci_stop = hc("stop")
            ci_track = hc("tracking")
            ci_arr = hc("arrival")
            ci_win = hc("time window")
            ci_addr = hc("customer address", "address")
            ci_post = hc("postal")
            ci_lat = hc("latitude")
            ci_lon = hc("longitude")
            ci_zone = hc("zone")

            stops = []
            for r in rows[2:]:
                if not r:
                    continue
                track = r[ci_track] if ci_track is not None else None
                if not track:  # saltar el depot (sin tracking)
                    continue
                lat = r[ci_lat] if ci_lat is not None else None
                lon = r[ci_lon] if ci_lon is not None else None
                stops.append({
                    "seq": int(r[ci_stop]) if ci_stop is not None and isinstance(r[ci_stop], (int, float)) else None,
                    "tracking": str(track),
                    "arrival": str(r[ci_arr]).strip() if ci_arr is not None and r[ci_arr] else "",
                    "window": str(r[ci_win]).strip() if ci_win is not None and r[ci_win] else "",
                    "address": str(r[ci_addr]).strip() if ci_addr is not None and r[ci_addr] else "",
                    "postal": str(r[ci_post]).strip() if ci_post is not None and r[ci_post] else "",
                    "lat": float(lat) if isinstance(lat, (int, float)) else None,
                    "lon": float(lon) if isinstance(lon, (int, float)) else None,
                    "zone": str(r[ci_zone]).strip() if ci_zone is not None and r[ci_zone] else "",
                })
            if not stops:
                continue
            await db.route_plans.insert_one({
                "id": str(uuid.uuid4()),
                "date": today, "center": center, "code": code,
                "distance_km": distance_km, "plan_time": plan_time,
                "stops_count": len(stops), "stops": stops,
                "created_at": datetime.now(timezone.utc).isoformat(),
            })
            saved += 1
            total_stops += len(stops)

        return {"success": True, "routes": saved, "stops": total_stops, "date": today}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Route plan: {e}")
        raise HTTPException(status_code=502, detail=f"No se pudo procesar el archivo: {e}")


@api_router.get("/metrics/routeplan")
async def get_route_plan(code: str, center: str = "OGA5", date: Optional[str] = None, _=Depends(require_admin)):
    """Devuelve las paradas de una ruta (la más reciente o de una fecha)."""
    q = {"center": center, "code": code}
    if date:
        q["date"] = date
    doc = await db.route_plans.find_one(q, {"_id": 0}, sort=[("date", -1)])
    if not doc:
        raise HTTPException(status_code=404, detail="No hay plan para esa ruta. Sube el archivo de la mañana.")
    return doc


@api_router.get("/metrics/routeplan-available")
async def routeplan_available(center: str = "OGA5", _=Depends(require_admin)):
    """Lista de códigos de ruta con plan disponible hoy (para enlazar el mapa)."""
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    docs = await db.route_plans.find({"center": center, "date": today}, {"_id": 0, "code": 1, "stops_count": 1, "distance_km": 1}).to_list(200)
    return {"date": today, "routes": {d["code"]: {"stops": d.get("stops_count"), "km": d.get("distance_km")} for d in docs}}


@api_router.post("/metrics/upload-report")
async def upload_amazon_report(file: UploadFile = File(...), center: str = Form("OGA5"), _=Depends(require_admin)):
    """Sube un report de Amazon (Scorecard, Daily, POD, CDF…) y lo analiza con IA."""
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Archivo vacío")
    if len(content) > 25 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Archivo demasiado grande (máx 25 MB)")
    try:
        # Mapa Driver ID (Amazon) -> nombre real
        drivers_db = await db.drivers.find(
            {"driver_id": {"$exists": True, "$ne": ""}},
            {"_id": 0, "name": 1, "driver_id": 1, "phone": 1}
        ).to_list(1000)
        driver_map = {d["driver_id"].strip().upper(): d.get("name", "")
                      for d in drivers_db if d.get("driver_id")}
        driver_info = {d["driver_id"].strip().upper(): {"name": d.get("name", ""), "phone": d.get("phone", "")}
                       for d in drivers_db if d.get("driver_id")}
        fn = (file.filename or "").lower()
        analysis = None
        # ¿Es el Excel de Rutas de Cortex? → panel operativo estructurado (sin IA)
        if fn.endswith(".xlsx") or fn.endswith(".xlsm"):
            analysis = _parse_routes_excel(content, driver_info, file.filename)
        if analysis is None:
            text, pdf_bytes = await _extract_report_text(content, file.filename or "report")
            analysis = await _analyze_report_with_gemini(text, pdf_bytes, driver_map)
        # Post-proceso: si algún 'name' sigue siendo un ID conocido, sustituir por el nombre
        for dr in (analysis.get("drivers_at_risk") or []):
            nm = (dr.get("name") or "").strip().upper()
            if nm in driver_map and driver_map[nm]:
                dr["transporter_id"] = dr.get("name")
                dr["name"] = driver_map[nm]
    except asyncio.TimeoutError:
        raise HTTPException(status_code=504, detail="El análisis tardó demasiado. Inténtalo de nuevo.")
    except Exception as e:
        logger.error(f"Análisis report: {e}")
        raise HTTPException(status_code=502, detail="No se pudo analizar el report. ¿Es un informe de Amazon válido?")

    doc = {
        "id": str(uuid.uuid4()),
        "filename": file.filename or "report",
        "center": center,
        "analysis": analysis,
        "uploaded_by": "admin",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.amazon_reports.insert_one(serialize_doc(dict(doc)))
    doc.pop("_id", None)

    # Acumular histórico por conductor/ruta (para construir el ritmo medio con el tiempo)
    if analysis and analysis.get("is_routes"):
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        for rt in analysis.get("routes", []):
            if not rt.get("transporter_id"):
                continue
            try:
                await db.route_history.insert_one({
                    "date": today, "center": center,
                    "transporter_id": rt["transporter_id"],
                    "driver_name": rt.get("driver_name"),
                    "code": rt.get("code"),
                    "service": rt.get("service"),
                    "stops_total": rt.get("stops_total"),
                    "stops_done": rt.get("stops_done"),
                    "rate": rt.get("rate"),
                    "rescue": rt.get("rescue"),
                    "snapshot_at": datetime.now(timezone.utc).isoformat(),
                })
            except Exception:
                pass

    return {"success": True, "report": doc}


@api_router.get("/metrics/driver-history/{transporter_id}")
async def driver_route_history(transporter_id: str, _=Depends(require_admin)):
    """Histórico acumulado de un conductor: ritmo medio, rutas hechas, rescates."""
    tid = transporter_id.strip().upper()
    snaps = await db.route_history.find(
        {"transporter_id": tid}, {"_id": 0}
    ).sort("snapshot_at", -1).to_list(500)
    if not snaps:
        return {"transporter_id": tid, "snapshots": 0}
    # último snapshot por día para no contar el mismo día varias veces en la media
    by_day = {}
    for s in snaps:
        by_day.setdefault(s.get("date"), s)
    days = list(by_day.values())
    rates = [d["rate"] for d in days if d.get("rate")]
    rescues = [d["rescue"] for d in days if d.get("rescue") is not None]
    return {
        "transporter_id": tid,
        "driver_name": snaps[0].get("driver_name"),
        "snapshots": len(snaps),
        "days": len(days),
        "avg_rate": round(sum(rates) / len(rates), 1) if rates else None,
        "avg_rescue": round(sum(rescues) / len(rescues), 1) if rescues else None,
        "days_needed_help": sum(1 for r in rescues if r > 0),
        "recent": days[:10],
    }


@api_router.get("/metrics/reports")
async def list_amazon_reports(center: Optional[str] = None, _=Depends(require_admin)):
    query = {}
    if center and center != "Todos":
        query["center"] = center
    docs = await db.amazon_reports.find(query, {"_id": 0}).sort("created_at", -1).to_list(200)
    return docs


@api_router.delete("/metrics/reports/all")
async def delete_all_reports(_=Depends(require_admin)):
    r = await db.amazon_reports.delete_many({})
    return {"success": True, "deleted": r.deleted_count}


@api_router.delete("/metrics/reports/{report_id}")
async def delete_amazon_report(report_id: str, _=Depends(require_admin)):
    await db.amazon_reports.delete_one({"id": report_id})
    return {"success": True}



@api_router.post("/drivers/import-ids")
async def import_driver_ids(data: dict, _=Depends(require_admin)):
    """Importa Amazon Transporter IDs en masa desde un export de Cortex (pegado).
    La IA extrae pares (nombre, transporter_id) de cualquier formato y los cruza
    con los conductores de la BD por nombre (insensible a acentos)."""
    raw_text = (data.get("text") or "")[:30000]
    if not raw_text.strip():
        raise HTTPException(status_code=400, detail="Pega el contenido del export de Cortex")

    TID_RE = re.compile(r"\bA[A-Z0-9]{8,16}\b")

    # 1) Parseo con Gemini (entiende cualquier formato)
    pairs = []
    try:
        from google import genai as genai_sdk
        from google.genai import types as genai_types
        use_vertex = os.environ.get("USE_VERTEX_AI", "").lower() in ("1", "true", "yes")
        if use_vertex:
            from google.oauth2 import service_account
            import json as _json, base64 as _b64
            sa = os.environ.get("GCP_SERVICE_ACCOUNT_JSON", "").strip()
            if sa and not sa.startswith("{"):
                sa = _b64.b64decode(sa).decode("utf-8")
            creds = service_account.Credentials.from_service_account_info(
                _json.loads(sa), scopes=["https://www.googleapis.com/auth/cloud-platform"]) if sa else None
            client = genai_sdk.Client(vertexai=True, project=os.environ.get("GCP_PROJECT", ""),
                                      location=os.environ.get("GCP_LOCATION", "us-central1"), credentials=creds)
        else:
            client = genai_sdk.Client(api_key=os.environ.get("GEMINI_API_KEY", ""))
        prompt = (
            "Este texto es un export de Amazon Cortex/Scorecard con conductores (DAs). "
            "Extrae TODAS las parejas de nombre del conductor y su Transporter ID "
            "(el ID empieza por 'A' y tiene letras y numeros, ej: A1W24EJAOPQ5F0). "
            "Responde SOLO JSON sin markdown: "
            '{"pairs":[{"name":"nombre tal cual","transporter_id":"AXXXX"}]}\n\nTEXTO:\n' + raw_text
        )
        cfg = genai_types.GenerateContentConfig(temperature=0.0, response_mime_type="application/json")
        loop = asyncio.get_running_loop()
        async with _gemini_sem:
            resp = await asyncio.wait_for(
                loop.run_in_executor(_executor, lambda: client.models.generate_content(
                    model=os.environ.get("GEMINI_MODEL", "gemini-2.5-flash"), contents=[prompt], config=cfg)),
                timeout=40.0)
        gd = json.loads(_strip_markdown_json(resp.text or "{}"))
        for p in gd.get("pairs", []):
            tid = (p.get("transporter_id") or "").strip().upper()
            nm = (p.get("name") or "").strip()
            if nm and TID_RE.match(tid):
                pairs.append({"name": nm, "tid": tid})
    except Exception as _e:
        logger.warning(f"Import IDs: Gemini fallo ({_e}), usando deteccion por lineas")
        pairs = []

    # 2) Respaldo: detectar por linea (un ID + el texto restante = nombre)
    if not pairs:
        for line in raw_text.splitlines():
            m = TID_RE.search(line.upper())
            if not m:
                continue
            tid = m.group(0)
            name = re.sub(TID_RE, "", line, flags=re.I)
            name = re.sub(r"[\t,;|]+", " ", name).strip()
            if len(name) > 3:
                pairs.append({"name": name, "tid": tid})

    if not pairs:
        raise HTTPException(status_code=400, detail="No se encontraron pares nombre + Transporter ID en el texto.")

    # 3) Cruce con la BD por nombre (insensible a acentos)
    import unicodedata as _ud

    def _fold(s):
        s = _ud.normalize("NFD", (s or "").lower())
        return "".join(c for c in s if _ud.category(c) != "Mn")

    def _words(s):
        return set(w for w in re.sub(r"[^a-z ]", " ", _fold(s)).split() if len(w) > 1)

    drivers = await db.drivers.find({"status": {"$ne": "deleted"}}, {"_id": 0, "id": 1, "name": 1}).to_list(1000)
    driver_words = [(d, _words(d.get("name", ""))) for d in drivers]

    matched, ambiguous, unmatched = 0, [], []
    for p in pairs:
        pw = _words(p["name"])
        scored = []
        for d, dw in driver_words:
            inter = pw & dw
            if inter:
                scored.append((sum(2 if len(w) >= 5 else 1 for w in inter), d))
        scored.sort(key=lambda x: -x[0])
        best = scored[0] if scored else None
        second = scored[1] if len(scored) > 1 else None
        ok = best and (best[0] >= 3 or (best[0] >= 2 and (not second or second[0] < best[0])))
        if ok and second and second[0] == best[0]:
            ok = False
            ambiguous.append(p["name"])
        if ok:
            await db.drivers.update_one({"id": best[1]["id"]}, {"$set": {"driver_id": p["tid"]}})
            matched += 1
        elif p["name"] not in ambiguous:
            unmatched.append(p["name"])

    logger.info(f"Import IDs Cortex: {matched}/{len(pairs)} asignados")
    return {"success": True, "total": len(pairs), "matched": matched,
            "ambiguous": ambiguous[:20], "unmatched": unmatched[:20]}


# =========================
# CUADRANTE DE TURNOS (shifts) + solicitudes de días
#   shifts:        {center, driver_id, driver_name, date "YYYY-MM-DD", type}
#                  type ∈ trabaja | libre | extra
#   shift_requests:{id, center, driver_id, driver_name, date, type(libre|extra),
#                   status(pendiente|aprobado|rechazado), created_at, resolved_by, note}
#   shift_settings:{center, min_cobertura}  -> para auto-aprobar si sobra gente
# =========================

VALID_SHIFT_TYPE = {"trabaja", "libre", "extra"}


async def _user_center(user: dict) -> Optional[str]:
    """Centro del usuario. Admin -> None (ve todos)."""
    if user.get("role") == "admin":
        return None
    d = await db.drivers.find_one({"id": user.get("sub")}, {"_id": 0, "center": 1})
    return d.get("center") if d else None


async def _min_cobertura(center: str) -> int:
    s = await db.shift_settings.find_one({"center": center}, {"_id": 0, "min_cobertura": 1})
    if s and isinstance(s.get("min_cobertura"), int):
        return s["min_cobertura"]
    return 0  # 0 = auto-aprobado desactivado (todo pasa a pendiente)


async def _coverage_for_date(center: str, date: str) -> int:
    """Conductores que ese día están 'trabaja' o 'extra'."""
    return await db.shifts.count_documents(
        {"center": center, "date": date, "type": {"$in": ["trabaja", "extra"]}}
    )


@api_router.get("/shifts")
async def get_shifts(center: Optional[str] = None, desde: Optional[str] = None,
                     hasta: Optional[str] = None, _=Depends(require_admin)):
    q = {}
    if center:
        q["center"] = center
    if desde or hasta:
        rng = {}
        if desde:
            rng["$gte"] = desde
        if hasta:
            rng["$lte"] = hasta
        q["date"] = rng
    docs = await db.shifts.find(q, {"_id": 0}).to_list(20000)
    return {"shifts": docs}


@api_router.post("/shifts/bulk")
async def save_shifts_bulk(data: dict = Body(...), _=Depends(require_admin)):
    """Guarda/actualiza varios turnos. body: {items:[{driver_id,driver_name,center,date,type}]}"""
    items = data.get("items") or []
    saved = 0
    for it in items:
        did = it.get("driver_id")
        date = it.get("date")
        center = it.get("center")
        typ = it.get("type")
        if not (did and date and center and typ in VALID_SHIFT_TYPE):
            continue
        await db.shifts.update_one(
            {"driver_id": did, "date": date},
            {"$set": {"driver_id": did, "driver_name": it.get("driver_name", ""),
                      "center": center, "date": date, "type": typ}},
            upsert=True,
        )
        saved += 1
    return {"success": True, "saved": saved}


@api_router.get("/shifts/coverage")
async def get_coverage(center: str, desde: str, hasta: str, _=Depends(require_admin)):
    """Nº de conductores disponibles (trabaja+extra) por día en el rango."""
    cur = db.shifts.find(
        {"center": center, "date": {"$gte": desde, "$lte": hasta},
         "type": {"$in": ["trabaja", "extra"]}},
        {"_id": 0, "date": 1},
    )
    counts: dict = {}
    async for s in cur:
        counts[s["date"]] = counts.get(s["date"], 0) + 1
    return {"coverage": counts, "min": await _min_cobertura(center)}


@api_router.post("/shifts/settings")
async def set_shift_settings(data: dict = Body(...), _=Depends(require_admin)):
    center = data.get("center")
    mn = data.get("min_cobertura")
    if not center or not isinstance(mn, int):
        raise HTTPException(status_code=400, detail="center y min_cobertura (int) requeridos")
    await db.shift_settings.update_one(
        {"center": center}, {"$set": {"center": center, "min_cobertura": mn}}, upsert=True
    )
    return {"success": True}


@api_router.get("/shifts/mine")
async def get_my_shifts(desde: Optional[str] = None, hasta: Optional[str] = None,
                        user: dict = Depends(require_any_auth)):
    """Calendario propio del conductor: sus turnos + sus solicitudes."""
    did = user.get("sub")
    q = {"driver_id": did}
    if desde or hasta:
        rng = {}
        if desde:
            rng["$gte"] = desde
        if hasta:
            rng["$lte"] = hasta
        q["date"] = rng
    shifts = await db.shifts.find(q, {"_id": 0}).to_list(2000)
    reqs = await db.shift_requests.find(
        {"driver_id": did}, {"_id": 0}
    ).sort("date", 1).to_list(500)
    return {"shifts": shifts, "requests": reqs}


@api_router.post("/shift-requests")
async def create_shift_request(data: dict = Body(...), user: dict = Depends(require_any_auth)):
    """El conductor solicita un día (libre o extra). Auto-aprueba si hay cobertura."""
    date = data.get("date")
    typ = data.get("type", "libre")
    if not date or typ not in {"libre", "extra"}:
        raise HTTPException(status_code=400, detail="date y type(libre|extra) requeridos")
    did = user.get("sub")
    drv = await db.drivers.find_one({"id": did}, {"_id": 0, "name": 1, "center": 1})
    if not drv:
        raise HTTPException(status_code=404, detail="Conductor no encontrado")
    center = drv.get("center")
    name = drv.get("name", "")
    note = (data.get("note") or "").strip()[:200]

    # ¿auto-aprobar? solo para 'libre' y si tras quitarlo aún queda cobertura mínima
    status = "pendiente"
    resolved_by = None
    mn = await _min_cobertura(center)
    if typ == "libre" and mn > 0:
        cov = await _coverage_for_date(center, date)
        if (cov - 1) >= mn:
            status = "aprobado"
            resolved_by = "auto"

    req = {
        "id": str(uuid.uuid4()), "center": center, "driver_id": did,
        "driver_name": name, "date": date, "type": typ, "status": status,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "resolved_by": resolved_by, "note": note,
    }
    await db.shift_requests.insert_one(req)
    # si se auto-aprobó, refleja el turno
    if status == "aprobado":
        await db.shifts.update_one(
            {"driver_id": did, "date": date},
            {"$set": {"driver_id": did, "driver_name": name, "center": center,
                      "date": date, "type": typ}},
            upsert=True,
        )
    req.pop("_id", None)
    return {"success": True, "request": req, "auto": status == "aprobado"}


@api_router.get("/shift-requests")
async def list_shift_requests(center: Optional[str] = None, status: Optional[str] = None,
                              _=Depends(require_admin)):
    q = {}
    if center:
        q["center"] = center
    if status:
        q["status"] = status
    reqs = await db.shift_requests.find(q, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return {"requests": reqs}


@api_router.post("/shift-requests/{req_id}/resolve")
async def resolve_shift_request(req_id: str, data: dict = Body(...),
                                admin: dict = Depends(require_admin)):
    """Aprobar o rechazar. body: {action: aprobar|rechazar}"""
    action = data.get("action")
    if action not in {"aprobar", "rechazar"}:
        raise HTTPException(status_code=400, detail="action debe ser aprobar|rechazar")
    req = await db.shift_requests.find_one({"id": req_id}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Solicitud no encontrada")
    new_status = "aprobado" if action == "aprobar" else "rechazado"
    await db.shift_requests.update_one(
        {"id": req_id},
        {"$set": {"status": new_status, "resolved_by": admin.get("name", "admin")}},
    )
    if new_status == "aprobado":
        await db.shifts.update_one(
            {"driver_id": req["driver_id"], "date": req["date"]},
            {"$set": {"driver_id": req["driver_id"], "driver_name": req.get("driver_name", ""),
                      "center": req["center"], "date": req["date"], "type": req["type"]}},
            upsert=True,
        )
    return {"success": True, "status": new_status}


@api_router.post("/shifts/import")
async def import_shifts(file: UploadFile = File(...), center: str = Form(...),
                        _=Depends(require_admin)):
    """Importa un cuadrante desde Excel.
    Formato esperado: 1ª columna = nombre del conductor; cabeceras de las
    siguientes columnas = fechas (YYYY-MM-DD o día del mes); celdas = T/L/E
    (trabaja/libre/extra). Tolerante: reconoce 't','trabaja','x' -> trabaja,
    'l','libre','-' -> libre, 'e','extra' -> extra."""
    content = await file.read()
    fname = (file.filename or "").lower()
    rows = []
    try:
        if fname.endswith(".xls"):
            import xlrd
            book = xlrd.open_workbook(file_contents=content)
            sh = book.sheet_by_index(0)
            for r in range(sh.nrows):
                rows.append([sh.cell_value(r, c) for c in range(sh.ncols)])
        else:
            import openpyxl, io
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"No se pudo leer el Excel: {e}")

    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel vacío o sin cabecera de fechas")

    def norm_type(v):
        s = str(v or "").strip().lower()
        if s in {"t", "trabaja", "x", "tr", "w", "1"}:
            return "trabaja"
        if s in {"e", "extra", "ex"}:
            return "extra"
        if s in {"l", "libre", "-", "0", "off", "d"}:
            return "libre"
        return None

    def norm_date(v):
        s = str(v).strip()
        if not s:
            return None
        if hasattr(v, "strftime"):
            return v.strftime("%Y-%m-%d")
        m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", s)
        if m:
            return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
        return None

    header = rows[0]
    date_cols = {}  # idx -> 'YYYY-MM-DD'
    for ci in range(1, len(header)):
        d = norm_date(header[ci])
        if d:
            date_cols[ci] = d
    if not date_cols:
        raise HTTPException(status_code=400,
                            detail="No encontré fechas válidas (YYYY-MM-DD) en la cabecera")

    drivers = await db.drivers.find({"center": center}, {"_id": 0, "id": 1, "name": 1}).to_list(1000)
    by_name = {(d.get("name") or "").strip().lower(): d for d in drivers}

    saved, unmatched = 0, []
    for r in rows[1:]:
        if not r or not str(r[0]).strip():
            continue
        nm = str(r[0]).strip()
        drv = by_name.get(nm.lower())
        if not drv:
            unmatched.append(nm)
            continue
        for ci, date in date_cols.items():
            if ci >= len(r):
                continue
            typ = norm_type(r[ci])
            if not typ:
                continue
            await db.shifts.update_one(
                {"driver_id": drv["id"], "date": date},
                {"$set": {"driver_id": drv["id"], "driver_name": drv["name"],
                          "center": center, "date": date, "type": typ}},
                upsert=True,
            )
            saved += 1
    return {"success": True, "saved": saved, "dias": len(date_cols),
            "unmatched": unmatched[:20]}


@api_router.get("/route-demand")
async def get_route_demand(center: str, desde: str, hasta: str, _=Depends(require_admin)):
    """Rutas que pide Amazon por día: objetivo y máximo."""
    docs = await db.route_demand.find(
        {"center": center, "date": {"$gte": desde, "$lte": hasta}}, {"_id": 0}
    ).to_list(400)
    demand = {d["date"]: {"objetivo": d.get("objetivo"), "maximo": d.get("maximo")} for d in docs}
    return {"demand": demand}


@api_router.post("/route-demand")
async def set_route_demand(data: dict = Body(...), _=Depends(require_admin)):
    """body: {center, items:[{date, objetivo, maximo}]}"""
    center = data.get("center")
    items = data.get("items") or []
    if not center:
        raise HTTPException(status_code=400, detail="center requerido")
    saved = 0
    for it in items:
        date = it.get("date")
        if not date:
            continue
        obj = it.get("objetivo")
        mx = it.get("maximo")
        await db.route_demand.update_one(
            {"center": center, "date": date},
            {"$set": {"center": center, "date": date,
                      "objetivo": int(obj) if obj not in (None, "") else None,
                      "maximo": int(mx) if mx not in (None, "") else None}},
            upsert=True,
        )
        saved += 1
    return {"success": True, "saved": saved}


def _date_range(desde: str, hasta: str):
    out = []
    d0 = datetime.strptime(desde, "%Y-%m-%d")
    d1 = datetime.strptime(hasta, "%Y-%m-%d")
    cur = d0
    while cur <= d1:
        out.append(cur.strftime("%Y-%m-%d"))
        cur += timedelta(days=1)
    return out


_SCHEDULER_PROMPT = """Eres el planificador de turnos de una empresa DSP de reparto Amazon.
Tu objetivo es montar el cuadrante semanal MÁS EFICIENTE posible (NO equitativo):
los mejores repartidores cubren el trabajo, garantizando que TODOS los días
queden cubiertos con al menos la cobertura mínima indicada.

REGLAS DE CONTRATO (obligatorias):
- contrato "empresa": trabaja SOLO de lunes a viernes. NUNCA sábado ni domingo
  (esos dos días van siempre 'libre' para ellos).
- contrato "ett": puede trabajar cualquier día de la semana, incluido finde.

NIVELES DE NOVATO (carga progresiva) — van PROTEGIDOS APARTE:
- Coloca PRIMERO a los novatos (L1/L2/L3) asegurándoles días de trabajo para que
  cojan experiencia. NO los metas en el ranking de eficiencia y NUNCA los dejes
  sin trabajar por falta de datos: es injusto, tienen que empezar a rodar.
- L1 = recién entra: rutas cortas/fáciles y acompañado, carga ligera.
- L2 = ruta normal en zona ya conocida.
- L3 = casi autónomo, carga casi completa.
- nivel "pleno" = ya formado: entra al ranking normal de eficiencia.

EFICIENCIA (solo para los 'pleno'):
- Usa SOLO datos reales. Dos señales: la "scorecard tier" (Fantastic+ > Fantastic >
  Great > Fair > Poor > At Risk; media 1-6, más alto = mejor) y el "ritmo" =
  paradas/hora reales. Da más peso a la scorecard si está; el ritmo desempata.
  Si un 'pleno' no tiene ninguna de las dos, trátalo como medio; NO inventes cifras.
- Prioriza a los de mayor ritmo para los días de trabajo; los de menor ritmo
  son los primeros en librar cuando sobra capacidad. No busques reparto equitativo.

GENERAL:
- Respeta SIEMPRE las instrucciones extra del usuario. Cruza por el nombre.
- Cada conductor ~5 días/semana salvo que el usuario o su contrato digan otra cosa.
- OBLIGATORIO: cada día debe tener en 'trabaja'+'extra' tantos conductores como
  rutas objetivo pide Amazon ese día (1 conductor = 1 ruta). No superes el máximo.
  Si no hay gente suficiente para el objetivo, avísalo en el resumen.
- Usa 'extra' para llegar al objetivo del día cuando con los habituales no alcanza.
- Tipos válidos: "trabaja", "libre", "extra".

Devuelve EXCLUSIVAMENTE JSON:
{"assignments":[{"driver_id":"<id>","date":"YYYY-MM-DD","type":"trabaja|libre|extra"}],
 "resumen":"2-3 frases: criterio seguido, cómo colocaste a los novatos y avisos de cobertura"}"""


async def _generate_schedule_with_gemini(context_text: str, user_prompt: str):
    from google import genai as genai_sdk
    from google.genai import types as genai_types
    use_vertex = os.environ.get("USE_VERTEX_AI", "").lower() in ("1", "true", "yes")
    if use_vertex:
        from google.oauth2 import service_account
        import json as _json, base64 as _b64
        sa = os.environ.get("GCP_SERVICE_ACCOUNT_JSON", "").strip()
        if sa and not sa.startswith("{"):
            sa = _b64.b64decode(sa).decode("utf-8")
        creds = service_account.Credentials.from_service_account_info(
            _json.loads(sa), scopes=["https://www.googleapis.com/auth/cloud-platform"]) if sa else None
        client = genai_sdk.Client(vertexai=True, project=os.environ.get("GCP_PROJECT", ""),
                                  location=os.environ.get("GCP_LOCATION", "us-central1"), credentials=creds)
    else:
        client = genai_sdk.Client(api_key=os.environ.get("GEMINI_API_KEY", ""))

    full = (_SCHEDULER_PROMPT + "\n\n=== INSTRUCCIONES DEL USUARIO ===\n" +
            (user_prompt or "(sin instrucciones extra: prioriza eficiencia y cobertura)") +
            "\n\n=== DATOS ===\n" + context_text)
    cfg = genai_types.GenerateContentConfig(temperature=0.2, response_mime_type="application/json")
    loop = asyncio.get_running_loop()
    async with _gemini_sem:
        resp = await asyncio.wait_for(
            loop.run_in_executor(_executor, lambda: client.models.generate_content(
                model=os.environ.get("GEMINI_MODEL", "gemini-2.5-flash"), contents=[full], config=cfg)),
            timeout=170.0)
    return json.loads(_strip_markdown_json(resp.text or "{}"))


@api_router.post("/shifts/generate")
async def generate_schedule(data: dict = Body(...), _=Depends(require_admin)):
    """Genera un cuadrante con IA a partir de un prompt + datos reales.
    No guarda nada: devuelve la propuesta para que el admin la revise y guarde."""
    center = data.get("center")
    desde = data.get("desde")
    hasta = data.get("hasta")
    user_prompt = (data.get("prompt") or "").strip()
    if not (center and desde and hasta):
        raise HTTPException(status_code=400, detail="center, desde y hasta requeridos")
    days = _date_range(desde, hasta)
    min_cov = data.get("min_cobertura")
    if not isinstance(min_cov, int):
        min_cov = await _min_cobertura(center)

    drivers = await db.drivers.find(
        {"center": {"$regex": re.escape(center), "$options": "i"}, "active": {"$ne": False}},
        {"_id": 0, "id": 1, "name": 1, "driver_id": 1, "contrato": 1, "zona": 1, "nivel": 1},
    ).to_list(1000)
    if not drivers:
        raise HTTPException(status_code=404, detail=f"No hay conductores en {center}")

    # ── ritmo real (paradas/hora) por transporter_id desde route_history ──
    tids = [d.get("driver_id") for d in drivers if d.get("driver_id")]
    pace = {}
    if tids:
        cur = db.route_history.find(
            {"transporter_id": {"$in": [t.upper() for t in tids]}},
            {"_id": 0, "transporter_id": 1, "date": 1, "rate": 1},
        )
        tmp = {}
        async for s in cur:
            t = s.get("transporter_id")
            if s.get("rate"):
                tmp.setdefault(t, {})[s.get("date")] = s["rate"]  # último por día
        for t, byday in tmp.items():
            vals = list(byday.values())
            if vals:
                pace[t] = round(sum(vals) / len(vals), 1)

    # ── días ya trabajados en los últimos 14 días (para no recargar siempre a los mismos) ──
    d14 = (datetime.strptime(desde, "%Y-%m-%d") - timedelta(days=14)).strftime("%Y-%m-%d")
    worked = {}
    cur2 = db.shifts.find(
        {"center": {"$regex": re.escape(center), "$options": "i"}, "date": {"$gte": d14, "$lt": desde},
         "type": {"$in": ["trabaja", "extra"]}},
        {"_id": 0, "driver_id": 1},
    )
    async for s in cur2:
        worked[s["driver_id"]] = worked.get(s["driver_id"], 0) + 1

    # ── rendimiento de scorecard (últimas semanas) por conductor ──
    sc = {}  # driver_id -> {"tier": último, "avg": media tier_score}
    sc_cur = db.driver_scorecard.find(
        {"$or": [{"driver_id": {"$in": [d["id"] for d in drivers]}},
                 {"transporter_id": {"$in": [t.upper() for t in tids]}}]},
        {"_id": 0, "driver_id": 1, "transporter_id": 1, "tier": 1, "tier_score": 1, "imported_at": 1},
    ).sort("imported_at", -1)
    tid_to_id = {(d.get("driver_id") or "").upper(): d["id"] for d in drivers if d.get("driver_id")}
    tmp_sc = {}
    async for s in sc_cur:
        did = s.get("driver_id") or tid_to_id.get((s.get("transporter_id") or "").upper())
        if not did:
            continue
        e = tmp_sc.setdefault(did, {"tiers": [], "scores": []})
        if len(e["tiers"]) < 4:
            if s.get("tier"):
                e["tiers"].append(s["tier"])
            if isinstance(s.get("tier_score"), (int, float)):
                e["scores"].append(s["tier_score"])
    for did, e in tmp_sc.items():
        sc[did] = {"tier": e["tiers"][0] if e["tiers"] else None,
                   "avg": round(sum(e["scores"]) / len(e["scores"]), 1) if e["scores"] else None}

    # ── demanda de Amazon por día (rutas objetivo / máximo) ──
    dem_docs = await db.route_demand.find(
        {"center": center, "date": {"$gte": desde, "$lte": hasta}}, {"_id": 0}
    ).to_list(400)
    demand = {d["date"]: d for d in dem_docs}
    if not any(demand.get(dd, {}).get("objetivo") for dd in days):
        raise HTTPException(
            status_code=400,
            detail="Falta la demanda de Amazon: pon las rutas objetivo de cada día antes de generar el cuadrante.")

    # ── construir contexto para la IA ──
    dem_lines = []
    for dd in days:
        o = demand.get(dd, {}).get("objetivo")
        mx = demand.get(dd, {}).get("maximo")
        dem_lines.append(f"  {dd}: objetivo {o if o is not None else '—'} rutas"
                         + (f", máximo {mx}" if mx is not None else ""))
    lines = [f"Centro: {center}", f"Días a planificar: {', '.join(days)}",
             "Rutas que pide Amazon por día (1 conductor = 1 ruta; cubre el objetivo, no superes el máximo):",
             "\n".join(dem_lines),
             f"Total conductores disponibles: {len(drivers)}", "",
             "Conductores (id | nombre | contrato | nivel | scorecard tier (media 1-6) | ritmo real p/h | días trabajados últimas 2 sem | zona):"]
    for d in drivers:
        tid = (d.get("driver_id") or "").upper()
        r = pace.get(tid)
        contrato = (d.get("contrato") or "ett").lower()
        nivel = (d.get("nivel") or "pleno")
        scd = sc.get(d["id"])
        sc_txt = (f"{scd['tier']} (med {scd['avg']})" if scd and scd.get("tier")
                  else (f"med {scd['avg']}" if scd and scd.get("avg") is not None else "sin scorecard"))
        lines.append(
            f"- {d['id']} | {d.get('name','')} | "
            f"{contrato} | {nivel} | {sc_txt} | "
            f"{(str(r)+' p/h') if r else 'sin ritmo'} | "
            f"{worked.get(d['id'],0)} días | {d.get('zona') or '—'}"
        )
    context_text = "\n".join(lines)

    try:
        result = await _generate_schedule_with_gemini(context_text, user_prompt)
    except Exception as e:
        logger.error(f"generate_schedule gemini error: {type(e).__name__}: {repr(e)}")
        raise HTTPException(status_code=502, detail="La IA no pudo generar el cuadrante; reintenta.")

    # validar/filtrar assignments contra ids y días reales
    valid_ids = {d["id"] for d in drivers}
    name_by_id = {d["id"]: d.get("name", "") for d in drivers}
    day_set = set(days)
    clean = []
    for a in (result.get("assignments") or []):
        did = a.get("driver_id")
        date = a.get("date")
        typ = a.get("type")
        if did in valid_ids and date in day_set and typ in VALID_SHIFT_TYPE:
            clean.append({"driver_id": did, "driver_name": name_by_id.get(did, ""),
                          "center": center, "date": date, "type": typ})

    # cobertura resultante por día (para avisar)
    cov = {dd: 0 for dd in days}
    for a in clean:
        if a["type"] in ("trabaja", "extra"):
            cov[a["date"]] = cov.get(a["date"], 0) + 1

    return {"success": True, "assignments": clean,
            "resumen": result.get("resumen", ""),
            "coverage": cov, "min_cobertura": min_cov,
            "con_datos_ritmo": len(pace)}


# Presupuesto de días/semana por nivel (carga progresiva de novatos)
_NIVEL_BUDGET = {"L1": 3, "L2": 4, "L3": 5, "pleno": 5}
_NOVATOS = {"L1", "L2", "L3"}


@api_router.post("/shifts/generate-auto")
async def generate_schedule_auto(data: dict = Body(...), _=Depends(require_admin)):
    """Genera el cuadrante con un ALGORITMO DETERMINISTA (sin IA): instantáneo,
    gratis y siempre consistente. Cubre la demanda de Amazon por eficiencia real,
    respeta empresa(L-V)/ETT y protege a los novatos. No guarda: devuelve propuesta."""
    center = data.get("center")
    desde = data.get("desde")
    hasta = data.get("hasta")
    if not (center and desde and hasta):
        raise HTTPException(status_code=400, detail="center, desde y hasta requeridos")
    days = _date_range(desde, hasta)

    drivers = await db.drivers.find(
        {"center": {"$regex": re.escape(center), "$options": "i"}, "active": {"$ne": False}},
        {"_id": 0, "id": 1, "name": 1, "driver_id": 1, "contrato": 1, "nivel": 1},
    ).to_list(1000)
    if not drivers:
        raise HTTPException(status_code=404, detail=f"No hay conductores en {center}")

    # demanda
    dem_docs = await db.route_demand.find(
        {"center": center, "date": {"$gte": desde, "$lte": hasta}}, {"_id": 0}
    ).to_list(400)
    demand = {d["date"]: d for d in dem_docs}
    if not any(demand.get(dd, {}).get("objetivo") for dd in days):
        raise HTTPException(status_code=400,
            detail="Falta la demanda de Amazon: pon las rutas objetivo de cada día antes de generar.")

    # ritmo real (paradas/hora) por transporter_id
    tids = [d.get("driver_id") for d in drivers if d.get("driver_id")]
    pace = {}
    if tids:
        cur = db.route_history.find(
            {"transporter_id": {"$in": [t.upper() for t in tids]}},
            {"_id": 0, "transporter_id": 1, "date": 1, "rate": 1})
        tmp = {}
        async for s in cur:
            if s.get("rate"):
                tmp.setdefault(s.get("transporter_id"), {})[s.get("date")] = s["rate"]
        for t, byday in tmp.items():
            vals = list(byday.values())
            if vals:
                pace[t] = sum(vals) / len(vals)

    # scorecard (media tier_score 1-6) por driver_id
    sc = {}
    sc_cur = db.driver_scorecard.find(
        {"$or": [{"driver_id": {"$in": [d["id"] for d in drivers]}},
                 {"transporter_id": {"$in": [t.upper() for t in tids]}}]},
        {"_id": 0, "driver_id": 1, "transporter_id": 1, "tier_score": 1, "imported_at": 1},
    ).sort("imported_at", -1)
    tid_to_id = {(d.get("driver_id") or "").upper(): d["id"] for d in drivers if d.get("driver_id")}
    tmp_sc = {}
    async for s in sc_cur:
        did = s.get("driver_id") or tid_to_id.get((s.get("transporter_id") or "").upper())
        if not did:
            continue
        e = tmp_sc.setdefault(did, [])
        if len(e) < 4 and isinstance(s.get("tier_score"), (int, float)):
            e.append(s["tier_score"])
    for did, vals in tmp_sc.items():
        if vals:
            sc[did] = sum(vals) / len(vals)

    # días ya trabajados últimas 2 semanas (para no recargar siempre a los mismos)
    d14 = (datetime.strptime(desde, "%Y-%m-%d") - timedelta(days=14)).strftime("%Y-%m-%d")
    worked = {}
    cur2 = db.shifts.find(
        {"center": {"$regex": re.escape(center), "$options": "i"}, "date": {"$gte": d14, "$lt": desde},
         "type": {"$in": ["trabaja", "extra"]}}, {"_id": 0, "driver_id": 1})
    async for s in cur2:
        worked[s["driver_id"]] = worked.get(s["driver_id"], 0) + 1

    # eficiencia por conductor: scorecard (1-6) + ritmo como desempate
    pace_vals = [v for v in pace.values() if v]
    pmax = max(pace_vals) if pace_vals else 1.0

    def eff(d):
        did = d["id"]
        base = sc.get(did)
        if base is None:
            base = 3.5  # sin datos = medio (no se infravalora)
        tid = (d.get("driver_id") or "").upper()
        pv = pace.get(tid)
        pn = (pv / pmax) if (pv and pmax) else 0.0
        return base + 0.3 * pn

    # agrupar en semanas (lunes inicia semana) para el cupo de ~5 días
    weeks, cur = [], []
    for dd in days:
        if datetime.strptime(dd, "%Y-%m-%d").weekday() == 0 and cur:
            weeks.append(cur); cur = []
        cur.append(dd)
    if cur:
        weeks.append(cur)

    assignments, faltas = [], []
    name_by_id = {d["id"]: d.get("name", "") for d in drivers}
    for wk in weeks:
        assigned = {d["id"]: 0 for d in drivers}   # días asignados esta semana
        for dd in wk:
            obj = demand.get(dd, {}).get("objetivo")
            if not obj:
                continue
            mx = demand.get(dd, {}).get("maximo") or obj
            need = min(int(obj), int(mx))
            is_weekend = datetime.strptime(dd, "%Y-%m-%d").weekday() >= 5

            elig = []
            for d in drivers:
                contrato = (d.get("contrato") or "ett").lower()
                if contrato == "empresa" and is_weekend:
                    continue   # empresa nunca trabaja en finde
                elig.append(d)

            def sortkey(d):
                nivel = d.get("nivel") or "pleno"
                budget = _NIVEL_BUDGET.get(nivel, 5)
                under = assigned[d["id"]] < budget
                is_nov = nivel in _NOVATOS
                grp = 0 if (is_nov and under) else (1 if under else 2)  # novatos protegidos primero
                return (grp, -eff(d), assigned[d["id"]], worked.get(d["id"], 0), d.get("name", ""))

            elig.sort(key=sortkey)
            picked = elig[:need]
            for d in picked:
                budget = _NIVEL_BUDGET.get(d.get("nivel") or "pleno", 5)
                typ = "extra" if assigned[d["id"]] >= budget else "trabaja"
                assignments.append({"driver_id": d["id"], "driver_name": name_by_id[d["id"]],
                                    "center": center, "date": dd, "type": typ})
                assigned[d["id"]] += 1
            if len(picked) < int(obj):
                faltas.append(f"{dd[8:]}/{dd[5:7]} faltan {int(obj) - len(picked)}")

    cov = {dd: 0 for dd in days}
    for a in assignments:
        if a["type"] in ("trabaja", "extra"):
            cov[a["date"]] = cov.get(a["date"], 0) + 1

    con_sc = len(sc)
    resumen = (f"Cuadrante automático por eficiencia real "
               f"({con_sc} con scorecard, {len(pace)} con ritmo). "
               f"Empresa solo L-V, novatos protegidos con carga progresiva.")
    if faltas:
        resumen += " ⚠️ Días sin cubrir el objetivo: " + ", ".join(faltas[:12])
    else:
        resumen += " Todos los días cubren el objetivo."

    return {"success": True, "assignments": assignments, "resumen": resumen,
            "coverage": cov, "con_datos_ritmo": len(pace), "con_scorecard": con_sc,
            "faltas": len(faltas)}


# =========================
# IMPORTAR SCORECARDS → rendimiento real por conductor (para el generador)
#   driver_scorecard: {center, semana, driver_name, transporter_id, tier,
#                      tier_score(1-6), posicion, score, metrics, imported_at}
# =========================

_TIER_SCORE = {"fantastic+": 6, "fantastic +": 6, "fantastic plus": 6, "fantastic": 5,
               "great": 4, "fair": 3, "poor": 2, "at risk": 1, "en riesgo": 1}


def _tier_to_score(tier):
    t = (tier or "").strip().lower()
    for k, v in _TIER_SCORE.items():
        if k in t:
            return v
    return None


_SCORECARD_EXTRACT_PROMPT = """Eres un analista de Amazon DSP. Te paso una Scorecard semanal.
Extrae el RENDIMIENTO INDIVIDUAL de CADA repartidor que aparezca (no solo los que fallan).
Para cada uno devuelve su nombre tal cual, su Transporter ID si aparece (formato tipo
A1W24EJAOPQ5F0), su tier/standing global de la semana y, si están, sus métricas clave.

Tiers posibles: "Fantastic+", "Fantastic", "Great", "Fair", "Poor", "At Risk".

Devuelve EXCLUSIVAMENTE JSON:
{"semana":"texto de la semana o fecha si aparece",
 "conductores":[
   {"name":"Nombre Apellido","transporter_id":"A1... o null",
    "tier":"Fantastic|Great|Fair|Poor|At Risk|...","posicion":<int o null>,
    "score":<número global 0-100 o null>,
    "dcr":<o null>,"dnr_dpmo":<o null>,"pod":<o null>,"cc":<o null>}
 ]}
No inventes conductores ni cifras: solo lo que aparezca en la Scorecard."""


async def _extract_scorecard_standings(text, pdf_bytes):
    from google import genai as genai_sdk
    from google.genai import types as genai_types
    use_vertex = os.environ.get("USE_VERTEX_AI", "").lower() in ("1", "true", "yes")
    if use_vertex:
        from google.oauth2 import service_account
        import json as _json, base64 as _b64
        sa = os.environ.get("GCP_SERVICE_ACCOUNT_JSON", "").strip()
        if sa and not sa.startswith("{"):
            sa = _b64.b64decode(sa).decode("utf-8")
        creds = service_account.Credentials.from_service_account_info(
            _json.loads(sa), scopes=["https://www.googleapis.com/auth/cloud-platform"]) if sa else None
        client = genai_sdk.Client(vertexai=True, project=os.environ.get("GCP_PROJECT", ""),
                                  location=os.environ.get("GCP_LOCATION", "us-central1"), credentials=creds)
    else:
        client = genai_sdk.Client(api_key=os.environ.get("GEMINI_API_KEY", ""))
    contents = [_SCORECARD_EXTRACT_PROMPT]
    if pdf_bytes:
        contents.append(genai_types.Part.from_bytes(data=pdf_bytes, mime_type="application/pdf"))
    else:
        contents.append("CONTENIDO DE LA SCORECARD:\n\n" + (text or ""))
    cfg = genai_types.GenerateContentConfig(temperature=0.1, response_mime_type="application/json")
    loop = asyncio.get_running_loop()
    async with _gemini_sem:
        resp = await asyncio.wait_for(
            loop.run_in_executor(_executor, lambda: client.models.generate_content(
                model=os.environ.get("GEMINI_MODEL", "gemini-2.5-flash"), contents=contents, config=cfg)),
            timeout=170.0)
    return json.loads(_strip_markdown_json(resp.text or "{}"))


@api_router.post("/scorecard/import")
async def import_scorecard(file: UploadFile = File(...), center: str = Form(...),
                           user: dict = Depends(require_admin)):
    """Sube una Scorecard; la IA extrae el rendimiento de cada repartidor y lo guarda
    para que el generador de cuadrantes lo use como señal de eficiencia real."""
    await _require_plan_feature(user, "scorecard")
    content = await file.read()
    text, pdf_bytes = await _extract_report_text(content, file.filename or "scorecard")
    try:
        data = await _extract_scorecard_standings(text, pdf_bytes)
    except Exception as e:
        logger.error(f"import_scorecard gemini error: {type(e).__name__}: {repr(e)}")
        raise HTTPException(status_code=502, detail="La IA no pudo leer la Scorecard; reintenta.")

    semana = (data.get("semana") or file.filename or "").strip()[:80]
    conductores = data.get("conductores") or []
    # mapa de nombres -> driver para cruzar transporter_id que falte
    drivers = await db.drivers.find({}, {"_id": 0, "id": 1, "name": 1, "driver_id": 1}).to_list(2000)
    by_tid = {(d.get("driver_id") or "").upper(): d for d in drivers if d.get("driver_id")}
    by_name = {(d.get("name") or "").strip().lower(): d for d in drivers}

    saved, matched = 0, 0
    for c in conductores:
        name = (c.get("name") or "").strip()
        tid = (c.get("transporter_id") or "").strip().upper()
        drv = by_tid.get(tid) or by_name.get(name.lower())
        if drv and not tid:
            tid = (drv.get("driver_id") or "").upper()
        if drv:
            matched += 1
        tier = c.get("tier")
        doc = {
            "center": center, "semana": semana,
            "driver_name": drv.get("name") if drv else name,
            "transporter_id": tid or None,
            "driver_id": drv.get("id") if drv else None,
            "tier": tier, "tier_score": _tier_to_score(tier),
            "posicion": c.get("posicion"), "score": c.get("score"),
            "metrics": {"dcr": c.get("dcr"), "dnr_dpmo": c.get("dnr_dpmo"),
                        "pod": c.get("pod"), "cc": c.get("cc")},
            "imported_at": datetime.now(timezone.utc).isoformat(),
        }
        # una entrada por (conductor, semana, centro)
        key = {"center": center, "semana": semana}
        if tid:
            key["transporter_id"] = tid
        else:
            key["driver_name"] = doc["driver_name"]
        await db.driver_scorecard.update_one(key, {"$set": doc}, upsert=True)
        saved += 1

    return {"success": True, "semana": semana, "conductores": saved,
            "cruzados": matched, "sin_cruzar": saved - matched}


@api_router.delete("/scorecard/week")
async def delete_scorecard_week(semana: str, center: Optional[str] = None,
                               _=Depends(require_admin)):
    """Borra todas las líneas de una semana de scorecard (para reimportar o limpiar)."""
    q = {"semana": semana}
    if center:
        q["center"] = center
    r = await db.driver_scorecard.delete_many(q)
    return {"success": True, "borradas": r.deleted_count}


@api_router.get("/scorecard/standings")
async def scorecard_standings(center: Optional[str] = None, user: dict = Depends(require_admin)):
    """Últimas semanas de scorecard guardadas (para ver qué hay importado)."""
    await _require_plan_feature(user, "scorecard")
    q = {}
    if center:
        q["center"] = center
    docs = await db.driver_scorecard.find(q, {"_id": 0}).sort("imported_at", -1).to_list(2000)
    semanas = {}
    cruzados = 0
    for d in docs:
        semanas.setdefault(d.get("semana", "?"), 0)
        semanas[d["semana"]] += 1
        if d.get("driver_id"):
            cruzados += 1
    return {"semanas": semanas, "total": len(docs), "cruzados": cruzados}



# =========================
# SCORECARD EN VIVO — reportes diarios de Cortex (por centro, día a día)
#   Parsea el "Daily Report" HTML de Cortex: fallos por conductor (RTS/DNR/POD/CC)
#   + motivos de POD para coaching. Acumula la semana → proyección de tier.
#   daily_dsp: {center, date, drivers:[{transporter_id,rts,dnr,pod,cc}],
#               totals:{...}, pod_reasons:{...}, cc_reasons:{...}, uploaded_at}
# =========================

def _parse_daily_report(content: bytes, filename: str):
    html = content.decode("utf-8", errors="ignore")
    center, date = None, None
    m = re.search(r"([A-Z]{3}\d)-Daily-Report_(\d{4}-\d{2}-\d{2})", filename or "")
    if m:
        center, date = m.group(1), m.group(2)
    if not center:
        tm = re.search(r"TDSL\s+([A-Z]{3}\d)\s+Daily Report\s+(\d{4}-\d{2}-\d{2})", html)
        if tm:
            center, date = tm.group(1), tm.group(2)

    tables = re.findall(r"<table.*?</table>", html, re.S | re.I)

    def parse_rows(tbl):
        out = []
        for tr in re.findall(r"<tr.*?</tr>", tbl, re.S | re.I):
            cells = re.findall(r"<t[hd].*?</t[hd]>", tr, re.S | re.I)
            cells = [re.sub(r"\s+", " ", re.sub(r"&nbsp;", " ", re.sub(r"<[^>]+>", " ", c))).strip()
                     for c in cells]
            if any(cells):
                out.append(cells)
        return out

    def num(x):
        x = (x or "").strip()
        return int(x) if x.isdigit() else 0

    drivers, pod_reasons, cc_reasons = [], {}, {}
    for tbl in tables:
        rows = parse_rows(tbl)
        if not rows:
            continue
        # la cabecera real es la 1ª fila con >=2 celdas (las tablas empiezan
        # con una fila "Download This Table" de 1 sola celda)
        hi = next((i for i, r in enumerate(rows) if len(r) >= 2), 0)
        hdr = [c.lower() for c in rows[hi]]
        rows = rows[hi:]
        hjoin = " ".join(hdr)
        # Tabla 1: fallos por conductor
        if "transporter id" in hdr and "pod fails" in hjoin and "rts" in hdr:
            for r in rows[1:]:
                if len(r) >= 5 and re.match(r"^A[A-Z0-9]{8,}$", r[0]):
                    drivers.append({"transporter_id": r[0], "rts": num(r[1]),
                                    "dnr": num(r[2]), "pod": num(r[3]), "cc": num(r[4])})
        # Tabla 4: motivos de POD
        elif "pod audit" in hjoin:
            idx = next((i for i, h in enumerate(hdr) if "pod audit" in h), len(hdr) - 1)
            for r in rows[1:]:
                if len(r) > idx and r[idx] and r[idx] != "-":
                    pod_reasons[r[idx]] = pod_reasons.get(r[idx], 0) + 1
        # Tabla 5: motivos de Contact Compliance
        elif "cc type" in hjoin or "call duration" in hjoin:
            idx = next((i for i, h in enumerate(hdr) if "reason" in h), 2)
            for r in rows[1:]:
                if len(r) > idx and r[idx]:
                    key = r[idx].split("DELIVERED")[0].strip() or r[idx]
                    cc_reasons[key[:40]] = cc_reasons.get(key[:40], 0) + 1

    totals = {k: sum(d[k] for d in drivers) for k in ("rts", "dnr", "pod", "cc")}
    return {"center": center, "date": date, "drivers": drivers, "totals": totals,
            "pod_reasons": pod_reasons, "cc_reasons": cc_reasons}


@api_router.post("/metrics/upload-daily")
async def upload_daily_report(file: UploadFile = File(...), _=Depends(require_admin)):
    """Sube un Daily Report de Cortex (HTML). Guarda fallos por conductor del día."""
    content = await file.read()
    parsed = _parse_daily_report(content, file.filename or "")
    if not parsed.get("center") or not parsed.get("date"):
        raise HTTPException(status_code=400,
            detail="No reconozco el reporte. ¿Es un Daily Report de Cortex (HTML)?")
    if not parsed.get("drivers"):
        raise HTTPException(status_code=400, detail="No encontré la tabla de fallos por conductor")
    doc = dict(parsed)
    doc["uploaded_at"] = datetime.now(timezone.utc).isoformat()
    await db.daily_dsp.update_one(
        {"center": parsed["center"], "date": parsed["date"]}, {"$set": doc}, upsert=True)
    return {"success": True, "center": parsed["center"], "date": parsed["date"],
            "conductores": len(parsed["drivers"]), "totales": parsed["totals"]}


# Objetivos del scorecard (umbrales del tier) — los pone el usuario a mano.
# Valores por defecto orientativos; cada DSP/centro ajusta los suyos.
_DEFAULT_TARGETS = {"dcr": 98.5, "dnr_dpmo": 1500, "pod": 97.5, "cc": 95.0,
                    "rts_pct": 1.5, "fdds": 98.5}


@api_router.get("/scorecard/targets")
async def get_scorecard_targets(center: Optional[str] = None, _=Depends(require_admin)):
    """Objetivos (umbrales) del scorecard que el usuario define a mano."""
    doc = await db.scorecard_targets.find_one(
        {"center": center or "GLOBAL"}, {"_id": 0}) if center else None
    if not doc:
        doc = await db.scorecard_targets.find_one({"center": "GLOBAL"}, {"_id": 0})
    targets = dict(_DEFAULT_TARGETS)
    if doc:
        for k in _DEFAULT_TARGETS:
            if doc.get(k) is not None:
                targets[k] = doc[k]
    return {"center": center or "GLOBAL", "targets": targets, "default": _DEFAULT_TARGETS}


@api_router.post("/scorecard/targets")
async def set_scorecard_targets(data: dict = Body(...), _=Depends(require_admin)):
    """body: {center?, dcr, dnr_dpmo, pod, cc, rts_pct, fdds} — los que envíes."""
    center = data.get("center") or "GLOBAL"
    upd = {"center": center}
    for k in _DEFAULT_TARGETS:
        if data.get(k) is not None:
            try:
                upd[k] = float(data[k])
            except Exception:
                pass
    upd["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.scorecard_targets.update_one({"center": center}, {"$set": upd}, upsert=True)
    return {"success": True, "center": center}


def _sun_sat_week(date_iso: str):
    """Devuelve (domingo, sábado) de la semana scorecard que contiene la fecha."""
    d = datetime.strptime(date_iso, "%Y-%m-%d")
    # weekday(): lunes=0 … domingo=6. Queremos retroceder al domingo anterior.
    back = (d.weekday() + 1) % 7   # domingo->0, lunes->1, … sábado->6
    sun = d - timedelta(days=back)
    sat = sun + timedelta(days=6)
    return sun.strftime("%Y-%m-%d"), sat.strftime("%Y-%m-%d")


@api_router.get("/scorecard/week-range")
async def scorecard_week_range(date: Optional[str] = None, _=Depends(require_admin)):
    """Rango dom-sáb de la semana scorecard. Si no se pasa fecha, usa el último
    día con dato disponible (hoy-2 por el desfase de Cortex)."""
    if not date:
        date = (datetime.now(timezone.utc) - timedelta(days=2)).strftime("%Y-%m-%d")
    sun, sat = _sun_sat_week(date)
    return {"desde": sun, "hasta": sat, "data_hasta": date}


@api_router.get("/metrics/daily-week")
async def daily_week(center: str, desde: str, hasta: str, _=Depends(require_admin)):
    """Acumulado de la semana (por centro): fallos por conductor, totales,
    motivos de POD y los conductores que más arrastran el scorecard."""
    docs = await db.daily_dsp.find(
        {"center": center, "date": {"$gte": desde, "$lte": hasta}}, {"_id": 0}
    ).sort("date", 1).to_list(40)

    # nombres por transporter_id
    drv_docs = await db.drivers.find(
        {"driver_id": {"$ne": None}}, {"_id": 0, "name": 1, "driver_id": 1}).to_list(2000)
    name_by_tid = {(d.get("driver_id") or "").upper(): d.get("name") for d in drv_docs}

    by_driver = {}
    totals = {"rts": 0, "dnr": 0, "pod": 0, "cc": 0}
    pod_reasons, cc_reasons = {}, {}
    days = []
    for doc in docs:
        days.append(doc["date"])
        for k in totals:
            totals[k] += doc.get("totals", {}).get(k, 0)
        for d in doc.get("drivers", []):
            tid = (d.get("transporter_id") or "").upper()
            e = by_driver.setdefault(tid, {"transporter_id": tid,
                "name": name_by_tid.get(tid), "rts": 0, "dnr": 0, "pod": 0, "cc": 0})
            for k in ("rts", "dnr", "pod", "cc"):
                e[k] += d.get(k, 0)
        for r, n in (doc.get("pod_reasons") or {}).items():
            pod_reasons[r] = pod_reasons.get(r, 0) + n
        for r, n in (doc.get("cc_reasons") or {}).items():
            cc_reasons[r] = cc_reasons.get(r, 0) + n

    ranking = sorted(by_driver.values(),
                     key=lambda e: -(e["pod"] * 2 + e["dnr"] * 3 + e["rts"] + e["cc"]))
    top_pod = sorted(pod_reasons.items(), key=lambda x: -x[1])
    return {"center": center, "dias": days, "totals": totals,
            "ranking": ranking, "pod_reasons": top_pod,
            "cc_reasons": sorted(cc_reasons.items(), key=lambda x: -x[1])}


# =========================
# SCORECARD COMPLETA EN VIVO — métricas reales de Amazon (Seguridad + Calidad)
#   Tiers calculados con umbrales sembrados de tu scorecard (editables y que se
#   afinan con cada scorecard que subas). Valores: manual + auto (daily).
# =========================

# 16 métricas reales del DSP Scorecard 3.0. dir(+1 más alto mejor / -1 más bajo mejor)
_SC_METRICS = [
    {"key": "fico",     "label": "Conducción segura (FICO)",          "group": "safety",   "unit": "score", "dir": 1,  "manual": True},
    {"key": "speeding", "label": "Eventos de velocidad /100",          "group": "safety",   "unit": "ratio", "dir": -1, "manual": True},
    {"key": "mentor",   "label": "Adopción del mentor",                "group": "safety",   "unit": "%",     "dir": 1,  "manual": True},
    {"key": "vsa",      "label": "Auditoría de vehículos (VSA)",       "group": "safety",   "unit": "%",     "dir": 1,  "manual": True},
    {"key": "whc",      "label": "Cumplimiento de horas (WHC)",        "group": "safety",   "unit": "%",     "dir": 1,  "manual": True},
    {"key": "dcr",      "label": "Finalización de entregas (DCR)",     "group": "quality",  "unit": "%",     "dir": 1,  "drill": "all"},
    {"key": "lor_dpmo", "label": "Perdido en ruta (LoR) DPMO",        "group": "quality",  "unit": "DPMO",  "dir": -1, "drill": "dnr"},
    {"key": "dsc_dpmo", "label": "Condiciones de entrega (DSC) DPMO", "group": "quality",  "unit": "DPMO",  "dir": -1},
    {"key": "cec_dpmo", "label": "Escalación del cliente (CEC) DPMO", "group": "quality",  "unit": "DPMO",  "dir": -1},
    {"key": "pod",      "label": "Foto en la entrega (POD)",           "group": "quality",  "unit": "%",     "dir": 1,  "drill": "pod"},
    {"key": "cc",       "label": "Normas de contacto (CC)",            "group": "quality",  "unit": "%",     "dir": 1,  "drill": "cc"},
    {"key": "ndcr",     "label": "Capacidad día siguiente",            "group": "capacity", "unit": "%",     "dir": 1,  "manual": True},
]

# Pesos oficiales por métrica — fuente: Excel "Peso metricas scorecard wk22.xlsx" DGA1.
# Suma = 100. DNR, BOC y CAS no aparecen como métricas ponderadas en Scorecard 3.0.
_SC_SEED_WEIGHTS = {
    "fico": 8.625, "speeding": 8.625, "mentor": 8.625, "vsa": 8.625, "whc": 5.625,
    "dcr": 15.625, "lor_dpmo": 5.625, "dsc_dpmo": 15.625,
    "cec_dpmo": 8.125, "pod": 5.625, "cc": 3.625,
    "ndcr": 5.625,
}

# Umbrales globales sembrados (se sobreescriben por centro con datos reales del PDF).
# Fantastic se ajusta cuando el usuario sube el PDF oficial — los valores de "great"/"fair"
# están basados en las proporciones habituales de Amazon Scorecard 3.0.
_SC_SEED_THR = {
    # Fuente: Excel "Targets scorecard nuevos" DGA1 — t0=Fantastic Plus, t1=Fantastic, t2=Great, t3=Fair
    "fico":     {"fantastic_plus": 850,  "fantastic": 810,  "great": 790,  "fair": 730},
    "speeding": {"fantastic_plus": 0,    "fantastic": 5,    "great": 7,    "fair": 10},
    "mentor":   {"fantastic_plus": 100,  "fantastic": 90,   "great": 82,   "fair": 75},
    "vsa":      {"fantastic_plus": 100,  "fantastic": 98.5, "great": 97,   "fair": 96},
    "whc":      {"fantastic_plus": 100,  "fantastic": 100,  "great": 97,   "fair": 95},
    "dcr":      {"fantastic_plus": 100,  "fantastic": 99,   "great": 98.5, "fair": 97},
    "lor_dpmo": {"fantastic_plus": 0,    "fantastic": 0,    "great": 40,   "fair": 85},
    "dsc_dpmo": {"fantastic_plus": 0,    "fantastic": 606,  "great": 906,  "fair": 1186},
    "cec_dpmo": {"fantastic_plus": 0,    "fantastic": 0.02, "great": 40,   "fair": 135},
    "pod":      {"fantastic_plus": 100,  "fantastic": 97,   "great": 95,   "fair": 90},
    "cc":       {"fantastic_plus": 100,  "fantastic": 98,   "great": 97,   "fair": 95},
    "ndcr":     {"fantastic_plus": 100,  "fantastic": 100,  "great": 95,   "fair": 90},
}
_TIER_ORDER = ["Poor", "Fair", "Great", "Fantastic", "Fantastic Plus"]

# Ancla semana Amazon ↔ domingo: la semana 23 de 2026 fue 31/05–06/06 (dom–sáb)
_SC_ANCHOR_SUN = "2026-05-31"
_SC_ANCHOR_WEEK = 23


def _week_num_to_sun(week_num):
    d = datetime.strptime(_SC_ANCHOR_SUN, "%Y-%m-%d") + timedelta(days=(int(week_num) - _SC_ANCHOR_WEEK) * 7)
    return d.strftime("%Y-%m-%d")


def _sun_to_week_num(sun):
    d = datetime.strptime(sun, "%Y-%m-%d")
    a = datetime.strptime(_SC_ANCHOR_SUN, "%Y-%m-%d")
    return _SC_ANCHOR_WEEK + round((d - a).days / 7)


def _sc_tier(value, thr, direction):
    if value is None or thr is None:
        return None
    fp, f, g, fa = thr.get("fantastic_plus"), thr.get("fantastic"), thr.get("great"), thr.get("fair")
    if None in (f, g, fa):
        return None
    if direction > 0:
        if fp is not None and value >= fp:
            return "Fantastic Plus"
        return "Fantastic" if value >= f else "Great" if value >= g else "Fair" if value >= fa else "Poor"
    if fp is not None and value <= fp:
        return "Fantastic Plus"
    return "Fantastic" if value <= f else "Great" if value <= g else "Fair" if value <= fa else "Poor"


def _interp(value, anchors):
    """Interpola linealmente sobre anclas (valor, puntos), clamp 0-100."""
    anchors = sorted(anchors)
    if value <= anchors[0][0]:
        (x0, y0), (x1, y1) = anchors[0], anchors[1]
        y = y0 if x1 == x0 else y0 + (value - x0) / (x1 - x0) * (y1 - y0)
        return max(0.0, min(100.0, y))
    for i in range(len(anchors) - 1):
        x0, y0 = anchors[i]
        x1, y1 = anchors[i + 1]
        if value <= x1:
            return y1 if x1 == x0 else y0 + (value - x0) / (x1 - x0) * (y1 - y0)
    (x0, y0), (x1, y1) = anchors[-2], anchors[-1]
    y = y1 if x1 == x0 else y1 + (value - x1) / (x1 - x0) * (y1 - y0)
    return max(0.0, min(100.0, y))


def _metric_subscore(value, thr, direction):
    """Sub-puntuación 0-100 de una métrica (modelo Amazon: t0=100, t1=90,
    t2=70, t3=50, interpolado). Calibrado con scorecards reales."""
    if value is None or thr is None:
        return None
    fp, f, g, fa = thr.get("fantastic_plus"), thr.get("fantastic"), thr.get("great"), thr.get("fair")
    if None in (f, g, fa):
        return None
    if fp is None:
        fp = f
    # Cuando fp == f (ej. NDCR=100, WHC=100), alcanzar ese valor es Fantastic Plus → 100 puntos.
    # La interpolación normal falla con dos anclas en el mismo x.
    if fp == f:
        if (direction > 0 and value >= fp) or (direction < 0 and value <= fp):
            return 100.0
        return _interp(value, [(f, 90), (g, 70), (fa, 50)])
    return _interp(value, [(fp, 100), (f, 90), (g, 70), (fa, 50)])


def _overall_score(out, weights):
    """Overall Score ponderado = Σ(peso × sub-puntuación) / Σpesos."""
    num = den = 0.0
    usados = 0
    for m in out:
        w = weights.get(m["key"], 0) or 0
        if not w:
            continue
        ss = _metric_subscore(m.get("value"), m.get("thr"), m["dir"])
        if ss is None:
            continue
        num += w * ss
        den += w
        usados += 1
    return (round(num / den, 2) if den else None), den, usados


_OVERALL_TIER_BANDS = [(93, "Fantastic Plus"), (85, "Fantastic"), (70, "Great"), (50, "Fair"), (0, "Poor")]


def _score_to_tier(score):
    if score is None:
        return None
    for thr, name in _OVERALL_TIER_BANDS:
        if score >= thr:
            return name
    return "Poor"


def _sc_next_target(value, tier, thr, direction):
    """Qué valor hace falta para subir al siguiente tier."""
    if value is None or thr is None or tier == "Fantastic Plus":
        return None
    nxt = {"Poor": "fair", "Fair": "great", "Great": "fantastic", "Fantastic": "fantastic_plus"}.get(tier)
    if not nxt:
        return None
    target = thr.get(nxt) or (thr.get("fantastic") if nxt == "fantastic_plus" else None)
    if target is None:
        return None
    gap = round((target - value) if direction > 0 else (value - target), 2)
    label = {"fair": "Fair", "great": "Great", "fantastic": "Fantastic", "fantastic_plus": "Fantastic Plus"}[nxt]
    return {"to_tier": label, "target": target, "gap": gap}


async def _sc_thresholds(center):
    """Umbrales: semilla + override del usuario (por centro o global)."""
    thr = {k: dict(v) for k, v in _SC_SEED_THR.items()}
    for c in ("GLOBAL", center):
        if not c:
            continue
        doc = await db.scorecard_thresholds.find_one({"center": c}, {"_id": 0})
        if doc:
            for k in thr:
                if isinstance(doc.get(k), dict):
                    thr[k].update({kk: vv for kk, vv in doc[k].items() if vv is not None})
    return thr


async def _latest_week_with_data(center):
    """Domingo de la semana que estás siguiendo. Manda la semana de tus REPORTES
    DIARIOS (lo que subes día a día); si no hay, la oficial; luego el resumen."""
    # 1º: la semana de los reportes diarios (señal de la semana en curso que sigues)
    d = await db.daily_dsp.find_one({"center": center}, {"date": 1}, sort=[("date", -1)])
    if d and d.get("date"):
        return _sun_sat_week(d["date"])[0]
    r = await db.daily_ratios.find_one({"center": center}, {"date": 1}, sort=[("date", -1)])
    if r and r.get("date"):
        return _sun_sat_week(r["date"])[0]
    # 2º: scorecard oficial más reciente
    o = await db.scorecard_official.find_one({"center": center}, {"week": 1}, sort=[("week", -1)])
    if o and o.get("week"):
        return _week_num_to_sun(int(o["week"]))
    # 3º: resumen semanal (puede ser la semana en curso a medias)
    w = await db.scorecard_weekly.find_one({"center": center}, {"week": 1}, sort=[("week", -1)])
    if w and w.get("week"):
        return _week_num_to_sun(int(w["week"]))
    return _sun_sat_week((datetime.now(timezone.utc) - timedelta(days=2)).strftime("%Y-%m-%d"))[0]


@api_router.get("/scorecard/full")
async def scorecard_full(center: str, week: Optional[str] = None, user: dict = Depends(require_admin)):
    """Scorecard completa de la semana (dom): cada métrica con valor + tier +
    qué falta para subir. Valores manuales (db) y auto donde haya."""
    await _require_plan_feature(user, "scorecard")
    if not week:
        week = await _latest_week_with_data(center)
    sun, sat = _sun_sat_week(week)
    thr = await _sc_thresholds(center)
    doc = await db.scorecard_live.find_one({"center": center, "week": sun}, {"_id": 0})
    values = (doc or {}).get("values", {})
    # ¿hay scorecard OFICIAL de esa semana? → valores+tiers REALES de Amazon
    wnum = _sun_to_week_num(sun)
    off = await db.scorecard_official.find_one({"center": center, "week": wnum}, {"_id": 0})
    off_metrics = {mm.get("key"): mm for mm in (off.get("metrics") if off else [])}
    has_official = bool(off)
    # ratios diarios subidos (para rellenar Calidad de la semana en vivo)
    ratio_vals, ratio_dias = await _ratios_week_values(center, sun, sat)
    # Resumen de entregas semanal (DSP Resumen Entregas): DCR/DNR/POD reales de la semana
    wk_doc = await db.scorecard_weekly.find_one({"center": center, "week": wnum}, {"_id": 0})
    week_vals = (wk_doc or {}).get("values", {})
    # Baseline: arrastra Seguridad/Capacidad de tu ÚLTIMA scorecard conocida.
    # Los reportes diarios solo traen Calidad; Seguridad cambia muy poco semana a
    # semana, así que la proyectamos con el último dato real (no inventado) para que
    # la nota de la semana en curso cubra TODO el peso sin volver a subir la scorecard.
    base = None
    estimar = bool((doc or {}).get("estimar"))  # OPT-IN: por defecto NO se estima nada
    if not has_official and estimar:
        base = await db.scorecard_official.find_one(
            {"center": center, "week": {"$lt": wnum}}, {"_id": 0}, sort=[("week", -1)])
    base_metrics = {mm.get("key"): mm for mm in (base.get("metrics") if base else [])}

    out = []
    counts = {}
    for m in _SC_METRICS:
        om = off_metrics.get(m["key"])
        if om and om.get("value") is not None:
            v = om.get("value")
            t = om.get("tier") or _sc_tier(v, thr.get(m["key"]), m["dir"])
            src = "oficial"
        elif values.get(m["key"]) is not None:
            v = values.get(m["key"])
            t = _sc_tier(v, thr.get(m["key"]), m["dir"])
            src = "manual"
        elif week_vals.get(m["key"]) is not None:
            v = week_vals.get(m["key"])
            t = _sc_tier(v, thr.get(m["key"]), m["dir"])
            src = "resumen"
        elif ratio_vals.get(m["key"]) is not None:
            v = ratio_vals.get(m["key"])
            t = _sc_tier(v, thr.get(m["key"]), m["dir"])
            src = "ratios"
        elif base_metrics.get(m["key"]) and base_metrics[m["key"]].get("value") is not None:
            v = base_metrics[m["key"]].get("value")
            t = _sc_tier(v, thr.get(m["key"]), m["dir"])
            src = "estimado"  # arrastrado de la última scorecard conocida
        else:
            v = None
            t = None
            src = None
        counts[t or "Sin datos"] = counts.get(t or "Sin datos", 0) + 1
        out.append({**m, "value": v, "tier": t, "thr": thr.get(m["key"]), "source": src,
                    "next": _sc_next_target(v, t, thr.get(m["key"]), m["dir"])})

    to_improve = sorted(
        [{"key": m["key"], "label": m["label"], "group": m["group"], "tier": m["tier"],
          "value": m["value"], "next": m["next"]}
         for m in out if m["tier"] and m["tier"] not in ("Fantastic", "Fantastic Plus") and m["next"]],
        key=lambda m: abs(m["next"]["gap"]) if m["next"] else 9e9)

    def _cat_tier(group):
        ts = [m["tier"] for m in out if m["group"] == group and m["tier"]]
        if not ts:
            return None
        cnt = {}
        for t in ts:
            cnt[t] = cnt.get(t, 0) + 1
        mx = max(cnt.values())
        return max([t for t in cnt if cnt[t] == mx], key=lambda t: _TIER_ORDER.index(t))
    safety_tier = _cat_tier("safety")
    quality_tier = _cat_tier("quality")
    capacity_tier = _cat_tier("capacity")

    weights = await _sc_weights(center)
    score_calc, wsum, nused = _overall_score(out, weights)
    peso_total = sum(v for v in weights.values() if isinstance(v, (int, float)))
    cobertura = round(wsum / peso_total * 100) if peso_total else 0

    if has_official:
        overall = off.get("overall_tier")
        overall_score = off.get("overall_score")
        cats = off.get("categories") or {}
        safety_tier = cats.get("compliance_safety") or safety_tier
        quality_tier = cats.get("quality_swc") or quality_tier
        capacity_tier = cats.get("capacity") or capacity_tier
        overall_method = "oficial — de tu scorecard de Amazon"
    else:
        overall_score = score_calc
        overall = _score_to_tier(score_calc)
        n_real = sum(1 for m in out if m["source"] in ("resumen", "ratios", "manual", "oficial"))
        n_est = sum(1 for m in out if m["source"] == "estimado")
        if score_calc is None:
            overall_method = "faltan datos para la nota"
        elif n_est:
            overall_method = (f"{n_real} métricas reales de tus archivos + {n_est} estimadas "
                              f"de la W{base.get('week') if base else '?'} · {cobertura}% del peso")
        else:
            overall_method = f"solo datos reales de tus archivos — {cobertura}% del peso ({n_real} métricas)"

    return {"center": center, "week": sun, "desde": sun, "hasta": sat, "week_num": wnum,
            "metrics": out, "counts": counts, "to_improve": to_improve,
            "safety_tier": safety_tier, "quality_tier": quality_tier, "capacity_tier": capacity_tier,
            "overall": overall, "overall_score": overall_score,
            "score_calculado": score_calc, "cobertura_peso": cobertura,
            "estimada_desde": (base.get("week") if base else None),
            "dias_ratios": len(ratio_dias), "estimacion_on": estimar,
            "has_official": has_official, "overall_method": overall_method}


@api_router.post("/scorecard/full")
async def scorecard_set_value(data: dict = Body(...), user: dict = Depends(require_admin)):
    """Guarda el valor de una métrica. body: {center, week(dom), key, value}"""
    await _require_plan_feature(user, "scorecard")
    center, week, key = data.get("center"), data.get("week"), data.get("key")
    if not (center and week and key):
        raise HTTPException(status_code=400, detail="center, week y key requeridos")
    sun, _sat = _sun_sat_week(week)
    val = data.get("value")
    try:
        val = float(val) if val not in (None, "") else None
    except Exception:
        val = None
    await db.scorecard_live.update_one(
        {"center": center, "week": sun},
        {"$set": {"center": center, "week": sun, f"values.{key}": val,
                  "updated_at": datetime.now(timezone.utc).isoformat()}}, upsert=True)
    return {"success": True}


@api_router.post("/scorecard/thresholds")
async def scorecard_set_thresholds(data: dict = Body(...), _=Depends(require_admin)):
    """body: {center?, key, fantastic, great, fair}"""
    center = data.get("center") or "GLOBAL"
    key = data.get("key")
    if key not in _SC_SEED_THR:
        raise HTTPException(status_code=400, detail="key inválida")
    band = {}
    for b in ("fantastic", "great", "fair"):
        if data.get(b) is not None:
            try:
                band[b] = float(data[b])
            except Exception:
                pass
    await db.scorecard_thresholds.update_one(
        {"center": center}, {"$set": {"center": center, key: band}}, upsert=True)
    return {"success": True}


@api_router.delete("/scorecard/thresholds")
async def scorecard_reset_thresholds(center: str, _=Depends(require_admin)):
    """Escribe los seeds oficiales de Amazon directamente en MongoDB para el centro.
    Sobreescribe cualquier calibración anterior con los valores verificados del Excel de targets."""
    doc = {"center": center}
    for key, bands in _SC_SEED_THR.items():
        doc[key] = {b: v for b, v in bands.items() if v is not None}
    await db.scorecard_thresholds.replace_one({"center": center}, doc, upsert=True)
    return {"success": True, "center": center, "metricas": list(_SC_SEED_THR.keys()),
            "mensaje": f"Umbrales de {center} escritos directamente desde los targets oficiales de Amazon."}


_OFFICIAL_SC_PROMPT = """Eres un analista de Amazon DSP. Te paso la scorecard semanal OFICIAL (Scorecard 3.0) de un DSP.
Extrae EXACTAMENTE lo que aparece, sin inventar nada. Si un valor no está o pone "None"/"N/A", déjalo null.

Devuelve SOLO JSON con esta estructura:
{
 "week": <número de semana, int>,
 "year": <año, int>,
 "center": "<código del centro, ej OGA5, DGA1>",
 "overall_score": <número, ej 91.99, o null>,
 "overall_tier": "<Fantastic Plus|Fantastic|Great|Fair|Poor|At Risk>",
 "categories": {
   "compliance_safety": "<tier>", "quality_swc": "<tier>", "capacity": "<tier>"
 },
 "metrics": [
   {"key":"fico","value":<num o null>,"tier":"<tier o null>"},
   {"key":"speeding","value":<num o null>,"tier":"<tier o null>"},
   {"key":"mentor","value":<num o null>,"tier":"<tier o null>"},
   {"key":"vsa","value":<num o null>,"tier":"<tier o null>"},
   {"key":"boc","value":<num o null>,"tier":"<tier o null>"},
   {"key":"whc","value":<num o null>,"tier":"<tier o null>"},
   {"key":"cas","value":<num o null>,"tier":"<tier o null>"},
   {"key":"dcr","value":<num o null>,"tier":"<tier o null>"},
   {"key":"dnr_dpmo","value":<num o null>,"tier":"<tier o null>"},
   {"key":"lor_dpmo","value":<num o null>,"tier":"<tier o null>"},
   {"key":"dsc_dpmo","value":<num o null>,"tier":"<tier o null>"},
   {"key":"cec_dpmo","value":<num o null>,"tier":"<tier o null>"},
   {"key":"cdf","value":<num o null>,"tier":"<tier o null>"},
   {"key":"pod","value":<num o null>,"tier":"<tier o null>"},
   {"key":"cc","value":<num o null>,"tier":"<tier o null>"},
   {"key":"ndcr","value":<num o null>,"tier":"<tier o null>"}
 ],
 "explicit_thresholds": {
   "fico":     {"fantastic": <num o null>, "dir": 1},
   "speeding": {"fantastic": <num o null>, "dir": -1},
   "mentor":   {"fantastic": <num o null>, "dir": 1},
   "vsa":      {"fantastic": <num o null>, "dir": 1},
   "whc":      {"fantastic": <num o null>, "dir": 1},
   "cas":      {"fantastic": <num o null>, "dir": 1},
   "dcr":      {"fantastic": <num o null>, "dir": 1},
   "dnr_dpmo": {"fantastic": <num o null>, "dir": -1},
   "lor_dpmo": {"fantastic": <num o null>, "dir": -1},
   "dsc_dpmo": {"fantastic": <num o null>, "dir": -1},
   "cec_dpmo": {"fantastic": <num o null>, "dir": -1},
   "cdf":      {"fantastic": <num o null>, "dir": -1},
   "pod":      {"fantastic": <num o null>, "dir": 1},
   "cc":       {"fantastic": <num o null>, "dir": 1},
   "ndcr":     {"fantastic": <num o null>, "dir": 1}
 }
}

REGLAS CRÍTICAS:
1. En "metrics": extrae los valores y tiers de la PÁGINA DE RESULTADOS (página 2).
   Para %, devuelve solo el número (98.33, no "98.33%"). Null si no aparece o es N/A.
2. En "explicit_thresholds": busca frases como "A DSP who has Fantastic standing would receive a X in METRIC"
   en las PÁGINAS DE DEFINICIONES (páginas 5-7). Extrae el número X.
   SOLO pon valores que estén EXPLÍCITAMENTE en el texto — no inventes ni inferias.
   Si el texto dice ">= 99%" para DCR, pon 99. Si dice "<= 1080" para DNR, pon 1080.
3. Mapeo nombres → key:
   Safe Driving Metric/FICO=fico; Speeding Event Rate=speeding; Mentor Adoption Rate=mentor;
   Vehicle Audit/VSA=vsa; Breach of Contract/BOC=boc; Working Hours Compliance/WHC=whc;
   Comprehensive Audit Score/CAS=cas; Delivery Completion Rate/DCR=dcr;
   Delivered Not Received/DNR DPMO=dnr_dpmo; Lost on Road/LoR DPMO=lor_dpmo;
   Delivery Success Conditions/DSC DPMO=dsc_dpmo; Customer escalation DPMO=cec_dpmo;
   Customer Delivery Feedback/CDF=cdf; Photo-On-Delivery/POD=pod;
   Contact Compliance=cc; Next Day Capacity Reliability=ndcr."""


async def _parse_official_scorecard(content: bytes, filename: str):
    text, pdf_bytes = await _extract_report_text(content, filename)
    from google import genai as genai_sdk
    from google.genai import types as genai_types
    use_vertex = os.environ.get("USE_VERTEX_AI", "").lower() in ("1", "true", "yes")
    if use_vertex:
        from google.oauth2 import service_account
        import json as _json, base64 as _b64
        sa = os.environ.get("GCP_SERVICE_ACCOUNT_JSON", "").strip()
        if sa and not sa.startswith("{"):
            sa = _b64.b64decode(sa).decode("utf-8")
        creds = service_account.Credentials.from_service_account_info(
            _json.loads(sa), scopes=["https://www.googleapis.com/auth/cloud-platform"]) if sa else None
        client = genai_sdk.Client(vertexai=True, project=os.environ.get("GCP_PROJECT", ""),
                                  location=os.environ.get("GCP_LOCATION", "us-central1"), credentials=creds)
    else:
        client = genai_sdk.Client(api_key=os.environ.get("GEMINI_API_KEY", ""))
    contents = [_OFFICIAL_SC_PROMPT]
    if pdf_bytes:
        contents.append(genai_types.Part.from_bytes(data=pdf_bytes, mime_type="application/pdf"))
    else:
        contents.append("SCORECARD:\n\n" + (text or ""))
    cfg = genai_types.GenerateContentConfig(temperature=0.0, response_mime_type="application/json")
    loop = asyncio.get_running_loop()
    async with _gemini_sem:
        resp = await asyncio.wait_for(
            loop.run_in_executor(_executor, lambda: client.models.generate_content(
                model=os.environ.get("GEMINI_MODEL", "gemini-2.5-flash"), contents=contents, config=cfg)),
            timeout=120.0)
    return json.loads(_strip_markdown_json(resp.text or "{}"))


@api_router.post("/scorecard/import-official")
async def import_official_scorecard(file: UploadFile = File(...), center: Optional[str] = Form(None),
                                    _=Depends(require_admin)):
    """Sube la scorecard OFICIAL (PDF). Extrae métricas+tiers reales y el Overall
    Score, y los guarda como observaciones para afinar umbrales (verídico)."""
    content = await file.read()
    try:
        sc = await _parse_official_scorecard(content, file.filename or "scorecard.pdf")
    except Exception as e:
        logger.error(f"import-official gemini: {type(e).__name__}: {repr(e)}")
        raise HTTPException(status_code=502, detail="La IA no pudo leer la scorecard; reintenta.")
    cen = (sc.get("center") or center or "").upper() or "OGA5"
    week = sc.get("week")
    year = sc.get("year")
    if not week:
        raise HTTPException(status_code=400, detail="No reconocí la semana en la scorecard")
    explicit_thr = sc.get("explicit_thresholds") or {}
    doc = {"center": cen, "week": int(week), "year": int(year) if year else None,
           "overall_score": sc.get("overall_score"), "overall_tier": sc.get("overall_tier"),
           "categories": sc.get("categories") or {}, "metrics": sc.get("metrics") or [],
           "explicit_thresholds": explicit_thr,
           "uploaded_at": datetime.now(timezone.utc).isoformat()}
    await db.scorecard_official.update_one(
        {"center": cen, "week": int(week)}, {"$set": doc}, upsert=True)

    # Guardar observaciones (valor→tier) para derivar umbrales por inferencia
    for m in (sc.get("metrics") or []):
        if m.get("key") and m.get("value") is not None and m.get("tier"):
            await db.scorecard_obs.update_one(
                {"center": cen, "week": int(week), "metric": m["key"]},
                {"$set": {"center": cen, "week": int(week), "metric": m["key"],
                          "value": m["value"], "tier": m["tier"]}}, upsert=True)

    # Si el PDF tenía umbrales explícitos, aplicarlos directamente como thresholds del centro
    thr_aplicados = []
    if explicit_thr:
        thr_update = {"center": cen}
        for key, info in explicit_thr.items():
            f_val = info.get("fantastic") if isinstance(info, dict) else None
            if f_val is not None and key in _SC_SEED_THR:
                # Solo actualiza el campo "fantastic"; mantiene great/fair existentes
                thr_update[f"{key}.fantastic"] = float(f_val)
                thr_aplicados.append(key)
        if thr_update:
            await db.scorecard_thresholds.update_one(
                {"center": cen}, {"$set": thr_update}, upsert=True)

    n_obs = await db.scorecard_official.count_documents({"center": cen})
    return {"success": True, "center": cen, "week": week, "year": year,
            "overall_score": sc.get("overall_score"), "overall_tier": sc.get("overall_tier"),
            "metricas": len(sc.get("metrics") or []), "scorecards_guardadas": n_obs,
            "umbrales_explicitos_aplicados": thr_aplicados}


@api_router.get("/scorecard/official")
async def list_official_scorecards(center: str, _=Depends(require_admin)):
    docs = await db.scorecard_official.find({"center": center}, {"_id": 0}).sort("week", -1).to_list(60)
    return {"scorecards": docs}


# =========================
# SUBIDA UNIFICADA + PREDICCIÓN DE SCORE (datos reales de los archivos)
# =========================

_MESES_ES = {"enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
             "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
             "noviembre": 11, "diciembre": 12}


def _read_table_any(content: bytes, filename: str):
    """Devuelve filas (listas de str) de xlsx/xls/csv/html."""
    fn = (filename or "").lower()
    rows = []
    if fn.endswith(".xlsx") or fn.endswith(".xlsm"):
        import openpyxl, io
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        for ws in wb.worksheets:
            for r in ws.iter_rows(values_only=True):
                rows.append(["" if c is None else str(c) for c in r])
    elif fn.endswith(".xls"):
        import xlrd
        book = xlrd.open_workbook(file_contents=content)
        for sh in book.sheets():
            for ri in range(sh.nrows):
                rows.append([str(sh.cell_value(ri, ci)) for ci in range(sh.ncols)])
    elif fn.endswith(".csv"):
        import csv, io
        txt = content.decode("utf-8", errors="ignore")
        for r in csv.reader(io.StringIO(txt)):
            rows.append(r)
    else:
        html = content.decode("utf-8", errors="ignore")
        for tbl in re.findall(r"<table.*?</table>", html, re.S | re.I):
            for tr in re.findall(r"<tr.*?</tr>", tbl, re.S | re.I):
                cells = [re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", c)).strip()
                         for c in re.findall(r"<t[hd].*?</t[hd]>", tr, re.S | re.I)]
                if any(cells):
                    rows.append(cells)
    return rows


def _es_date(s: str, year: int):
    s = (s or "").lower().strip()
    m = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    for nombre, num in _MESES_ES.items():
        m = re.search(nombre + r"\s+(\d{1,2})", s) or re.search(r"(\d{1,2})\s+" + nombre, s)
        if m:
            return f"{year}-{num:02d}-{int(m.group(1)):02d}"
    return None


def _num(s):
    s = str(s).strip()
    if s in ("", "-", "Sin datos", "None", "N/A"):
        return None
    s = s.replace("%", "").strip()
    s = s.replace(",", "") if (s.count(",") and "." in s) else s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


# Mapa de filas de "Descripción general" (Cortex, ES) → claves. COINCIDENCIA
# EXACTA del nombre de fila (evita que "paquetes entregados no recibidos"
# colisione con "paquetes entregados").
_RATIO_ROWS = [
    (["entrega correcta (%) - dsp", "entrega correcta (%)-dsp"], "dcr"),
    (["dpmo de dnr"], "dnr_dpmo"),
    (["tasa de éxito de pod", "tasa de exito de pod"], "pod"),
    (["paquetes devueltos al centro (rts) %"], "rts_pct"),
    (["éxito de entrega en el primer día (%)", "exito de entrega en el primer día (%)",
      "éxito de entrega en el primer dia (%)"], "fdds"),
    (["paquetes enviados"], "c_enviados"),
    (["paquetes entregados no recibidos (dnr)", "paquetes entregados no recibidos"], "c_dnr"),
    (["paquetes entregados"], "c_entregados"),
    (["oportunidades de pod"], "c_pod_opp"),
    (["éxito de pod", "exito de pod"], "c_pod_succ"),
]


def _parse_ratios(content: bytes, filename: str):
    """Parsea la 'Descripción general' (ratios diarios) → {days:{date:{dcr,pod,...,conteos}}}."""
    rows = _read_table_any(content, filename)
    if not rows:
        return None
    year = datetime.now(timezone.utc).year
    # localizar fila cabecera: la que tenga >=2 fechas reconocibles
    hdr_i, date_cols = None, {}
    for i, r in enumerate(rows[:8]):
        cols = {}
        for ci in range(1, len(r)):
            d = _es_date(r[ci], year)
            if d:
                cols[ci] = d
        if len(cols) >= 2:
            hdr_i, date_cols = i, cols
            break
    if not date_cols:
        return None
    days = {d: {} for d in date_cols.values()}
    for r in rows[hdr_i + 1:]:
        if not r:
            continue
        label = (r[0] or "").lower().strip()
        key = next((k for names, k in _RATIO_ROWS if label in names), None)
        if not key:
            continue
        for ci, date in date_cols.items():
            if ci < len(r):
                v = _num(r[ci])
                if v is not None:
                    days[date][key] = v
    return {"days": days} if any(days.values()) else None


def _parse_resumen_semanal(content: bytes, filename: str):
    """Parsea el 'DSP Resumen Entregas' de Cortex: filas = ratios, columnas =
    'Semana 22/23/24/25'. Devuelve {week_num: {dcr,dnr_dpmo,pod,...}} con métricas reales."""
    rows = _read_table_any(content, filename)
    if not rows:
        return None
    # cabecera: fila con >=2 columnas tipo "Semana NN"
    hdr_i, week_cols = None, {}
    for i, r in enumerate(rows[:6]):
        cols = {}
        for ci in range(1, len(r)):
            m = re.search(r"semana\s*(\d{1,2})", str(r[ci]).lower())
            if m:
                cols[ci] = int(m.group(1))
        if len(cols) >= 1:
            hdr_i, week_cols = i, cols
            break
    if not week_cols:
        return None
    out = {wn: {} for wn in week_cols.values()}
    for r in rows[hdr_i + 1:]:
        if not r:
            continue
        label = (r[0] or "").lower().strip()
        key = next((k for names, k in _RATIO_ROWS if label in names), None)
        if not key:
            continue
        for ci, wn in week_cols.items():
            if ci < len(r):
                v = _num(r[ci])
                if v is not None:
                    out[wn][key] = v
    # deja solo semanas con métricas de scorecard útiles (dcr/dnr/pod)
    res = {wn: v for wn, v in out.items() if any(k in v for k in ("dcr", "dnr_dpmo", "pod"))}
    return res or None


def _parse_contact_compliance(content: bytes, filename: str):
    """Contact Compliance Report (HTML): tabla 'Driver Summary' con Total Addresses /
    Total Contacts por conductor. CC del DSP = Σcontactos / Σdirecciones × 100."""
    html = content.decode("utf-8", "ignore")
    mw = re.search(r"(\d{4})[-\s]+(\d{1,2})\b", html)
    week = int(mw.group(2)) if mw else None
    # filas de 4 celdas: ID, direcciones, contactos, %  (solo la tabla Driver Summary)
    rows = re.findall(
        r"<td>\s*([A-Z0-9]{8,})\s*</td>\s*<td>\s*(\d+)\s*</td>\s*<td>\s*(\d+)\s*</td>\s*<td>\s*[\d.]+\s*%",
        html)
    if not rows or not week:
        return None
    tot_addr = sum(int(r[1]) for r in rows)
    tot_con = sum(int(r[2]) for r in rows)
    if not tot_addr:
        return None
    return {"week": week, "key": "cc", "value": round(tot_con / tot_addr * 100, 2),
            "detalle": f"{tot_con}/{tot_addr} contactos ({len(rows)} conductores)"}


async def _store_weekly_metric(center, week, key, value):
    """Guarda UNA métrica semanal (de un reporte por métrica) en scorecard_weekly."""
    cen = (center or "OGA5").upper()
    await db.scorecard_weekly.update_one(
        {"center": cen, "week": int(week)},
        {"$set": {"center": cen, "week": int(week), f"values.{key}": value,
                  "uploaded_at": datetime.now(timezone.utc).isoformat()}}, upsert=True)


@api_router.post("/scorecard/upload")
async def scorecard_upload(file: UploadFile = File(...), center: Optional[str] = Form(None),
                           _=Depends(require_admin)):
    """Subida UNIFICADA: detecta el tipo (PDF scorecard / HTML reporte diario /
    Excel-CSV ratios), valida, parsea, guarda y devuelve estado claro."""
    content = await file.read()
    fn = (file.filename or "").lower()
    if not content:
        raise HTTPException(status_code=400, detail="Archivo vacío")
    import hashlib
    fhash = hashlib.md5(content).hexdigest()
    ext = fn.rsplit(".", 1)[-1] if "." in fn else ""
    if ext not in ("pdf", "html", "htm", "xlsx", "xls", "xlsm", "csv"):
        raise HTTPException(status_code=400, detail=f"Formato no soportado: .{ext}")
    logger.info(f"[upload] {fn} ({len(content)}B, {ext})")

    # Reportes de UNA métrica suelta (no son la scorecard ni el Resumen): de momento
    # esas métricas se meten a mano. Mensaje claro en vez de guardarlas mal.
    # Reporte de Customer Escalation (PDF de imagen) → métrica CEC. De momento a mano.
    if "escalation" in fn:
        raise HTTPException(status_code=422, detail="El reporte de Customer Escalation es un PDF de imagen (sin texto). Mete el CEC a mano: clic en su número en 'Calidad'. (Pronto lo leeré por OCR.)")
    # Reporte de Contact Compliance → métrica CC (calculada de la tabla por conductor)
    if "compliance" in fn and ("contact" in fn or "concession" in fn):
        cc = _parse_contact_compliance(content, fn)
        if not cc:
            raise HTTPException(status_code=422, detail="No pude leer la tabla del Contact Compliance.")
        await _store_weekly_metric(center, cc["week"], cc["key"], cc["value"])
        return {"tipo": "metrica", "ok": True, "center": (center or "OGA5").upper(),
                "mensaje": f"Contact Compliance W{cc['week']}: CC = {cc['value']}% ({cc['detalle']})"}

    try:
        # PDF → scorecard oficial
        if ext == "pdf":
            sc = await _parse_official_scorecard(content, file.filename or "")
            cen = (sc.get("center") or center or "").upper() or "OGA5"
            week = sc.get("week")
            if not week:
                raise HTTPException(status_code=422, detail="No reconocí la semana en el PDF (¿es una scorecard oficial?)")
            if sc.get("overall_score") is None and sc.get("overall_tier") in (None, "None") and not (sc.get("metrics") or []):
                raise HTTPException(status_code=422, detail=f"El PDF no parece una scorecard completa (no trae nota global). Si es un reporte de una métrica suelta, mete ese valor a mano.")
            doc = {"center": cen, "week": int(week), "year": sc.get("year"),
                   "overall_score": sc.get("overall_score"), "overall_tier": sc.get("overall_tier"),
                   "categories": sc.get("categories") or {}, "metrics": sc.get("metrics") or [],
                   "hash": fhash, "uploaded_at": datetime.now(timezone.utc).isoformat()}
            await db.scorecard_official.update_one({"center": cen, "week": int(week)}, {"$set": doc}, upsert=True)
            for m in (sc.get("metrics") or []):
                if m.get("key") and m.get("value") is not None and m.get("tier"):
                    await db.scorecard_obs.update_one(
                        {"center": cen, "week": int(week), "metric": m["key"]},
                        {"$set": {"center": cen, "week": int(week), "metric": m["key"],
                                  "value": m["value"], "tier": m["tier"]}}, upsert=True)
            return {"tipo": "scorecard", "ok": True, "center": cen,
                    "mensaje": f"Scorecard W{week}: {sc.get('overall_tier')} ({sc.get('overall_score')})",
                    "metricas": len(sc.get("metrics") or [])}

        # HTML → reporte diario (fallos) o ratios (Descripción general)
        if ext in ("html", "htm"):
            parsed = _parse_daily_report(content, file.filename or "")
            if parsed.get("center") and parsed.get("date") and parsed.get("drivers"):
                cen = parsed["center"]
                doc = dict(parsed); doc["hash"] = fhash
                doc["uploaded_at"] = datetime.now(timezone.utc).isoformat()
                await db.daily_dsp.update_one({"center": cen, "date": parsed["date"]}, {"$set": doc}, upsert=True)
                return {"tipo": "reporte_diario", "ok": True, "center": cen,
                        "mensaje": f"Reporte {parsed['date']}: {len(parsed['drivers'])} conductores",
                        "fecha": parsed["date"]}
            ratios = _parse_ratios(content, file.filename or "")
            if ratios:
                return await _store_ratios(ratios, center)
            raise HTTPException(status_code=422,
                                detail="HTML no reconocido (ni reporte diario ni Descripción general)")

        # Excel/CSV → 1º Resumen semanal (columnas Semana NN), 2º Descripción general (fechas)
        sem = _parse_resumen_semanal(content, file.filename or "")
        if sem:
            return await _store_resumen_semanal(sem, center)
        ratios = _parse_ratios(content, file.filename or "")
        if not ratios:
            raise HTTPException(status_code=422,
                                detail="Excel/CSV no reconocido. ¿Es la 'Descripción general' o el 'Resumen de entregas' de Cortex?")
        return await _store_ratios(ratios, center)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"[upload] {fn} ERROR: {type(e).__name__}: {repr(e)}")
        raise HTTPException(status_code=422, detail=f"No se pudo procesar {fn}: {type(e).__name__}")


@api_router.post("/scorecard/reset")
async def scorecard_reset(data: dict = Body(...), _=Depends(require_admin)):
    """Pone la semana A CERO: borra ratios, reportes diarios, valores a mano y la
    scorecard oficial de ESA semana, y desactiva la estimación de W21. A partir de
    ahí la nota se calcula SOLO con lo que subas/edites (datos reales)."""
    center = data.get("center")
    week = data.get("week")
    if not center:
        raise HTTPException(status_code=400, detail="center requerido")
    if not week:
        week = (datetime.now(timezone.utc) - timedelta(days=2)).strftime("%Y-%m-%d")
    sun, sat = _sun_sat_week(week)
    wnum = _sun_to_week_num(sun)
    r1 = await db.daily_ratios.delete_many({"center": center, "date": {"$gte": sun, "$lte": sat}})
    r2 = await db.daily_dsp.delete_many({"center": center, "date": {"$gte": sun, "$lte": sat}})
    r3 = await db.scorecard_official.delete_one({"center": center, "week": wnum})
    await db.scorecard_obs.delete_many({"center": center, "week": wnum})
    await db.scorecard_live.update_one(
        {"center": center, "week": sun},
        {"$set": {"center": center, "week": sun, "values": {}, "estimar": False,
                  "updated_at": datetime.now(timezone.utc).isoformat()}}, upsert=True)
    return {"ok": True, "semana": sun,
            "borrados": {"ratios": r1.deleted_count, "diarios": r2.deleted_count,
                         "oficial": r3.deleted_count}}


@api_router.post("/scorecard/estimacion")
async def scorecard_estimacion(data: dict = Body(...), _=Depends(require_admin)):
    """Activa/desactiva el relleno de huecos con la última scorecard conocida.
    on=True → estima lo que falta (nota completa). on=False → solo datos reales."""
    center = data.get("center")
    week = data.get("week")
    on = bool(data.get("on"))
    if not (center and week):
        raise HTTPException(status_code=400, detail="center y week requeridos")
    sun, _ = _sun_sat_week(week)
    await db.scorecard_live.update_one(
        {"center": center, "week": sun},
        {"$set": {"center": center, "week": sun, "estimar": on}}, upsert=True)
    return {"ok": True, "estimacion": on}


@api_router.get("/scorecard/daily-trend")
async def scorecard_daily_trend(center: str, week: Optional[str] = None, _=Depends(require_admin)):
    """Evolución DÍA A DÍA de la semana (del Resumen diario): cada día su DCR/DNR/POD
    y el ACUMULADO hasta ese día. Para saber 'cómo vamos' un miércoles."""
    if not week:
        week = await _latest_week_with_data(center)
    sun, sat = _sun_sat_week(week)
    docs = await db.daily_ratios.find(
        {"center": center, "date": {"$gte": sun, "$lte": sat}}, {"_id": 0}).sort("date", 1).to_list(10)
    dias = []
    cEnv = cEnt = cDnr = cOpp = cSucc = 0
    for d in docs:
        env = d.get("c_enviados") or 0
        ent = d.get("c_entregados") or 0
        dnr = d.get("c_dnr") or 0
        opp = d.get("c_pod_opp") or 0
        suc = d.get("c_pod_succ") or 0
        cEnv += env; cEnt += ent; cDnr += dnr; cOpp += opp; cSucc += suc
        dias.append({
            "fecha": d["date"],
            "dia": {"dcr": d.get("dcr"), "dnr_dpmo": d.get("dnr_dpmo"), "pod": d.get("pod"),
                    "entregados": ent or None},
            "acum": {
                "dcr": round(cEnt / cEnv * 100, 2) if cEnv else None,
                "dnr_dpmo": round(cDnr / cEnt * 1e6) if cEnt else None,
                "pod": round(cSucc / cOpp * 100, 2) if cOpp else None,
                "entregados": cEnt or None}})
    return {"center": center, "desde": sun, "hasta": sat, "dias": dias,
            "n_dias": len(dias),
            "acumulado": (dias[-1]["acum"] if dias else None)}


@api_router.get("/scorecard/sources")
async def scorecard_sources(center: str, week: Optional[str] = None, _=Depends(require_admin)):
    """Lista TODO lo cargado de la semana (para verlo y poder borrarlo):
    días de ratios, reportes diarios, scorecard oficial y valores metidos a mano."""
    if not week:
        last = (datetime.now(timezone.utc) - timedelta(days=2)).strftime("%Y-%m-%d")
        week, _s = _sun_sat_week(last)
    sun, sat = _sun_sat_week(week)
    wnum = _sun_to_week_num(sun)
    items = []
    async for d in db.daily_ratios.find({"center": center, "date": {"$gte": sun, "$lte": sat}}, {"_id": 0}):
        items.append({"kind": "ratios", "fecha": d["date"], "ref": d["date"],
                      "label": "Descripción general · " + d["date"],
                      "detalle": "envíos/entregas del día", "subido": d.get("uploaded_at")})
    async for d in db.daily_dsp.find({"center": center, "date": {"$gte": sun, "$lte": sat}}, {"_id": 0}):
        items.append({"kind": "daily", "fecha": d["date"], "ref": d["date"],
                      "label": "Reporte diario · " + d["date"],
                      "detalle": str(len(d.get("drivers") or [])) + " conductores", "subido": d.get("uploaded_at")})
    wk = await db.scorecard_weekly.find_one({"center": center, "week": wnum}, {"_id": 0})
    if wk:
        vv = wk.get("values") or {}
        items.append({"kind": "resumen", "fecha": None, "ref": str(wnum),
                      "label": "Resumen de entregas · W" + str(wnum),
                      "detalle": "DCR " + str(vv.get("dcr", "?")) + "% · POD " + str(vv.get("pod", "?")) + "%",
                      "subido": wk.get("uploaded_at")})
    off = await db.scorecard_official.find_one({"center": center, "week": wnum}, {"_id": 0})
    if off:
        items.append({"kind": "official", "fecha": None, "ref": str(wnum),
                      "label": "Scorecard oficial · W" + str(wnum),
                      "detalle": (off.get("overall_tier") or "") + " " + str(off.get("overall_score") or ""),
                      "subido": off.get("uploaded_at")})
    live = await db.scorecard_live.find_one({"center": center, "week": sun}, {"_id": 0})
    manual = [{"kind": "manual", "fecha": None, "ref": k,
               "label": "Valor a mano · " + next((m["label"] for m in _SC_METRICS if m["key"] == k), k),
               "detalle": str(v), "subido": (live or {}).get("updated_at")}
              for k, v in ((live or {}).get("values") or {}).items() if v is not None]
    items = sorted(items, key=lambda x: (x["fecha"] or "", x["kind"]))
    return {"center": center, "week": sun, "desde": sun, "hasta": sat,
            "items": items + manual, "total": len(items) + len(manual)}


@api_router.delete("/scorecard/source")
async def scorecard_delete_source(center: str, kind: str, ref: str,
                                  week: Optional[str] = None, _=Depends(require_admin)):
    """Borra un archivo/dato cargado. kind: ratios|daily|official|manual.
    ref = fecha (ratios/daily), nº de semana (official) o clave de métrica (manual)."""
    if kind == "ratios":
        r = await db.daily_ratios.delete_one({"center": center, "date": ref})
    elif kind == "daily":
        r = await db.daily_dsp.delete_one({"center": center, "date": ref})
    elif kind == "official":
        try:
            wn = int(ref)
        except Exception:
            raise HTTPException(status_code=400, detail="ref debe ser nº de semana")
        await db.scorecard_obs.delete_many({"center": center, "week": wn})
        r = await db.scorecard_official.delete_one({"center": center, "week": wn})
    elif kind == "resumen":
        try:
            wn = int(ref)
        except Exception:
            raise HTTPException(status_code=400, detail="ref debe ser nº de semana")
        r = await db.scorecard_weekly.delete_one({"center": center, "week": wn})
    elif kind == "manual":
        if week:
            wsun, _ = _sun_sat_week(week)
        else:
            last = (datetime.now(timezone.utc) - timedelta(days=2)).strftime("%Y-%m-%d")
            wsun, _ = _sun_sat_week(last)
        r = await db.scorecard_live.update_one(
            {"center": center, "week": wsun}, {"$unset": {f"values.{ref}": ""}})
    else:
        raise HTTPException(status_code=400, detail=f"kind inválido: {kind}")
    borrados = getattr(r, "deleted_count", None)
    if borrados is None:
        borrados = getattr(r, "modified_count", 0)
    return {"ok": True, "borrados": borrados}


async def _store_resumen_semanal(sem, center):
    cen = (center or "OGA5").upper()
    guardadas = []
    for wn, vals in sem.items():
        await db.scorecard_weekly.update_one(
            {"center": cen, "week": int(wn)},
            {"$set": {"center": cen, "week": int(wn), "values": vals,
                      "uploaded_at": datetime.now(timezone.utc).isoformat()}}, upsert=True)
        guardadas.append(int(wn))
    guardadas.sort()
    return {"tipo": "resumen", "ok": True, "center": cen,
            "mensaje": f"Resumen de entregas: semanas {', '.join('W'+str(w) for w in guardadas)} (DCR/DNR/POD)",
            "semanas": guardadas}


async def _store_ratios(ratios, center):
    cen = (center or "OGA5").upper()
    n = 0
    for date, vals in ratios["days"].items():
        if not vals:
            continue
        await db.daily_ratios.update_one(
            {"center": cen, "date": date},
            {"$set": {"center": cen, "date": date, **vals,
                      "uploaded_at": datetime.now(timezone.utc).isoformat()}}, upsert=True)
        n += 1
    return {"tipo": "ratios", "ok": True, "center": cen,
            "mensaje": f"Ratios de {n} día(s) cargados", "dias": n}


async def _ratios_week_values(center, sun, sat):
    """% de calidad de la semana a partir de los ratios diarios (conteos reales)."""
    docs = await db.daily_ratios.find(
        {"center": center, "date": {"$gte": sun, "$lte": sat}}, {"_id": 0}).to_list(10)
    if not docs:
        return {}, []

    def s(k):
        return sum(d.get(k, 0) or 0 for d in docs)
    enviados, entregados = s("c_enviados"), s("c_entregados")
    pod_opp, pod_succ, dnr = s("c_pod_opp"), s("c_pod_succ"), s("c_dnr")
    vals = {}
    if enviados and entregados:
        vals["dcr"] = round(entregados / enviados * 100, 2)
        vals["dnr_dpmo"] = round(dnr / entregados * 1e6)
    if pod_opp:
        vals["pod"] = round(pod_succ / pod_opp * 100, 2)
    for k in ("dcr", "pod", "fdds", "rts_pct"):
        if k not in vals:
            xs = [d[k] for d in docs if d.get(k) is not None]
            if xs:
                vals[k] = round(sum(xs) / len(xs), 2)
    return vals, sorted([d["date"] for d in docs])


@api_router.get("/scorecard/ratios-raw")
async def ratios_raw(center: str, desde: str, hasta: str, _=Depends(require_admin)):
    docs = await db.daily_ratios.find(
        {"center": center, "date": {"$gte": desde, "$lte": hasta}}, {"_id": 0}).sort("date", 1).to_list(40)
    return {"count": len(docs), "docs": docs}


@api_router.get("/scorecard/predict")
async def scorecard_predict(center: str, week: Optional[str] = None, _=Depends(require_admin)):
    """Predicción completa de la semana: usa TODOS los datos disponibles
    (oficial PDF > manual > resumen semanal > ratios diarios) para las 16 métricas.
    Calcula el score global exacto y el gap al siguiente tier."""
    if not week:
        last = (datetime.now(timezone.utc) - timedelta(days=2)).strftime("%Y-%m-%d")
        week, _s = _sun_sat_week(last)
    sun, sat = _sun_sat_week(week)
    wnum = _sun_to_week_num(sun)
    thr = await _sc_thresholds(center)
    weights = await _sc_weights(center)

    # Todas las fuentes disponibles (mismo orden de prioridad que /scorecard/full)
    ratio_vals, dias = await _ratios_week_values(center, sun, sat)
    live_doc = await db.scorecard_live.find_one({"center": center, "week": sun}, {"_id": 0})
    live_vals = (live_doc or {}).get("values", {})
    wk_doc = await db.scorecard_weekly.find_one({"center": center, "week": wnum}, {"_id": 0})
    week_vals = (wk_doc or {}).get("values", {})
    off = await db.scorecard_official.find_one({"center": center, "week": wnum}, {"_id": 0})
    off_metrics = {mm.get("key"): mm for mm in (off.get("metrics") if off else [])}

    # Baseline estimado de la última scorecard conocida (Safety/Capacity)
    base = await db.scorecard_official.find_one(
        {"center": center, "week": {"$lt": wnum}}, {"_id": 0}, sort=[("week", -1)])
    base_metrics = {mm.get("key"): mm for mm in (base.get("metrics") if base else [])}

    metrics_out = []
    fuentes_usadas = set()
    for m in _SC_METRICS:
        if off_metrics.get(m["key"]) and off_metrics[m["key"]].get("value") is not None:
            v = off_metrics[m["key"]].get("value")
            src = "oficial"
        elif live_vals.get(m["key"]) is not None:
            v = live_vals[m["key"]]
            src = "manual"
        elif week_vals.get(m["key"]) is not None:
            v = week_vals[m["key"]]
            src = "resumen"
        elif ratio_vals.get(m["key"]) is not None:
            v = ratio_vals[m["key"]]
            src = "ratios"
        elif base_metrics.get(m["key"]) and base_metrics[m["key"]].get("value") is not None:
            # Para Safety/Capacity: arrastra última scorecard conocida
            v = base_metrics[m["key"]]["value"]
            src = "estimado"
        else:
            v = None
            src = None
        if src:
            fuentes_usadas.add(src)
        t = _sc_tier(v, thr.get(m["key"]), m["dir"]) if v is not None else None
        metrics_out.append({"key": m["key"], "label": m["label"], "group": m["group"],
                            "unit": m.get("unit"), "dir": m["dir"],
                            "value": v, "tier": t, "source": src,
                            "thr": thr.get(m["key"]),
                            "next": _sc_next_target(v, t, thr.get(m["key"]), m["dir"]) if t else None})

    # Score global ponderado (igual que /scorecard/full)
    score_calc, wsum, nused = _overall_score(metrics_out, weights)
    predicted_tier = _score_to_tier(score_calc) if score_calc is not None else None

    # Gap al siguiente tier (cuánto score falta para subir)
    gap_to_next = None
    next_tier_name = None
    if score_calc is not None:
        for threshold, name in _OVERALL_TIER_BANDS:
            if score_calc < threshold:
                gap_to_next = round(threshold - score_calc, 2)
                next_tier_name = name
                break  # primer tier por encima

    # Cobertura del peso total
    peso_total = sum(v for v in weights.values() if isinstance(v, (int, float)))
    cobertura = round(wsum / peso_total * 100) if peso_total else 0

    # Confianza: métricas con dato real (no estimado) / total métricas con peso
    n_real = sum(1 for m in metrics_out if m["source"] in ("oficial", "manual", "resumen", "ratios") and m["value"] is not None)
    n_total = sum(1 for m in metrics_out if (weights.get(m["key"]) or 0) > 0)
    confidence = round(n_real / max(1, n_total) * 100)

    helps = [m["label"] for m in metrics_out if m["tier"] in ("Fantastic", "Fantastic Plus")]
    hurts = [{"label": m["label"], "tier": m["tier"], "value": m["value"], "unit": m.get("unit")}
             for m in metrics_out if m["tier"] in ("Fair", "Poor")]
    faltan = [m["label"] for m in metrics_out if m["value"] is None]

    # Delta vs scorecard oficial anterior
    prev = await db.scorecard_official.find_one(
        {"center": center, "week": {"$lt": wnum}}, {"_id": 0}, sort=[("week", -1)])
    delta = None
    if prev and prev.get("overall_tier"):
        delta = {"week": prev.get("week"), "tier": prev.get("overall_tier"),
                 "score": prev.get("overall_score")}

    return {
        "center": center, "desde": sun, "hasta": sat, "week_num": wnum,
        "dias_con_datos": dias,
        "predicted_tier": predicted_tier,
        "predicted_score": score_calc,
        "gap_to_next": gap_to_next,
        "next_tier": next_tier_name,
        "confidence": confidence,
        "cobertura_peso": cobertura,
        "metrics": metrics_out,
        "ayudan": helps, "empeoran": hurts, "faltan_datos": faltan,
        "delta_anterior": delta,
        "fuentes": sorted(fuentes_usadas),
        "estimado_desde": (base.get("week") if base and "estimado" in fuentes_usadas else None),
    }


# Columna del Excel de umbrales → (clave métrica, dirección +1/-1)
_XLSX_THR_MAP = {
    "dcr": ("dcr", 1), "dsc_dpmo": ("dsc_dpmo", -1), "lor_dpmo": ("lor_dpmo", -1),
    "pod": ("pod", 1), "cc": ("cc", 1), "capacity_reliability": ("ndcr", 1),
    "ce_dpmo": ("cec_dpmo", -1), "cdf_dpmo": ("cdf", -1), "speeding_event": ("speeding", -1),
    "fico": ("fico", 1), "ementor_adoption": ("mentor", 1), "vsa": ("vsa", 1),
    "dvic": ("dvic", 1), "dex": ("dex", 1), "uwh": ("whc", 1),
}


@api_router.post("/scorecard/calibrate-thresholds")
async def calibrate_thresholds(data: dict = Body(...), _=Depends(require_admin)):
    """Calibra umbrales del centro con 2 fuentes en orden de prioridad:
    1. Umbrales EXPLÍCITOS del texto del PDF ("Fantastic >= 99% in DCR") — más fiables.
    2. Inferencia desde pares valor+tier observados en múltiples scorecards — solo cuando
       hay ≥2 tiers distintos observados para la misma métrica (evita confundir el valor
       observado con el umbral cuando solo hay 1 punto de datos)."""
    center = data.get("center")
    if not center:
        raise HTTPException(status_code=400, detail="center requerido")
    docs = await db.scorecard_official.find({"center": center}, {"_id": 0}).to_list(200)
    if not docs:
        raise HTTPException(status_code=404, detail=f"No hay scorecards oficiales importadas para {center}. Sube primero el PDF.")

    # 1. Recopilar umbrales explícitos de los PDFs (campo explicit_thresholds)
    explicit = {}  # key -> {"fantastic": valor}
    for doc in docs:
        for key, info in (doc.get("explicit_thresholds") or {}).items():
            if isinstance(info, dict) and info.get("fantastic") is not None:
                explicit.setdefault(key, []).append(float(info["fantastic"]))

    # 2. Acumular pares (valor, tier) para inferencia
    pairs = {}
    for doc in docs:
        for m in (doc.get("metrics") or []):
            key = m.get("key")
            val = m.get("value")
            tier = m.get("tier")
            if key and val is not None and tier and tier in ("Fantastic Plus", "Fantastic", "Great", "Fair", "Poor"):
                pairs.setdefault(key, []).append((float(val), tier))

    thr_update = {}
    explicitos_usados = []
    inferidos_usados = []

    for m_cfg in _SC_METRICS:
        key = m_cfg["key"]
        direction = m_cfg["dir"]
        thr_doc = {}

        # Fuente 1: umbral explícito del texto del PDF
        if key in explicit:
            # Media de los umbrales explícitos de todas las scorecards (suelen ser iguales)
            thr_doc["fantastic"] = round(sum(explicit[key]) / len(explicit[key]), 4)
            explicitos_usados.append(key)

        # Fuente 2: inferencia desde pares valor+tier
        # Solo inferimos si hay ≥2 tiers distintos observados (para poder acotar el umbral)
        obs = pairs.get(key, [])
        if obs:
            by_tier = {}
            for val, tier in obs:
                by_tier.setdefault(tier, []).append(val)
            tiers_vistos = set(by_tier.keys())

            # Umbral Great: solo inferimos si vimos Great Y (Fantastic o Fair)
            # Esto evita confundir "98.83% Great" con el umbral Great cuando no vemos Fantastic
            if "Great" in tiers_vistos and ("Fantastic" in tiers_vistos or "Fantastic Plus" in tiers_vistos or "Fair" in tiers_vistos):
                thr_doc["great"] = min(by_tier["Great"]) if direction == 1 else max(by_tier["Great"])

            # Umbral Fair: solo si vimos Fair Y (Great o Poor)
            if "Fair" in tiers_vistos and ("Great" in tiers_vistos or "Poor" in tiers_vistos):
                thr_doc["fair"] = min(by_tier["Fair"]) if direction == 1 else max(by_tier["Fair"])

            # Fantastic solo desde inferencia si no hay explícito Y hay ≥2 tiers con Fantastic
            if "fantastic" not in thr_doc and ("Fantastic" in tiers_vistos or "Fantastic Plus" in tiers_vistos):
                if "Great" in tiers_vistos:  # Acotamos: vimos ambos Fantastic y Great
                    f_vals = by_tier.get("Fantastic", []) + by_tier.get("Fantastic Plus", [])
                    thr_doc["fantastic"] = min(f_vals) if direction == 1 else max(f_vals)
                    inferidos_usados.append(key)

        if thr_doc:
            thr_update[key] = thr_doc

    if not thr_update:
        raise HTTPException(status_code=422, detail="No se encontraron métricas con umbrales. Asegúrate de haber subido el PDF de la scorecard oficial.")

    update_doc = {"center": center}
    update_doc.update(thr_update)
    await db.scorecard_thresholds.update_one({"center": center}, {"$set": update_doc}, upsert=True)
    return {"success": True, "center": center,
            "calibradas": list(thr_update.keys()),
            "desde_texto_pdf": explicitos_usados,
            "inferidas": inferidos_usados,
            "desde_scorecards": len(docs)}


@api_router.post("/scorecard/import-thresholds")
async def import_thresholds(file: UploadFile = File(...), _=Depends(require_admin)):
    """Sube el Excel de umbrales de Amazon (t0/t1/t2/t3 por métrica y semana).
    Guarda, por estación, los umbrales de la ÚLTIMA semana = vigentes."""
    content = await file.read()
    rows = _read_table_any(content, file.filename or "umbrales.xlsx")
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel vacío o sin filas")
    hdr = [str(c).strip().lower() for c in rows[0]]
    col = {name: i for i, name in enumerate(hdr)}

    def cell(r, name):
        i = col.get(name)
        return r[i] if (i is not None and i < len(r)) else None

    best = {}  # station -> (week, row)
    for r in rows[1:]:
        st = str(cell(r, "station") or "").upper().strip()
        wk = cell(r, "week")
        if not st or wk in (None, ""):
            continue
        try:
            wk = int(float(wk))
        except Exception:
            continue
        if st not in best or wk > best[st][0]:
            best[st] = (wk, r)

    saved = []
    for st, (wk, r) in best.items():
        thr_doc = {"center": st, "week": wk}
        n = 0
        for xcol, (key, _dir) in _XLSX_THR_MAP.items():
            band = {}
            for lvl, name in (("t0", "fantastic_plus"), ("t1", "fantastic"),
                              ("t2", "great"), ("t3", "fair")):
                v = cell(r, xcol + "_" + lvl)
                if v not in (None, ""):
                    try:
                        band[name] = float(v)
                    except Exception:
                        pass
            if band:
                thr_doc[key] = band
                n += 1
        await db.scorecard_thresholds.update_one({"center": st}, {"$set": thr_doc}, upsert=True)
        saved.append({"center": st, "week": wk, "metricas": n})
    return {"success": True, "guardadas": saved}


# Columna de pesos (CSV ..._wt_final) → clave métrica
_XLSX_WT_MAP = {
    "dcr": "dcr", "dnr": "dnr_dpmo", "dsc_dpmo": "dsc_dpmo", "lor_dpmo": "lor_dpmo",
    "pod": "pod", "cc": "cc", "capacity_reliability": "ndcr", "whc": "whc", "cas": "cas",
    "ce_dpmo": "cec_dpmo", "positive_dex": "dex", "cdf_dpmo": "cdf",
    "speeding_event_rate": "speeding", "fico": "fico", "ementor_adoption_rate": "mentor",
    "vsa": "vsa", "dvic": "dvic",
}


@api_router.post("/scorecard/import-weights")
async def import_weights(file: UploadFile = File(...), _=Depends(require_admin)):
    """Sube el CSV/Excel de pesos de Amazon (..._wt_final por métrica y semana).
    Guarda, por estación, los pesos de la ÚLTIMA semana = vigentes."""
    content = await file.read()
    rows = _read_table_any(content, file.filename or "pesos.csv")
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Archivo vacío")
    hdr = [str(c).strip().lower() for c in rows[0]]
    col = {name: i for i, name in enumerate(hdr)}

    def cell(r, name):
        i = col.get(name)
        return r[i] if (i is not None and i < len(r)) else None

    best = {}
    for r in rows[1:]:
        st = str(cell(r, "station") or "").upper().strip()
        wk = cell(r, "week")
        if not st or wk in (None, ""):
            continue
        try:
            wk = int(float(wk))
        except Exception:
            continue
        if st not in best or wk > best[st][0]:
            best[st] = (wk, r)

    saved = []
    for st, (wk, r) in best.items():
        doc = {"center": st, "week": wk}
        total = 0.0
        for xcol, key in _XLSX_WT_MAP.items():
            v = cell(r, xcol + "_wt_final")
            if v not in (None, ""):
                try:
                    fv = float(v)
                    doc[key] = fv
                    total += fv
                except Exception:
                    pass
        await db.scorecard_weights.update_one({"center": st}, {"$set": doc}, upsert=True)
        saved.append({"center": st, "week": wk, "suma_pesos": round(total, 2)})
    return {"success": True, "guardados": saved}


async def _sc_weights(center):
    """Pesos por métrica. Prioridad: centro > GLOBAL > semilla uniforme por pilar."""
    weights = dict(_SC_SEED_WEIGHTS)
    for c in ("GLOBAL", center):
        if not c:
            continue
        doc = await db.scorecard_weights.find_one({"center": c}, {"_id": 0})
        if doc:
            weights.update({k: v for k, v in doc.items() if k not in ("center", "week") and isinstance(v, (int, float))})
    return weights


# ──────────────────────────────────────────────
# BLOQUE 4 – Generador de plantilla de turno
# ──────────────────────────────────────────────

_PROMPT_CORTEX = """
Eres un extractor de datos de una captura de pantalla de Cortex (plataforma Amazon DSP).
La imagen muestra una lista de rutas con sus conductores (DAs) y horas de salida.

Devuelve ÚNICAMENTE un JSON válido, sin texto adicional, sin markdown, sin bloques de código.

Formato exacto:
{
  "week": <número de semana ISO como entero, o null si no aparece>,
  "date": "<DD/MM/YYYY visible en la imagen, o null si no aparece>",
  "rutas": [
    {
      "ruta": "<código exacto de ruta tal como aparece, p.ej. CA_A44>",
      "conductor": "<nombre del PRIMER conductor (DA principal) en MAYÚSCULAS>",
      "h_salida": "<hora de salida HH:MM exacta — null si no aparece>"
    }
  ]
}

REGLAS:
1. Extrae ÚNICAMENTE lo que puedas leer con total claridad. Null si no está claro.
2. Si una ruta tiene VARIOS conductores listados (conductor principal + ayuda o compartido),
   coge SOLO EL PRIMERO que aparece. Ignora completamente el segundo y siguientes.
3. No dupliques rutas.
4. Ordena por código de ruta.
5. La hora de salida suele aparecer como HH:MM cerca del nombre de la ruta.
"""

_PROMPT_PLATAFORMA = """
Eres un extractor de datos de una captura de pantalla de una plataforma de gestión de furgonetas DSP.
La imagen muestra filas con conductores (DAs) y su furgoneta asignada.

Los nombres en la imagen aparecen en formato "Apellido(s), Nombre" (apellidos primero, luego coma, luego nombre).
Debes invertirlos y devolverlos en formato "NOMBRE APELLIDO(S)" en MAYÚSCULAS.
Ejemplo: "Rodriguez Arias, Andrea" → conductor: "ANDREA RODRIGUEZ ARIAS"

La tabla tiene columnas. La ÚLTIMA columna de cada fila contiene el código de matrícula
de la furgoneta (formato: 4 dígitos + 3 letras, p.ej. 7906NFX o 7906 NFX o 0116 MBP).
Puede haber también una columna con tipo de contrato (ETT11, ETT20, ETT28, ETT30) — eso NO es la matrícula.

Devuelve ÚNICAMENTE un JSON válido, sin texto adicional, sin markdown, sin bloques de código.

Formato exacto:
{
  "asignaciones": [
    {
      "conductor": "<NOMBRE APELLIDO(S) en MAYÚSCULAS, con el nombre primero>",
      "furgo": "<código de la ÚLTIMA columna: 4 dígitos + 3 letras, p.ej. 7906NFX>"
    }
  ]
}

REGLAS:
1. Extrae ÚNICAMENTE lo que puedas leer con claridad.
2. Invierte siempre el orden: la imagen dice "Apellido, Nombre" → tú devuelves "NOMBRE APELLIDO".
3. La matrícula siempre tiene formato 4 dígitos + 3 letras. No confundas con ETT11/ETT20/ETT28/ETT30.
4. No dupliques conductores. Extrae TODOS los visibles.
"""


def _gemini_client_plantilla():
    """Construye cliente Gemini reutilizando el mismo patrón que el resto del backend."""
    from google import genai as genai_sdk
    gemini_key = os.environ.get("GEMINI_API_KEY", "")
    use_vertex = os.environ.get("USE_VERTEX_AI", "").lower() in ("1", "true", "yes")
    if use_vertex:
        vertex_project  = os.environ.get("GCP_PROJECT", "")
        vertex_location = os.environ.get("GCP_LOCATION", "us-central1")
        sa_json = os.environ.get("GCP_SERVICE_ACCOUNT_JSON", "")
        if sa_json:
            from google.oauth2 import service_account
            import base64 as _b64
            sa_clean = sa_json.strip()
            if not sa_clean.startswith("{"):
                try:
                    sa_clean = _b64.b64decode(sa_clean).decode("utf-8")
                except Exception:
                    pass
            creds_info = json.loads(sa_clean)
            credentials = service_account.Credentials.from_service_account_info(
                creds_info, scopes=["https://www.googleapis.com/auth/cloud-platform"]
            )
            return genai_sdk.Client(
                vertexai=True, project=vertex_project,
                location=vertex_location, credentials=credentials
            )
        return genai_sdk.Client(vertexai=True, project=vertex_project, location=vertex_location)
    if not gemini_key:
        raise HTTPException(503, "GEMINI_API_KEY no configurada")
    return genai_sdk.Client(api_key=gemini_key)


def _mime_type(data: bytes) -> str:
    if data[:4] == b"\x89PNG":
        return "image/png"
    return "image/jpeg"


def _calc_horas(h_salida: str | None) -> tuple[str, str]:
    """Dado h_salida (HH:MM de Cortex) calcula bajada y llegada. Devuelve ("", "") si no hay hora."""
    if not h_salida:
        return "", ""
    from datetime import datetime as _dt, timedelta as _td
    for fmt in ("%H:%M", "%H.%M", "%I:%M %p", "%I:%M%p"):
        try:
            t = _dt.strptime(h_salida.strip(), fmt)
            bajada  = (_dt.combine(_dt.today(), t.time()) - _td(minutes=10)).strftime("%H:%M")
            llegada = (_dt.combine(_dt.today(), t.time()) - _td(minutes=30)).strftime("%H:%M")
            return llegada, bajada
        except ValueError:
            continue
    return "", ""


def _build_plantilla_excel(rows: list, red_routes: set, pink_furgos: set, week_num, fecha_str: str, yellow_routes: set = None, marked_conductors: set = None) -> bytes:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime as _dt
    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Turno"

    # Colores exactos de la plantilla de referencia
    C_TITLE_BG   = "F2F2F2"   # fila título: gris muy claro
    C_TITLE_FG   = "000000"
    C_HEADER_BG  = "FFD966"   # cabeceras columnas: amarillo
    C_HEADER_FG  = "000000"   # texto negro
    C_WHITE      = "FFFFFF"   # fila normal
    C_BLUE       = "BDD7EE"   # fila ola tardía (H.WAVE >= 12:20)
    C_RED_BG     = "FF0000"   # fila "no vino"
    C_RED_FG     = "FFFFFF"
    C_PINK       = "F4CCCC"   # celda FURGO especial
    C_BORDER     = "BFBFBF"

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def brd():
        s = Side(style="thin", color=C_BORDER)
        return Border(left=s, right=s, top=s, bottom=s)

    ca = Alignment(horizontal="center", vertical="center")
    la = Alignment(horizontal="left",   vertical="center")

    # ── Fila 1: WEEK xx (izq) · fecha (der) ──
    ws.merge_cells("A1:D1")
    ws.merge_cells("E1:H1")
    for cell, val, align in [
        (ws["A1"], f"WEEK {week_num}", la),
        (ws["E1"], fecha_str,          ca),
    ]:
        cell.value = val
        cell.fill  = fill(C_TITLE_BG)
        cell.font  = Font(bold=True, color=C_TITLE_FG, size=11)
        cell.alignment = align
        cell.border = brd()
    ws.row_dimensions[1].height = 22

    # ── Fila 2: cabeceras ──
    col_defs = [
        ("RUTA",               13),
        ("CONDUCTOR",          34),
        ("MOVIL",               9),
        ("FURGO",              12),
        ("H. LLEGADA A NAVE",  18),
        ("H. BAJADA AL YARD",  18),
        ("H. WAVE",            11),
        ("OBSERVACIONES",      32),
    ]
    for i, (h, w) in enumerate(col_defs, start=1):
        cell = ws.cell(row=2, column=i, value=h)
        cell.fill  = fill(C_HEADER_BG)
        cell.font  = Font(bold=True, color=C_HEADER_FG, size=9)
        cell.alignment = ca
        cell.border = brd()
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 21

    C_YELLOW      = "FFF2CC"   # amarillo clarito — marca manual
    C_MARK        = "FCE4D6"   # naranja clarito — conductor marcado (sin batch, etc.)
    C_MARK_FG     = "7B3F00"
    # Paleta de pasteles para distinguir olas — máx 6 waves distintas
    WAVE_PALETTE  = ["BDD7EE", "C6EFCE", "FFEB9C", "FCE4D6", "E2EFDA", "EAD1DC"]
    _yellow_routes     = {r.upper() for r in (yellow_routes or [])}
    _marked_conductors = {c.upper() for c in (marked_conductors or [])}

    # Mapear cada hora de wave a un color (ordenadas cronológicamente)
    _wave_times = sorted(set(
        (row.get("h_salida") or "").strip()
        for row in rows if (row.get("h_salida") or "").strip()
    ))
    _wave_color = {wt: WAVE_PALETTE[i % len(WAVE_PALETTE)] for i, wt in enumerate(_wave_times)}

    # ── Filas de datos ──
    for idx, row in enumerate(rows):
        r         = idx + 3
        ruta      = row.get("ruta")          or ""
        conductor = row.get("conductor")     or ""
        movil     = row.get("movil")         or ""
        furgo     = row.get("furgo")         or ""
        h_llegada = row.get("h_llegada")     or ""
        h_bajada  = row.get("h_bajada")      or ""
        h_wave    = row.get("h_salida") or row.get("h_wave") or ""
        obs       = row.get("observaciones") or ""

        row_key   = ruta.upper() if ruta.strip() else conductor.upper()
        is_red    = row_key in red_routes
        is_yellow = row_key in _yellow_routes

        # Ola tardía: h_llegada >= 11:50 (solo si no tiene otra marca)
        is_blue = False
        if not is_red and not is_yellow and h_llegada:
            try:
                t = _dt.strptime(h_llegada.strip(), "%H:%M").time()
                is_blue = t >= _dt.strptime("11:50", "%H:%M").time()
            except ValueError:
                pass

        if is_red:
            row_bg, row_fg = C_RED_BG, C_RED_FG
        elif is_yellow:
            row_bg, row_fg = C_YELLOW, "000000"
        elif is_blue:
            row_bg, row_fg = C_BLUE, "000000"
        else:
            row_bg, row_fg = C_WHITE, "000000"

        is_pink_furgo   = furgo.upper() in {f.upper() for f in pink_furgos}
        is_marked_cond  = conductor.upper() in _marked_conductors
        wave_time_color = _wave_color.get(h_wave.strip(), None) if h_wave else None

        cells_vals = [ruta, conductor, movil, furgo, h_llegada, h_bajada, h_wave, obs]
        for ci, val in enumerate(cells_vals, start=1):
            cell = ws.cell(row=r, column=ci, value=val)
            if ci == 4 and is_pink_furgo and not is_red:
                # FURGO especial — rosa
                cell.fill = fill(C_PINK)
                cell.font = Font(color="000000", size=9)
            elif ci == 2 and is_marked_cond and not is_red:
                # Conductor marcado — naranja clarito
                cell.fill = fill(C_MARK)
                cell.font = Font(color=C_MARK_FG, size=9, bold=True)
            elif ci in (5, 6, 7) and wave_time_color and not is_red and not is_yellow:
                # Hora de ola — color de la wave correspondiente
                cell.fill = fill(wave_time_color)
                cell.font = Font(color="000000", size=9)
            else:
                cell.fill = fill(row_bg)
                cell.font = Font(color=row_fg, size=9, bold=(ci == 1))
            cell.alignment = ca if ci not in (2, 8) else la
            cell.border = brd()
        ws.row_dimensions[r].height = 16

    ws.freeze_panes = "A3"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _normalize_name(name: str) -> str:
    """Mayúsculas, sin tildes, sin puntuación ni puntos suspensivos."""
    import unicodedata, re
    s = unicodedata.normalize("NFD", name.upper())
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^A-Z0-9 ]", " ", s)  # quita puntos, comas, etc.
    return s


def _name_tokens(name: str) -> list:
    """Tokens de palabras >= 3 letras, ordenados por longitud desc."""
    return sorted([w for w in _normalize_name(name).split() if len(w) >= 3], key=len, reverse=True)


def _match_score(name_a: str, name_b: str) -> float:
    """
    Score entre dos nombres. Coincidencia exacta = 1.0, prefijo (>=4 chars) = 0.8.
    Maneja nombres truncados con '...' porque _normalize_name ya elimina los puntos.
    """
    tokens_a = _name_tokens(name_a)
    tokens_b = _name_tokens(name_b)
    score = 0.0
    used_b = set()
    for ta in tokens_a:
        for i, tb in enumerate(tokens_b):
            if i in used_b:
                continue
            if ta == tb:
                score += 1.0
                used_b.add(i)
                break
            short, long_ = (ta, tb) if len(ta) <= len(tb) else (tb, ta)
            if len(short) >= 4 and long_.startswith(short):
                score += 0.8
                used_b.add(i)
                break
    return score


async def _gemini_extract(client, model_name: str, prompt: str, img_bytes: bytes) -> dict:
    from google.genai import types as genai_types
    resp = await asyncio.wait_for(
        asyncio.to_thread(
            lambda: client.models.generate_content(
                model=model_name,
                contents=[prompt, genai_types.Part.from_bytes(data=img_bytes, mime_type=_mime_type(img_bytes))],
                config=genai_types.GenerateContentConfig(temperature=0.0),
            )
        ),
        timeout=120,
    )
    raw = _strip_markdown_json((resp.text or "").strip())
    return json.loads(raw)


# ── Paso 1: extraer datos con Gemini, devuelve JSON para preview ──
@app.post("/api/tools/plantilla-extraer", dependencies=[Depends(require_admin)])
async def plantilla_extraer(
    plataforma: List[UploadFile] = File(...),
    cortex:     Optional[List[UploadFile]] = File(default=None),
):
    if not plataforma:
        raise HTTPException(400, "Sube al menos las capturas de plataforma.")
    cortex_imgs = [await f.read() for f in (cortex or [])]
    plat_imgs   = [await f.read() for f in plataforma]

    try:
        client     = _gemini_client_plantilla()
        model_name = os.environ.get("GEMINI_MODEL", "gemini-2.5-flash")

        tasks = (
            [_gemini_extract(client, model_name, _PROMPT_CORTEX,     b) for b in cortex_imgs] +
            [_gemini_extract(client, model_name, _PROMPT_PLATAFORMA, b) for b in plat_imgs]
        )
        results = await asyncio.gather(*tasks)

    except asyncio.TimeoutError:
        raise HTTPException(504, "Gemini tardó demasiado. Intenta de nuevo.")
    except json.JSONDecodeError as e:
        raise HTTPException(422, f"Gemini no devolvió JSON válido: {e}")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"plantilla_extraer error: {e}", exc_info=True)
        raise HTTPException(500, str(e))

    nc = len(cortex_imgs)
    cortex_results = results[:nc]
    plat_results   = results[nc:]

    from datetime import date as _date

    week_num  = _date.today().isocalendar()[1]
    fecha_str = _date.today().strftime("%d/%m/%Y")

    # Combinar datos Cortex (dedup por ruta)
    rutas_map: dict = {}
    for cd in cortex_results:
        if cd.get("week"):
            week_num = cd["week"]
        if cd.get("date"):
            fecha_str = cd["date"]
        for r in cd.get("rutas", []):
            key = (r.get("ruta") or "").strip()
            if key and key not in rutas_map:
                rutas_map[key] = r

    # Combinar datos plataforma (dedup por conductor)
    asignaciones: list = []
    seen_conductores: set = set()
    for pd in plat_results:
        for a in pd.get("asignaciones", []):
            nombre = _normalize_name(a.get("conductor") or "")
            if nombre and nombre not in seen_conductores:
                seen_conductores.add(nombre)
                asignaciones.append(a)

    rows_out = []

    if rutas_map:
        # MODO COMPLETO: Cortex + Plataforma → cruce por nombre
        for ruta_key in sorted(rutas_map):
            ruta_row  = rutas_map[ruta_key]
            conductor = ruta_row.get("conductor") or ""
            h_salida  = ruta_row.get("h_salida") or ""

            furgo = ""
            movil = ""
            best_score = 0.0
            best_asig  = None
            for asig in asignaciones:
                sc = _match_score(conductor, asig.get("conductor") or "")
                if sc > best_score:
                    best_score = sc
                    best_asig  = asig

            if best_asig and best_score >= 0.8:
                furgo = (best_asig.get("furgo") or "").replace(" ", "")
                movil = best_asig.get("movil") or ""

            h_llegada, h_bajada = _calc_horas(h_salida)
            rows_out.append({
                "ruta":          ruta_key,
                "conductor":     conductor,
                "movil":         movil,
                "furgo":         furgo,
                "h_salida":      h_salida,
                "h_bajada":      h_bajada,
                "h_llegada":     h_llegada,
                "observaciones": "",
            })
    else:
        # MODO SOLO PLATAFORMA: una fila por conductor, ruta vacía para rellenar a mano
        for asig in asignaciones:
            rows_out.append({
                "ruta":          "",
                "conductor":     asig.get("conductor") or "",
                "movil":         asig.get("movil") or "",
                "furgo":         (asig.get("furgo") or "").replace(" ", ""),
                "h_salida":      "",
                "h_bajada":      "",
                "h_llegada":     "",
                "observaciones": "",
            })

    return {"week": week_num, "date": fecha_str, "rows": rows_out}


# ── Paso 2: generar Excel con rutas rojas elegidas por el usuario ──
@app.post("/api/tools/plantilla-excel", dependencies=[Depends(require_admin)])
async def plantilla_excel(body: dict = Body(...), admin=Depends(require_admin)):
    from datetime import date as _date

    rows               = body.get("rows", [])
    red_routes         = {r.upper() for r in body.get("red_routes",         [])}
    pink_furgos        = {f.upper() for f in body.get("pink_furgos",        [])}
    yellow_routes      = {r.upper() for r in body.get("yellow_routes",      [])}
    marked_conductors  = {c.upper() for c in body.get("marked_conductors",  [])}
    week_num           = body.get("week")  or _date.today().isocalendar()[1]
    fecha_str          = body.get("date")  or _date.today().strftime("%d/%m/%Y")
    save_to_hist       = body.get("save", False)
    center             = body.get("center", "")

    # Verificar acceso al centro antes de guardar en historial
    if center and center != "Todos":
        allowed = admin.get("allowed_centers") or []
        if allowed and center not in allowed:
            raise HTTPException(403, "Sin acceso a este centro")

    import io
    xlsx = _build_plantilla_excel(rows, red_routes, pink_furgos, week_num, fecha_str, yellow_routes, marked_conductors)
    # Sanitizar fecha para uso seguro en nombre de archivo y header HTTP
    safe_fecha = re.sub(r'[^a-zA-Z0-9_\-]', '-', fecha_str)
    safe_week  = re.sub(r'[^a-zA-Z0-9_\-]', '-', str(week_num))
    filename = f"plantilla_turno_w{safe_week}_{safe_fecha}.xlsx"

    # Guardar en historial si se pide y hay centro especificado
    if save_to_hist and center and center != "Todos":
        try:
            from uuid import uuid4 as _uuid4
            r2 = get_r2()
            if r2:
                r2_key = f"plantillas/{center}/{filename}"
                r2.put_object(
                    Bucket=R2_BUCKET, Key=r2_key, Body=xlsx,
                    ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                meta = {
                    "id": str(_uuid4()),
                    "center": center,
                    "date": fecha_str,
                    "week": str(week_num),
                    "r2_key": r2_key,
                    "filename": filename,
                    "uploaded_at": datetime.now(timezone.utc).isoformat(),
                }
                await db.plantillas_diarias.insert_one(meta)
        except Exception as _e:
            logger.warning(f"plantilla_excel: no se pudo guardar historial: {_e}")

    return StreamingResponse(
        io.BytesIO(xlsx),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""},
    )


# ── Historial de plantillas ──
@api_router.get("/plantillas")
async def list_plantillas(center: str = None, admin=Depends(require_admin)):
    allowed = admin.get("allowed_centers") or []
    query: dict = {}
    if center and center != "Todos":
        if allowed and center not in allowed:
            raise HTTPException(403, "Sin acceso a este centro")
        query["center"] = center
    elif allowed:
        query["center"] = {"$in": allowed}
    docs = await db.plantillas_diarias.find(query).sort("uploaded_at", -1).to_list(500)
    for d in docs:
        d.pop("_id", None)
    return docs


@api_router.get("/plantillas/{plantilla_id}/download")
async def download_plantilla(plantilla_id: str, _=Depends(require_admin)):
    doc = await db.plantillas_diarias.find_one({"id": plantilla_id})
    if not doc:
        raise HTTPException(404, "Plantilla no encontrada")
    r2 = get_r2()
    if not r2:
        raise HTTPException(503, "R2 no disponible")
    try:
        obj = r2.get_object(Bucket=R2_BUCKET, Key=doc["r2_key"])
        content = obj["Body"].read()
    except Exception as e:
        raise HTTPException(500, f"Error al recuperar el archivo: {e}")
    import io
    return StreamingResponse(
        io.BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"{doc['filename']}\""},
    )


@api_router.delete("/plantillas/{plantilla_id}")
async def delete_plantilla(plantilla_id: str, _=Depends(require_admin)):
    doc = await db.plantillas_diarias.find_one({"id": plantilla_id})
    if not doc:
        raise HTTPException(404, "Plantilla no encontrada")
    try:
        r2 = get_r2()
        if r2:
            r2.delete_object(Bucket=R2_BUCKET, Key=doc["r2_key"])
    except Exception as _e:
        logger.warning(f"delete_plantilla R2 error: {_e}")
    await db.plantillas_diarias.delete_one({"id": plantilla_id})
    return {"success": True}


# =========================
# MONETIZACIÓN — reservas de plaza fundador (página de planes) y ofertas
# patrocinadas del portal conductor. Ambas viven en global_db (negocio de la
# plataforma, no de un tenant) y notifican por Telegram al owner.
# =========================

FOUNDER_TOTAL_SLOTS = 10


class FounderReserve(BaseModel):
    name: str
    email: str
    phone: str = ""
    fleet_size: str = ""


@api_router.get("/founder/slots")
async def founder_slots():
    """Público: plazas fundador restantes. Contador real, no marketing falso."""
    used = await global_db.founder_reservations.count_documents({})
    return {"total": FOUNDER_TOTAL_SLOTS, "left": max(0, FOUNDER_TOTAL_SLOTS - used)}


@api_router.post("/founder/reserve")
async def founder_reserve(data: FounderReserve, request: Request):
    """Público: reserva una plaza fundador sin pago. El owner cierra la venta
    en persona (Telegram avisa al momento con los datos de contacto)."""
    _rl_public_action(f"founder:{_rl_key_ip(request)}", max_count=3, window_s=3600,
                      detail="Demasiadas reservas desde esta conexión")
    name = data.name.strip()[:80]
    email = data.email.strip().lower()[:120]
    if not name or "@" not in email or "." not in email.split("@")[-1]:
        raise HTTPException(400, "Nombre y email válidos son obligatorios")
    if await global_db.founder_reservations.find_one({"email": email}):
        return {"success": True, "already": True}
    used = await global_db.founder_reservations.count_documents({})
    if used >= FOUNDER_TOTAL_SLOTS:
        raise HTTPException(409, "No quedan plazas fundador")
    await global_db.founder_reservations.insert_one({
        "id": str(uuid.uuid4()), "name": name, "email": email,
        "phone": data.phone.strip()[:40], "fleet_size": data.fleet_size.strip()[:40],
        "status": "pending", "created_at": datetime.now(timezone.utc).isoformat(),
    })
    await send_telegram_alert(
        "💰 RESERVA FUNDADOR — llámale hoy",
        f"👤 {name}\n📧 {email}\n📞 {data.phone or '—'}\n🚐 Flota: {data.fleet_size or '—'}\n"
        f"Plazas restantes: {FOUNDER_TOTAL_SLOTS - used - 1}/{FOUNDER_TOTAL_SLOTS}",
        severity="critico",
    )
    return {"success": True, "left": FOUNDER_TOTAL_SLOTS - used - 1}


# Oferta por defecto del portal conductor mientras no haya patrocinadores:
# bucle de crecimiento (el conductor recomienda FlotaDSP a su jefe).
_DEFAULT_DRIVER_OFFER = {
    "id": "ref-flotadsp",
    "emoji": "🎁",
    "title": "¿Conoces otro DSP que sufra con las furgonetas?",
    "description": "Recomiéndale FlotaDSP: inspecciones con IA en 30 segundos, como las tuyas.",
    "cta": "Ver cómo funciona",
    "url": "https://flotadsp.com/?utm_source=driver-portal&utm_medium=referral",
}


class DriverOfferIn(BaseModel):
    title: str
    description: str = ""
    cta: str = ""
    url: str
    emoji: str = "🎁"
    active: bool = True


@api_router.get("/driver-offers")
async def list_driver_offers():
    """Público (portal conductor): ofertas patrocinadas activas.
    Cuenta impresiones para poder vender el espacio con métricas reales."""
    docs = await global_db.driver_offers.find(
        {"active": True}, {"_id": 0}).sort("created_at", -1).to_list(4)
    if docs:
        await global_db.driver_offers.update_many(
            {"id": {"$in": [d["id"] for d in docs]}}, {"$inc": {"views": 1}})
        return {"offers": docs}
    return {"offers": [_DEFAULT_DRIVER_OFFER]}


@api_router.post("/driver-offers/{offer_id}/click")
async def click_driver_offer(offer_id: str, request: Request):
    """Público: registra un clic en una oferta (métrica de venta del espacio)."""
    _rl_public_action(f"offer:{_rl_key_ip(request)}", max_count=30, window_s=600)
    await global_db.driver_offers.update_one(
        {"id": offer_id[:64]},
        {"$inc": {"clicks": 1},
         "$setOnInsert": {"active": False, "title": "(clics de la oferta por defecto)"}},
        upsert=True,
    )
    return {"success": True}


@api_router.get("/admin/driver-offers")
async def admin_list_driver_offers(_=Depends(require_superadmin)):
    """Super-admin: todas las ofertas con sus métricas (views/clicks)."""
    docs = await global_db.driver_offers.find({}, {"_id": 0}).sort("created_at", -1).to_list(50)
    return {"offers": docs}


@api_router.get("/admin/founder-reservations")
async def admin_list_founder_reservations(_=Depends(require_superadmin)):
    """Super-admin: reservas de plaza fundador (para cerrarlas por teléfono)."""
    docs = await global_db.founder_reservations.find({}, {"_id": 0}).sort("created_at", -1).to_list(50)
    return {"reservations": docs, "total_slots": FOUNDER_TOTAL_SLOTS}


@api_router.post("/admin/driver-offers")
async def admin_create_driver_offer(data: DriverOfferIn, _=Depends(require_superadmin)):
    url = data.url.strip()
    if not url.startswith("https://"):
        raise HTTPException(400, "La URL de la oferta debe empezar por https://")
    doc = {
        "id": str(uuid.uuid4()), "title": data.title.strip()[:120],
        "description": data.description.strip()[:240], "cta": data.cta.strip()[:60],
        "url": url[:300], "emoji": data.emoji.strip()[:8] or "🎁",
        "active": data.active, "views": 0, "clicks": 0,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await global_db.driver_offers.insert_one(dict(doc))
    doc.pop("_id", None)
    return {"success": True, "offer": doc}


@api_router.patch("/admin/driver-offers/{offer_id}")
async def admin_toggle_driver_offer(offer_id: str, data: dict, _=Depends(require_superadmin)):
    """Solo permite activar/desactivar (whitelist de 1 campo a propósito)."""
    if "active" not in data:
        raise HTTPException(400, "Falta el campo active")
    r = await global_db.driver_offers.update_one(
        {"id": offer_id}, {"$set": {"active": bool(data["active"])}})
    if not r.matched_count:
        raise HTTPException(404, "Oferta no encontrada")
    return {"success": True}


@api_router.delete("/admin/driver-offers/{offer_id}")
async def admin_delete_driver_offer(offer_id: str, _=Depends(require_superadmin)):
    await global_db.driver_offers.delete_one({"id": offer_id})
    return {"success": True}


# =========================================================================
# PACKAGE INTELLIGENCE CENTER (Cortex) — investigación de incidencias de paquetes
# =========================================================================
# La extensión de navegador intercepta el JSON real de Cortex (route-summaries /
# route-details) usando la sesión ya autenticada del usuario y lo envía aquí
# normalizado. El backend reconstruye el histórico (solo cambios de estado),
# el timeline por paquete, aplica reglas de transición, prioriza y genera la
# ficha del investigador. Datos aislados por org (colecciones en la BD tenant).

_CORTEX_STATES = {
    # crudo (Cortex/es) → canónico
    "loaded": "LOADED", "out_for_delivery": "LOADED", "on_road": "LOADED", "en_route": "LOADED",
    "picked_up": "LOADED", "in_transit": "LOADED",
    # taskState reales de Cortex que significan "aún por hacer / en la furgoneta"
    "not_started": "LOADED", "not_attempted": "LOADED", "assigned": "LOADED",
    "pending": "LOADED", "ready": "LOADED", "incomplete": "LOADED", "in_progress": "LOADED",
    "reattemptable": "ATTEMPTED",
    "arrived": "ARRIVED", "arrived_at_stop": "ARRIVED", "at_stop": "ARRIVED",
    "attempted": "ATTEMPTED", "delivery_attempted": "ATTEMPTED", "attempt": "ATTEMPTED",
    "rejected": "ATTEMPTED", "business_closed": "ATTEMPTED", "unable_to_access": "ATTEMPTED",
    "missing": "MISSING", "not_on_van": "MISSING", "not_found": "MISSING",
    "delivered": "DELIVERED", "completed": "DELIVERED",
    "recovered": "RECOVERED", "rescued": "RECOVERED",
    "returned": "RETURNED", "returned_to_station": "RETURNED", "rts": "RETURNED",
    "lost": "LOST",
    "uncollected": "UNCOLLECTED", "pickup_failed": "UNCOLLECTED",
}
# palabras clave (texto libre en español) → canónico, por si llega evidencia de texto
_CORTEX_STATE_TEXT = [
    ("MISSING", ("falta", "perdido", "missing", "no está en la furgoneta")),
    ("RECOVERED", ("recuperado", "recovered")),
    ("RETURNED", ("devuelto a la estación", "devuelto", "returned")),
    ("LOST", ("lost", "extraviado definit")),
    ("ATTEMPTED", ("se ha intentado", "no se ha podido entregar", "no se puede entregar", "intento de entrega")),
    ("UNCOLLECTED", ("no se ha podido recoger", "recogida pendiente")),
    ("DELIVERED", ("entregado", "buzón", "puerta principal", "recibido", "delivered")),
    ("ARRIVED", ("en la parada", "llegada a la parada", "arrived")),
    ("LOADED", ("cargado", "en ruta", "loaded")),
]
_CORTEX_ORDER = {s: i for i, s in enumerate(
    ["LOADED", "ARRIVED", "ATTEMPTED", "MISSING", "RECOVERED", "DELIVERED", "RETURNED", "LOST",
     "UNCOLLECTED", "OBSERVED"])}


def _cortex_canon_state(raw) -> str:
    if not raw:
        return "OBSERVED"
    s = str(raw).strip().lower().replace(" ", "_").replace("-", "_")
    if s.upper() in _CORTEX_ORDER:
        return s.upper()
    if s in _CORTEX_STATES:
        return _CORTEX_STATES[s]
    low = str(raw).strip().lower()
    for canon, kws in _CORTEX_STATE_TEXT:
        if any(k in low for k in kws):
            return canon
    # No reconocido: si parece un código de estado (letras/underscore), lo
    # mostramos tal cual (verdad de Cortex) en vez de "OBSERVED" genérico.
    if re.fullmatch(r"[A-Za-z_]{2,28}", s):
        return s.upper()
    return "OBSERVED"


def _cortex_parse_dt(v):
    if not v:
        return None
    try:
        s = str(v).replace("Z", "+00:00")
        d = datetime.fromisoformat(s)
        return d if d.tzinfo else d.replace(tzinfo=timezone.utc)
    except Exception:
        try:
            return datetime.fromtimestamp(float(v) / (1000 if float(v) > 1e12 else 1), tz=timezone.utc)
        except Exception:
            return None


def _cortex_ingest_org(request: Request) -> str:
    """Autentica la extensión por su token de ingesta y fija la BD del DSP.
    Devuelve el org_id. El token es un JWT de larga duración con scope propio."""
    token = request.headers.get("x-ingest-token") or request.query_params.get("ingest_token", "")
    if not token:
        raise HTTPException(status_code=401, detail="Falta el token de ingesta (X-Ingest-Token).")
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[JWT_ALGORITHM])
    except Exception:
        raise HTTPException(status_code=401, detail="Token de ingesta inválido o caducado.")
    if payload.get("scope") != "cortex_ingest":
        raise HTTPException(status_code=403, detail="Token sin permiso de ingesta.")
    set_current_org_db(payload.get("db_name"))
    return payload.get("org_id", "")


def _cortex_evaluate(pkg: dict) -> dict:
    """Motor de reglas: prioridad + ficha del investigador a partir del timeline.
    Determinista y sin IA generativa: es un peritaje de secuencia de estados."""
    tl = pkg.get("timeline") or []
    seq = [e.get("state") for e in tl]
    state = pkg.get("state") or (seq[-1] if seq else "OBSERVED")
    now = datetime.now(timezone.utc)

    def _at(i):
        return _cortex_parse_dt(tl[i].get("at")) if 0 <= i < len(tl) else None

    def _mins_since(dt):
        return int((now - dt).total_seconds() // 60) if dt else None

    # ¿Hubo Attempted inmediatamente antes de un Missing?
    attempted_before_missing = False
    idx_missing = next((i for i, s in enumerate(seq) if s == "MISSING"), None)
    if idx_missing is not None:
        prev = [s for s in seq[:idx_missing] if s in ("ATTEMPTED", "ARRIVED", "LOADED")]
        attempted_before_missing = bool(prev) and prev[-1] == "ATTEMPTED"

    missing_at = _at(idx_missing) if idx_missing is not None else None
    mins_missing = _mins_since(missing_at)

    priority, reason = "low", ""
    if state in ("RECOVERED", "DELIVERED"):
        priority = "low"
        reason = "Resuelto" if state == "DELIVERED" else "Recuperado"
    elif state == "LOST":
        priority, reason = "critical", "Paquete dado por perdido (Lost)."
    elif state == "MISSING" and attempted_before_missing:
        priority, reason = "critical", "Transición Attempted → Missing: probablemente estaba en la furgoneta."
    elif state == "MISSING" and mins_missing is not None and mins_missing >= 15:
        priority, reason = "high", f"Missing desde hace {mins_missing} min sin resolver."
    elif state == "MISSING":
        priority, reason = "high", "Missing reciente sin resolver."
    elif state == "ATTEMPTED":
        priority, reason = "medium", "Intento de entrega sin resolver."

    # Ficha del investigador (recomendación)
    investigator = None
    if state in ("MISSING", "LOST"):
        container = pkg.get("container_id")
        rec_type, confidence, text = "review", 0.6, ""
        if attempted_before_missing:
            rec_type, confidence = "vehicle_or_tote", 0.9
            att_at = _at(seq.index("ATTEMPTED")) if "ATTEMPTED" in seq else None
            gap = None
            if att_at and missing_at:
                gap = int((missing_at - att_at).total_seconds() // 60)
            text = (f"El paquete tuvo un Attempted{f' {gap} minutos' if gap else ''} antes del Missing. "
                    "Es muy probable que estuviera físicamente en la furgoneta. "
                    + (f"Revisa el contenedor {container}." if container else "Revisa el tote del stop u organización de la furgoneta."))
        elif seq and seq[0] == "LOADED" and "ATTEMPTED" not in seq and "ARRIVED" not in seq:
            rec_type, confidence = "loaded_never_attempted", 0.75
            text = ("El paquete se cargó pero nunca se intentó entregar ni llegó a la parada. "
                    "Puede seguir en la furgoneta o no haberse escaneado correctamente al cargar. "
                    + (f"Revisa el contenedor {container}." if container else "Revisa la carga y el tote."))
        else:
            text = ("Missing sin un Attempted previo claro. Revisa el último escaneo del paquete y "
                    "el tote del stop antes de darlo por perdido.")
        investigator = {
            "type": rec_type, "confidence": confidence, "text": text,
            "container": container, "mins_since_missing": mins_missing,
        }

    return {"priority": priority, "reason": reason,
            "attempted_before_missing": attempted_before_missing,
            "investigator": investigator}


async def _cortex_apply_observation(obs: dict, captured_at) -> str:
    """Aplica una observación canónica: crea o actualiza el paquete guardando
    solo los cambios de estado en el histórico. Devuelve 'new'|'changed'|'same'."""
    tba = (obs.get("tba") or obs.get("package_id") or "").strip().upper()
    if not tba:
        return "same"
    state = _cortex_canon_state(obs.get("state") or obs.get("raw_state") or obs.get("taskState"))
    observed_at = _cortex_parse_dt(obs.get("observed_at")) or _cortex_parse_dt(captured_at) or datetime.now(timezone.utc)
    ev = {
        "state": state, "at": observed_at.isoformat(),
        "raw": (obs.get("raw_state") or obs.get("taskState") or "")[:80],
        "stop_id": obs.get("stop_id"), "container_id": obs.get("container_id"),
    }
    common = {
        "tba": tba, "reference_id": obs.get("reference_id") or tba,
        "route_code": obs.get("route_code") or obs.get("route_id"),
        "route_id": obs.get("route_id"),
        "driver_name": obs.get("driver_name"), "driver_id": obs.get("driver_id"),
        "stop_id": obs.get("stop_id"), "stop_address": obs.get("stop_address") or obs.get("address"),
        "container_id": obs.get("container_id"), "station": obs.get("station"),
        "lat": obs.get("lat"), "lng": obs.get("lng"),
        "state": state, "updated_at": observed_at.isoformat(),
    }
    # Día de servicio (el que el usuario tiene seleccionado en Cortex). Se guarda
    # una sola vez y no se pisa con null: cada paquete pertenece a un día.
    service_day = str(obs.get("service_day") or "").strip()[:10]
    if service_day:
        common["service_day"] = service_day
    pkg = await db.cortex_packages.find_one({"tba": tba}, {"_id": 0})
    if not pkg:
        # Sembrar timeline con recentTaskEvents si vienen, más el estado actual
        seed = []
        for e in (obs.get("events") or []):
            st = _cortex_canon_state(e.get("state") or e.get("raw"))
            at = _cortex_parse_dt(e.get("at") or e.get("time"))
            if at:
                seed.append({"state": st, "at": at.isoformat(), "raw": (e.get("raw") or "")[:80]})
        seed.sort(key=lambda e: e["at"])
        if not seed or seed[-1]["state"] != state:
            seed.append(ev)
        doc = {**common, "id": str(uuid.uuid4()), "first_seen": observed_at.isoformat(),
               "timeline": seed}
        evalr = _cortex_evaluate(doc)
        doc.update({"priority": evalr["priority"], "reason": evalr["reason"]})
        await db.cortex_packages.insert_one(serialize_doc(doc))
        await db.cortex_events.insert_one(serialize_doc({"id": str(uuid.uuid4()), **ev, "tba": tba}))
        return "new"

    last = (pkg.get("timeline") or [{}])[-1]
    if last.get("state") == state:
        await db.cortex_packages.update_one({"tba": tba}, {"$set": common})
        return "same"
    # Cambio de estado real → histórico + timeline
    doc = {**pkg, **common, "timeline": (pkg.get("timeline") or []) + [ev]}
    evalr = _cortex_evaluate(doc)
    await db.cortex_packages.update_one(
        {"tba": tba},
        {"$set": {**common, "priority": evalr["priority"], "reason": evalr["reason"]},
         "$push": {"timeline": ev}})
    await db.cortex_events.insert_one(serialize_doc({"id": str(uuid.uuid4()), **ev, "tba": tba}))
    return "changed"


@api_router.get("/cortex/ingest-token")
async def cortex_ingest_token(user: dict = Depends(require_admin)):
    """Genera el token que la extensión usa para enviar datos (1 año)."""
    payload = {
        "scope": "cortex_ingest", "org_id": user.get("org_id", ""),
        "db_name": user.get("db_name"), "name": user.get("name", ""),
        "exp": datetime.now(timezone.utc) + timedelta(days=365),
    }
    return {"token": jwt.encode(payload, SECRET_KEY, algorithm=JWT_ALGORITHM),
            "ingest_url": f"{PUBLIC_BASE_URL}/api/cortex/ingest"}


@api_router.post("/cortex/ingest")
async def cortex_ingest(request: Request):
    """Recibe observaciones canónicas de la extensión (JSON real de Cortex ya
    normalizado). No requiere JWT de usuario: se autentica por token de ingesta."""
    _cortex_ingest_org(request)
    body = await request.json()
    packages = body.get("packages") or []
    if not isinstance(packages, list):
        raise HTTPException(status_code=400, detail="Formato inválido: se espera 'packages': [...]")
    captured_at = body.get("captured_at")
    stats = {"new": 0, "changed": 0, "same": 0}
    for obs in packages[:2000]:
        try:
            stats[await _cortex_apply_observation(obs, captured_at)] += 1
        except Exception as e:
            logger.warning(f"Cortex ingest obs: {e}")
    logger.info(f"Cortex ingest: {stats['new']} nuevos, {stats['changed']} cambios de {len(packages)} obs")
    return {"ok": True, **stats, "received": len(packages)}


def _cortex_day_query(day: str) -> dict:
    """Filtro Mongo por día de servicio. Los paquetes nuevos traen service_day;
    los antiguos (sin él) se ubican por la fecha de updated_at como fallback."""
    if not day:
        return {}
    return {"$or": [
        {"service_day": day},
        {"$and": [{"service_day": {"$in": [None, ""]}},
                  {"updated_at": {"$regex": f"^{re.escape(day)}"}}]},
    ]}


@api_router.get("/cortex/days")
async def cortex_days(_=Depends(require_admin)):
    """Días con datos, para el selector del panel (más reciente primero)."""
    pkgs = await db.cortex_packages.find({}, {"_id": 0, "service_day": 1, "updated_at": 1}).to_list(20000)
    counts = {}
    for p in pkgs:
        d = (p.get("service_day") or str(p.get("updated_at") or "")[:10]) or None
        if d:
            counts[d] = counts.get(d, 0) + 1
    days = sorted(({"day": k, "n": v} for k, v in counts.items()), key=lambda x: x["day"], reverse=True)
    return {"days": days, "today": datetime.now(timezone.utc).strftime("%Y-%m-%d")}


@api_router.get("/cortex/overview")
async def cortex_overview(day: str = "", _=Depends(require_admin)):
    today = (day or datetime.now(timezone.utc).strftime("%Y-%m-%d"))
    pkgs = await db.cortex_packages.find(_cortex_day_query(today), {"_id": 0}).to_list(5000)
    n = len(pkgs)
    missing = [p for p in pkgs if p.get("state") == "MISSING"]
    recovered_today, lost, missing_today, rec_times, attempts_pre = [], [], [], [], []
    by_driver, by_route = {}, {}
    for p in pkgs:
        tl = p.get("timeline") or []
        states = [e.get("state") for e in tl]
        if p.get("state") == "LOST":
            lost.append(p)
        # Missing hoy
        m_ev = next((e for e in tl if e.get("state") == "MISSING"), None)
        if m_ev and str(m_ev.get("at", ""))[:10] == today:
            missing_today.append(p)
            if p.get("driver_name"):
                by_driver[p["driver_name"]] = by_driver.get(p["driver_name"], 0) + 1
            if p.get("route_code"):
                by_route[p["route_code"]] = by_route.get(p["route_code"], 0) + 1
            attempts_pre.append(sum(1 for s in states[:states.index("MISSING")] if s == "ATTEMPTED") if "MISSING" in states else 0)
        # Recuperación hoy + tiempo medio
        if "RECOVERED" in states or (p.get("state") == "DELIVERED" and "MISSING" in states):
            i_m = states.index("MISSING") if "MISSING" in states else None
            i_r = next((i for i in range(len(states)) if states[i] in ("RECOVERED", "DELIVERED") and (i_m is None or i > i_m)), None)
            if i_m is not None and i_r is not None:
                t0, t1 = _cortex_parse_dt(tl[i_m].get("at")), _cortex_parse_dt(tl[i_r].get("at"))
                if t0 and t1 and str(tl[i_r].get("at", ""))[:10] == today:
                    recovered_today.append(p)
                    rec_times.append((t1 - t0).total_seconds() / 60)
    resolved = len(recovered_today) + len(lost)
    recovery_pct = round(100 * len(recovered_today) / resolved) if resolved else None
    health = 100 - min(60, len(missing) * 6) - min(20, len(lost) * 10)
    return {
        "tracked": n, "missing_now": len(missing),
        "missing_today": len(missing_today), "recovered_today": len(recovered_today),
        "lost": len(lost), "recovery_pct": recovery_pct,
        "avg_recovery_min": round(sum(rec_times) / len(rec_times)) if rec_times else None,
        "avg_attempts_before_missing": round(sum(attempts_pre) / len(attempts_pre), 1) if attempts_pre else None,
        "health": max(0, health),
        "by_driver": sorted([{"name": k, "n": v} for k, v in by_driver.items()], key=lambda x: -x["n"])[:8],
        "by_route": sorted([{"route": k, "n": v} for k, v in by_route.items()], key=lambda x: -x["n"])[:8],
    }


@api_router.get("/cortex/packages")
async def cortex_packages(q: str = "", state: str = "", priority: str = "", day: str = "", limit: int = 200,
                          _=Depends(require_admin)):
    ands = []
    dq = _cortex_day_query(day)
    if dq:
        ands.append(dq)
    if q:
        rx = {"$regex": re.escape(q), "$options": "i"}
        ands.append({"$or": [{"tba": rx}, {"reference_id": rx}, {"route_code": rx},
                             {"driver_name": rx}, {"stop_address": rx}, {"stop_id": rx}]})
    query = {}
    if state:
        query["state"] = state.upper()
    if priority:
        query["priority"] = priority.lower()
    if ands:
        query["$and"] = ands
    pkgs = await db.cortex_packages.find(query, {"_id": 0, "timeline": 0}).sort("updated_at", -1).to_list(limit)
    return {"packages": pkgs}


@api_router.get("/cortex/package/{tba}")
async def cortex_package_detail(tba: str, _=Depends(require_admin)):
    pkg = await db.cortex_packages.find_one({"tba": tba.upper()}, {"_id": 0})
    if not pkg:
        raise HTTPException(status_code=404, detail="Paquete no encontrado")
    pkg["timeline"] = sorted(pkg.get("timeline") or [], key=lambda e: str(e.get("at", "")))
    evalr = _cortex_evaluate(pkg)
    # Paquetes del mismo stop (para la pista "otros del stop se entregaron")
    same_stop = []
    if pkg.get("stop_id") and pkg.get("route_code"):
        same_stop = await db.cortex_packages.find(
            {"stop_id": pkg["stop_id"], "route_code": pkg["route_code"], "tba": {"$ne": pkg["tba"]}},
            {"_id": 0, "tba": 1, "state": 1}).to_list(20)
    return {"package": pkg, "evaluation": evalr, "same_stop": same_stop}


@api_router.get("/cortex/alerts")
async def cortex_alerts(day: str = "", _=Depends(require_admin)):
    q = {"state": {"$in": ["MISSING", "LOST"]}}
    dq = _cortex_day_query(day)
    if dq:
        q = {"$and": [q, dq]}
    pkgs = await db.cortex_packages.find(q, {"_id": 0}).sort("updated_at", -1).to_list(500)
    alerts = []
    rank = {"critical": 0, "high": 1, "medium": 2, "low": 3}
    for p in pkgs:
        ev = _cortex_evaluate(p)
        if ev["priority"] in ("critical", "high"):
            alerts.append({
                "tba": p["tba"], "priority": ev["priority"], "reason": ev["reason"],
                "route_code": p.get("route_code"), "driver_name": p.get("driver_name"),
                "stop_id": p.get("stop_id"), "state": p.get("state"),
                "confidence": (ev.get("investigator") or {}).get("confidence"),
                "recommendation": (ev.get("investigator") or {}).get("text"),
            })
    alerts.sort(key=lambda a: rank.get(a["priority"], 9))
    return {"alerts": alerts}


@api_router.get("/cortex/heatmap")
async def cortex_heatmap(day: str = "", _=Depends(require_admin)):
    q = {"lat": {"$ne": None}, "state": {"$in": ["MISSING", "LOST", "ATTEMPTED"]}}
    dq = _cortex_day_query(day)
    if dq:
        q = {"$and": [q, dq]}
    pkgs = await db.cortex_packages.find(
        q, {"_id": 0, "lat": 1, "lng": 1, "state": 1, "tba": 1}).to_list(2000)
    return {"points": [p for p in pkgs if p.get("lat") and p.get("lng")]}


@api_router.post("/cortex/seed-demo")
async def cortex_seed_demo(_=Depends(require_admin)):
    """Siembra paquetes de demostración para ver el módulo funcionando."""
    base = datetime.now(timezone.utc).replace(microsecond=0)

    def iso(mins):
        return (base - timedelta(minutes=mins)).isoformat()
    demos = [
        {"tba": "TBADEMO001", "reference_id": "REF-948201", "route_code": "XA_C1", "driver_name": "A. García",
         "stop_id": "42", "stop_address": "Rúa do Vilar 12, Santiago", "container_id": "B14",
         "lat": 42.881, "lng": -8.545, "state": "MISSING",
         "timeline": [{"state": "LOADED", "at": iso(390)}, {"state": "ARRIVED", "at": iso(40)},
                      {"state": "ATTEMPTED", "at": iso(30)}, {"state": "MISSING", "at": iso(13)}]},
        {"tba": "TBADEMO014", "reference_id": "REF-948214", "route_code": "XA_C1", "driver_name": "A. García",
         "stop_id": "43", "stop_address": "Avda de Lugo 5, Santiago", "container_id": "B14",
         "lat": 42.878, "lng": -8.537, "state": "RECOVERED",
         "timeline": [{"state": "LOADED", "at": iso(400)}, {"state": "ATTEMPTED", "at": iso(70)},
                      {"state": "MISSING", "at": iso(55)}, {"state": "RECOVERED", "at": iso(18)}]},
        {"tba": "TBADEMO032", "reference_id": "REF-948232", "route_code": "XA_C4", "driver_name": "M. López",
         "stop_id": "12", "stop_address": "Calle Real 88, A Coruña", "container_id": "A03",
         "lat": 43.370, "lng": -8.396, "state": "DELIVERED",
         "timeline": [{"state": "LOADED", "at": iso(410)}, {"state": "ARRIVED", "at": iso(60)},
                      {"state": "DELIVERED", "at": iso(50)}]},
        {"tba": "TBADEMO007", "reference_id": "REF-948207", "route_code": "XA_C7", "driver_name": "G. Portomeñe",
         "stop_id": "77", "stop_address": "Praza Maior 3, Lugo", "container_id": "C21",
         "lat": 43.012, "lng": -7.556, "state": "MISSING",
         "timeline": [{"state": "LOADED", "at": iso(420)}, {"state": "ATTEMPTED", "at": iso(95)},
                      {"state": "MISSING", "at": iso(80)}]},
    ]
    for d in demos:
        d["id"] = str(uuid.uuid4())
        d["first_seen"] = d["timeline"][0]["at"]
        d["updated_at"] = d["timeline"][-1]["at"]
        ev = _cortex_evaluate(d)
        d["priority"], d["reason"] = ev["priority"], ev["reason"]
        await db.cortex_packages.update_one({"tba": d["tba"]}, {"$set": serialize_doc(d)}, upsert=True)
    return {"ok": True, "seeded": len(demos)}


@api_router.post("/cortex/clear-demo")
async def cortex_clear_demo(_=Depends(require_admin)):
    """Borra los paquetes de demostración (TBADEMO*) y su historial de eventos,
    para que no contaminen KPIs ni alertas cuando ya llegan datos reales."""
    q = {"tba": {"$regex": "^TBADEMO"}}
    p = await db.cortex_packages.delete_many(q)
    e = await db.cortex_events.delete_many(q)
    return {"ok": True, "packages_deleted": p.deleted_count, "events_deleted": e.deleted_count}


@api_router.post("/cortex/reset")
async def cortex_reset(_=Depends(require_admin)):
    """Borra TODOS los paquetes y eventos de Cortex (empezar de cero). Útil tras
    pruebas que ensuciaron los días. La extensión los repuebla al capturar."""
    p = await db.cortex_packages.delete_many({})
    e = await db.cortex_events.delete_many({})
    return {"ok": True, "packages_deleted": p.deleted_count, "events_deleted": e.deleted_count}


app.include_router(auth_router)
app.include_router(api_router)
